import json
import os

import openpyxl

wb = openpyxl.load_workbook(r"D:\Documents\NAUMUR\SF modifié.xlsx", data_only=False)
ws = wb.active

# Get headers
headers = [cell.value for cell in ws[1]]

data = []
# Iterate through rows starting from 2
for row in ws.iter_rows(min_row=2):
    row_data = {}
    for cell, header in zip(row, headers):
        if not header:
            continue

        # Check for hyperlink
        val = cell.value
        if cell.hyperlink:
            val = cell.hyperlink.target

        row_data[header] = val

    # Only add if row has some data
    if any(row_data.values()):
        data.append(row_data)

# Save to JSON
output_path = os.path.join(os.getcwd(), "scratch", "sf_modifie.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump({"columns": headers, "data": data}, f, ensure_ascii=False, indent=2)

print(f"Extracted {len(data)} rows with hyperlinks.")
