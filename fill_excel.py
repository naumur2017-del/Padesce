"""
fill_excel.py — Remplit automatiquement le fichier DESCENTES CLASSES Vxx.xlsx
en interrogeant l'API du site call.naumur.com.

Feuille SA (CTR PUB MAN) :
  - Colonne C  : "Lien des enquêtes"  (hyperlien vers la page classe du site)
  - Colonne G  : code de la classe (ex. CLA001)
  - Colonne H  : "SA"
  - Colonne J  : "Achevé"
  - Colonne P  : "Liens CSV"          (hyperlien vers le CSV de la classe)

Feuille SF (créée si absente) :
  - Même structure mais pour les prestations formateurs (SF)

Usage :
  python fill_excel.py \
      --excel "C:/Users/Jack Brayan/Downloads/DESCENTES CLASSES  V17.xlsx" \
      --api-key VOTRE_CLE_EXPORT_API_KEY \
      --site-url https://call.naumur.com

Prérequis :  pip install openpyxl requests
"""

import argparse
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl manquant : pip install openpyxl")

try:
    import requests
except ImportError:
    sys.exit("requests manquant : pip install requests")

# ---------------------------------------------------------------------------
# Couleurs de style pour les nouvelles lignes
# ---------------------------------------------------------------------------
FILL_SA = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_SF = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
FONT_LINK = Font(color="0563C1", underline="single")
FONT_BOLD = Font(bold=True)

# En-têtes de la feuille SF (formateurs par prestation)
SF_HEADERS = [
    "N°", "Code Prestataire", "Lien des enquêtes", "Prestataire",
    "Bénéficiaire", "Formation", "Cohorte", None, "SF", None, "Achevé",
    None, None, None, None, "Liens CSV",
]

# Colonnes cibles dans CTR PUB MAN (index 1-based)
COL_NUM   = 1   # A
COL_LIEN  = 3   # C  ← "Lien des enquêtes"
COL_CLASSE= 7   # G  ← code classe
COL_TYPE  = 8   # H  ← "SA"
COL_STATUT= 10  # J  ← "Achevé"
COL_CSV   = 16  # P  ← "Liens CSV"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fetch_json(url: str, api_key: str) -> dict:
    resp = requests.get(url, headers={"X-Export-Api-Key": api_key}, timeout=30)
    if resp.status_code == 403:
        sys.exit(f"[ERREUR 403] Clé API refusée par le site : {url}")
    resp.raise_for_status()
    return resp.json()


def write_hyperlink(ws, row: int, col: int, url: str, label: str, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.value = f'=HYPERLINK("{url}", "{label}")'
    cell.font = FONT_LINK
    if fill:
        cell.fill = fill


def find_last_data_row(ws, key_col: int = COL_CLASSE) -> int:
    """Renvoie le dernier numéro de ligne non vide dans la colonne clé."""
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=key_col).value:
            return row
    return 1


def collect_existing_codes(ws, col: int) -> set:
    """Collecte toutes les valeurs non-None de la colonne `col`."""
    codes = set()
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if v:
            codes.add(str(v).strip().upper())
    return codes


# ---------------------------------------------------------------------------
# Remplissage feuille SA (CTR PUB MAN)
# ---------------------------------------------------------------------------

def fill_sa(ws, classes: list, fill_color=FILL_SA):
    existing = collect_existing_codes(ws, COL_CLASSE)
    last_row = find_last_data_row(ws)

    new_classes = [c for c in classes if c["code"].upper() not in existing]
    if not new_classes:
        print("  [SA] Aucune nouvelle classe à ajouter.")
        return 0

    for cls in new_classes:
        last_row += 1
        code = cls["code"]
        enquete_url = cls["enquete_url"]
        csv_url = cls["csv_url"]

        # N° (numéro auto)
        prev = ws.cell(row=last_row - 1, column=COL_NUM).value
        if prev is not None:
            ws.cell(row=last_row, column=COL_NUM).value = f"=1+A{last_row - 1}"

        # C — lien enquête
        write_hyperlink(ws, last_row, COL_LIEN, enquete_url, "Lien des enquêtes", fill_color)

        # G — code classe
        c_classe = ws.cell(row=last_row, column=COL_CLASSE)
        c_classe.value = code
        if fill_color:
            c_classe.fill = fill_color

        # H — type SA
        c_type = ws.cell(row=last_row, column=COL_TYPE)
        c_type.value = "SA"
        if fill_color:
            c_type.fill = fill_color

        # J — statut Achevé
        c_statut = ws.cell(row=last_row, column=COL_STATUT)
        c_statut.value = "Achevé"
        if fill_color:
            c_statut.fill = fill_color

        # P — lien CSV
        write_hyperlink(ws, last_row, COL_CSV, csv_url, "Liens CSV", fill_color)

        print(f"  [SA] + ligne {last_row} : {code}")

    return len(new_classes)


# ---------------------------------------------------------------------------
# Remplissage / création feuille SF
# ---------------------------------------------------------------------------

def ensure_sf_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    if "SF" in wb.sheetnames:
        return wb["SF"]
    ws = wb.create_sheet("SF")
    # En-têtes
    for col_idx, h in enumerate(SF_HEADERS, start=1):
        if h:
            cell = ws.cell(row=1, column=col_idx)
            cell.value = h
            cell.font = FONT_BOLD
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    # Largeurs approximatives
    col_widths = {1: 5, 2: 14, 3: 20, 4: 30, 5: 30, 6: 35, 7: 20, 9: 6, 11: 12, 16: 14}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return ws


def fill_sf(ws, prestations: list, fill_color=FILL_SF):
    # Col B = code prestataire (col index 2)
    COL_CODE_SF = 2
    existing = collect_existing_codes(ws, COL_CODE_SF)
    last_row = find_last_data_row(ws, key_col=COL_CODE_SF)

    new_prestas = [p for p in prestations if p["code"].upper() not in existing]
    if not new_prestas:
        print("  [SF] Aucune nouvelle prestation à ajouter.")
        return 0

    for presta in new_prestas:
        last_row += 1
        code = presta["code"]
        enquete_url = presta["enquete_url"]
        csv_url = presta["csv_url"]

        # A — N°
        prev = ws.cell(row=last_row - 1, column=1).value
        if prev is not None:
            ws.cell(row=last_row, column=1).value = f"=1+A{last_row - 1}"
        else:
            ws.cell(row=last_row, column=1).value = last_row - 1

        # B — code prestataire
        c = ws.cell(row=last_row, column=COL_CODE_SF)
        c.value = code
        c.fill = fill_color

        # C — lien enquête
        write_hyperlink(ws, last_row, 3, enquete_url, "Lien des enquêtes", fill_color)

        # D — prestataire
        c = ws.cell(row=last_row, column=4)
        c.value = presta.get("prestataire", code)
        c.fill = fill_color

        # E — bénéficiaire
        c = ws.cell(row=last_row, column=5)
        c.value = presta.get("beneficiaire", "")
        c.fill = fill_color

        # F — formation
        c = ws.cell(row=last_row, column=6)
        c.value = presta.get("formation", "")
        c.fill = fill_color

        # G — cohorte
        c = ws.cell(row=last_row, column=7)
        c.value = presta.get("cohorte", "")
        c.fill = fill_color

        # I — SF
        c = ws.cell(row=last_row, column=9)
        c.value = "SF"
        c.fill = fill_color

        # K — Achevé
        c = ws.cell(row=last_row, column=11)
        c.value = "Achevé"
        c.fill = fill_color

        # P — lien CSV
        write_hyperlink(ws, last_row, 16, csv_url, "Liens CSV", fill_color)

        print(f"  [SF] + ligne {last_row} : {code}")

    return len(new_prestas)


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Remplit le fichier Excel DESCENTES CLASSES")
    parser.add_argument(
        "--excel",
        default=r"C:\Users\Jack Brayan\Downloads\DESCENTES CLASSES  V17.xlsx",
        help="Chemin vers le fichier Excel",
    )
    parser.add_argument(
        "--api-key",
        required=True,
        help="Valeur de EXPORT_API_KEY configurée dans le .env du site",
    )
    parser.add_argument(
        "--site-url",
        default="https://call.naumur.com",
        help="URL de base du site (sans slash final)",
    )
    parser.add_argument(
        "--sa-sheet",
        default="CTR PUB MAN",
        help="Nom de la feuille SA dans le fichier Excel",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Afficher les changements sans sauvegarder",
    )
    args = parser.parse_args()

    site = args.site_url.rstrip("/")
    api_key = args.api_key
    excel_path = Path(args.excel)

    if not excel_path.exists():
        sys.exit(f"Fichier introuvable : {excel_path}")

    print(f"\n=== Récupération des données depuis {site} ===")
    classes_data = fetch_json(f"{site}/satisfaction-apprenants/api/classes-excel/", api_key)
    prestations_data = fetch_json(f"{site}/satisfaction-formateurs/api/prestations-excel/", api_key)

    classes = classes_data.get("classes", [])
    prestations = prestations_data.get("prestations", [])
    print(f"  Classes reçues    : {len(classes)}")
    print(f"  Prestations reçues: {len(prestations)}")

    print(f"\n=== Ouverture de {excel_path.name} ===")
    wb = openpyxl.load_workbook(str(excel_path))

    if args.sa_sheet not in wb.sheetnames:
        sys.exit(f"Feuille '{args.sa_sheet}' introuvable. Feuilles disponibles : {wb.sheetnames}")

    ws_sa = wb[args.sa_sheet]
    ws_sf = ensure_sf_sheet(wb)

    print(f"\n=== Remplissage feuille SA ({args.sa_sheet}) ===")
    n_sa = fill_sa(ws_sa, classes)

    print(f"\n=== Remplissage feuille SF ===")
    n_sf = fill_sf(ws_sf, prestations)

    if args.dry_run:
        print(f"\n[DRY RUN] Aucune sauvegarde. SA={n_sa} lignes, SF={n_sf} lignes.")
        return

    out_path = excel_path.with_stem(excel_path.stem + "_filled")
    wb.save(str(out_path))
    print(f"\n✓ Fichier sauvegardé : {out_path}")
    print(f"  SA : {n_sa} nouvelles lignes ajoutées")
    print(f"  SF : {n_sf} nouvelles lignes ajoutées")


if __name__ == "__main__":
    main()
