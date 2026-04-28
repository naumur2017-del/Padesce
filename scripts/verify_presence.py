#!/usr/bin/env python3
"""Verify presence controls in the provided Excel workbook.

Usage:
  python scripts/verify_presence.py "path/to/Fichier pour plateforme de satisfaction (1).xlsx" --apprenant APP847

Outputs a JSON and CSV report in the current folder.
"""
from pathlib import Path
import sys
import re
import json
import csv
from datetime import datetime
import argparse

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import sqlite3
import shutil


def get_db_conn(db_path):
    conn = sqlite3.connect(db_path)
    return conn


def ensure_db_has_c_columns(conn, table_name='apprenants_apprenant'):
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info('{table_name}')")
    cols = [r[1] for r in cur.fetchall()]
    added = []
    for c in ('c1', 'c2', 'c3', 'c4'):
        if c not in cols:
            cur.execute(f"ALTER TABLE {table_name} ADD COLUMN {c} TEXT DEFAULT ''")
            added.append(c)
    if added:
        conn.commit()
    return added


def find_classe_id(conn, classe_code):
    cur = conn.cursor()
    cur.execute("SELECT id FROM formations_classe WHERE code = ?", (classe_code,))
    r = cur.fetchone()
    return r[0] if r else None


def find_prestation_code_for_classe(conn, classe_id):
    cur = conn.cursor()
    cur.execute("SELECT prestation_id FROM formations_classe WHERE id = ?", (classe_id,))
    r = cur.fetchone()
    if not r:
        return None
    prestation_id = r[0]
    cur.execute("SELECT code FROM formations_prestation WHERE id = ?", (prestation_id,))
    rr = cur.fetchone()
    return rr[0] if rr else None


def find_apprenant_row(conn, code):
    cur = conn.cursor()
    cur.execute("SELECT id FROM apprenants_apprenant WHERE code = ?", (code,))
    r = cur.fetchone()
    return r[0] if r else None


def sync_apprenants_to_db(conn, report, input_path, dry_run=True):
    cur = conn.cursor()
    table = 'apprenants_apprenant'
    ensure_db_has_c_columns(conn, table)
    updates = []
    for a in report.get('apprenants', []):
        code = a.get('apprenant_id')
        if not code:
            continue
        db_id = find_apprenant_row(conn, code)
        if not db_id:
            updates.append({'code': code, 'status': 'missing_in_db'})
            continue
        # prepare set clause for c1..c4 and maybe taux
        set_pairs = []
        params = []
        for c in ('C1','C2','C3','C4'):
            # map header names to lowercase column names c1..c4
            val = a.get('C_suggested', {}).get(c)
            if val is not None:
                col = c.lower()
                set_pairs.append(f"{col} = ?")
                params.append(val)
        if not set_pairs:
            updates.append({'code': code, 'status': 'no_changes'})
            continue
        params.append(db_id)
        sql = f"UPDATE {table} SET {', '.join(set_pairs)} WHERE id = ?"
        if dry_run:
            updates.append({'code': code, 'status': 'would_update', 'sql': sql, 'params': params})
        else:
            cur.execute(sql, params)
            updates.append({'code': code, 'status': 'updated'})
    if not dry_run:
        conn.commit()
    return updates


def read_decompte_global(wb):
    # find sheet name that contains 'decompte'
    dg_name = None
    for s in wb.sheetnames:
        if 'decompte' in s.lower():
            dg_name = s
            break
    if dg_name is None:
        return {}, None
    ws = wb[dg_name]
    col_B_idx = column_index_from_string('B')
    col_AN_idx = column_index_from_string('AN')
    mapping = {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        presta = row[col_B_idx - 1] if len(row) >= col_B_idx else None
        taux = row[col_AN_idx - 1] if len(row) >= col_AN_idx else None
        if presta not in (None, ''):
            mapping[str(presta).strip()] = taux
    return mapping, dg_name


def find_taux_column(cols):
    # return column name if header contains 'taux' or 'presence'
    for c in cols:
        if c is None:
            continue
        lab = str(c).lower()
        if 'taux' in lab or 'presence' in lab:
            return c
    return None


def is_blank_value(v):
    # treat None, empty-string, pandas NaN, 'nan'/'NaN' as blank
    import pandas as _pd
    if v is None:
        return True
    if _pd.isna(v):
        return True
    s = str(v).strip()
    if s == '':
        return True
    if s.lower() == 'nan':
        return True
    return False


def normalize_cell(v):
    if is_blank_value(v):
        return ''
    return str(v).strip()


def analyze_workbook(path, apprenant_id=None):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(path)

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    xls = pd.ExcelFile(p)

    decompte_map, dg_name = read_decompte_global(wb)

    apprenants_report = []
    presta_reports = []

    for sheet in xls.sheet_names:
        try:
            df = pd.read_excel(p, sheet_name=sheet, dtype=str)
        except Exception:
            continue
        if df.empty:
            continue

        # columns matching C1..C4
        c_cols = [c for c in df.columns if isinstance(c, str) and re.match(r'^\s*C[1-4]\s*$', c.strip(), re.I)]

        # detect apprenant codes in the sheet (APP followed by digits)
        df_str = df.fillna('').astype(str)
        app_pattern = re.compile(r'(APP\d+)', re.I)
        found_apps = {}
        for ridx, row in df_str.iterrows():
            combined = ' '.join(row.values)
            m = app_pattern.findall(combined)
            if m:
                # prefer first occurrence
                code = m[0].upper()
                found_apps[ridx] = code

        for idx, code in found_apps.items():
            row = df.loc[idx]
            row_dict = {str(k): normalize_cell(row[k]) for k in df.columns}
            c_values = {c: normalize_cell(row.get(c, '')) for c in c_cols} if c_cols else {}
            # determine recommended display for C1..C4
            all_blank = all(is_blank_value(row.get(c, None)) for c in c_cols) if c_cols else True
            has_absent = any((not is_blank_value(row.get(c, None)) and str(row.get(c)).strip().lower() == 'absent') for c in c_cols)
            suggested = {}
            for c in c_cols:
                raw = row.get(c, None)
                if is_blank_value(raw):
                    suggested[c] = '-'
                elif str(raw).strip().lower() == 'absent':
                    suggested[c] = 'Absent'
                else:
                    suggested[c] = str(raw).strip()

            # find taux in this sheet if any
            taux_col = find_taux_column(df.columns)
            taux_val = normalize_cell(row[taux_col]) if taux_col and taux_col in df.columns else ''
            if all_blank:
                taux_suggested = '-'
            else:
                # keep existing taux unless blank
                taux_suggested = taux_val if taux_val != '' else '-'

            apprenants_report.append({
                'sheet': sheet,
                'excel_row_index': int(idx) + 2,
                'apprenant_id': code,
                'C_columns': c_values,
                'C_suggested': suggested,
                'taux_column_found': taux_col,
                'taux_value': taux_val,
                'taux_suggested': taux_suggested,
            })

        # For each prestation id found in this sheet, compare with Decompte Global
        # detect prestation ids by looking for pattern PRESTA followed by digits
        df_str = df.fillna('').astype(str)
        for col in df.columns:
            # search each cell for PRESTA pattern
            series = df_str[col]
            for i, cell in series.items():
                m = re.search(r'(PRESTA\d+)', cell, re.I)
                if m:
                    presta_id = m.group(1).upper()
                    # locate taux in this sheet (header)
                    taux_col = find_taux_column(df.columns)
                    taux_val = normalize_cell(df.loc[i, taux_col]) if taux_col in df.columns else ''
                    decompte_val = decompte_map.get(presta_id)
                    if decompte_val is None:
                        status = 'missing_in_decompte'
                    else:
                        # compare as strings after normalization
                        dv = normalize_cell(decompte_val)
                        if dv == '':
                            status = 'decompte_taux_empty'
                        elif taux_val == '':
                            status = 'site_taux_empty'
                        elif dv != taux_val:
                            status = 'mismatch'
                        else:
                            status = 'ok'
                    presta_reports.append({
                        'sheet': sheet,
                        'excel_row_index': int(i) + 2,
                        'presta_id': presta_id,
                        'site_taux_col': taux_col,
                        'site_taux_val': taux_val,
                        'decompte_taux_val': decompte_val,
                        'status': status,
                    })

    # dedupe presta_reports by (presta_id, sheet, row)
    return {
        'apprenants': apprenants_report,
        'prestas': presta_reports,
        'decompte_sheet': dg_name,
        'decompte_map_sample': dict(list(decompte_map.items())[:30])
    }


def apply_fixes(report, input_path, output_path):
    """Write suggested fixes into a new workbook copy.

    Only handles suggested fixes for C columns and taux_suggested where needed.
    """
    wb = openpyxl.load_workbook(input_path)
    for a in report.get('apprenants', []):
        sheet = a['sheet']
        row_idx = a['excel_row_index']
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # update C columns by header name -> find column letters
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # map header name to column index
        hdr_map = {str(v).strip(): i + 1 for i, v in enumerate(header) if v is not None}
        for col_name, new_val in a.get('C_suggested', {}).items():
            if col_name in hdr_map:
                col_idx = hdr_map[col_name]
                cell = ws.cell(row=row_idx, column=col_idx)
                old = cell.value
                if str(old).strip() != new_val:
                    cell.value = new_val
        # update taux column if suggested '-' and taux column exists
        taux_col = a.get('taux_column_found')
        if taux_col and taux_col in hdr_map:
            col_idx = hdr_map[taux_col]
            cell = ws.cell(row=row_idx, column=col_idx)
            old = cell.value
            new = a.get('taux_suggested')
            if new is not None and str(old).strip() != str(new):
                cell.value = new

    wb.save(output_path)
    return output_path


def save_reports(report, out_prefix=None):
    if out_prefix is None:
        t = datetime.now().strftime('%Y%m%d_%H%M%S')
        out_prefix = f'verification_report_{t}'
    json_path = Path(f'{out_prefix}.json')
    csv_path = Path(f'{out_prefix}_prestas.csv')
    with json_path.open('w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # write prestas CSV
    keys = ['sheet', 'excel_row_index', 'presta_id', 'site_taux_col', 'site_taux_val', 'decompte_taux_val', 'status']
    with csv_path.open('w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        for r in report.get('prestas', []):
            writer.writerow({k: r.get(k, '') for k in keys})

    return json_path, csv_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('file', help='Excel file to check')
    parser.add_argument('--apprenant', help='Apprenant id to look for (e.g. APP847)', default=None)
    parser.add_argument('--write-fixed', help='Path to write a fixed copy of the workbook with suggested changes', default=None)
    parser.add_argument('--sync-to-db', help='Path to sqlite DB to sync apprenant presence values (dry-run unless --apply-db)', default=None)
    parser.add_argument('--apply-db', help='If set, actually write changes to the sqlite DB (use with --sync-to-db)', action='store_true')
    args = parser.parse_args()

    report = analyze_workbook(args.file, apprenant_id=args.apprenant)
    json_path, csv_path = save_reports(report)
    print('Report saved:', json_path, csv_path)
    if args.write_fixed:
        out = Path(args.write_fixed)
        apply_fixes(report, args.file, out)
        print('Fixed workbook written to', out)
    # DB sync options
    if args.sync_to_db:
        db_path = Path(args.sync_to_db)
        if not db_path.exists():
            print('DB file not found:', db_path)
        else:
            conn = get_db_conn(str(db_path))
            if args.apply_db:
                # create a timestamped backup of the DB before applying
                bak = db_path.with_suffix('.sqlite3.bak')
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                bak = db_path.with_name(f"{db_path.stem}_pre_sync_{timestamp}{db_path.suffix}")
                shutil.copy2(db_path, bak)
                print('DB backup created at', bak)
                updates = sync_apprenants_to_db(conn, report, args.file, dry_run=False)
            else:
                updates = sync_apprenants_to_db(conn, report, args.file, dry_run=True)
            # save updates summary
            summary_path = Path(f"{json_path.stem}_db_sync_summary.json")
            with summary_path.open('w', encoding='utf-8') as f:
                json.dump(updates, f, ensure_ascii=False, indent=2)
            print('DB sync summary saved to', summary_path)


if __name__ == '__main__':
    main()
