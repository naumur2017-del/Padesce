#!/usr/bin/env python
"""Analyze the Classeur.xlsx file structure"""

import sys

import openpyxl

try:
    wb = openpyxl.load_workbook("Classeur.xlsx", data_only=True)
    print(f"Sheets in workbook: {wb.sheetnames}\n")

    # Get the first sheet (usually Sheet1)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}\n")

    # Read headers
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").strip())

    print(f"Headers ({len(headers)} columns):")
    for i, h in enumerate(headers, 1):
        print(f"  {i}: {h}")

    # Read first 10 data rows
    print("\nFirst 10 data rows:")
    row_count = 0
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2
    ):
        if row_count >= 10:
            break
        # Only process if first cell (class code) has content
        if row[0]:
            row_count += 1
            print(f"\nRow {row_idx}:")
            for i, (header, value) in enumerate(zip(headers, row)):
                if header and value:
                    print(f"  {header}: {value}")

    # Count total rows
    total_rows = ws.max_row - 1
    print(f"\nTotal data rows: {total_rows}")

except Exception as e:
    print(f"Error: {e}", file=sys.stderr)
    import traceback

    traceback.print_exc()
    sys.exit(1)
