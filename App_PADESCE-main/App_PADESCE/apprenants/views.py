import csv
import io
import json
import logging
import os
import re
import random
import string
import unicodedata
import urllib.parse
import urllib.request
from typing import List, Tuple

from django.contrib import messages
from django.db import IntegrityError, transaction
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_POST
from django.conf import settings
from openpyxl import load_workbook

from App_PADESCE.apprenants.forms import ImportApprenantsForm
from App_PADESCE.apprenants.models import Apprenant, SmsLog
from App_PADESCE.formations.models import Classe

logger = logging.getLogger(__name__)

COLUMN_DEFS: List[Tuple[str, str]] = [
    ("numero", "N° / No"),
    ("nom_complet", "Nom complet"),
    ("beneficiaire", "Beneficiaire"),
    ("genre", "Genre"),
    ("age", "Age"),
    ("fonction", "Fonction"),
    ("qualification", "Diplome"),
    ("nb_annees_experience", "Annees d'experience"),
    ("ville_residence", "Ville de residence"),
    ("prestataire", "Prestataire"),
    ("intitule_formation_solicitee", "Intitule sollicite"),
    ("intitule_formation_dispensee", "Intitule dispensee"),
    ("fenetre", "Fenetre"),
    ("ville_formation", "Ville de formation"),
    ("arrondissement", "Arrondissement"),
    ("departement", "Departement"),
    ("region", "Region"),
    ("lieu_formation", "Lieu de formation"),
    ("precision_lieu", "Precision lieu"),
    ("longitude", "Longitude"),
    ("latitude", "Latitude"),
    ("telephone1", "1er numero"),
    ("telephone2", "2e numero"),
    ("cohorte", "Cohorte"),
    ("tel_formateur", "Tel formateur"),
    ("_code", "Code"),
]

HEADER_ALIASES = {
    # Primary labels (normalized) from the default column definitions
    # normalized dynamically later to stay in sync with COLUMN_DEFS
}

def generate_code(existing: set[str]) -> str:
    """Generate a 4-char alphanumeric uppercase code, unique against existing."""
    charset = string.ascii_uppercase + string.digits
    for _ in range(1000):
        code = "".join(random.choices(charset, k=4))
        if code not in existing:
            return code
    raise ValueError("Impossible de generer un code unique")


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
    return str(value).strip()


def _to_int(value, default=0):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalize_header_name(value: str) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", str(value))
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_only = ascii_only.lower().replace("&", "and")
    ascii_only = re.sub(r"[^a-z0-9]+", " ", ascii_only)
    return " ".join(ascii_only.split())


def _build_header_aliases() -> dict:
    aliases = {_normalize_header_name(label): key for key, label in COLUMN_DEFS}
    aliases.update(
        {
            "nom": "nom_complet",
            "nom et prenom": "nom_complet",
            "name first name": "nom_complet",
            "numero": "numero",
            "n": "numero",
            "beneficiaires": "beneficiaire",
            "beneficiaires beneficiary": "beneficiaire",
            "genre h f gender m f": "genre",
            "diplome 1 0 1 2 etc diploma 1 0 1 2 etc": "qualification",
            "nb d annees d experience 0 1 2 etc years of experience 0 1 2 etc": "nb_annees_experience",
            "fonction d c e m b function d c e m b": "fonction",
            "ville de residence de l apprenant learner s city of residence": "ville_residence",
            "intitule de la formation sollicitee title of requested training": "intitule_formation_solicitee",
            "intitule de formation sollicitee title of requested training": "intitule_formation_solicitee",
            "intitule de la formation dispensee title of training delivered": "intitule_formation_dispensee",
            "intitule de formation dispensee title of training delivered": "intitule_formation_dispensee",
            "ville de la formation training city": "ville_formation",
            "arrondissement de la formation training district": "arrondissement",
            "departement de la formation training division": "departement",
            "region de la formation training region": "region",
            "denomination du lieu de la formation training venue name": "lieu_formation",
            "denomination du lieu de formation training venue name": "lieu_formation",
            "precision sur le lieu quartier de formation training area or neighborhood details": "precision_lieu",
            "coordonnees gps du lieu de formation longitude training venue gps longitude": "longitude",
            "coordonnees gps du lieu de formation latitude training venue gps latitude": "latitude",
            "telephone": "telephone1",
            "tel1": "telephone1",
            "tel 1": "telephone1",
            "tel2": "telephone2",
            "tel 2": "telephone2",
            "1er no tel apprenant learner s 1st phone number": "telephone1",
            "2e no tel apprenant si disponible learner s 2nd phone number if available": "telephone2",
            "cohorte cohort": "cohorte",
            "cohort": "cohorte",
            "tel formateur point focal sur place trainer or local focal point phone number": "tel_formateur",
            "code apprenant": "_code",
        }
    )
    return aliases


HEADER_ALIASES.update(_build_header_aliases())
IMPORT_FIELDS = [field for field, _ in COLUMN_DEFS]


def _default_column_defs() -> List[dict]:
    return [{"field": key, "label": label} for key, label in COLUMN_DEFS]


def _read_csv_rows(data: str) -> List[List[str]]:
    first_line = data.splitlines()[0] if data else ""
    delimiter = ";" if ";" in first_line else ","
    reader = csv.reader(io.StringIO(data), delimiter=delimiter)
    return [[cell.strip() for cell in row] for row in reader]


def _read_xlsx_rows(file_obj) -> List[List[str]]:
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_normalize_cell(cell) for cell in row])
    return rows


def _is_header_row(row: List[str]) -> bool:
    if len(row) < 2:
        return False
    header = (row[1] or "").lower()
    return "nom" in header or "prenom" in header


def _build_header_mapping(header_row: List[str]) -> List[Tuple[int, str, str]]:
    mapping: List[Tuple[int, str, str]] = []
    seen = set()
    for idx, raw in enumerate(header_row):
        normalized = _normalize_header_name(raw)
        key = HEADER_ALIASES.get(normalized)
        if not key or key in seen:
            continue
        label = raw.strip() or dict(COLUMN_DEFS).get(key, key)
        mapping.append((idx, key, label))
        seen.add(key)
    return mapping


def _clean_phones(tel1: str, tel2: str) -> tuple[str, str]:
    tel1_clean = (tel1 or "").strip()
    tel2_clean = (tel2 or "").strip()
    if "/" in tel1_clean:
        parts = [p.strip() for p in tel1_clean.split("/") if p.strip()]
        if parts:
            tel1_clean = parts[0]
            if len(parts) > 1 and not tel2_clean:
                tel2_clean = parts[1]
    return tel1_clean, tel2_clean


def _rows_from_payload(payload: list) -> List[dict]:
    rows: List[dict] = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        base = {field: _normalize_cell(item.get(field)) for field in IMPORT_FIELDS}
        tel1, tel2 = _clean_phones(base.get("telephone1"), base.get("telephone2"))
        base["telephone1"] = tel1 or None
        base["telephone2"] = tel2 or None
        base["age"] = _to_int(base.get("age"), None) if base.get("age") else None
        base["nb_annees_experience"] = _to_int(base.get("nb_annees_experience") or 0, 0)
        rows.append(base)
    return rows


def _rows_from_table(rows: List[List[str]], header_map: List[Tuple[int, str, str]]) -> List[dict]:
    preview_rows: List[dict] = []
    if not header_map:
        header_map = [
            (idx, key, label) for idx, (key, label) in enumerate(COLUMN_DEFS) if idx < len(rows[0] if rows else [])
        ]
    for row in rows:
        normalized_cells = [_normalize_cell(cell) for cell in row]
        if not normalized_cells or all(not cell for cell in normalized_cells):
            continue
        if _is_header_row(normalized_cells):
            continue
        base = {field: "" for field in IMPORT_FIELDS}
        for idx, key, _label in header_map:
            if idx < len(normalized_cells):
                base[key] = normalized_cells[idx]
        tel1, tel2 = _clean_phones(base.get("telephone1"), base.get("telephone2"))
        base["telephone1"] = tel1 or None
        base["telephone2"] = tel2 or None
        base["age"] = _to_int(base.get("age"), None) if base.get("age") else None
        base["nb_annees_experience"] = _to_int(base.get("nb_annees_experience") or 0, 0)
        preview_rows.append(base)
    return preview_rows


def _header_map_to_defs(header_map: List[Tuple[int, str, str]]) -> List[dict]:
    if not header_map:
        return _default_column_defs()
    return [{"field": key, "label": label} for _, key, label in header_map]


def _validate_preview(preview_rows: List[dict], formation, generate_codes: bool, errors: List[str]) -> None:
    noms = [r["nom_complet"] for r in preview_rows if r.get("nom_complet")]
    tels = [r["telephone1"] for r in preview_rows if r.get("telephone1")]
    numeros = [r["numero"] for r in preview_rows if r.get("numero")]
    
    if len(noms) != len(set(noms)):
        errors.append("Noms en double detectes dans le fichier.")
    if len(tels) != len(set(tels)):
        errors.append("Telephones en double detectes dans le fichier.")
    if len(numeros) != len(set(numeros)):
        errors.append("Numeros en double detectes dans le fichier.")

    # Validation: Valeurs manquantes (Age)
    missing_age_count = sum(1 for r in preview_rows if r.get("age") is None)
    if missing_age_count > 0:
        errors.append(f"Valeurs manquantes detectees: {missing_age_count} ligne(s).")

    # Validation: Format telephone (9 chiffres) et valeurs manquantes
    invalid_tel_count = 0
    missing_tel_count = 0
    for r in preview_rows:
        t = r.get("telephone1")
        if t:
            # On verifie si c'est numerique et fait 9 caracteres
            if not (t.isdigit() and len(t) == 9):
                invalid_tel_count += 1
        else:
            missing_tel_count += 1
    
    if invalid_tel_count > 0 or missing_tel_count > 0:
        msg = f"Telephone apprenant 1 invalide: {invalid_tel_count} ligne(s). Format attendu: 9 chiffres."
        if missing_tel_count > 0:
            msg += f", {missing_tel_count} valeurs manquantes"
        errors.append(msg)

    existing_tels = set(
        Apprenant.objects.filter(formation=formation).values_list("telephone1", flat=True)
    )
    duplicate_existing = existing_tels.intersection(set(tels))
    if duplicate_existing:
        errors.append(
            "Telephones deja utilises pour cette formation: "
            + ", ".join(sorted(duplicate_existing))
        )

    if not generate_codes:
        codes = [r["_code"] for r in preview_rows if r["_code"]]
        if len(codes) != len(set(codes)):
            errors.append("Codes en double detectes dans le fichier.")
        existing_codes = set(Apprenant.objects.values_list("code", flat=True))
        duplicate_codes = existing_codes.intersection(set(codes))
        if duplicate_codes:
            errors.append("Codes deja utilises: " + ", ".join(sorted(duplicate_codes)))


def import_csv(request, classe_id: int):
    classe = get_object_or_404(Classe.objects.select_related("formation"), pk=classe_id)
    formation = classe.formation
    existing_tels = list(
        Apprenant.objects.filter(formation=formation).values_list("telephone1", flat=True)
    )
    form = ImportApprenantsForm(request.POST or None, request.FILES or None)
    preview_rows: List[dict] = []
    errors: List[str] = []
    column_defs = _default_column_defs()

    if request.method == "POST" and form.is_valid():
        fichier = form.cleaned_data["fichier"]
        generate_codes = form.cleaned_data.get("generate_codes", False)
        edited_rows = form.cleaned_data.get("edited_rows") or ""
        ext = os.path.splitext(getattr(fichier, "name", ""))[1].lower()
        rows: List[List[str]] = []
        header_map: List[Tuple[int, str, str]] = []
        if edited_rows:
            try:
                payload = json.loads(edited_rows)
                preview_rows = _rows_from_payload(payload)
            except json.JSONDecodeError:
                errors.append("Impossible de lire les lignes modifiees.")
        if not preview_rows and not errors:
            try:
                if ext in {".xlsx", ".xlsm"}:
                    rows = _read_xlsx_rows(fichier)
                else:
                    data = fichier.read().decode("utf-8-sig")
                    rows = _read_csv_rows(data)
            except UnicodeDecodeError:
                errors.append("Le fichier doit etre encode en UTF-8.")
                rows = []

        if rows:
            header_row = rows[0] if rows else []
            data_rows = rows[1:] if rows else []
            header_map = _build_header_mapping(header_row)
            if not header_map:
                header_map = [
                    (idx, key, label)
                    for idx, (key, label) in enumerate(COLUMN_DEFS)
                    if idx < len(header_row)
                ]
            column_defs = _header_map_to_defs(header_map)
            preview_rows = _rows_from_table(data_rows, header_map)

        if preview_rows and not errors:
            _validate_preview(preview_rows, formation, generate_codes, errors)

        if not errors and preview_rows:
            try:
                with transaction.atomic():
                    existing_codes = set(Apprenant.objects.values_list("code", flat=True))
                    new_objects = []
                    for row in preview_rows:
                        if generate_codes or not row.get("_code"):
                            code = generate_code(existing_codes)
                            existing_codes.add(code)
                        else:
                            code = row["_code"]
                            existing_codes.add(code)
                        new_objects.append(
                            Apprenant(
                                code=code,
                                classe=classe,
                                formation=formation,
                                **{k: v for k, v in row.items() if k != "_code"},
                            )
                        )
                    Apprenant.objects.bulk_create(new_objects, ignore_conflicts=False)
                messages.success(request, f"{len(preview_rows)} apprenants importes pour {classe.code}.")
            except IntegrityError:
                errors.append(
                    "Telephones en double detectes dans le fichier ou deja en base. "
                    "Veuillez re-valider apres correction."
                )

    return render(
        request,
        "apprenants/import.html",
        {
            "form": form,
            "classe": classe,
            "preview_rows": preview_rows,
            "errors": errors,
            "existing_tels": existing_tels,
            "column_defs": column_defs,
            "header_aliases": HEADER_ALIASES,
        },
    )



def api_codes(request):
    """Expose the list of existing apprenant codes for front-end checks."""
    codes = list(Apprenant.objects.values_list("code", flat=True))
    return JsonResponse({"codes": codes})


def _parse_json_payload(request):
    if request.content_type and request.content_type.startswith("application/json"):
        try:
            return json.loads(request.body.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError):
            return {}
    return {}


def _parse_ids(value) -> List[int]:
    ids: List[int] = []
    if isinstance(value, (list, tuple)):
        raw_list = value
    elif isinstance(value, str):
        raw_list = [item for item in value.split(",") if item.strip()]
    else:
        raw_list = []
    for item in raw_list:
        try:
            ids.append(int(item))
        except (TypeError, ValueError):
            continue
    return ids


def _normalize_phone(value: str) -> str:
    if not value:
        return ""
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    if len(digits) > 9:
        if digits.startswith(settings.OBIT_COUNTRY) and len(digits) == len(settings.OBIT_COUNTRY) + 9:
            digits = digits[len(settings.OBIT_COUNTRY):]
        else:
            digits = digits[-9:]
    return digits if len(digits) == 9 else ""


def _build_sms_message(apprenant: Apprenant) -> str:
    return f"Bonjour Mr/Mlle, votre code PADESCE est: {apprenant.code}. "


def _send_obit_sms(local_number: str, message: str) -> tuple[bool, str]:
    if not settings.OBIT_API_KEY:
        return False, "API key manquante"
    params = {
        "key_api": settings.OBIT_API_KEY,
        "sender": settings.OBIT_SENDER,
        "destination": f"{settings.OBIT_COUNTRY}{local_number}",
        "message": message,
    }
    url = f"{settings.OBIT_API_URL}?{urllib.parse.urlencode(params)}"
    try:
        with urllib.request.urlopen(url, timeout=15) as resp:
            body = resp.read().decode("utf-8", "ignore")
            if getattr(resp, "status", 200) != 200:
                return False, f"HTTP {resp.status}"
    except Exception as exc:  # pragma: no cover - network errors
        logger.exception("SMS request failed")
        return False, f"Exception: {exc}"

    try:
        data = json.loads(body)
        if isinstance(data, dict) and data.get("success") is True:
            return True, "OK"
    except json.JSONDecodeError:
        pass
    if '"success":true' in body.replace(" ", "").lower():
        return True, "OK"
    if "error" in body.lower():
        return False, "Erreur API"
    return False, "Reponse inconnue"


@require_POST
def update_appartenance(request, apprenant_id: int):
    apprenant = get_object_or_404(Apprenant, pk=apprenant_id)
    value = None
    payload = _parse_json_payload(request)
    if payload:
        value = payload.get("value")
    if value is None:
        value = request.POST.get("value")
    value_str = str(value).strip().lower()
    apprenant.appartenance_beneficiaire = value_str in {"1", "true", "on", "yes"}
    apprenant.save(update_fields=["appartenance_beneficiaire"])
    return JsonResponse({"ok": True, "value": apprenant.appartenance_beneficiaire})


@require_POST
def update_appartenance_bulk(request):
    payload = _parse_json_payload(request)
    ids = _parse_ids(payload.get("ids")) if payload else _parse_ids(request.POST.getlist("ids"))
    if not ids:
        ids = _parse_ids(request.POST.get("ids"))
    classe_id = payload.get("classe_id") if payload else request.POST.get("classe_id")
    value = payload.get("value") if payload else request.POST.get("value")
    if value is None:
        value = request.POST.get("value")
    value_str = str(value).strip().lower()
    value_bool = value_str in {"1", "true", "on", "yes"}
    qs = Apprenant.objects.filter(id__in=ids)
    if classe_id:
        qs = qs.filter(classe_id=classe_id)
    updated = qs.update(appartenance_beneficiaire=value_bool)
    return JsonResponse({"ok": True, "updated": updated, "value": value_bool})


@require_POST
def delete_apprenants(request):
    payload = _parse_json_payload(request)
    ids = _parse_ids(payload.get("ids")) if payload else _parse_ids(request.POST.getlist("ids"))
    if not ids:
        ids = _parse_ids(request.POST.get("ids"))
    classe_id = payload.get("classe_id") if payload else request.POST.get("classe_id")
    qs = Apprenant.objects.filter(id__in=ids)
    if classe_id:
        qs = qs.filter(classe_id=classe_id)
    deleted, _ = qs.delete()
    return JsonResponse({"ok": True, "deleted": deleted})


@require_POST
def send_sms(request):
    payload = _parse_json_payload(request)
    ids = _parse_ids(payload.get("ids")) if payload else _parse_ids(request.POST.getlist("ids"))
    if not ids:
        ids = _parse_ids(request.POST.get("ids"))
    classe_id = payload.get("classe_id") if payload else request.POST.get("classe_id")
    qs = Apprenant.objects.filter(id__in=ids)
    if classe_id:
        qs = qs.filter(classe_id=classe_id)

    results = []
    logs = []
    sent = 0
    failed = 0
    for apprenant in qs:
        phone = _normalize_phone(apprenant.telephone1 or "")
        message = _build_sms_message(apprenant)
        if not phone:
            logs.append(
                SmsLog(
                    apprenant=apprenant,
                    classe=apprenant.classe,
                    telephone=apprenant.telephone1 or "",
                    message=message,
                    status="failed",
                    detail="Numero invalide",
                )
            )
            results.append({"id": apprenant.id, "ok": False, "detail": "Numero invalide"})
            failed += 1
            continue
        ok, detail = _send_obit_sms(phone, message)
        logs.append(
            SmsLog(
                apprenant=apprenant,
                classe=apprenant.classe,
                telephone=phone,
                message=message,
                status="sent" if ok else "failed",
                detail=detail,
            )
        )
        results.append({"id": apprenant.id, "ok": ok, "detail": detail})
        if ok:
            sent += 1
        else:
            failed += 1
    if logs:
        SmsLog.objects.bulk_create(logs)

    return JsonResponse({"ok": True, "sent": sent, "failed": failed, "results": results})
