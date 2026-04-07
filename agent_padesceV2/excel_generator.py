# agent_padesce/excel_generator.py
"""Génération du rapport Excel formaté."""

import math
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import BLACK, GRAY_LIGHT, GREEN_DARK, GREEN_MED, PURPLE, TEAL, WHITE, YELLOW_HDR

# ═══════════════════════════════════════════════════════════════
#  FONCTIONS UTILITAIRES EXCEL
# ═══════════════════════════════════════════════════════════════


def _font(bold=False, color=BLACK, size=10, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_purple():
    s = Side(style="thin", color=PURPLE)
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_border_range(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _border_purple()


def _set_cell(ws, row, col, value="", bold=False, fg=BLACK, bg=None, h_align="left", wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=fg)
    cell.alignment = _align(h=h_align, wrap=wrap, v="center")
    if bg:
        cell.fill = _fill(bg)
    cell.border = _border_purple()
    return cell


def _merge(ws, r1, c1, r2, c2, value="", bold=False, fg=BLACK, bg=None, h_align="left", wrap=False):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = _font(bold=bold, color=fg)
    cell.alignment = _align(h=h_align, wrap=wrap, v="center")
    if bg:
        cell.fill = _fill(bg)
    _apply_border_range(ws, r1, c1, r2, c2)
    return cell


def _safe(val, default=""):
    if val is None:
        return default
    try:
        if math.isnan(float(val)):
            return default
    except (TypeError, ValueError):
        pass
    return val


def _pct_str(val):
    if val == "" or val is None:
        return ""
    s = str(val).strip().rstrip("%")
    try:
        f = float(s)
        return f"{round(f)}%" if f > 1 else f"{round(f * 100)}%"
    except (TypeError, ValueError):
        return str(val)


# ═══════════════════════════════════════════════════════════════
#  GÉNÉRATION DU RAPPORT
# ═══════════════════════════════════════════════════════════════


def generer_rapport_excel(prestations_data: list, output_path: str) -> dict:
    """
    Génère un rapport Excel formaté avec les prestations et leurs classes.
    """
    wb = Workbook()
    wb.remove(wb.active)

    prestataires_impliques = set()
    beneficiaires_impliques = set()
    nb_classes_total = 0

    for idx, data in enumerate(prestations_data):
        infos_df = data["infos"]
        classes = data["classes"]

        if infos_df.empty:
            continue

        infos = infos_df.iloc[0]

        presta_id = str(_safe(infos.get("prestation_id", f"Prestation_{idx + 1}")))
        sheet_name = re.sub(r"[\\/*?:\[\]]", "_", presta_id)[:31]
        ws = wb.create_sheet(title=sheet_name)

        prestataires_impliques.add(str(_safe(infos.get("prestataire", ""))))
        beneficiaires_impliques.add(str(_safe(infos.get("bénéficaire", ""))))
        nb_classes_total += len(classes) if classes is not None else 0

        for i, w in enumerate([28, 22, 18, 22, 22], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # En-tête
        _set_cell(ws, row, 1, "Statut", bold=True, fg=WHITE, bg=PURPLE, h_align="center")
        _set_cell(ws, row, 2, "Prestation ID", bold=True, fg=WHITE, bg=PURPLE, h_align="center")
        _set_cell(ws, row, 3, "Prestaire", bold=True, fg=WHITE, bg=TEAL, h_align="center")
        _merge(
            ws, row, 4, row, 5, "Bénéficiaire", bold=True, fg=BLACK, bg=YELLOW_HDR, h_align="center"
        )
        row += 1

        statut = _safe(infos.get("statut", ""))
        prestataire = _safe(infos.get("prestataire", ""))
        beneficiaire = _safe(infos.get("bénéficaire", ""))

        _set_cell(ws, row, 1, statut, fg=BLACK)
        _set_cell(ws, row, 2, presta_id, fg=BLACK, bold=True)
        _set_cell(ws, row, 3, prestataire, fg=BLACK)
        _merge(ws, row, 4, row, 5, beneficiaire, fg=BLACK)
        row += 1

        # Images placeholder
        ws.row_dimensions[row].height = 60
        ws.row_dimensions[row + 1].height = 60
        _merge(ws, row, 1, row + 1, 2, "Image", h_align="center", fg=BLACK)
        _merge(ws, row, 3, row + 1, 5, "Image", h_align="center", fg=BLACK)
        row += 2

        # Thème
        theme = _safe(infos.get("formation", ""))
        _merge(ws, row, 1, row, 2, "Thème de la formation", bold=True, fg=BLACK)
        _merge(ws, row, 3, row, 5, theme, fg=BLACK)
        row += 1

        # Fenêtre
        fenetre = _safe(infos.get("fénetre", ""))
        _set_cell(ws, row, 1, "Fenêtre", bold=True, fg=BLACK)
        _set_cell(ws, row, 2, fenetre, fg=BLACK, h_align="right")
        for c in [3, 4, 5]:
            _set_cell(ws, row, c, "", fg=BLACK)
        row += 1

        # Objectifs / Inscrits
        obj_padesce = _safe(infos.get("objectifs_padesce_t", ""))
        nb_inscrits = _safe(infos.get("apprenants_inscrits_t", ""))

        _set_cell(ws, row, 1, "Objectifs d'inscription PADESCE", bold=True, fg=BLACK)
        _set_cell(ws, row, 2, obj_padesce, fg=BLACK, h_align="right")
        _set_cell(ws, row, 3, "Nombre d'inscrits", bold=True, fg=BLACK)
        _merge(ws, row, 4, row, 5, nb_inscrits, fg=BLACK, h_align="right")
        row += 1

        # En-tête classes
        _merge(ws, row, 1, row, 2, "CLASSES", bold=True, fg=WHITE, bg=GREEN_DARK)
        _set_cell(ws, row, 3, "Suivi", bold=True, fg=WHITE, bg=GREEN_MED, h_align="center")
        _set_cell(
            ws, row, 4, "Nbre d'inscrits", bold=True, fg=WHITE, bg=GREEN_MED, h_align="center"
        )
        _set_cell(ws, row, 5, "Ville", bold=True, fg=WHITE, bg=GREEN_MED, h_align="center")
        row += 1

        # Lignes classes
        if classes is not None and len(classes) > 0:
            for _, cl_row in classes.iterrows():
                classe_val = _safe(cl_row.get("classe", ""))
                inscrits = _safe(cl_row.get("Nb personnes", ""))
                ville_val = _safe(cl_row.get("villes", ""))
                _merge(ws, row, 1, row, 2, classe_val, fg=BLACK)
                _set_cell(ws, row, 3, "", fg=BLACK)
                _set_cell(ws, row, 4, inscrits, fg=BLACK, h_align="center")
                _set_cell(ws, row, 5, ville_val, fg=BLACK)
                row += 1
        else:
            _merge(ws, row, 1, row, 5, "Aucune classe", fg=BLACK, h_align="center")
            row += 1

        # Performance
        _set_cell(ws, row, 1, "Performance", bold=True, fg=BLACK)
        _set_cell(ws, row, 2, "De la prestation", bold=True, fg=BLACK, h_align="center", wrap=True)
        _set_cell(ws, row, 3, "Moyenne fenêtre", bold=True, fg=BLACK, h_align="center", wrap=True)
        _merge(ws, row, 4, row, 5, "Moyenne générale", bold=True, fg=BLACK, h_align="center")
        row += 1

        taux_insc = _pct_str(_safe(infos.get("taux_inscription_calcule", "")))
        taux_assiste = _pct_str(
            _safe(
                infos.get("apprenants_suivis_ayant_assisté_au_moins_une_fois_à_la_formation_t", "")
            )
        )
        taux_presence = _pct_str(_safe(infos.get("taux__de_présence_moyen__nan", "")))
        taux_formes = _pct_str(_safe(infos.get("taux__de_personnes_formées_nan", "")))
        sat_form = _pct_str(_safe(infos.get("taux_de_satisfaction_formateurs_nan", "")))
        sat_app = _pct_str(_safe(infos.get("taux_de_satisfaction_appreannts_nan", "")))

        perf_rows = [
            ("Taux d'inscription", taux_insc),
            ("% de personnes ayant assisté au moins une fois", taux_assiste),
            ("Taux moyen de présence", taux_presence),
            ("Taux de personnes formés", taux_formes),
            ("Taux de satifaction formateur", sat_form),
            ("Taux de satifaction apprenants", sat_app),
            ("Respect du thème de formation", "100%"),
            ("Respect du volume horaire", "0%"),
        ]

        for label, val_presta in perf_rows:
            _set_cell(ws, row, 1, label, fg=BLACK)
            _set_cell(ws, row, 2, val_presta, fg=BLACK, h_align="right")
            _set_cell(ws, row, 3, "", fg=BLACK)
            _merge(ws, row, 4, row, 5, "", fg=BLACK)
            row += 1

        # Incidents
        _set_cell(ws, row, 1, "Inidents", bold=True, fg=BLACK)
        _set_cell(ws, row, 2, "Référence", bold=True, fg=BLACK, h_align="center")
        _merge(ws, row, 3, row, 5, "Titre", bold=True, fg=BLACK, h_align="center")
        row += 1

        for inc_label, inc_ref in [
            ("Mélange de cohorte", "PAD236"),
            ("Mélange de cohorte", "PAD237"),
            ("Mélange de cohorte", "PAD238"),
            ("Mélange de cohorte", "PAD239"),
        ]:
            _set_cell(ws, row, 1, inc_label, fg=BLACK)
            _set_cell(ws, row, 2, inc_ref, fg=BLACK, h_align="center")
            _merge(ws, row, 3, row, 5, "", fg=BLACK)
            row += 1

        # Commentaires
        _merge(
            ws,
            row,
            1,
            row,
            5,
            "Commentaires généraux",
            bold=True,
            fg=BLACK,
            bg=GRAY_LIGHT,
            h_align="center",
        )
        row += 1
        ws.row_dimensions[row].height = 60
        _merge(ws, row, 1, row, 5, "", fg=BLACK)

    wb.save(output_path)
    print(f"Classeur sauvegardé : {output_path}")

    return {
        "statut": "succes",
        "fichier": output_path,
        "nb_prestations": len(prestations_data),
        "prestataires_impliques": list(prestataires_impliques - {""}),
        "beneficiaires_impliques": list(beneficiaires_impliques - {""}),
        "nb_classes_total": nb_classes_total,
    }
