import datetime
import io
import logging
import threading
import zipfile
import unicodedata
import csv
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Case, Count, DecimalField, ExpressionWrapper, F, Q, When
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from App_PADESCE.apprenants.models import Apprenant
from App_PADESCE.appels.models import (
    Appel,
    AppelAnswers,
    AppelCGA,
    AppelFormateur,
    AppelImportArchive,
    appel_answers_completed_q,
    appel_answers_modified_completion_q,
    padesce_form_tracking_cutoff,
)
from App_PADESCE.formations.models import Classe

logger = logging.getLogger(__name__)
from App_PADESCE.satisfaction_apprenants.models import SatisfactionApprenant


APPEL_QUESTION_FIELDS = [
    "q1_clarte_exposes",
    "q2_interaction_formateur",
    "q3_maitrise_contenu",
    "q4_salle_adequate",
    "q5_materiel_disponible",
    "q6_organisation_temps",
    "q7_utilite_formation",
    "q8_adequation_besoins",
    "q9_satisfaction_globale",
]





def _normalize_header(value):
    if value is None:
        return ""
    text = " ".join(str(value).strip().lower().split())
    # Strip accents for stable matching across encodings.
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_phone(value: str) -> str:
    return "".join(ch for ch in str(value or "") if ch.isdigit())


def _normalize_name(value: str) -> str:
    text = " ".join(str(value or "").strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _parse_bool_flag(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    return s in ("1", "true", "on", "yes")


def _resolve_audio_file_state(file_field) -> tuple[bool, str, bool]:
    if not file_field:
        return False, "", False
    name = getattr(file_field, "name", "") or ""
    if not name:
        return False, "", False
    try:
        exists = file_field.storage.exists(name)
    except Exception:
        exists = False
    if not exists:
        return False, "", True
    try:
        return True, file_field.url, False
    except Exception:
        return True, "", False


def _safe_audio_url(instance) -> str:
    return _resolve_audio_file_state(getattr(instance, "audio_file", None))[1]


def _has_audio_file(instance) -> bool:
    return _resolve_audio_file_state(getattr(instance, "audio_file", None))[0]


def _bind_audio_state(rows):
    for row in rows:
        has_audio_file, audio_file_url, audio_file_missing = _resolve_audio_file_state(
            getattr(row, "audio_file", None)
        )
        row.has_audio_file = has_audio_file
        row.audio_file_url = audio_file_url
        row.audio_file_missing = audio_file_missing
    return rows


def _coerce_note_value(value, default=None):
    try:
        note = int(value)
    except (TypeError, ValueError):
        return default
    return note if 1 <= note <= 5 else default


def _save_appel_answers(appel: Appel, user, payload: dict, *, apply_defaults: bool = False):
    defaults = {}
    for index, field in enumerate(APPEL_QUESTION_FIELDS, 1):
        raw_value = payload.get(field, payload.get(f"q{index}"))
        note = _coerce_note_value(raw_value, 3 if apply_defaults else None)
        if note is not None:
            defaults[field] = note

    if "commentaire" in payload or apply_defaults:
        defaults["commentaire"] = str(payload.get("commentaire", "") or "").strip() or "RAS"
    if "recommandations" in payload or apply_defaults:
        defaults["recommandations"] = str(payload.get("recommandations", "") or "").strip() or "RAS"

    defaults["modified_by"] = user if getattr(user, "is_authenticated", False) else None
    defaults["modified_at"] = timezone.now()
    answers, _created = AppelAnswers.objects.update_or_create(appel=appel, defaults=defaults)
    return answers


def _build_satisfaction_defaults(payload: dict, user, *, now=None) -> dict:
    now = now or timezone.localtime()
    defaults = {
        "inspecteur": None,
        "enqueteur": user if getattr(user, "is_authenticated", False) else None,
        "date": now.date(),
        "heure": now.time().replace(microsecond=0),
        "commentaire": str(payload.get("commentaire", "") or "").strip() or "RAS",
        "recommandations": str(payload.get("recommandations", "") or "").strip() or "RAS",
        "transcription": "",
    }
    for index, field in enumerate(APPEL_QUESTION_FIELDS, 1):
        defaults[field] = _coerce_note_value(payload.get(field, payload.get(f"q{index}")), 3) or 3
    return defaults


def _appel_reference_details(appel: Appel) -> str:
    classe_ref = str(getattr(getattr(appel, "classe", None), "code", "") or appel.classe_label or "-").strip() or "-"
    tel_ref = str(appel.telephone1 or appel.telephone2 or "").strip() or "-"
    return (
        f"(ligne={appel.id}, code={appel.code or '-'}, nom={appel.nom or '-'}, "
        f"tel={tel_ref}, classe={classe_ref})"
    )


def _find_existing_satisfaction_for_appel(appel: Appel, apprenant: Apprenant | None, survey_date):
    satisfaction = SatisfactionApprenant.objects.filter(appel=appel).first()
    if satisfaction:
        return satisfaction

    legacy_qs = SatisfactionApprenant.objects.filter(appel__isnull=True, date=survey_date)
    if apprenant:
        legacy_qs = legacy_qs.filter(apprenant=apprenant)
        if apprenant.classe_id:
            legacy_qs = legacy_qs.filter(Q(classe_id=apprenant.classe_id) | Q(classe__isnull=True))
    elif appel.classe_id:
        legacy_qs = legacy_qs.filter(apprenant__isnull=True, classe_id=appel.classe_id)
    else:
        return None

    candidate_ids = list(legacy_qs.order_by("-updated_at", "-created_at", "-id").values_list("id", flat=True)[:2])
    if len(candidate_ids) == 1:
        return SatisfactionApprenant.objects.filter(pk=candidate_ids[0]).first()
    return None


def _save_satisfaction_for_appel(appel: Appel, user, payload: dict, apprenant: Apprenant | None):
    defaults = _build_satisfaction_defaults(payload, user)
    defaults["apprenant"] = apprenant
    defaults["classe"] = getattr(apprenant, "classe", None) or appel.classe

    satisfaction = _find_existing_satisfaction_for_appel(appel, apprenant, defaults["date"])
    if satisfaction:
        for field, value in defaults.items():
            setattr(satisfaction, field, value)
        satisfaction.appel = appel
        satisfaction.save()
        return satisfaction

    return SatisfactionApprenant.objects.create(appel=appel, **defaults)


def _attach_appel_audio_to_satisfaction(satisfaction: SatisfactionApprenant | None, appel: Appel):
    if not satisfaction or not appel.audio_file:
        return
    satisfaction.audio_appel = appel.audio_file
    satisfaction.save(update_fields=["audio_appel", "updated_at"])


def _status_rank(value: str) -> int:
    ranks = {
        "termine": 4,
        "en_cours": 3,
        "pause": 2,
        "a_rappeler": 1,
        "en_attente": 0,
    }
    return ranks.get(str(value or "").strip(), -1)


def _best_duplicate_winner(rows):
    def sort_key(row):
        audio_name = getattr(getattr(row, "audio_file", None), "name", "") or ""
        updated_at = getattr(row, "updated_at", None)
        created_at = getattr(row, "created_at", None)
        return (
            1 if audio_name else 0,
            _status_rank(getattr(row, "status", "")),
            1 if getattr(row, "locked_by_id", None) else 0,
            updated_at.isoformat() if updated_at else "",
            created_at.isoformat() if created_at else "",
            -(getattr(row, "id", 0) or 0),
        )

    return sorted(rows, key=sort_key, reverse=True)[0]


def _deactivate_duplicate_rows(queryset, key_builder):
    groups = {}
    for row in queryset.iterator(chunk_size=2000):
        dedupe_key = key_builder(row)
        if not dedupe_key:
            continue
        groups.setdefault(dedupe_key, []).append(row)

    duplicates = []
    for rows in groups.values():
        if len(rows) < 2:
            continue
        winner = _best_duplicate_winner(rows)
        duplicates.extend(row.id for row in rows if row.id != winner.id)

    if duplicates:
        queryset.model.objects.filter(pk__in=duplicates).update(is_active=False)
    return len(duplicates)


def _appel_duplicate_key(row: Appel):
    phone = _normalize_phone(row.telephone1) or _normalize_phone(row.telephone2)
    if not phone:
        return None
    return (
        phone,
        _normalize_name(row.nom),
        _normalize_name(row.prestataire),
        _normalize_name(row.beneficiaire),
        _normalize_name(row.classe_label),
    )


def _cga_duplicate_key(row: AppelCGA):
    phone = _normalize_phone(row.telephone)
    if not phone:
        return None
    return phone


def _formateur_duplicate_key(row: AppelFormateur):
    phone = _normalize_phone(row.telephone)
    if not phone:
        return None
    return phone


def _normalize_dashboard_fenetre(value):
    text = str(value or "").strip()
    if not text:
        return ""
    if text in {"2", "3"}:
        return text
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits if digits in {"2", "3"} else ""


def _build_progress_metrics(queryset):
    stats = queryset.aggregate(
        total=Count("id"),
        termines=Count("id", filter=Q(status="termine")),
        rappels=Count("id", filter=Q(status="a_rappeler")),
        audios=Count("id", filter=Q(audio_file__isnull=False)),
    )
    total = int(stats.get("total") or 0)
    termines = int(stats.get("termines") or 0)
    threshold_target = (total + 1) // 2 if total else 0
    completion_rate = round((termines / total) * 100, 1) if total else 0.0
    threshold_reached = total > 0 and termines >= threshold_target
    stats.update(
        {
            "remaining": max(total - termines, 0),
            "completion_rate": completion_rate,
            "completion_label": f"{termines} / {total}" if total else "0 / 0",
            "threshold_target": threshold_target,
            "threshold_remaining": max(threshold_target - termines, 0),
            "threshold_reached": threshold_reached,
            "threshold_message": (
                "Seuil de 50% atteint. Vous pouvez passer a autre chose."
                if threshold_reached
                else (
                    f"Encore {max(threshold_target - termines, 0)} appel(s) termine(s) pour atteindre 50%."
                    if total
                    else "Aucun appel dans ce filtre."
                )
            ),
        }
    )
    return stats



















def _find_apprenant_for_appel(base_qs, appel: Appel):
    """
    Cherche l'apprenant correspondant à l'appel.
    Utilise plusieurs stratégies : code, téléphone, nom, combinaisons.
    """
    apprenant = None
    
    # 1. Chercher par code (exact)
    if appel.code:
        apprenant = base_qs.filter(code__iexact=str(appel.code or "").strip()).first()
        if apprenant:
            logger.debug("Apprenant trouvé par code exact: %s", appel.code)
            return apprenant
    
    # 2. Chercher par code (partiel)
    if appel.code:
        apprenant = base_qs.filter(code__icontains=str(appel.code).strip()).order_by("id").first()
        if apprenant:
            logger.debug("Apprenant trouvé par code partiel: %s", appel.code)
            return apprenant

    # 3. Chercher par téléphone
    tel_candidates = {_normalize_phone(appel.telephone1), _normalize_phone(appel.telephone2)}
    tel_candidates = {t for t in tel_candidates if t}
    for tel in tel_candidates:
        apprenant = (
            base_qs.filter(Q(telephone1__icontains=tel) | Q(telephone2__icontains=tel))
            .order_by("id")
            .first()
        )
        if apprenant:
            logger.debug("Apprenant trouvé par téléphone: %s", tel)
            return apprenant

    # 4. Chercher par nom complet
    if appel.nom:
        nom_norm = _normalize_name(appel.nom)
        for candidate in base_qs.only("id", "nom_complet").iterator(chunk_size=2000):
            if _normalize_name(candidate.nom_complet) == nom_norm:
                apprenant = candidate
                logger.debug("Apprenant trouvé par nom: %s (%s)", appel.nom, apprenant.id)
                return apprenant
    
    # 5. Chercher par nom + prénom (plus flexible)
    if appel.nom:
        # Essayer de chercher les mots-clés du nom dans le nom_complet
        nom_parts = str(appel.nom).strip().lower().split()
        matching_apprenants = []
        for candidate in base_qs.only("id", "nom_complet").iterator(chunk_size=2000):
            nom_complet_lower = candidate.nom_complet.lower()
            # Compter combien de mots du nom correspondent
            matching_words = sum(1 for part in nom_parts if part in nom_complet_lower)
            if matching_words > 0:
                matching_apprenants.append((matching_words, candidate))
        
        if matching_apprenants:
            # Retourner celui avec le plus de mots correspondants
            matching_apprenants.sort(key=lambda x: x[0], reverse=True)
            apprenant = matching_apprenants[0][1]
            logger.debug("Apprenant trouvé par correspondance de nom flexible: %s (%s)", appel.nom, apprenant.id)
            return apprenant
    
    logger.warning(
        "Apprenant NON trouvé. code=%s nom=%s tel=%s",
        appel.code or "-",
        appel.nom or "-",
        appel.telephone1 or appel.telephone2 or "-"
    )
    return None

def _parse_excel_sheet(file_obj, sheet_name: str):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{sheet_name} introuvable dans le fichier.")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            idx = header_map.get(_normalize_header(key))
            if idx is None or idx >= len(row):
                continue
            return row[idx] or ""
        return ""

    data = []
    for row in rows:
        nom = get(
            row,
            "nom et prenom 0 name & first name",
            "nom et prénom 0 name & first name",
            "nom et prenom",
            "nom et prénom",
            "nom",
        )
        code = get(row, "code")
        if not nom or not code:
            continue
        prestataire = get(row, "prestataire")
        beneficiaire = get(row, "beneficiaires", "bénéficiaires", "beneficiaire")
        lieu = get(row, "lieux")
        classe_label = get(row, "classe")
        taux_presence = get(row, "taux de presence", "taux de présence", "présence (%)") or 0
        tel1 = get(row, "1er no tel 0 tel no apprenant", "1er no tel 0 tel no", "1er no tel 0 tel no apprenant")
        tel2 = get(
            row,
            "2e no tel 0 tel no apprenant (si disponible)",
            "2e no tel 0 tel no",
            "2e no tel 0 tel no apprenant (si disponible)",
        )
        type_formation = get(row, "type de formation declaree", "type de formation déclarée")
        formation_padesce = get(row, "formation padesce")
        fenetre = get(row, "fenêtre", "fenetre")
        try:
            taux_presence = Decimal(str(taux_presence))
        except Exception:
            taux_presence = Decimal("0")
        pct = Decimal("0")
        try:
            pct = taux_presence * 100 if taux_presence <= 1 else taux_presence
        except Exception:
            pct = Decimal("0")
        data.append(
            {
                "code": str(code).strip(),
                "nom": str(nom).strip(),
                "prestataire": str(prestataire).strip(),
                "beneficiaire": str(beneficiaire).strip(),
                "lieu": str(lieu).strip(),
                "classe_label": str(classe_label).strip(),
                "fenetre": str(fenetre).strip(),
                "taux_presence": pct,
                "telephone1": str(tel1).strip() if tel1 not in (None, "") else "",
                "telephone2": str(tel2).strip() if tel2 not in (None, "") else "",
                "type_formation_declaree": str(type_formation).strip(),
                "formation_padesce": str(formation_padesce).strip(),
            }
        )
    return data


def _parse_excel(file_obj):
    return _parse_excel_sheet(file_obj, "Sheet1")


def _parse_excel_non_forme(file_obj):
    return _parse_excel_sheet(file_obj, "Feuil2")


def _parse_bool_flag(value):
    if value is None:
        return None
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _archive_before_import_overwrite(appel: Appel, import_mode: str):
    AppelImportArchive.objects.create(
        appel=appel,
        import_mode=import_mode or "",
        source_code=appel.code or "",
        snapshot={
            "id": appel.id,
            "code": appel.code,
            "nom": appel.nom,
            "prestataire": appel.prestataire,
            "beneficiaire": appel.beneficiaire,
            "lieu": appel.lieu,
            "classe_label": appel.classe_label,
            "fenetre": appel.fenetre,
            "is_active": appel.is_active,
            "classe_id": appel.classe_id,
            "telephone1": appel.telephone1,
            "telephone2": appel.telephone2,
            "taux_presence": str(appel.taux_presence or 0),
            "status": appel.status,
            "rappel_at": appel.rappel_at.isoformat() if appel.rappel_at else None,
            "type_formation_declaree": appel.type_formation_declaree,
            "formation_padesce": appel.formation_padesce,
            "deja_forme": appel.deja_forme,
            "audio_file": appel.audio_file.name if appel.audio_file else "",
            "locked_by_id": appel.locked_by_id,
            "locked_at": appel.locked_at.isoformat() if appel.locked_at else None,
            "created_at": appel.created_at.isoformat() if getattr(appel, "created_at", None) else None,
            "updated_at": appel.updated_at.isoformat() if getattr(appel, "updated_at", None) else None,
        },
    )


def _build_filtered_appels_queryset(request):
    appels_qs = Appel.objects.filter(is_active=True)
    completed_answers_filter = appel_answers_completed_q("answers__")
    modified_answers_filter = appel_answers_modified_completion_q("answers__")
    tracked_audio_filter = Q(audio_file__isnull=False) & ~Q(audio_file="")
    status_filter = request.GET.get("status") or ""
    prestataire_filter = request.GET.get("prestataire") or ""
    beneficiaire_filter = request.GET.get("beneficiaire") or ""
    classe_filter = request.GET.get("classe") or ""
    fenetre_filter = request.GET.get("fenetre") or ""
    agent_filter = request.GET.get("agent") or ""
    formulaire_filter = request.GET.get("formulaire") or ""
    modified_by_filter = request.GET.get("modified_by") or ""
    tracking_termine_filter = (request.GET.get("tracking_termine") or "").strip()
    tracking_user_filter = (request.GET.get("tracking_user") or "").strip()
    taux_filter = request.GET.get("taux_min", "").strip()
    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str = request.GET.get("date_to", "").strip()
    search = request.GET.get("q", "").strip()

    if status_filter:
        appels_qs = appels_qs.filter(status=status_filter)
    if prestataire_filter:
        appels_qs = appels_qs.filter(prestataire__icontains=prestataire_filter)
    if beneficiaire_filter:
        appels_qs = appels_qs.filter(beneficiaire__icontains=beneficiaire_filter)
    if classe_filter:
        appels_qs = appels_qs.filter(classe_label__icontains=classe_filter)
    normalized_fenetre_filter = _normalize_dashboard_fenetre(fenetre_filter)
    if fenetre_filter:
        if normalized_fenetre_filter:
            matching_ids = [
                pk
                for pk, value in appels_qs.values_list("pk", "fenetre")
                if _normalize_dashboard_fenetre(value) == normalized_fenetre_filter
            ]
            appels_qs = appels_qs.filter(pk__in=matching_ids)
        else:
            appels_qs = appels_qs.filter(fenetre__icontains=fenetre_filter)
    if agent_filter:
        appels_qs = appels_qs.filter(locked_by__username__iexact=agent_filter)
    if formulaire_filter == "rempli":
        appels_qs = appels_qs.filter(completed_answers_filter)
    elif formulaire_filter == "vide":
        appels_qs = appels_qs.exclude(completed_answers_filter)
    elif formulaire_filter == "modifie":
        appels_qs = appels_qs.filter(modified_answers_filter)
    if modified_by_filter:
        appels_qs = appels_qs.filter(answers__modified_by__username__iexact=modified_by_filter)
    if tracking_termine_filter:
        tracking_username = tracking_user_filter or agent_filter or modified_by_filter
        if tracking_username:
            form_tracking_cutoff = padesce_form_tracking_cutoff()
            tracking_filter = (
                Q(locked_by__username__iexact=tracking_username, status="termine", updated_at__lt=form_tracking_cutoff)
                | (Q(locked_by__username__iexact=tracking_username) & completed_answers_filter)
                | (
                    Q(
                        locked_by__username__iexact=tracking_username,
                        status="termine",
                        updated_at__gte=form_tracking_cutoff,
                    )
                    & tracked_audio_filter
                )
                | (Q(answers__modified_by__username__iexact=tracking_username) & modified_answers_filter)
            )
            appels_qs = appels_qs.filter(tracking_filter).distinct()
    if taux_filter:
        try:
            seuil = Decimal(taux_filter)
            appels_qs = appels_qs.filter(taux_presence__gte=seuil)
        except Exception:
            pass
    if search:
        appels_qs = appels_qs.filter(
            Q(nom__icontains=search)
            | Q(code__icontains=search)
            | Q(telephone1__icontains=search)
            | Q(telephone2__icontains=search)
            | Q(prestataire__icontains=search)
            | Q(beneficiaire__icontains=search)
            | Q(classe_label__icontains=search)
            | Q(formation_padesce__icontains=search)
            | Q(type_formation_declaree__icontains=search)
        )
    if date_from_str:
        try:
            naive_date_from = datetime.datetime.fromisoformat(date_from_str)
            date_from = timezone.make_aware(naive_date_from)
            appels_qs = appels_qs.filter(created_at__gte=date_from)
        except (ValueError, TypeError):
            pass
    if date_to_str:
        try:
            naive_date_to = datetime.datetime.fromisoformat(date_to_str)
            date_to = timezone.make_aware(naive_date_to)
            appels_qs = appels_qs.filter(created_at__lte=date_to)
        except (ValueError, TypeError):
            pass

    appels_qs = appels_qs.annotate(
        taux_presence_display=Case(
            When(
                taux_presence__lte=1,
                then=ExpressionWrapper(F("taux_presence") * 100, output_field=DecimalField(max_digits=7, decimal_places=2)),
            ),
            default=F("taux_presence"),
            output_field=DecimalField(max_digits=7, decimal_places=2),
        )
    )

    filters = {
        "status": status_filter,
        "prestataire": prestataire_filter,
        "beneficiaire": beneficiaire_filter,
        "classe": classe_filter,
        "fenetre": normalized_fenetre_filter or fenetre_filter,
        "agent": agent_filter,
        "formulaire": formulaire_filter,
        "modified_by": modified_by_filter,
        "q": search,
        "prestataires": sorted(
            {p.strip() for p in appels_qs.exclude(prestataire="").values_list("prestataire", flat=True) if p}
        ),
        "beneficiaires": sorted(
            {b.strip() for b in appels_qs.exclude(beneficiaire="").values_list("beneficiaire", flat=True) if b}
        ),
        "classes": sorted(
            {c.strip() for c in appels_qs.exclude(classe_label="").values_list("classe_label", flat=True) if c}
        ),
        "fenetres": sorted(
            {
                normalized
                for normalized in (
                    _normalize_dashboard_fenetre(value)
                    for value in appels_qs.exclude(fenetre="").values_list("fenetre", flat=True)
                )
                if normalized
            }
        ),
        "agents": sorted(
            {
                u.strip()
                for u in appels_qs.exclude(locked_by__isnull=True).values_list("locked_by__username", flat=True)
                if u
            }
        ),
        "modified_bys": sorted(
            {
                u.strip()
                for u in appels_qs.exclude(answers__modified_by__isnull=True).values_list(
                    "answers__modified_by__username",
                    flat=True,
                )
                if u
            }
        ),
        "taux_min": taux_filter,
        "date_from": date_from_str,
        "date_to": date_to_str,
    }
    return appels_qs, filters


@login_required
def appels_export_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        "Code",
        "Nom et prenom 0 name & first name",
        "Beneficiaires",
        "Prestataire",
        "Lieux",
        "Classe",
        "1er no tel 0 tel no apprenant",
        "2e no tel 0 tel no apprenant (si disponible)",
        "Taux de presence",
        "Type formation declaree",
        "Formation Padesce",
        "Fenetre",
        "Statut",
        "Rappel at",
        "Deja forme",
        "Locked by",
        "Locked at",
        "Created at",
        "Updated at",
        "Audio file",
    ]
    ws.append(headers)
    qs = Appel.objects.select_related("locked_by").all().order_by("nom")
    for appel in qs:
        ws.append(
            [
                appel.code,
                appel.nom,
                appel.beneficiaire,
                appel.prestataire,
                appel.lieu,
                appel.classe_label,
                appel.telephone1,
                appel.telephone2,
                float(appel.taux_presence or 0),
                appel.type_formation_declaree,
                appel.formation_padesce,
                appel.fenetre,
                appel.status,
                appel.rappel_at.isoformat() if appel.rappel_at else "",
                "1" if appel.deja_forme else "0",
                appel.locked_by.username if appel.locked_by else "",
                appel.locked_at.isoformat() if appel.locked_at else "",
                appel.created_at.isoformat() if getattr(appel, "created_at", None) else "",
                appel.updated_at.isoformat() if getattr(appel, "updated_at", None) else "",
                _safe_audio_url(appel),
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="appels-export.xlsx"'
    wb.save(response)
    return response


@login_required
@transaction.atomic
def appels_index(request):
    if request.method == "POST" and request.FILES.get("file"):
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut importer des fichiers d'appels.")
            return redirect(request.path_info)
        f = request.FILES["file"]
        mode = request.POST.get("update_mode", "replace")
        try:
            file_bytes = io.BytesIO(f.read())
            if mode == "non_forme_feuil2":
                payload = _parse_excel_non_forme(file_bytes)
            else:
                payload = _parse_excel(file_bytes)
        except Exception as exc:
            messages.error(request, f"Impossible de lire le fichier : {exc}")
            return redirect(request.path_info)

        if mode in {"replace", "non_forme_feuil2", "capef_sheet1"}:
            Appel.objects.update(is_active=False)
            created = 0
            updated = 0
            for item in payload:
                code = item.get("code")
                if not code:
                    continue
                classe_obj = None
                if item["classe_label"]:
                    classe_obj = Classe.objects.filter(code=item["classe_label"]).first()
                appel = Appel.objects.filter(code=code.strip()).first()
                if appel:
                    _archive_before_import_overwrite(appel, mode)
                    for key, value in item.items():
                        setattr(appel, key, value)
                    appel.classe = classe_obj
                    appel.is_active = True
                    appel.save()
                    updated += 1
                else:
                    Appel.objects.create(
                        **item,
                        classe=classe_obj,
                        is_active=True,
                    )
                    created += 1
            deduped = _deactivate_duplicate_rows(Appel.objects.filter(is_active=True), _appel_duplicate_key)
            messages.success(
                request,
                (
                    f"Fichier importe. {created} nouveau(x) appel(s), {updated} appel(s) mis a jour. "
                    f"Affichage remplace. {deduped} doublon(s) desactive(s)."
                ),
            )
        elif mode == "append":
            created = 0
            updated = 0
            for item in payload:
                code = item.get("code")
                if not code:
                    continue
                classe_obj = None
                if item["classe_label"]:
                    classe_obj = Classe.objects.filter(code=item["classe_label"]).first()
                appel = Appel.objects.filter(code=code.strip()).first()
                if appel:
                    _archive_before_import_overwrite(appel, mode)
                    for key, value in item.items():
                        setattr(appel, key, value)
                    appel.classe = classe_obj
                    appel.is_active = True
                    appel.save()
                    updated += 1
                else:
                    Appel.objects.create(
                        **item,
                        classe=classe_obj,
                    )
                    created += 1
            _deactivate_duplicate_rows(Appel.objects.filter(is_active=True), _appel_duplicate_key)
            messages.success(request, f"Fichier importe. {created} nouveau(x) appel(s), {updated} appel(s) mis a jour.")
        else:
            updated = 0
            for item in payload:
                code = item.get("code")
                if not code:
                    continue
                appel = Appel.objects.filter(code=code.strip()).first()
                if not appel:
                    continue
                _archive_before_import_overwrite(appel, mode)
                appel.type_formation_declaree = str(item.get("type_formation_declaree") or "").strip()
                appel.formation_padesce = str(item.get("formation_padesce") or "").strip()
                appel.save(update_fields=["type_formation_declaree", "formation_padesce", "updated_at"])
                updated += 1
            messages.success(
                request,
                f"Fichier importe. {updated} appel(s) mis a jour avec le type de formation et la formation Padesce."
            )
        return redirect(request.path_info)

    appels_qs, filters = _build_filtered_appels_queryset(request)

    appels_count = appels_qs.count()
    stats = _build_progress_metrics(appels_qs)
    appels_qs = appels_qs.order_by("status", "nom")
    
    # ── Pagination: 30 lignes par page ──
    paginator = Paginator(appels_qs, 30)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    
    appels = _bind_audio_state(list(page_obj.object_list))
    page_obj.object_list = appels

    # ── per-class 50 % threshold ──
    from django.db.models import Count, Q as DQ
    classe_progress_qs = (
        Appel.objects.filter(is_active=True)
        .exclude(classe_label="")
        .values("classe_label")
        .annotate(
            total=Count("id"),
            termines=Count("id", filter=DQ(status="termine")),
        )
        .order_by("classe_label")
    )
    classe_progress = []
    for row in classe_progress_qs:
        total = row["total"]
        termines = row["termines"]
        target = (total + 1) // 2 if total else 0
        reached = total > 0 and termines >= target
        classe_progress.append({
            "classe": row["classe_label"],
            "total": total,
            "termines": termines,
            "target": target,
            "reached": reached,
            "pct": round((termines / total) * 100, 1) if total else 0,
        })

    # Map progress to class filter options
    progress_map = {item["classe"]: item for item in classe_progress}
    enriched_classes = []
    for c_label in filters["classes"]:
        prog = progress_map.get(c_label)
        label_with_prog = c_label
        if prog:
            if prog["reached"]:
                label_with_prog = f"{c_label} (Objectif Atteint ✅)"
            else:
                label_with_prog = f"{c_label} ({prog['termines']}/{prog['total']} - {prog['pct']}%)"
        enriched_classes.append({"value": c_label, "label": label_with_prog})
    filters["classes_enriched"] = enriched_classes

    import json as _json
    return render(
        request,
        "appels/index.html",
        {
            "appels": appels,
            "page_obj": page_obj,
            "paginator": paginator,
            "filters": filters,
            "appels_count": appels_count,
            "stats": stats,
            "classe_progress": classe_progress,
            "classe_progress_json": _json.dumps(classe_progress),
        },
    )


@login_required
def appels_export_filtered_csv(request):
    appels_qs, _ = _build_filtered_appels_queryset(request)
    appels = appels_qs.select_related("locked_by").order_by("status", "nom")

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="appels-filtres.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(
        [
            "Nom",
            "Code",
            "Prestataire",
            "Beneficiaire",
            "Lieu",
            "Classe",
            "Fenetre",
            "Telephone",
            "Taux presence",
            "Statut",
            "Agent",
            "Rappel at",
            "Audio URL",
            "Created at",
            "Updated at",
        ]
    )
    for appel in appels:
        writer.writerow(
            [
                appel.nom,
                appel.code,
                appel.prestataire,
                appel.beneficiaire,
                appel.lieu,
                appel.classe_label,
                appel.fenetre,
                appel.telephone1 or appel.telephone2 or "",
                float(getattr(appel, "taux_presence_display", appel.taux_presence or 0)),
                appel.get_status_display(),
                appel.locked_by.username if appel.locked_by else "",
                appel.rappel_at.isoformat() if appel.rappel_at else "",
                _safe_audio_url(appel),
                appel.created_at.isoformat() if getattr(appel, "created_at", None) else "",
                appel.updated_at.isoformat() if getattr(appel, "updated_at", None) else "",
            ]
        )
    return response




def _cleanup_stale_locks(timeout_minutes=3):
    """Release locks older than timeout_minutes to prevent stuck calls.
    
    Lock timeouts (very aggressive to handle network failures):
    - en_cours: 3 minutes max (dropped connections should release within 3min)
    - pause: 2 minutes max (paused calls should resume or terminate quickly)
    """
    from App_PADESCE.appels.models import Appel, AppelCGA, AppelFormateur
    
    # Cleanup for en_cours locks (3 min)
    cutoff_active = timezone.now() - datetime.timedelta(minutes=timeout_minutes)
    Appel.objects.filter(
        locked_by__isnull=False,
        locked_at__lt=cutoff_active,
        status='en_cours'
    ).update(locked_by=None, locked_at=None, status='pause')
    
    AppelCGA.objects.filter(
        locked_by__isnull=False,
        locked_at__lt=cutoff_active,
        status='en_cours'
    ).update(locked_by=None, locked_at=None, status='pause')
    
    AppelFormateur.objects.filter(
        locked_by__isnull=False,
        locked_at__lt=cutoff_active,
        status='en_cours'
    ).update(locked_by=None, locked_at=None, status='pause')
    
    # Cleanup for pause locks (2 min)
    cutoff_pause = timezone.now() - datetime.timedelta(minutes=2)
    Appel.objects.filter(
        locked_by__isnull=False,
        locked_at__lt=cutoff_pause,
        status='pause'
    ).update(locked_by=None, locked_at=None)
    
    AppelCGA.objects.filter(
        locked_by__isnull=False,
        locked_at__lt=cutoff_pause,
        status='pause'
    ).update(locked_by=None, locked_at=None)
    
    AppelFormateur.objects.filter(
        locked_by__isnull=False,
        locked_at__lt=cutoff_pause,
        status='pause'
    ).update(locked_by=None, locked_at=None)


def _handle_lock_conflict(instance, current_user, action):
    """Handle lock conflicts with stale lock recovery.
    
    After cleanup runs, any remaining lock is less than 3 min old.
    Allow takeover of very recent locks (< 2 min) to recover from network failures.
    
    Returns: (is_ok, error_message)
    """
    if not instance.locked_by or instance.locked_by == current_user:
        return True, None
    
    if instance.status not in ('en_cours', 'pause'):
        return True, None
    
    # Check if lock is recent enough to takeover
    now = timezone.now()
    if instance.locked_at:
        age = now - instance.locked_at
        
        # Ultra-aggressive: takeover any lock older than 90 seconds
        # This prevents stuck calls from blocking other agents
        very_stale_delta = datetime.timedelta(seconds=90)
        if age > very_stale_delta:
            return True, None
        
        # Lock is very recent - block with helpful message
        remaining_secs = max(1, int(90 - age.total_seconds()))
        return False, f"Appel actif. Veuillez essayer dans {remaining_secs}s ou contacter l'agent: {instance.locked_by.username or 'N/A'}"
    
    # No lock timestamp - allow
    return True, None


def _check_other_active_calls(user, current_instance):
    from App_PADESCE.appels.models import Appel, AppelCGA, AppelFormateur
    has_padesce = Appel.objects.filter(locked_by=user, status__in=['en_cours', 'pause']).exclude(pk=current_instance.pk if isinstance(current_instance, Appel) else None).exists()
    has_cga = AppelCGA.objects.filter(locked_by=user, status__in=['en_cours', 'pause']).exclude(pk=current_instance.pk if isinstance(current_instance, AppelCGA) else None).exists()
    has_formateur = AppelFormateur.objects.filter(locked_by=user, status__in=['en_cours', 'pause']).exclude(pk=current_instance.pk if isinstance(current_instance, AppelFormateur) else None).exists()
    return has_padesce or has_cga or has_formateur

@login_required
@require_POST
def appel_action(request, pk: int):
    # Clean up stale locks that may be blocking access
    _cleanup_stale_locks()
    _cleanup_stale_locks()
    
    appel = get_object_or_404(Appel, pk=pk)
    action = request.POST.get("action")
    rappel_at = request.POST.get("rappel_at")
    deja_forme_flag = _parse_bool_flag(request.POST.get("deja_forme"))

    now = timezone.now()
    satisfaction_saved = False
    satisfaction_message = ""

    if action == "start":
        appel.status = "en_cours"
        appel.locked_by = request.user
        appel.locked_at = now
    elif action == "pause":
        appel.status = "pause"
    elif action == "resume":
        appel.status = "en_cours"
        appel.locked_by = request.user
        appel.locked_at = now
    elif action == "rappeler":
        appel.status = "a_rappeler"
        appel.locked_by = request.user
        appel.locked_at = now
        if rappel_at:
            try:
                naive_dt = datetime.datetime.fromisoformat(rappel_at)
                appel.rappel_at = timezone.make_aware(naive_dt)
            except (ValueError, TypeError):
                appel.rappel_at = None
    elif action == "terminer":
        appel.status = "termine"
    else:
        return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)

    if deja_forme_flag is not None:
        appel.deja_forme = deja_forme_flag

    # Save flag fields
    update_fields = ["status", "locked_by", "locked_at", "rappel_at", "deja_forme", "updated_at"]
    for flag_name in ("flag_pas_forme", "flag_faux_nom", "flag_vrai_nom", "flag_deja_appele", "flag_numero_double"):
        val = request.POST.get(flag_name)
        if val is not None:
            if flag_name == "flag_vrai_nom":
                setattr(appel, flag_name, val)
            else:
                setattr(appel, flag_name, _parse_bool_flag(val) or False)
            update_fields.append(flag_name)

    appel.save(update_fields=update_fields)
    # Progress info for the UI
    class_info = None
    if appel.classe_label:
        q_count = Appel.objects.filter(classe_label=appel.classe_label, is_active=True)
        total_c = q_count.count()
        termines_c = q_count.filter(status="termine").count()
        target_c = (total_c + 1) // 2
        reached_c = total_c > 0 and termines_c >= target_c
        class_info = {
            "label": appel.classe_label,
            "total": total_c,
            "termines": termines_c,
            "target": target_c,
            "reached": reached_c,
            "pct": round((termines_c / total_c) * 100, 1) if total_c else 0,
        }

    return JsonResponse(
        {
            "ok": True,
            "status": appel.status,
            "status_label": appel.get_status_display(),
            "locked_by": appel.locked_by.username if appel.locked_by else "",
            "rappel_at": appel.rappel_at.isoformat() if appel.rappel_at else "",
            "class_progress": class_info,
        }
    )


@login_required
@require_POST
def finalize_appel(request, pk: int):
    """Save questionnaire answers first, then attach audio if one was provided."""
    _cleanup_stale_locks()
    
    try:
        with transaction.atomic():
            appel = Appel.objects.select_for_update().get(pk=pk)
            action = request.POST.get("action", "terminer")
            file_obj = request.FILES.get("audio")

            if action == "terminer":
                appel.status = "termine"
                appel.rappel_at = None
            elif action == "rappeler":
                appel.status = "a_rappeler"
                rappel_at = request.POST.get("rappel_at")
                if rappel_at:
                    try:
                        naive_dt = datetime.datetime.fromisoformat(rappel_at)
                        appel.rappel_at = timezone.make_aware(naive_dt)
                    except (ValueError, TypeError):
                        appel.rappel_at = None
                else:
                    appel.rappel_at = None

            update_fields = ["status", "updated_at", "rappel_at"]
            deja_forme_flag = _parse_bool_flag(request.POST.get("deja_forme")) or False
            appel.deja_forme = deja_forme_flag
            update_fields.append("deja_forme")

            for flag_name in ("flag_pas_forme", "flag_faux_nom", "flag_vrai_nom", "flag_deja_appele", "flag_numero_double"):
                val = request.POST.get(flag_name)
                if val is not None:
                    if flag_name == "flag_vrai_nom":
                        setattr(appel, flag_name, val)
                    else:
                        setattr(appel, flag_name, _parse_bool_flag(val) or False)
                    update_fields.append(flag_name)

            appel.save(update_fields=list(dict.fromkeys(update_fields)))

            satisfaction_saved = False
            satisfaction_message = ""
            if action == "terminer":
                commentaire_val = request.POST.get("commentaire", "") or "RAS"
                recommandations_val = request.POST.get("recommandations", "") or "RAS"
                # IMPORTANT: Toujours envoyer les réponses (même avec valeurs défaut)
                manual_data = {
                    "q1_clarte_exposes": request.POST.get("q1") or "3",
                    "q2_interaction_formateur": request.POST.get("q2") or "3",
                    "q3_maitrise_contenu": request.POST.get("q3") or "3",
                    "q4_salle_adequate": request.POST.get("q4") or "3",
                    "q5_materiel_disponible": request.POST.get("q5") or "3",
                    "q6_organisation_temps": request.POST.get("q6") or "3",
                    "q7_utilite_formation": request.POST.get("q7") or "3",
                    "q8_adequation_besoins": request.POST.get("q8") or "3",
                    "q9_satisfaction_globale": request.POST.get("q9") or "3",
                    "commentaire": commentaire_val.strip() or "RAS",
                    "recommandations": recommandations_val.strip() or "RAS",
                }
                auto_result = _auto_process_satisfaction_from_appel(appel, request.user, manual_data=manual_data)
                satisfaction_saved = auto_result.get("satisfaction_saved", False)
                satisfaction_message = auto_result.get("message", "")
                satisfaction_id = auto_result.get("satisfaction_id")
            else:
                satisfaction_id = None

            if file_obj:
                appel.audio_file = file_obj
                appel.save(update_fields=["audio_file", "updated_at"])
                if satisfaction_id:
                    satisfaction = SatisfactionApprenant.objects.filter(pk=satisfaction_id).first()
                    _attach_appel_audio_to_satisfaction(satisfaction, appel)
            
            # 5. Get class progress
            class_info = None
            if appel.classe_label:
                q_count = Appel.objects.filter(classe_label=appel.classe_label, is_active=True)
                total_c = q_count.count()
                termines_c = q_count.filter(status="termine").count()
                target_c = (total_c + 1) // 2
                reached_c = total_c > 0 and termines_c >= target_c
                class_info = {
                    "label": appel.classe_label,
                    "total": total_c,
                    "termines": termines_c,
                    "target": target_c,
                    "reached": reached_c,
                    "pct": round((termines_c / total_c) * 100, 1) if total_c else 0,
                }
            
            audio_url = _safe_audio_url(appel)
            return JsonResponse({
                "ok": True,
                "status": appel.status,
                "status_label": appel.get_status_display(),
                "audio_saved": bool(file_obj),
                "audio_url": audio_url,
                "satisfaction_saved": satisfaction_saved,
                "satisfaction_message": satisfaction_message,
                "class_progress": class_info,
            })
    except Appel.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Appel introuvable."}, status=404)
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Erreur lors de la finalisation: {str(e)}"}, status=500)


@login_required
@require_POST
def appel_upload_audio(request, pk: int):
    appel = get_object_or_404(Appel, pk=pk)
    file_obj = request.FILES.get("audio")
    if not file_obj:
        return JsonResponse({"ok": False, "error": "Aucun fichier audio."}, status=400)
    appel.audio_file = file_obj

    # Save flag fields
    update_fields = ["audio_file", "updated_at"]
    for flag_name in ("flag_pas_forme", "flag_faux_nom", "flag_vrai_nom", "flag_deja_appele", "flag_numero_double"):
        val = request.POST.get(flag_name)
        if val is not None:
            if flag_name == "flag_vrai_nom":
                setattr(appel, flag_name, val)
            else:
                setattr(appel, flag_name, _parse_bool_flag(val) or False)
            update_fields.append(flag_name)

    appel.save(update_fields=update_fields)
    payload = {"ok": True, "audio_saved": True, "audio_url": _safe_audio_url(appel)}
    if appel.status == "termine":
        commentaire_val = request.POST.get("commentaire", "") or ""
        recommandations_val = request.POST.get("recommandations", "") or ""
        manual_data = {
            "q1_clarte_exposes": request.POST.get("q1"),
            "q2_interaction_formateur": request.POST.get("q2"),
            "q3_maitrise_contenu": request.POST.get("q3"),
            "q4_salle_adequate": request.POST.get("q4"),
            "q5_materiel_disponible": request.POST.get("q5"),
            "q6_organisation_temps": request.POST.get("q6"),
            "q7_utilite_formation": request.POST.get("q7"),
            "q8_adequation_besoins": request.POST.get("q8"),
            "q9_satisfaction_globale": request.POST.get("q9"),
            "commentaire": commentaire_val.strip() or "RAS",
            "recommandations": recommandations_val.strip() or "RAS",
        }
        manual_data = {k: v for k, v in manual_data.items() if v is not None and v != ""}
        auto_result = _auto_process_satisfaction_from_appel(appel, request.user, manual_data=manual_data)
        payload.update(
            {
                "satisfaction_saved": auto_result.get("satisfaction_saved", auto_result.get("ok", False)),
                "satisfaction_message": auto_result["message"],
            }
        )
    # Progress info for the UI
    class_info = None
    if appel.classe_label:
        q_count = Appel.objects.filter(classe_label=appel.classe_label, is_active=True)
        total_c = q_count.count()
        termines_c = q_count.filter(status="termine").count()
        target_c = (total_c + 1) // 2
        reached_c = total_c > 0 and termines_c >= target_c
        class_info = {
            "label": appel.classe_label,
            "total": total_c,
            "termines": termines_c,
            "target": target_c,
            "reached": reached_c,
            "pct": round((termines_c / total_c) * 100, 1) if total_c else 0,
        }
    payload["class_progress"] = class_info

    return JsonResponse(payload)


def _auto_process_satisfaction_from_appel(appel: Appel, user, manual_data: dict = None) -> dict:
    """
    Enregistrer les réponses au questionnaire de satisfaction de l'apprenant.
    Priorité: rattacher d'abord les réponses à la ligne d'appel, puis compléter la fiche de satisfaction.
    """
    if manual_data is None:
        manual_data = {}
    _save_appel_answers(appel, user, manual_data, apply_defaults=True)
    base_qs = Apprenant.objects.all()
    scoped_qs = base_qs
    if appel.classe_id:
        scoped_qs = scoped_qs.filter(classe_id=appel.classe_id)
    elif appel.classe_label:
        scoped_qs = scoped_qs.filter(classe__code__iexact=str(appel.classe_label).strip())

    apprenant = _find_apprenant_for_appel(scoped_qs, appel)
    if not apprenant:
        apprenant = _find_apprenant_for_appel(base_qs, appel)

    satisfaction = _save_satisfaction_for_appel(appel, user, manual_data, apprenant)
    if not apprenant:
        return {
            "ok": True,
            "satisfaction_saved": True,
            "satisfaction_id": satisfaction.id,
            "message": (
                f"Reponses enregistrees et rattachees a la ligne d'appel {_appel_reference_details(appel)}. "
                "Referentiel apprenant introuvable."
            ),
        }

    # Log pour analyse avec ID apprenant, classe, prestataire, bénéficiaire
    logger.info(
        "Satisfaction enregistrée. apprenant_id=%s classe=%s prestataire=%s beneficiaire=%s",
        apprenant.id if apprenant else "N/A",
        getattr(satisfaction.classe, "code", "N/A"),
        appel.prestataire or "N/A",
        appel.beneficiaire or "N/A"
    )
    
    return {
        "ok": True,
        "satisfaction_saved": True,
        "satisfaction_id": satisfaction.id,
        "message": "Questionnaire apprenant enregistre.",
    }


@login_required
@require_POST
def download_appel_audios(request):
    ids = request.POST.getlist("ids")
    if not ids:
        return JsonResponse({"ok": False, "error": "Aucun appel sélectionné."}, status=400)
    try:
        ids = [int(val) for val in ids]
    except ValueError:
        return JsonResponse({"ok": False, "error": "Identifiants invalides."}, status=400)

    appels = list(
        Appel.objects.filter(pk__in=ids, audio_file__isnull=False)
        .order_by("nom")
    )
    if not appels:
        return JsonResponse({"ok": False, "error": "Aucun audio disponible pour les appels sélectionnés."}, status=404)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        written = 0
        for appel in appels:
            if not _has_audio_file(appel):
                continue
            try:
                with appel.audio_file.open("rb") as audio:
                    suffix = Path(appel.audio_file.name).suffix or ".mp3"
                    safe_name = f"{slugify(appel.code) or 'code'}-{slugify(appel.nom) or 'appel'}{suffix}"
                    archive.writestr(safe_name, audio.read())
                    written += 1
            except Exception:
                continue
        if written == 0:
            return JsonResponse({"ok": False, "error": "Pas d'audio récupérable."}, status=404)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="appels-audios.zip"'
    return response


@login_required
@require_POST
@transaction.atomic
def deduplicate_all_call_tables(request):
    if not request.user.is_superuser:
        messages.error(request, "Seul un superadmin peut nettoyer les doublons.")
        return redirect(request.META.get("HTTP_REFERER") or "/dashboard/")

    appels_removed = _deactivate_duplicate_rows(Appel.objects.filter(is_active=True), _appel_duplicate_key)
    cga_removed = _deactivate_duplicate_rows(AppelCGA.objects.filter(is_active=True), _cga_duplicate_key)
    formateurs_removed = _deactivate_duplicate_rows(
        AppelFormateur.objects.filter(is_active=True),
        _formateur_duplicate_key,
    )
    total_removed = appels_removed + cga_removed + formateurs_removed
    messages.success(
        request,
        (
            f"Nettoyage doublons termine. {total_removed} ligne(s) desactivee(s) "
            f"(PADESCE: {appels_removed}, CGA: {cga_removed}, Formateurs: {formateurs_removed})."
        ),
    )
    return redirect(request.META.get("HTTP_REFERER") or "/dashboard/")








@login_required
@require_POST
def appel_transcription_detail(request, pk: int):
    appel = get_object_or_404(Appel, pk=pk)
    try:
        obj, generated = _ensure_appel_transcription(appel)
        return JsonResponse({"ok": True, "generated": generated, "transcription": _transcription_to_payload(obj)})
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)



@login_required
def appel_answers_detail(request, pk: int):
    """GET: return appel info + existing answers. POST: update answers with audit trail."""
    appel = get_object_or_404(Appel, pk=pk)
    if request.method == "POST":
        answers, _ = AppelAnswers.objects.get_or_create(appel=appel)
        q_fields = [
            "q1_clarte_exposes", "q2_interaction_formateur", "q3_maitrise_contenu",
            "q4_salle_adequate", "q5_materiel_disponible", "q6_organisation_temps",
            "q7_utilite_formation", "q8_adequation_besoins", "q9_satisfaction_globale",
        ]
        for i, field in enumerate(q_fields, 1):
            val = request.POST.get(f"q{i}")
            if val is not None and val != "":
                try:
                    setattr(answers, field, int(val))
                except (ValueError, TypeError):
                    pass
        comm = request.POST.get("commentaire", "")
        reco = request.POST.get("recommandations", "")
        answers.commentaire = comm.strip() or "RAS"
        answers.recommandations = reco.strip() or "RAS"
        answers.modified_by = request.user
        answers.modified_at = timezone.now()
        answers.save()

        # Also save flags on the appel
        update_fields = ["updated_at"]
        for flag_name in ("flag_pas_forme", "flag_faux_nom", "flag_vrai_nom", "flag_deja_appele", "flag_numero_double", "deja_forme"):
            val = request.POST.get(flag_name)
            if val is not None:
                if flag_name in ("flag_vrai_nom",):
                    setattr(appel, flag_name, val)
                else:
                    setattr(appel, flag_name, _parse_bool_flag(val) or False)
                update_fields.append(flag_name)
        if len(update_fields) > 1:
            appel.save(update_fields=update_fields)

        return JsonResponse({"ok": True, "message": "Réponses mises à jour."})

    # GET
    answers = getattr(appel, "answers", None)
    data = {
        "ok": True,
        "appel": {
            "id": appel.id,
            "nom": appel.nom,
            "code": appel.code,
            "telephone1": appel.telephone1,
            "telephone2": appel.telephone2,
            "prestataire": appel.prestataire,
            "beneficiaire": appel.beneficiaire,
            "lieu": appel.lieu,
            "classe_label": appel.classe_label,
            "status": appel.status,
            "status_label": appel.get_status_display(),
            "deja_forme": appel.deja_forme,
            "flag_pas_forme": appel.flag_pas_forme,
            "flag_faux_nom": appel.flag_faux_nom,
            "flag_vrai_nom": appel.flag_vrai_nom,
            "flag_deja_appele": appel.flag_deja_appele,
            "flag_numero_double": appel.flag_numero_double,
            "locked_by": appel.locked_by.username if appel.locked_by else "",
        },
        "answers": None,
    }
    if answers:
        data["answers"] = {
            "q1": answers.q1_clarte_exposes,
            "q2": answers.q2_interaction_formateur,
            "q3": answers.q3_maitrise_contenu,
            "q4": answers.q4_salle_adequate,
            "q5": answers.q5_materiel_disponible,
            "q6": answers.q6_organisation_temps,
            "q7": answers.q7_utilite_formation,
            "q8": answers.q8_adequation_besoins,
            "q9": answers.q9_satisfaction_globale,
            "commentaire": answers.commentaire or "",
            "recommandations": answers.recommandations or "",
            "modified_by": answers.modified_by.username if answers.modified_by else "",
            "modified_at": answers.modified_at.isoformat() if answers.modified_at else "",
        }
    return JsonResponse(data)
