import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
wb = openpyxl.load_workbook(
    r"docs\data\network_excel_cache\network-fichier-consolide.xlsm", read_only=True, data_only=True
)

target_phones = ["687379404", "697030975", "670594810"]

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- Scanning Sheet: {sheet_name} ---")
    headers = [str(c.value) if c.value else "" for c in next(sheet.iter_rows())]
    for i, header in enumerate(headers):
        if header:
            print(f"Col {i}: {header}")

    match_count = 0
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        row_str = " | ".join([str(c) for c in row if c is not None])
        if any(phone in row_str for phone in target_phones):
            print(f"Row {row_idx + 1}: {row_str}")
            match_count += 1
            if match_count > 5:
                break
    print(f"Found {match_count} matches in {sheet_name}")
