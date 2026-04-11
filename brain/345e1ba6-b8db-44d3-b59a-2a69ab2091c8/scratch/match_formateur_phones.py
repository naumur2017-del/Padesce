import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")
wb = openpyxl.load_workbook(
    r"docs\data\network_excel_cache\network-fichier-consolide.xlsm", read_only=True, data_only=True
)

# Collect all Beneficiaires phones
benef_sheet = wb["Beneficiaires"]
phone_to_benef = {}
for row in benef_sheet.iter_rows(min_row=2, values_only=True):
    benef_name = row[1]
    phone = str(row[3]).strip() if row[3] else None
    if phone and benef_name:
        phone_to_benef[phone] = benef_name

print("Phones in Beneficiaires:", len(phone_to_benef))

# Collect Formateur phones from Consolidation
cons_sheet = wb["Consolidation"]
formateurs = set()
for row in cons_sheet.iter_rows(min_row=2, values_only=True):
    tel = str(row[24]).strip() if row[24] else None
    if tel and tel != "None":
        formateurs.add(tel)

print("Unique formateur phones in Consolidation:", len(formateurs))

# Match
matches = 0
for tel in formateurs:
    if tel in phone_to_benef:
        print(f"Match: {tel} => {phone_to_benef[tel]}")
        matches += 1

print(f"Found {matches} matches out of {len(formateurs)} formateurs")
