from __future__ import annotations

from collections import Counter
from pathlib import Path

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from App_PADESCE.appels.models import Appel, AppelAnswers, CALL_SUCCESS_STATUSES


QUESTION_FIELDS = (
    "q1_clarte_exposes",
    "q2_interaction_formateur",
    "q3_maitrise_contenu",
    "q4_salle_adequate",
    "q5_materiel_disponible",
    "q6_organisation_temps",
    "q7_utilite_formation",
    "q8_adequation_besoins",
    "q9_satisfaction_globale",
)

QUESTION_HEADERS = (
    "Q1 clarte exposes",
    "Q2 interaction formateur",
    "Q3 maitrise contenu",
    "Q4 salle adequate",
    "Q5 materiel disponible",
    "Q6 organisation temps",
    "Q7 utilite formation",
    "Q8 adequation besoins",
    "Q9 satisfaction globale",
)

SUCCESS_HEADERS = (
    "ID",
    "Code",
    "Nom apprenant",
    "Telephone 1",
    "Telephone 2",
    "Prestataire",
    "Beneficiaire",
    "Classe",
    "Fenetre",
    "Lieu",
    "Formation PADESCE",
    "Type formation declaree",
    "Statut appel",
    "Actif",
    "Date creation",
    "Derniere mise a jour",
    "Enqueteur / modif",
    "Date formulaire",
    *QUESTION_HEADERS,
    "Note moyenne / 5",
    "Commentaire",
    "Recommandations",
    "Transcription",
    "Audio",
    "Rapport complet",
)

FAILURE_HEADERS = (
    "ID",
    "Code",
    "Nom apprenant",
    "Telephone 1",
    "Telephone 2",
    "Prestataire",
    "Beneficiaire",
    "Classe",
    "Fenetre",
    "Lieu",
    "Formation PADESCE",
    "Type formation declaree",
    "Statut appel",
    "Actif",
    "Date creation",
    "Derniere mise a jour",
    "Vrai nom",
    "Deja forme",
    "Pas forme",
    "Numero double",
    "Deja appele",
    "Formulaire complet",
    "Formulaire RAS",
    "Motif echec",
    *QUESTION_HEADERS,
    "Note moyenne / 5",
    "Commentaire",
    "Recommandations",
    "Transcription",
    "Audio",
    "Rapport complet",
)

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
KPI_FILL = PatternFill("solid", fgColor="DCE6F1")
SUCCESS_FILL = PatternFill("solid", fgColor="D9EAD3")
FAIL_FILL = PatternFill("solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
SUBHEADER_FILL = PatternFill("solid", fgColor="FCE5CD")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7C0CE"),
    right=Side(style="thin", color="B7C0CE"),
    top=Side(style="thin", color="B7C0CE"),
    bottom=Side(style="thin", color="B7C0CE"),
)


def _clean_text(value) -> str:
    return str(value or "").strip()


def _safe_audio_name(appel: Appel) -> str:
    audio = getattr(appel, "audio_file", None)
    return getattr(audio, "name", "") or ""


def _safe_related(obj, attr_name: str):
    try:
        return getattr(obj, attr_name)
    except Exception:
        return None


def _is_ras_or_blank(value) -> bool:
    text = _clean_text(value)
    return not text or text.upper() == "RAS"


def _answer_complete(answer: AppelAnswers | None) -> bool:
    return bool(answer) and all(
        getattr(answer, field, None) is not None for field in QUESTION_FIELDS
    )


def _answer_values(answer: AppelAnswers | None) -> list[int | None]:
    if not answer:
        return [None] * len(QUESTION_FIELDS)
    return [getattr(answer, field, None) for field in QUESTION_FIELDS]


def _average_score(answer: AppelAnswers | None) -> float | None:
    values = [value for value in _answer_values(answer) if value is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def _is_form_ras(answer: AppelAnswers | None) -> bool:
    return (
        bool(answer)
        and _is_ras_or_blank(answer.commentaire)
        and _is_ras_or_blank(answer.recommandations)
    )


def _comment_without_answers(answer: AppelAnswers | None) -> bool:
    if not answer or _answer_complete(answer):
        return False
    return not _is_ras_or_blank(answer.commentaire) or not _is_ras_or_blank(answer.recommandations)


def _failure_reasons(appel: Appel, answer: AppelAnswers | None) -> list[str]:
    reasons: list[str] = []
    if appel.deja_forme:
        reasons.append("Deja forme")
    if appel.flag_pas_forme:
        reasons.append("Pas forme")
    if appel.flag_faux_nom or _clean_text(appel.flag_vrai_nom):
        vrai_nom = _clean_text(appel.flag_vrai_nom)
        reasons.append(f"Faux nom / vrai nom: {vrai_nom}" if vrai_nom else "Faux nom")
    if appel.flag_numero_double:
        reasons.append("Numero double")
    if appel.status not in CALL_SUCCESS_STATUSES:
        reasons.append(f"Statut {appel.get_status_display()}")
    if not answer:
        reasons.append("Formulaire absent")
    elif not _answer_complete(answer):
        reasons.append("Formulaire incomplet")
    if _comment_without_answers(answer):
        reasons.append("Commentaire sans reponses")
    if _is_form_ras(answer):
        reasons.append("Formulaire RAS")
    if not reasons:
        reasons.append("A controler")
    return reasons


def _is_success(appel: Appel, answer: AppelAnswers | None) -> bool:
    if appel.status not in CALL_SUCCESS_STATUSES:
        return False
    if not _answer_complete(answer):
        return False
    if (
        appel.deja_forme
        or appel.flag_pas_forme
        or appel.flag_faux_nom
        or _clean_text(appel.flag_vrai_nom)
    ):
        return False
    if appel.flag_numero_double:
        return False
    if _is_form_ras(answer):
        return False
    return True


def _build_common_prefix(appel: Appel) -> list:
    modified_by = ""
    answer = _safe_related(appel, "answers")
    if answer and getattr(answer, "modified_by", None):
        modified_by = answer.modified_by.username

    return [
        appel.id,
        appel.code,
        appel.nom,
        appel.telephone1,
        appel.telephone2,
        appel.prestataire,
        appel.beneficiaire,
        appel.classe_label,
        appel.fenetre,
        appel.lieu,
        appel.formation_padesce,
        appel.type_formation_declaree,
        appel.get_status_display(),
        "Oui" if appel.is_active else "Non",
        timezone.localtime(appel.created_at).strftime("%Y-%m-%d %H:%M") if appel.created_at else "",
        timezone.localtime(appel.updated_at).strftime("%Y-%m-%d %H:%M") if appel.updated_at else "",
        modified_by,
    ]


def _build_success_report(appel: Appel, answer: AppelAnswers | None, transcription: str) -> str:
    avg = _average_score(answer)
    comment = _clean_text(getattr(answer, "commentaire", ""))
    reco = _clean_text(getattr(answer, "recommandations", ""))
    parts = [
        f"Appel PADESCE reussi pour {appel.nom} ({appel.code}).",
        f"Classe {appel.classe_label or '-'} / prestataire {appel.prestataire or '-'} / beneficiaire {appel.beneficiaire or '-'}.",
        f"Contacts: {appel.telephone1 or '-'} / {appel.telephone2 or '-'}.",
        f"Formulaire complet enregistre avec une note moyenne de {avg if avg is not None else '-'} / 5.",
        f"Commentaire: {comment or '-'}.",
        f"Recommandations: {reco or '-'}.",
    ]
    if transcription:
        parts.append(f"Transcription disponible ({len(transcription)} caracteres).")
    return " ".join(parts)


def _build_failure_report(appel: Appel, answer: AppelAnswers | None, transcription: str) -> str:
    reasons = "; ".join(_failure_reasons(appel, answer))
    comment = _clean_text(getattr(answer, "commentaire", ""))
    reco = _clean_text(getattr(answer, "recommandations", ""))
    vrai_nom = _clean_text(appel.flag_vrai_nom)
    parts = [
        f"Appel PADESCE classe en echec pour {appel.nom} ({appel.code}).",
        f"Motifs: {reasons}.",
        f"Classe {appel.classe_label or '-'} / prestataire {appel.prestataire or '-'} / beneficiaire {appel.beneficiaire or '-'}.",
        f"Contacts: {appel.telephone1 or '-'} / {appel.telephone2 or '-'}.",
    ]
    if vrai_nom:
        parts.append(f"Vrai nom renseigne: {vrai_nom}.")
    if comment or reco:
        parts.append(f"Commentaire: {comment or '-'} / recommandations: {reco or '-'}.")
    if transcription:
        parts.append(f"Transcription disponible ({len(transcription)} caracteres).")
    return " ".join(parts)


def _sheet_title(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:F2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(color="404040", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws, row_idx: int, fill: PatternFill = HEADER_FILL) -> None:
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_table(ws, header_row: int, display_name: str) -> None:
    if ws.max_row <= header_row:
        return
    ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _autosize_columns(ws, fixed_widths: dict[str, float] | None = None) -> None:
    fixed_widths = fixed_widths or {}
    for column_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(column_idx)
        if letter in fixed_widths:
            ws.column_dimensions[letter].width = fixed_widths[letter]
            continue
        max_len = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 28)


def _style_data_rows(ws, header_row: int, wrap_columns: set[int] | None = None) -> None:
    wrap_columns = wrap_columns or set()
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column in wrap_columns:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")


def _build_dashboard(
    ws,
    *,
    total_calls: int,
    successful_calls: int,
    failed_calls: int,
    complete_forms: int,
    absent_forms: int,
    incomplete_forms: int,
    ras_forms: int,
    finished_calls: int,
    active_calls: int,
    audio_calls: int,
    reason_counter: Counter,
) -> None:
    _sheet_title(
        ws,
        "Rapport PADESCE - Recap appels",
        f"Genere le {timezone.localtime().strftime('%Y-%m-%d %H:%M')} a partir de toute l'historique des appels.",
    )

    ws["A4"] = "Indicateur"
    ws["B4"] = "Valeur"
    _style_header_row(ws, 4, SUBHEADER_FILL)

    summary_rows = [
        ("Total appels", total_calls),
        ("Appels reussis", successful_calls),
        ("Appels echoues", failed_calls),
        ("Taux de reussite", None),
        ("Taux d'echec", None),
        ("Appels termines", finished_calls),
        ("Appels actifs", active_calls),
        ("Formulaires complets", complete_forms),
        ("Formulaires absents", absent_forms),
        ("Formulaires incomplets", incomplete_forms),
        ("Formulaires RAS", ras_forms),
        ("Appels avec audio", audio_calls),
    ]
    for offset, (label, value) in enumerate(summary_rows, start=5):
        ws[f"A{offset}"] = label
        if value is not None:
            ws[f"B{offset}"] = value
        ws[f"A{offset}"].border = THIN_BORDER
        ws[f"B{offset}"].border = THIN_BORDER
        if offset in (5, 6, 7):
            ws[f"A{offset}"].fill = KPI_FILL
            ws[f"B{offset}"].fill = KPI_FILL
            ws[f"A{offset}"].font = Font(bold=True)
            ws[f"B{offset}"].font = Font(bold=True)

    ws["B8"] = (successful_calls / total_calls) if total_calls else 0
    ws["B9"] = (failed_calls / total_calls) if total_calls else 0
    ws["B8"].number_format = "0.0%"
    ws["B9"].number_format = "0.0%"

    ws["D4"] = "Motif d'echec"
    ws["E4"] = "Nombre"
    _style_header_row(ws, 4, SUBHEADER_FILL)
    failure_labels = [
        "Deja forme",
        "Pas forme",
        "Faux nom",
        "Numero double",
        "Statut non termine",
        "Formulaire absent",
        "Formulaire incomplet",
        "Commentaire sans reponses",
        "Formulaire RAS",
    ]
    for index, reason in enumerate(failure_labels, start=5):
        ws[f"D{index}"] = reason
        ws[f"E{index}"] = int(reason_counter.get(reason, 0))
        ws[f"D{index}"].border = THIN_BORDER
        ws[f"E{index}"].border = THIN_BORDER

    pie = PieChart()
    pie.title = "Reussis vs echoues"
    pie.add_data(Reference(ws, min_col=2, min_row=6, max_row=7), titles_from_data=False)
    pie.set_categories(Reference(ws, min_col=1, min_row=6, max_row=7))
    pie.height = 8
    pie.width = 10
    ws.add_chart(pie, "G4")

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Principaux motifs d'echec"
    bar.y_axis.title = "Motifs"
    bar.x_axis.title = "Nombre"
    bar.add_data(Reference(ws, min_col=5, min_row=4, max_row=13), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=4, min_row=5, max_row=13))
    bar.height = 9
    bar.width = 12
    ws.add_chart(bar, "G20")

    ws["A19"] = "Regle de classement"
    ws["A20"] = (
        "Reussi = appel termine + formulaire complet exploitable + sans deja forme / faux nom / numero double / RAS."
    )
    ws["A21"] = (
        "Echoue = tous les autres cas, avec le ou les motifs detailles dans la feuille Appels echoues."
    )
    ws["A19"].font = Font(bold=True)
    ws["A20"].alignment = Alignment(wrap_text=True)
    ws["A21"].alignment = Alignment(wrap_text=True)

    ws.freeze_panes = "A5"
    _autosize_columns(ws, {"A": 28, "B": 16, "D": 28, "E": 14})


def _fill_success_sheet(ws, rows: list[list]) -> None:
    _sheet_title(ws, "Appels PADESCE reussis", "Toutes les lignes classees comme reussies.")
    ws.append([])
    ws.append(list(SUCCESS_HEADERS))
    header_row = 4
    _style_header_row(ws, header_row, SUCCESS_FILL)
    for row in rows:
        ws.append(row)
    _style_data_rows(ws, header_row, wrap_columns={29, 30, 31, 33})
    _apply_table(ws, header_row, "AppelsReussisTable")
    ws.freeze_panes = "A5"
    _autosize_columns(
        ws,
        {
            "C": 28,
            "F": 22,
            "G": 22,
            "J": 20,
            "K": 24,
            "L": 24,
            "AC": 20,
            "AD": 20,
            "AE": 40,
            "AF": 32,
            "AG": 65,
        },
    )


def _fill_failure_sheet(ws, rows: list[list]) -> None:
    _sheet_title(
        ws, "Appels PADESCE echoues", "Tous les appels classes en echec avec motifs detailes."
    )
    ws.append([])
    ws.append(list(FAILURE_HEADERS))
    header_row = 4
    _style_header_row(ws, header_row, FAIL_FILL)
    for row in rows:
        ws.append(row)
    _style_data_rows(ws, header_row, wrap_columns={24, 35})
    _apply_table(ws, header_row, "AppelsEchouesTable")
    ws.freeze_panes = "A5"
    _autosize_columns(
        ws,
        {
            "C": 28,
            "F": 22,
            "G": 22,
            "J": 20,
            "K": 24,
            "L": 24,
            "X": 36,
            "AH": 20,
            "AI": 20,
            "AJ": 40,
            "AK": 32,
            "AL": 65,
        },
    )


def build_padesce_calls_report(output_path: str | Path) -> dict:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    dashboard_ws = workbook.active
    dashboard_ws.title = "Tableau de bord"
    success_ws = workbook.create_sheet("Appels reussis")
    failure_ws = workbook.create_sheet("Appels echoues")

    success_rows: list[list] = []
    failure_rows: list[list] = []
    reason_counter: Counter = Counter()
    complete_forms = 0
    absent_forms = 0
    incomplete_forms = 0
    ras_forms = 0
    finished_calls = 0
    active_calls = 0
    audio_calls = 0

    queryset = Appel.objects.select_related(
        "answers",
        "answers__modified_by",
        "satisfaction_apprenant",
        "satisfaction_apprenant__enqueteur",
        "locked_by",
    ).order_by("created_at", "id")

    for appel in queryset.iterator(chunk_size=500):
        answer = _safe_related(appel, "answers")
        satisfaction = _safe_related(appel, "satisfaction_apprenant")
        transcription = _clean_text(getattr(satisfaction, "transcription", ""))
        answer_values = _answer_values(answer)
        avg_score = _average_score(answer)
        if appel.status in CALL_SUCCESS_STATUSES:
            finished_calls += 1
        if appel.is_active:
            active_calls += 1
        if _safe_audio_name(appel):
            audio_calls += 1
        if answer:
            if _answer_complete(answer):
                complete_forms += 1
            else:
                incomplete_forms += 1
            if _is_form_ras(answer):
                ras_forms += 1
        else:
            absent_forms += 1

        if _is_success(appel, answer):
            report_text = _build_success_report(appel, answer, transcription)
            success_rows.append(
                _build_common_prefix(appel)
                + [
                    timezone.localtime(answer.modified_at).strftime("%Y-%m-%d %H:%M")
                    if answer and answer.modified_at
                    else "",
                    *answer_values,
                    avg_score,
                    _clean_text(getattr(answer, "commentaire", "")),
                    _clean_text(getattr(answer, "recommandations", "")),
                    transcription,
                    _safe_audio_name(appel),
                    report_text,
                ]
            )
            continue

        reasons = _failure_reasons(appel, answer)
        for reason in reasons:
            if reason.startswith("Faux nom"):
                reason_counter["Faux nom"] += 1
            elif reason.startswith("Statut "):
                reason_counter["Statut non termine"] += 1
            else:
                reason_counter[reason] += 1
        report_text = _build_failure_report(appel, answer, transcription)
        failure_rows.append(
            [
                appel.id,
                appel.code,
                appel.nom,
                appel.telephone1,
                appel.telephone2,
                appel.prestataire,
                appel.beneficiaire,
                appel.classe_label,
                appel.fenetre,
                appel.lieu,
                appel.formation_padesce,
                appel.type_formation_declaree,
                appel.get_status_display(),
                "Oui" if appel.is_active else "Non",
                timezone.localtime(appel.created_at).strftime("%Y-%m-%d %H:%M")
                if appel.created_at
                else "",
                timezone.localtime(appel.updated_at).strftime("%Y-%m-%d %H:%M")
                if appel.updated_at
                else "",
                _clean_text(appel.flag_vrai_nom),
                "Oui" if appel.deja_forme else "Non",
                "Oui" if appel.flag_pas_forme else "Non",
                "Oui" if appel.flag_numero_double else "Non",
                "Oui" if appel.flag_deja_appele else "Non",
                "Oui" if _answer_complete(answer) else "Non",
                "Oui" if _is_form_ras(answer) else "Non",
                " | ".join(reasons),
                *answer_values,
                avg_score,
                _clean_text(getattr(answer, "commentaire", "")),
                _clean_text(getattr(answer, "recommandations", "")),
                transcription,
                _safe_audio_name(appel),
                report_text,
            ]
        )

    _build_dashboard(
        dashboard_ws,
        total_calls=len(success_rows) + len(failure_rows),
        successful_calls=len(success_rows),
        failed_calls=len(failure_rows),
        complete_forms=complete_forms,
        absent_forms=absent_forms,
        incomplete_forms=incomplete_forms,
        ras_forms=ras_forms,
        finished_calls=finished_calls,
        active_calls=active_calls,
        audio_calls=audio_calls,
        reason_counter=reason_counter,
    )
    _fill_success_sheet(success_ws, success_rows)
    _fill_failure_sheet(failure_ws, failure_rows)

    workbook.save(output_path)
    workbook.close()

    return {
        "path": str(output_path),
        "total_calls": len(success_rows) + len(failure_rows),
        "successful_calls": len(success_rows),
        "failed_calls": len(failure_rows),
        "complete_forms": complete_forms,
        "absent_forms": absent_forms,
        "incomplete_forms": incomplete_forms,
        "ras_forms": ras_forms,
    }
