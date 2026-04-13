from __future__ import annotations

import json
import math
import os
import re
import shutil
import unicodedata
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Iterator
from urllib.parse import parse_qs, unquote, urlparse

from django.conf import settings
from django.core.cache import cache
from django.http import JsonResponse
import logging

from django.shortcuts import render
from openpyxl import Workbook, load_workbook

from App_PADESCE.core.access import require_analysis_access

logger = logging.getLogger(__name__)

NETWORK_WORKBOOK_DIRECTORY = Path(
    r"\\192.168.1.162\naumur - travaux en cours\Projets\PROJET PADESCE\Cellule informatique\Operation"
)
NETWORK_WORKBOOK_NAME = "fichier consolidé.xlsm"
NETWORK_WORKBOOK_CUTOFF_DIRECTORY = NETWORK_WORKBOOK_DIRECTORY / "CutOff"
NETWORK_WORKBOOK_CUTOFF_NAME = "fichier consolidé CutOff.xlsm"
DEFAULT_WORKBOOK_SOURCE = "main"
PAGE_SIZE = 20

CACHE_DIRECTORY = Path(settings.BASE_DIR) / "data" / "network_excel_cache"
LOCAL_WORKBOOK_COPY = CACHE_DIRECTORY / "network-fichier-consolide.xlsm"
LOCAL_CUTOFF_WORKBOOK_COPY = CACHE_DIRECTORY / "network-fichier-consolide-cutoff.xlsm"
BUNDLED_DIRECTORY = Path(settings.BASE_DIR) / "data" / "network_excel_bundle"
BUNDLED_WORKBOOK_COPY = BUNDLED_DIRECTORY / "network-fichier-consolide.xlsm"
BUNDLED_CUTOFF_WORKBOOK_COPY = BUNDLED_DIRECTORY / "network-fichier-consolide-cutoff.xlsm"
LEGACY_BUNDLED_DIRECTORY = Path(settings.BASE_DIR) / "docs" / "data" / "network_excel_cache"
LEGACY_BUNDLED_WORKBOOK_COPY = LEGACY_BUNDLED_DIRECTORY / "network-fichier-consolide.xlsm"
LEGACY_BUNDLED_CUTOFF_WORKBOOK_COPY = (
    LEGACY_BUNDLED_DIRECTORY / "network-fichier-consolide-cutoff.xlsm"
)
DEFAULT_DESCENTE_WORKBOOK_GLOB = "DESCENTES CLASSES*.xlsx"
DESCENTE_CHANNEL_SNAPSHOT = Path(settings.BASE_DIR) / "data" / "teams_channel_links.json"
SOURCE_CACHE_TIMEOUT = int(str(os.getenv("PADESCE_SOURCE_CACHE_TIMEOUT", "900") or "900"))
SOURCE_METADATA_CACHE_TIMEOUT = int(
    str(os.getenv("PADESCE_SOURCE_METADATA_CACHE_TIMEOUT", "120") or "120")
)


@dataclass(frozen=True)
class WorkbookSource:
    key: str
    label: str
    name: str
    source_path: Path
    cached_path: Path
    size: int
    modified_at: datetime


def _workbook_source_cache_key(source_key: str) -> str:
    normalized_source_key = normalize_workbook_source_key(source_key)
    return f"padesce-workbook-source:{normalized_source_key}"


def _serialize_workbook_source(source: WorkbookSource) -> dict[str, object]:
    return {
        "key": source.key,
        "label": source.label,
        "name": source.name,
        "source_path": str(source.source_path),
        "cached_path": str(source.cached_path),
        "size": int(source.size),
        "modified_at": source.modified_at.isoformat(),
    }


def _deserialize_workbook_source(payload: object) -> WorkbookSource | None:
    if not isinstance(payload, dict):
        return None

    try:
        cached_path = Path(str(payload["cached_path"]))
        source_path = Path(str(payload["source_path"]))
        modified_at = datetime.fromisoformat(str(payload["modified_at"]))
        size = int(payload["size"])
    except (KeyError, TypeError, ValueError):
        return None

    if not cached_path.exists() or not cached_path.is_file():
        return None

    return WorkbookSource(
        key=str(payload.get("key") or DEFAULT_WORKBOOK_SOURCE),
        label=str(payload.get("label") or ""),
        name=str(payload.get("name") or cached_path.name),
        source_path=source_path,
        cached_path=cached_path,
        size=size,
        modified_at=modified_at,
    )


def _cached_workbook_source(
    source_key: str = DEFAULT_WORKBOOK_SOURCE,
) -> WorkbookSource | None:
    payload = cache.get(_workbook_source_cache_key(source_key))
    return _deserialize_workbook_source(payload)


def _normalize_lookup(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    without_marks = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return without_marks.casefold().strip()


def normalize_network_lookup(value: str) -> str:
    return _normalize_lookup(value)


def _workbook_source_configs() -> dict[str, dict]:
    return {
        "main": {
            "key": "main",
            "label": "Fichier consolide",
            "directory": NETWORK_WORKBOOK_DIRECTORY,
            "name": NETWORK_WORKBOOK_NAME,
            "env_var": "PADESCE_MAIN_WORKBOOK_PATH",
            "cached_path": LOCAL_WORKBOOK_COPY,
            "bundled_paths": [
                BUNDLED_WORKBOOK_COPY,
                LEGACY_BUNDLED_WORKBOOK_COPY,
            ],
        },
        "cutoff": {
            "key": "cutoff",
            "label": "Fichier consolide CutOff",
            "directory": NETWORK_WORKBOOK_CUTOFF_DIRECTORY,
            "name": NETWORK_WORKBOOK_CUTOFF_NAME,
            "env_var": "PADESCE_CUTOFF_WORKBOOK_PATH",
            "cached_path": LOCAL_CUTOFF_WORKBOOK_COPY,
            "bundled_paths": [
                BUNDLED_CUTOFF_WORKBOOK_COPY,
                LEGACY_BUNDLED_CUTOFF_WORKBOOK_COPY,
            ],
        },
    }


def normalize_workbook_source_key(value: str | None) -> str:
    normalized = _normalize_lookup(str(value or ""))
    for key, config in _workbook_source_configs().items():
        if normalized in {
            _normalize_lookup(key),
            _normalize_lookup(config["label"]),
            _normalize_lookup(config["name"]),
        }:
            return key
    return DEFAULT_WORKBOOK_SOURCE


def get_workbook_source_options() -> list[dict[str, str]]:
    return [
        {
            "value": config["key"],
            "label": config["label"],
            "name": config["name"],
        }
        for config in _workbook_source_configs().values()
    ]


def _workbook_source_config(source_key: str | None) -> dict:
    key = normalize_workbook_source_key(source_key)
    return _workbook_source_configs()[key]


def _display_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%d/%m/%Y")
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _resolve_configured_workbook_path(env_var_name: str) -> Path | None:
    configured_path = str(os.getenv(env_var_name, "") or "").strip()
    if not configured_path:
        return None
    candidate = Path(configured_path).expanduser()
    if candidate.exists() and candidate.is_file():
        return candidate
    return None


def _minimal_fallback_workbook_path(source_key: str = DEFAULT_WORKBOOK_SOURCE) -> Path:
    """Fichier .xlsx minimal généré localement si aucun consolidé n’est disponible."""
    normalized = normalize_workbook_source_key(source_key)
    tag = "cutoff" if normalized == "cutoff" else "main"
    return BUNDLED_DIRECTORY / f"network-fichier-consolide-{tag}-fallback.xlsx"


def _write_minimal_consolidated_workbook(path: Path) -> None:
    """Crée un classeur avec les feuilles attendues par build_padesce_source_index (en-têtes seuls)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_classes = wb.create_sheet("Classes", 0)
    ws_classes.append(
        [
            "Classe ID",
            "Prestation ID",
            "Nom du Prestataire",
            "Nom du bénéficiaire",
            "Cohorte",
            "Lieux",
            "Ville",
            "FORMATION",
            "Region",
            "Statut de la prestation",
        ]
    )

    ws_prest = wb.create_sheet("Prestations", 1)
    ws_prest.append(
        [
            "ID Prestation",
            "Prestataire",
            "Nom du bénéficiaire",
            "Formation",
            "Fenetre",
            "Region",
            "Statut de la prestation",
        ]
    )

    ws_app = wb.create_sheet("Apprenants", 2)
    ws_app.append(
        [
            "Code",
            "Classe ID",
            "Apprenant ID",
            "Nom_Individu",
            "Cohorte",
            "Telephone1",
            "Telephone2",
            "Prestation ID",
        ]
    )

    wb.save(path)


def _ensure_minimal_fallback_workbook(path: Path) -> None:
    """Garantit un classeur lisible avec Apprenants / Classes / Prestations."""
    need_write = True
    if path.exists() and path.stat().st_size > 0:
        try:
            with path.open("rb") as fh:
                probe = load_workbook(fh, read_only=True, data_only=True, keep_links=False)
                names = set(probe.sheetnames)
                probe.close()
            if {"Apprenants", "Classes", "Prestations"}.issubset(names):
                need_write = False
        except Exception as exc:
            logger.warning("Classeur de secours illisible, recréation (%s): %s", path, exc)
    if need_write:
        _write_minimal_consolidated_workbook(path)


def _resolve_network_workbook(source_key: str = DEFAULT_WORKBOOK_SOURCE) -> Path:
    config = _workbook_source_config(source_key)
    expected_name = _normalize_lookup(config["name"])
    workbook_directory = config["directory"]
    cached_path = config["cached_path"]
    configured_path = _resolve_configured_workbook_path(config["env_var"])

    if configured_path is not None:
        return configured_path

    if workbook_directory.exists():
        for child in sorted(
            workbook_directory.iterdir(),
            key=lambda item: item.name.casefold(),
        ):
            if not child.is_file():
                continue
            if child.name.startswith("~$"):
                continue
            if child.suffix.lower() not in {".xlsm", ".xlsx"}:
                continue
            if _normalize_lookup(child.name) == expected_name:
                return child

    if cached_path.exists():
        return cached_path

    for bundled_path in config.get("bundled_paths", []):
        if bundled_path.exists() and bundled_path.is_file():
            return bundled_path

    try:
        fallback_path = _minimal_fallback_workbook_path(source_key)
        _ensure_minimal_fallback_workbook(fallback_path)
        if fallback_path.is_file():
            logger.warning(
                "Classeur consolidé « %s » introuvable pour la source %s. "
                "Gabarit vide utilisé : %s. Définissez la variable %s ou placez le fichier "
                "dans le cache (data/network_excel_cache/) pour les données réelles.",
                config["name"],
                normalize_workbook_source_key(source_key),
                fallback_path,
                config["env_var"],
            )
            return fallback_path
    except Exception:
        logger.exception(
            "Impossible de créer le classeur de secours pour la source %s",
            normalize_workbook_source_key(source_key),
        )

    raise FileNotFoundError(
        f"Le fichier Excel consolide attendu ({config['name']}) n'a ete trouve ni sur le partage reseau, ni dans le cache local, ni dans le fallback embarque."
    )


def _ensure_cached_workbook(
    source_key: str = DEFAULT_WORKBOOK_SOURCE,
    force_refresh: bool = False,
) -> WorkbookSource:
    normalized_source_key = normalize_workbook_source_key(source_key)
    if force_refresh:
        cache.delete(_workbook_source_cache_key(normalized_source_key))
    else:
        cached_source = _cached_workbook_source(normalized_source_key)
        if cached_source is not None:
            return cached_source

    config = _workbook_source_config(normalized_source_key)
    cached_path = config["cached_path"]
    source_path = _resolve_network_workbook(source_key=normalized_source_key)
    source_stat = source_path.stat()

    if source_path != cached_path:
        needs_copy = force_refresh or not cached_path.exists()
        if not needs_copy:
            try:
                local_stat = cached_path.stat()
                needs_copy = (
                    local_stat.st_size != source_stat.st_size
                    or local_stat.st_mtime_ns != source_stat.st_mtime_ns
                )
            except OSError:
                needs_copy = True

        if needs_copy:
            try:
                CACHE_DIRECTORY.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source_path, cached_path)
            except Exception as copy_exc:
                logger.warning(
                    "Impossible de copier le classeur vers le cache (%s → %s): %s. "
                    "Ouverture directe depuis la source.",
                    source_path,
                    cached_path,
                    copy_exc,
                )
                # Utiliser la source directement si la copie ou la création du
                # répertoire cache échoue (droits, filesystem en lecture seule…).
                cached_path = source_path

    workbook_source = WorkbookSource(
        key=config["key"],
        label=config["label"],
        name=config["name"] if source_path == cached_path else source_path.name,
        source_path=source_path,
        cached_path=cached_path,
        size=source_stat.st_size,
        modified_at=datetime.fromtimestamp(source_stat.st_mtime),
    )
    cache.set(
        _workbook_source_cache_key(normalized_source_key),
        _serialize_workbook_source(workbook_source),
        timeout=SOURCE_METADATA_CACHE_TIMEOUT,
    )
    return workbook_source


@contextmanager
def _open_cached_workbook(
    source_key: str = DEFAULT_WORKBOOK_SOURCE,
    force_refresh: bool = False,
) -> Iterator[tuple[WorkbookSource, object]]:
    source = _ensure_cached_workbook(source_key=source_key, force_refresh=force_refresh)
    workbook_file = source.cached_path.open("rb")
    workbook = None
    try:
        workbook = load_workbook(
            workbook_file,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        yield source, workbook
    finally:
        if workbook is not None:
            workbook.close()
        workbook_file.close()


def _build_headers(raw_header: tuple | list | None, column_count: int) -> list[str]:
    raw_header = list(raw_header or [])
    headers: list[str] = []
    for index in range(column_count):
        label = _display_cell(raw_header[index]) if index < len(raw_header) else ""
        headers.append(label or f"Colonne {index + 1}")
    return headers


def _build_row(raw_row: tuple | list | None, column_count: int) -> list[str]:
    raw_row = list(raw_row or [])
    cells = [
        _display_cell(raw_row[index]) if index < len(raw_row) else ""
        for index in range(column_count)
    ]
    return cells


def _sheet_header_lookup(raw_header: tuple | list | None) -> dict[str, int]:
    header = list(raw_header or [])
    lookup: dict[str, int] = {}
    for index, value in enumerate(header):
        normalized = _normalize_lookup(_display_cell(value))
        if normalized:
            lookup[normalized] = index
    return lookup


def _sheet_get(row: tuple | list | None, header_lookup: dict[str, int], *headers: str) -> str:
    cells = list(row or [])
    for header in headers:
        index = header_lookup.get(_normalize_lookup(header))
        if index is None or index >= len(cells):
            continue
        return _display_cell(cells[index])
    return ""


def _sheet_get_first(
    row: tuple | list | None, raw_header: tuple | list | None, *headers: str
) -> str:
    cells = list(row or [])
    normalized_targets = {_normalize_lookup(header) for header in headers if header}
    for index, header_value in enumerate(list(raw_header or [])):
        if _normalize_lookup(_display_cell(header_value)) not in normalized_targets:
            continue
        if index >= len(cells):
            continue
        return _display_cell(cells[index])
    return ""


def _resolve_descente_workbook() -> Path | None:
    configured_path = str(os.getenv("PADESCE_DESCENTE_WORKBOOK_PATH", "") or "").strip()
    if configured_path:
        candidate = Path(configured_path).expanduser()
        if candidate.exists() and candidate.is_file():
            return candidate

    downloads_dir = Path.home() / "Downloads"
    if not downloads_dir.exists():
        return None

    candidates = [
        item
        for item in downloads_dir.glob(DEFAULT_DESCENTE_WORKBOOK_GLOB)
        if item.is_file() and not item.name.startswith("~$")
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda item: (item.stat().st_mtime_ns, item.name.casefold()))


def _extract_teams_channel_metadata(channel_url: str) -> dict[str, str]:
    raw_url = str(channel_url or "").strip()
    if not raw_url:
        return {
            "channel_url": "",
            "team_id": "",
            "tenant_id": "",
            "channel_id": "",
        }

    parsed = urlparse(raw_url)
    query = parse_qs(parsed.query)
    decoded_path = unquote(parsed.path or "")
    channel_match = re.search(r"/l/(?:message|channel)/([^/?]+)", decoded_path)

    return {
        "channel_url": raw_url,
        "team_id": str((query.get("groupId") or [""])[0] or "").strip(),
        "tenant_id": str((query.get("tenantId") or [""])[0] or "").strip(),
        "channel_id": str(unquote(channel_match.group(1)) if channel_match else "").strip(),
    }


def _load_descente_channel_snapshot() -> dict[str, dict]:
    if not DESCENTE_CHANNEL_SNAPSHOT.exists():
        return {}
    try:
        payload = json.loads(DESCENTE_CHANNEL_SNAPSHOT.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return {}
    if not isinstance(payload, dict):
        return {}
    channels = payload.get("classes", payload)
    if not isinstance(channels, dict):
        return {}
    return {
        _normalize_lookup(key): value
        for key, value in channels.items()
        if _normalize_lookup(key) and isinstance(value, dict)
    }


@lru_cache(maxsize=4)
def _build_descente_channel_index_cached(cache_key: tuple[str, int, str]) -> dict[str, dict]:
    workbook_path = Path(cache_key[0])
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        if "CTR PUB MAN" not in workbook.sheetnames:
            return {}

        classes: dict[str, dict] = {}
        rows = workbook["CTR PUB MAN"].iter_rows(values_only=True)
        raw_header = next(rows, ())

        for row in rows:
            classe_id = _sheet_get_first(row, raw_header, "CLASSE 1", "Classe 1", "Classe")
            classe_key = _normalize_lookup(classe_id)
            channel_url = _sheet_get_first(row, raw_header, "LIEN DU CANAL")
            if not classe_key or not channel_url:
                continue

            metadata = _extract_teams_channel_metadata(channel_url)
            if not metadata["channel_url"]:
                continue

            channel_name = _sheet_get_first(row, raw_header, "NOM  DU CANAL", "NOM DU CANAL")
            responsable = _sheet_get_first(
                row,
                raw_header,
                "Responsable ",
                "Responsable",
                "RESPONSABLE ",
                "RESPONSABLE",
            )
            controle = _sheet_get_first(
                row,
                raw_header,
                "TYPE  DE CONTRÔLE",
                "TYPE DE CONTRÔLE",
            )

            entry = {
                **metadata,
                "channel_name": channel_name,
                "responsable": responsable,
                "controle_type": controle,
            }

            bucket = classes.setdefault(
                classe_key,
                {
                    "classe_id": classe_id,
                    "teams_channel_url": "",
                    "teams_channel_name": "",
                    "teams_team_id": "",
                    "teams_tenant_id": "",
                    "teams_channel_id": "",
                    "teams_responsable": "",
                    "teams_controle_type": "",
                    "teams_channels": [],
                    "_seen_urls": set(),
                },
            )

            if entry["channel_url"] in bucket["_seen_urls"]:
                continue

            bucket["_seen_urls"].add(entry["channel_url"])
            bucket["teams_channels"].append(entry)

            if not bucket["teams_channel_url"]:
                bucket["teams_channel_url"] = entry["channel_url"]
                bucket["teams_channel_name"] = entry["channel_name"]
                bucket["teams_team_id"] = entry["team_id"]
                bucket["teams_tenant_id"] = entry["tenant_id"]
                bucket["teams_channel_id"] = entry["channel_id"]
                bucket["teams_responsable"] = entry["responsable"]
                bucket["teams_controle_type"] = entry["controle_type"]

        for bucket in classes.values():
            bucket["teams_channel_count"] = len(bucket["teams_channels"])
            bucket.pop("_seen_urls", None)
        return classes
    finally:
        workbook.close()


def build_descente_channel_index(force_refresh: bool = False) -> dict[str, dict]:
    workbook_path = _resolve_descente_workbook()
    if workbook_path is None:
        return _load_descente_channel_snapshot()
    workbook_stat = workbook_path.stat()
    cache_key = (
        str(workbook_path),
        workbook_stat.st_size,
        datetime.fromtimestamp(workbook_stat.st_mtime).isoformat(),
    )
    if force_refresh:
        _build_descente_channel_index_cached.cache_clear()
    return _build_descente_channel_index_cached(cache_key)


def _sheet_page_payload(worksheet, page: int, search: str) -> dict:
    requested_page = max(page, 1)
    search_term = search.strip()
    search_lookup = _normalize_lookup(search_term)

    row_iterator = worksheet.iter_rows(values_only=True)
    header_row = next(row_iterator, ())
    column_count = max(len(header_row), worksheet.max_column or 0)
    headers = _build_headers(header_row, column_count)

    if not search_lookup:
        total_rows = max((worksheet.max_row or 1) - 1, 0)
        total_pages = max(1, math.ceil(total_rows / PAGE_SIZE)) if total_rows else 1
        current_page = min(requested_page, total_pages)
        start_index = (current_page - 1) * PAGE_SIZE
        end_index = start_index + PAGE_SIZE

        rows = []
        for row_index, row in enumerate(row_iterator):
            if row_index < start_index:
                continue
            if row_index >= end_index:
                break
            rows.append(_build_row(row, column_count))

        start_label = start_index + 1 if total_rows else 0
        end_label = min(start_index + len(rows), total_rows)
        return {
            "name": worksheet.title,
            "headers": headers,
            "rows": rows,
            "page": current_page,
            "page_size": PAGE_SIZE,
            "filtered": False,
            "search": "",
            "total_rows": total_rows,
            "total_pages": total_pages,
            "has_previous": current_page > 1,
            "has_next": current_page < total_pages,
            "summary": f"Lignes {start_label} a {end_label} sur {total_rows}",
        }

    start_index = (requested_page - 1) * PAGE_SIZE
    rows = []
    matches_seen = 0
    has_next = False

    for row in row_iterator:
        cells = _build_row(row, column_count)
        if not any(search_lookup in _normalize_lookup(cell) for cell in cells if cell):
            continue

        if matches_seen < start_index:
            matches_seen += 1
            continue

        if len(rows) < PAGE_SIZE:
            rows.append(cells)
            matches_seen += 1
            continue

        has_next = True
        break

    current_count = len(rows)
    start_label = start_index + 1 if current_count else 0
    end_label = start_index + current_count
    return {
        "name": worksheet.title,
        "headers": headers,
        "rows": rows,
        "page": requested_page,
        "page_size": PAGE_SIZE,
        "filtered": True,
        "search": search_term,
        "total_rows": None,
        "total_pages": None,
        "has_previous": requested_page > 1,
        "has_next": has_next,
        "summary": f"20 premiers resultats du filtre ({start_label} a {end_label})",
    }


@lru_cache(maxsize=8)
def _build_padesce_source_index_cached(cache_key: tuple[str, int, str, str]) -> dict:
    source_key = cache_key[0]
    with _open_cached_workbook(force_refresh=False, source_key=source_key) as (source, workbook):
        missing_sheets = [
            sheet_name
            for sheet_name in ("Apprenants", "Classes", "Prestations")
            if sheet_name not in workbook.sheetnames
        ]
        if missing_sheets:
            raise ValueError(
                "Feuilles source manquantes dans le classeur reseau: " + ", ".join(missing_sheets)
            )

        classes_by_id: dict[str, dict] = {}
        prestations_by_id: dict[str, dict] = {}
        apprenants_by_code: dict[str, dict] = {}
        inspecteurs_by_code: dict[str, dict] = {}
        class_inspecteurs: dict[str, dict] = {}
        class_enquetes: dict[str, list[str]] = {}
        duplicate_codes: set[str] = set()
        descente_channels = build_descente_channel_index()

        class_rows = workbook["Classes"].iter_rows(values_only=True)
        class_headers = _sheet_header_lookup(next(class_rows, ()))
        for row in class_rows:
            classe_id = _sheet_get(row, class_headers, "Classe ID", "Classe", "Class ID")
            classe_key = _normalize_lookup(classe_id)
            if not classe_key:
                continue
            classes_by_id[classe_key] = {
                "classe_id": classe_id,
                "prestation_id": _sheet_get(row, class_headers, "Prestation ID", "ID Prestation"),
                "prestataire": _sheet_get(row, class_headers, "Nom du Prestataire", "Prestataire"),
                "beneficiaire": _sheet_get(
                    row,
                    class_headers,
                    "Nom du beneficiaire",
                    "Nom du bénéficiaire",
                    "Beneficiaire",
                ),
                "cohorte": _sheet_get(row, class_headers, "Cohorte"),
                "lieu": _sheet_get(row, class_headers, "Lieux", "Lieu"),
                "ville": _sheet_get(row, class_headers, "Ville"),
                "formation": _sheet_get(row, class_headers, "FORMATION", "Formation"),
                "region": _sheet_get(row, class_headers, "Region"),
                "statut_prestation": _sheet_get(
                    row,
                    class_headers,
                    "Statut de la prestation",
                    "Statut prestation",
                ),
            }

        prestation_rows = workbook["Prestations"].iter_rows(values_only=True)
        prestation_headers = _sheet_header_lookup(next(prestation_rows, ()))
        for row in prestation_rows:
            prestation_id = _sheet_get(row, prestation_headers, "ID Prestation", "PRESTATION ID")
            prestation_key = _normalize_lookup(prestation_id)
            if not prestation_key:
                continue
            prestations_by_id[prestation_key] = {
                "prestation_id": prestation_id,
                "prestataire": _sheet_get(
                    row, prestation_headers, "Prestataire", "Nom du prestataire"
                ),
                "beneficiaire": _sheet_get(
                    row,
                    prestation_headers,
                    "Nom du bénéficiaire",
                    "Nom du beneficiaire",
                    "Beneficiaire",
                ),
                "formation": _sheet_get(row, prestation_headers, "Formation", "NomFormation"),
                "fenetre": _sheet_get(row, prestation_headers, "Fenetre", "Fenêtre"),
                "region": _sheet_get(row, prestation_headers, "Region"),
                "statut_prestation": _sheet_get(
                    row,
                    prestation_headers,
                    "Statut de la prestation",
                    "Statut prestation",
                    "Statut avec Taux",
                ),
            }

        if "Inspecteurs" in workbook.sheetnames:
            inspecteur_rows = workbook["Inspecteurs"].iter_rows(values_only=True)
            inspecteur_headers = _sheet_header_lookup(next(inspecteur_rows, ()))
            for row in inspecteur_rows:
                inspecteur_id = _sheet_get(row, inspecteur_headers, "ID Inspecteur", "InspecteurID")
                inspecteur_key = _normalize_lookup(inspecteur_id)
                if not inspecteur_key:
                    continue
                inspecteurs_by_code[inspecteur_key] = {
                    "inspecteur_id": inspecteur_id,
                    "inspecteur_label": _sheet_get(
                        row, inspecteur_headers, "Nom et Prenom", "Nom et Prénom"
                    ),
                    "telephone": _sheet_get(
                        row, inspecteur_headers, "Numéro de Téléphone", "Numero de Telephone"
                    ),
                    "ville": _sheet_get(row, inspecteur_headers, "Ville"),
                    "departement": _sheet_get(
                        row, inspecteur_headers, "Departement", "Département"
                    ),
                }

        if "ListesDV" in workbook.sheetnames:
            list_rows = workbook["ListesDV"].iter_rows(values_only=True)
            list_headers = _sheet_header_lookup(next(list_rows, ()))
            for row in list_rows:
                classe_id = _sheet_get(row, list_headers, "Classes_IDs", "Classe ID", "Classes IDs")
                classe_key = _normalize_lookup(classe_id)
                if not classe_key:
                    continue
                inspecteur_display = _sheet_get(
                    row, list_headers, "Inspecteurs_Display", "Inspecteur"
                )
                inspecteur_code = (
                    inspecteur_display.split(" - ", 1)[0].strip() if inspecteur_display else ""
                )
                inspecteur_info = inspecteurs_by_code.get(_normalize_lookup(inspecteur_code), {})
                class_inspecteurs[classe_key] = {
                    "inspecteur_id": inspecteur_code or inspecteur_info.get("inspecteur_id", ""),
                    "inspecteur_label": inspecteur_info.get("inspecteur_label", "")
                    or inspecteur_display,
                }

        if "PATH Presence" in workbook.sheetnames:
            presence_rows = workbook["PATH Presence"].iter_rows(values_only=True)
            presence_headers = _sheet_header_lookup(next(presence_rows, ()))
            for row in presence_rows:
                classe_id = _sheet_get(
                    row,
                    presence_headers,
                    "PrestationID",
                    "Classe ID",
                    "Class ID",
                )
                classe_key = _normalize_lookup(classe_id)
                enquete_id = _sheet_get(row, presence_headers, "EnqueteID", "EnquêteID")
                if not classe_key or not enquete_id:
                    continue
                class_enquetes.setdefault(classe_key, [])
                if enquete_id not in class_enquetes[classe_key]:
                    class_enquetes[classe_key].append(enquete_id)

        for classe_key, descente_info in descente_channels.items():
            class_bucket = classes_by_id.setdefault(
                classe_key,
                {
                    "classe_id": descente_info.get("classe_id", ""),
                    "prestation_id": "",
                    "prestataire": "",
                    "beneficiaire": "",
                    "cohorte": "",
                    "lieu": "",
                    "ville": "",
                    "formation": "",
                    "region": "",
                    "statut_prestation": "",
                },
            )
            for field in (
                "teams_channel_url",
                "teams_channel_name",
                "teams_team_id",
                "teams_tenant_id",
                "teams_channel_id",
                "teams_responsable",
                "teams_controle_type",
                "teams_channel_count",
                "teams_channels",
            ):
                class_bucket[field] = descente_info.get(field, class_bucket.get(field))

        apprenant_rows = workbook["Apprenants"].iter_rows(values_only=True)
        apprenant_headers = _sheet_header_lookup(next(apprenant_rows, ()))
        apprenant_count = 0
        for row in apprenant_rows:
            code = _sheet_get(row, apprenant_headers, "Code")
            code_key = _normalize_lookup(code)
            if not code_key:
                continue
            apprenant_count += 1
            if code_key in apprenants_by_code:
                duplicate_codes.add(code)
                continue

            classe_id = _sheet_get(row, apprenant_headers, "Classe ID", "Classe", "Class ID")
            classe_info = classes_by_id.get(_normalize_lookup(classe_id), {})
            prestation_id = classe_info.get("prestation_id", "")
            prestation_info = prestations_by_id.get(_normalize_lookup(prestation_id), {})
            inspecteur_info = class_inspecteurs.get(_normalize_lookup(classe_id), {})

            apprenants_by_code[code_key] = {
                "code": code,
                "apprenant_id": _sheet_get(row, apprenant_headers, "ApprenantID", "Apprenant ID"),
                "individu_id": _sheet_get(row, apprenant_headers, "ID_Individu", "IndividuID"),
                "beneficiaire_id": _sheet_get(
                    row, apprenant_headers, "ID_Beneficiaire", "BeneficiaireID"
                ),
                "nom_individu": _sheet_get(row, apprenant_headers, "Nom_Individu", "Nom Individu"),
                "beneficiaire": _sheet_get(
                    row, apprenant_headers, "Nom_Beneficiaire", "Nom Beneficiaire"
                )
                or classe_info.get("beneficiaire", "")
                or prestation_info.get("beneficiaire", ""),
                "classe_id": classe_id,
                "cohorte": _sheet_get(row, apprenant_headers, "Cohorte")
                or classe_info.get("cohorte", ""),
                "statut_apprenant": _sheet_get(
                    row,
                    apprenant_headers,
                    "Statut Apprenant",
                    "Statut",
                ),
                "numero": _sheet_get(row, apprenant_headers, "N°", "No", "N"),
                "telephone1": _sheet_get(
                    row,
                    apprenant_headers,
                    "N° Tél",
                    "N° Tel",
                    "N°Tél",
                    "N°Tel",
                    "Téléphone",
                    "Telephone",
                    "Tel",
                    "Tél",
                    "Numéro de téléphone",
                    "Numero de telephone",
                    "N° de téléphone",
                    "N° de telephone",
                    "1er No tél 0 Tel No Apprenant",
                    "1er No tel 0 Tel No Apprenant",
                    "1er No tél Apprenant",
                    "1er No tel Apprenant",
                    "N°",
                    "No",
                    "N",
                ),
                "telephone2": _sheet_get(
                    row,
                    apprenant_headers,
                    "2e No tél 0 Tel No Apprenant (si disponible)",
                    "2e No tel 0 Tel No Apprenant (si disponible)",
                    "2e No tél Apprenant",
                    "2e No tel Apprenant",
                    "2e No tél Apprenant (si disponible)",
                    "2e No tel Apprenant (si disponible)",
                    "Téléphone 2",
                    "Telephone 2",
                    "Tel 2",
                    "Tél 2",
                ),
                "sexe": _sheet_get(row, apprenant_headers, "Sexe", "Genre"),
                "prestation_id": prestation_id or prestation_info.get("prestation_id", ""),
                "prestataire": classe_info.get("prestataire", "")
                or prestation_info.get("prestataire", ""),
                "formation": classe_info.get("formation", "")
                or prestation_info.get("formation", ""),
                "lieu": classe_info.get("lieu", ""),
                "ville": classe_info.get("ville", ""),
                "fenetre": prestation_info.get("fenetre", ""),
                "region": classe_info.get("region", "") or prestation_info.get("region", ""),
                "statut_prestation": classe_info.get("statut_prestation", "")
                or prestation_info.get("statut_prestation", ""),
                "inspecteur_id": inspecteur_info.get("inspecteur_id", ""),
                "inspecteur_label": inspecteur_info.get("inspecteur_label", ""),
                "enquete_ids": class_enquetes.get(_normalize_lookup(classe_id), []),
            }

        return {
            "source": {
                "key": source.key,
                "label": source.label,
                "name": source.name,
                "path": str(source.source_path),
                "cached_path": str(source.cached_path),
                "modified_at": source.modified_at.isoformat(),
                "modified_label": source.modified_at.strftime("%d/%m/%Y a %H:%M"),
            },
            "counts": {
                "apprenants": apprenant_count,
                "classes": len(classes_by_id),
                "prestations": len(prestations_by_id),
            },
            "classes": classes_by_id,
            "prestations": prestations_by_id,
            "records": apprenants_by_code,
            "class_enquetes": class_enquetes,
            "duplicate_codes": sorted(code for code in duplicate_codes if code),
        }


def load_bundled_source_meta(source_key: str = "cutoff") -> dict:
    """Load precomputed metadata for the bundled workbook (avoids openpyxl at runtime)."""
    normalized = normalize_workbook_source_key(source_key)
    meta_path = BUNDLED_DIRECTORY / f"network-fichier-consolide-{normalized}-meta.json"
    if meta_path.exists():
        try:
            return json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def build_padesce_source_index(
    force_refresh: bool = False,
    source_key: str = DEFAULT_WORKBOOK_SOURCE,
) -> dict:
    normalized_source_key = normalize_workbook_source_key(source_key)
    source = _ensure_cached_workbook(source_key=normalized_source_key, force_refresh=force_refresh)
    cache_key = (
        normalized_source_key,
        str(source.cached_path),
        source.size,
        source.modified_at.isoformat(),
    )
    shared_cache_key = (
        "padesce-source-index:"
        f"{normalized_source_key}:{source.size}:{source.modified_at.isoformat()}"
    )
    if force_refresh:
        _build_padesce_source_index_cached.cache_clear()
        cache.delete(shared_cache_key)

    cached_payload = cache.get(shared_cache_key)
    if cached_payload is not None:
        return cached_payload

    payload = _build_padesce_source_index_cached(cache_key)
    cache.set(shared_cache_key, payload, timeout=SOURCE_CACHE_TIMEOUT)
    return payload


@lru_cache(maxsize=8)
def _build_consolidation_call_candidates_cached(cache_key: tuple[str, int, str, str]) -> dict:
    source_key = cache_key[0]
    with _open_cached_workbook(force_refresh=False, source_key=source_key) as (source, workbook):
        if "Consolidation" not in workbook.sheetnames:
            raise ValueError("La feuille Consolidation est introuvable dans le classeur reseau.")

        worksheet = workbook["Consolidation"]
        row_iterator = worksheet.iter_rows(values_only=True)
        header_row = next(row_iterator, ())
        header_lookup = _sheet_header_lookup(header_row)

        records_by_code: dict[str, dict] = {}
        duplicate_codes: set[str] = set()
        total_rows = 0

        for row_index, row in enumerate(row_iterator, start=2):
            code = _sheet_get(row, header_lookup, "Code")
            nom = _sheet_get(
                row,
                header_lookup,
                "Nom et prenom 0 Name & first name",
                "Nom et prénom 0 Name & first name",
                "Nom et prenom",
                "Nom et prénom",
                "Nom",
            )
            if not code or not nom:
                continue

            total_rows += 1
            code_key = _normalize_lookup(code)
            if code_key in records_by_code:
                duplicate_codes.add(code)
                continue

            records_by_code[code_key] = {
                "row_number": row_index,
                "numero": _sheet_get(row, header_lookup, "N°", "No", "N"),
                "code": code,
                "nom": nom,
                "beneficiaire": _sheet_get(
                    row,
                    header_lookup,
                    "Bénéficiaires",
                    "Beneficiaires",
                    "Beneficiaire",
                ),
                "prestataire": _sheet_get(row, header_lookup, "Prestataire"),
                "type_formation_declaree": _sheet_get(
                    row,
                    header_lookup,
                    "Type de formation declarée",
                    "Type de formation declaree",
                ),
                "formation_padesce": _sheet_get(row, header_lookup, "Formation Padesce"),
                "fenetre": _sheet_get(row, header_lookup, "fenêtre", "Fenetre", "Fenêtre"),
                "ville_formation": _sheet_get(row, header_lookup, "Ville de la formation"),
                "lieu": _sheet_get(row, header_lookup, "Lieux", "Lieu"),
                "telephone1": _sheet_get(
                    row,
                    header_lookup,
                    "1er No tél Apprenant",
                    "1er No tel Apprenant",
                    "1er No tél 0 Tel No Apprenant",
                    "1er No tél 0 Tel No\nApprenant",
                    "1er No tel 0 Tel No Apprenant",
                    "1er No tel 0 Tel No\nApprenant",
                    "Téléphone",
                    "Telephone",
                    "Tel",
                    "Tél",
                    "N° Tél",
                    "N° Tel",
                    "N°Tél",
                    "N°Tel",
                ),
                "telephone2": _sheet_get(
                    row,
                    header_lookup,
                    "2e No tél Apprenant",
                    "2e No tel Apprenant",
                    "2e No tél Apprenant (si disponible)",
                    "2e No tel Apprenant (si disponible)",
                    "2e No tél 0 Tel No Apprenant (si disponible)",
                    "2e No tél 0 Tel No\nApprenant (si disponible)",
                    "2e No tel 0 Tel No Apprenant (si disponible)",
                    "2e No tel 0 Tel No\nApprenant (si disponible)",
                    "Téléphone 2",
                    "Telephone 2",
                    "Tel 2",
                    "Tél 2",
                ),
                "cohorte": _sheet_get(row, header_lookup, "cohorte", "Cohorte"),
                "statut_prestation": _sheet_get(
                    row,
                    header_lookup,
                    "Statut de la prestation",
                    "Statut prestation",
                ),
                "classe_label": _sheet_get(row, header_lookup, "Classe", "Classe ID"),
                "prestation_id": _sheet_get(
                    row, header_lookup, "ID Prestation", "Prestation ID", "PRESTATION ID"
                ),
                "arrondissement": _sheet_get(row, header_lookup, "Arrondissement"),
                "departement": _sheet_get(row, header_lookup, "Département", "Departement"),
                "region": _sheet_get(row, header_lookup, "Région", "Region"),
            }

        return {
            "source": {
                "key": source.key,
                "label": source.label,
                "name": source.name,
                "path": str(source.source_path),
                "cached_path": str(source.cached_path),
                "modified_at": source.modified_at.isoformat(),
                "modified_label": source.modified_at.strftime("%d/%m/%Y a %H:%M"),
            },
            "sheet_name": "Consolidation",
            "count": len(records_by_code),
            "source_rows": total_rows,
            "records": list(records_by_code.values()),
            "duplicate_codes": sorted(code for code in duplicate_codes if code),
        }


def build_consolidation_call_candidates(
    force_refresh: bool = False,
    source_key: str = DEFAULT_WORKBOOK_SOURCE,
) -> dict:
    normalized_source_key = normalize_workbook_source_key(source_key)
    source = _ensure_cached_workbook(source_key=normalized_source_key, force_refresh=force_refresh)
    cache_key = (
        normalized_source_key,
        str(source.cached_path),
        source.size,
        source.modified_at.isoformat(),
    )
    shared_cache_key = (
        "padesce-consolidation-candidates:"
        f"{normalized_source_key}:{source.size}:{source.modified_at.isoformat()}"
    )
    if force_refresh:
        _build_consolidation_call_candidates_cached.cache_clear()
        cache.delete(shared_cache_key)

    cached_payload = cache.get(shared_cache_key)
    if cached_payload is not None:
        return cached_payload

    payload = _build_consolidation_call_candidates_cached(cache_key)
    cache.set(shared_cache_key, payload, timeout=SOURCE_CACHE_TIMEOUT)
    return payload


def build_network_excel_payload(
    selected_sheet: str = "",
    page: int = 1,
    search: str = "",
    force_refresh: bool = False,
    source_key: str = DEFAULT_WORKBOOK_SOURCE,
) -> dict:
    normalized_source_key = normalize_workbook_source_key(source_key)
    with _open_cached_workbook(source_key=normalized_source_key, force_refresh=force_refresh) as (
        source,
        workbook,
    ):
        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise ValueError("Le classeur reseau ne contient aucune feuille exploitable.")

        active_sheet = selected_sheet if selected_sheet in sheet_names else sheet_names[0]
        sheets = []
        for name in sheet_names:
            worksheet = workbook[name]
            sheets.append(
                {
                    "name": name,
                    "rows": max((worksheet.max_row or 1) - 1, 0),
                    "columns": worksheet.max_column or 0,
                    "active": name == active_sheet,
                }
            )

        sheet_payload = _sheet_page_payload(workbook[active_sheet], page=page, search=search)
        total_rows = sum(sheet["rows"] for sheet in sheets)

        return {
            "source": {
                "key": source.key,
                "label": source.label,
                "name": source.name,
                "path": str(source.source_path),
                "cached_path": str(source.cached_path),
                "size": source.size,
                "modified_at": source.modified_at.isoformat(),
                "modified_label": source.modified_at.strftime("%d/%m/%Y a %H:%M"),
            },
            "page_size": PAGE_SIZE,
            "sheet_count": len(sheets),
            "total_rows": total_rows,
            "sheets": sheets,
            "sheet": sheet_payload,
        }


@require_analysis_access
def network_excel_view(request):
    return render(
        request,
        "reporting/network_excel.html",
        {
            "page_size": PAGE_SIZE,
        },
    )


@require_analysis_access
def network_excel_api(request):
    raw_page = request.GET.get("page", "1").strip()
    try:
        page = max(int(raw_page), 1)
    except ValueError:
        page = 1

    search = request.GET.get("search", "")
    selected_sheet = request.GET.get("sheet", "")
    selected_source = request.GET.get("source", "")
    force_refresh = request.GET.get("refresh") == "1"

    try:
        payload = build_network_excel_payload(
            selected_sheet=selected_sheet,
            page=page,
            search=search,
            force_refresh=force_refresh,
            source_key=selected_source,
        )
    except FileNotFoundError as exc:
        return JsonResponse({"error": str(exc)}, status=404)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    except Exception as exc:
        return JsonResponse(
            {"error": f"Impossible de lire le classeur reseau: {exc}"},
            status=500,
        )

    return JsonResponse(payload)
