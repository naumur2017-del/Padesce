import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook

from App_PADESCE.reporting import app_report, network_excel


class NetworkExcelApiTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="reporting-tester",
            password="test-pass-123",
        )
        manager_group, _ = Group.objects.get_or_create(name="manager_padesce")
        self.user.groups.add(manager_group)
        self.client.force_login(self.user)

        self.regular_user = user_model.objects.create_user(
            username="reporting-regular",
            password="test-pass-123",
        )

        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.source_path = self.root / "source-workbook.xlsm"
        self.cutoff_source_path = self.root / "source-cutoff-workbook.xlsm"
        self.cache_dir = self.root / "cache"
        self.cache_path = self.cache_dir / "network-fichier-consolide.xlsm"
        self.cutoff_cache_path = self.cache_dir / "network-fichier-consolide-cutoff.xlsm"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Consolidation"
        sheet.append(["Code", "Nom", "Ville"])
        for index in range(1, 46):
            keyword = "ALPHA" if index <= 30 else "BETA"
            sheet.append([f"C{index:03d}", f"Nom {index}", f"{keyword} Ville"])

        second_sheet = workbook.create_sheet("Beneficiaires")
        second_sheet.append(["Beneficiaire", "Region"])
        second_sheet.append(["CAPEF", "Nord"])
        second_sheet.append(["IRISA", "Sud"])

        apprenants_sheet = workbook.create_sheet("Apprenants")
        apprenants_sheet.append(
            [
                "ApprenantID",
                "ID_Individu",
                "ID_Beneficiaire",
                "Nom_Individu",
                "Nom_Beneficiaire",
                "Classe ID",
                "Code",
                "Cohorte",
                "Statut Apprenant",
                "N°",
                "Sexe",
            ]
        )
        apprenants_sheet.append(
            ["APP001", "IND001", "BEN001", "Nom 1", "CAPEF", "CLA001", "C001", "1", "Actif", "1", "F"]
        )

        classes_sheet = workbook.create_sheet("Classes")
        classes_sheet.append(
            [
                "Classe ID",
                "Prestation ID",
                "Nom du Prestataire",
                "Nom du beneficiaire",
                "Cohorte",
                "Lieux",
                "Ville",
                "FORMATION",
                "Region",
            ]
        )
        classes_sheet.append(
            ["CLA001", "PRESTA001", "Prestataire Alpha", "CAPEF", "1", "Site A", "Garoua", "Formation X", "Nord"]
        )

        prestations_sheet = workbook.create_sheet("Prestations")
        prestations_sheet.append(
            [
                "ID Prestation",
                "Prestataire",
                "Nom du bénéficiaire",
                "Formation",
                "Fenetre",
                "Region",
            ]
        )
        prestations_sheet.append(
            ["PRESTA001", "Prestataire Alpha", "CAPEF", "Formation X", "2", "Nord"]
        )
        workbook.save(self.source_path)
        workbook.close()

        cutoff_workbook = Workbook()
        cutoff_sheet = cutoff_workbook.active
        cutoff_sheet.title = "Consolidation"
        cutoff_sheet.append(["Code", "Nom", "Ville"])
        cutoff_sheet.append(["CUT001", "Nom CutOff", "Maroua"])

        cutoff_apprenants_sheet = cutoff_workbook.create_sheet("Apprenants")
        cutoff_apprenants_sheet.append(
            [
                "ApprenantID",
                "ID_Individu",
                "ID_Beneficiaire",
                "Nom_Individu",
                "Nom_Beneficiaire",
                "Classe ID",
                "Code",
                "Cohorte",
                "Statut Apprenant",
                "NÂ°",
                "Sexe",
            ]
        )
        cutoff_apprenants_sheet.append(
            ["APPCUT001", "INDCUT001", "BENCUT001", "Nom CutOff", "CAPEF", "CLACUT001", "CUT001", "1", "Actif", "1", "F"]
        )

        cutoff_classes_sheet = cutoff_workbook.create_sheet("Classes")
        cutoff_classes_sheet.append(
            [
                "Classe ID",
                "Prestation ID",
                "Nom du Prestataire",
                "Nom du beneficiaire",
                "Cohorte",
                "Lieux",
                "Ville",
                "FORMATION",
                "Region",
            ]
        )
        cutoff_classes_sheet.append(
            ["CLACUT001", "PRESTACUT001", "Prestataire CutOff", "CAPEF", "1", "Site CutOff", "Maroua", "Formation CutOff", "Extreme-Nord"]
        )

        cutoff_prestations_sheet = cutoff_workbook.create_sheet("Prestations")
        cutoff_prestations_sheet.append(
            [
                "ID Prestation",
                "Prestataire",
                "Nom du bÃ©nÃ©ficiaire",
                "Formation",
                "Fenetre",
                "Region",
            ]
        )
        cutoff_prestations_sheet.append(
            ["PRESTACUT001", "Prestataire CutOff", "CAPEF", "Formation CutOff", "3", "Extreme-Nord"]
        )
        cutoff_workbook.save(self.cutoff_source_path)
        cutoff_workbook.close()

        self.resolve_patch = patch.object(
            network_excel,
            "_resolve_network_workbook",
            side_effect=lambda source_key="main": (
                self.cutoff_source_path
                if network_excel.normalize_workbook_source_key(source_key) == "cutoff"
                else self.source_path
            ),
        )
        self.cache_dir_patch = patch.object(
            network_excel,
            "CACHE_DIRECTORY",
            self.cache_dir,
        )
        self.cache_path_patch = patch.object(
            network_excel,
            "LOCAL_WORKBOOK_COPY",
            self.cache_path,
        )
        self.cutoff_cache_path_patch = patch.object(
            network_excel,
            "LOCAL_CUTOFF_WORKBOOK_COPY",
            self.cutoff_cache_path,
        )

        self.resolve_patch.start()
        self.cache_dir_patch.start()
        self.cache_path_patch.start()
        self.cutoff_cache_path_patch.start()

        self.addCleanup(self.resolve_patch.stop)
        self.addCleanup(self.cache_dir_patch.stop)
        self.addCleanup(self.cache_path_patch.stop)
        self.addCleanup(self.cutoff_cache_path_patch.stop)
        self.addCleanup(self.temp_dir.cleanup)

    def test_api_returns_first_page_with_20_rows(self):
        response = self.client.get(
            reverse("reporting_network_excel_api"),
            {"sheet": "Consolidation"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["sheet_count"], 5)
        self.assertEqual(payload["sheet"]["name"], "Consolidation")
        self.assertEqual(payload["sheet"]["page"], 1)
        self.assertEqual(payload["sheet"]["page_size"], 20)
        self.assertEqual(payload["sheet"]["total_pages"], 3)
        self.assertEqual(len(payload["sheet"]["rows"]), 20)
        self.assertEqual(payload["sheet"]["rows"][0][0], "C001")
        self.assertEqual(payload["sheet"]["rows"][-1][0], "C020")

    def test_api_moves_to_next_page(self):
        response = self.client.get(
            reverse("reporting_network_excel_api"),
            {"sheet": "Consolidation", "page": 2},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["sheet"]["page"], 2)
        self.assertEqual(len(payload["sheet"]["rows"]), 20)
        self.assertEqual(payload["sheet"]["rows"][0][0], "C021")
        self.assertEqual(payload["sheet"]["rows"][-1][0], "C040")
        self.assertTrue(payload["sheet"]["has_previous"])
        self.assertTrue(payload["sheet"]["has_next"])

    def test_api_filters_and_returns_first_20_matches(self):
        response = self.client.get(
            reverse("reporting_network_excel_api"),
            {"sheet": "Consolidation", "search": "alpha"},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertTrue(payload["sheet"]["filtered"])
        self.assertEqual(payload["sheet"]["page"], 1)
        self.assertEqual(len(payload["sheet"]["rows"]), 20)
        self.assertEqual(payload["sheet"]["rows"][0][0], "C001")
        self.assertEqual(payload["sheet"]["rows"][-1][0], "C020")
        self.assertTrue(payload["sheet"]["has_next"])

    def test_build_padesce_source_index_returns_joined_apprenant_data(self):
        payload = network_excel.build_padesce_source_index(force_refresh=True)

        record = payload["records"][network_excel.normalize_network_lookup("C001")]

        self.assertEqual(record["apprenant_id"], "APP001")
        self.assertEqual(record["classe_id"], "CLA001")
        self.assertEqual(record["prestation_id"], "PRESTA001")
        self.assertEqual(record["prestataire"], "Prestataire Alpha")
        self.assertEqual(record["fenetre"], "2")

    def test_build_padesce_source_index_uses_cutoff_source_when_requested(self):
        payload = network_excel.build_padesce_source_index(force_refresh=True, source_key="cutoff")

        record = payload["records"][network_excel.normalize_network_lookup("CUT001")]

        self.assertEqual(payload["source"]["key"], "cutoff")
        self.assertEqual(payload["source"]["label"], "Fichier consolide CutOff")
        self.assertEqual(record["apprenant_id"], "APPCUT001")
        self.assertEqual(record["prestation_id"], "PRESTACUT001")
        self.assertEqual(record["fenetre"], "3")

    def test_api_forbids_authenticated_user_without_analysis_role(self):
        self.client.force_login(self.regular_user)

        response = self.client.get(reverse("reporting_network_excel_api"))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(
            response.json()["error"],
            "Acces reserve aux superadmins et aux managers PADESCE/CGA.",
        )


class ReportEmailDeliveryTests(SimpleTestCase):
    @staticmethod
    def _sample_report() -> dict:
        return {"start_date": date(2026, 3, 28), "end_date": date(2026, 3, 28)}

    @override_settings(
        REPORT_EMAIL_TO="to1@example.com,to2@example.com",
        EMAIL_BACKEND="django.core.mail.backends.console.EmailBackend",
        EMAIL_HOST="smtp.gmail.com",
        EMAIL_PORT=587,
        EMAIL_USE_TLS=True,
        EMAIL_HOST_USER="sender@example.com",
        EMAIL_HOST_PASSWORD="smtp-secret",
        DEFAULT_FROM_EMAIL="sender@example.com",
    )
    @patch("App_PADESCE.reporting.app_report._get_report_logo_path", return_value=None)
    @patch("App_PADESCE.reporting.app_report.EmailMessage")
    @patch("App_PADESCE.reporting.app_report.get_connection")
    @patch("App_PADESCE.reporting.app_report.build_report_email_html", return_value="<p>rapport</p>")
    def test_send_report_by_email_forces_smtp_when_console_backend_is_active(
        self,
        _mock_html,
        mock_get_connection,
        mock_email_message,
        _mock_logo,
    ):
        connection = object()
        mock_get_connection.return_value = connection
        mock_email_message.return_value.send.return_value = 1

        with patch.object(app_report, "HAS_DOCX", False):
            result = app_report.send_report_by_email(self._sample_report())

        self.assertTrue(result["ok"])
        mock_get_connection.assert_called_once_with(
            backend="django.core.mail.backends.smtp.EmailBackend",
            host="smtp.gmail.com",
            port=587,
            username="sender@example.com",
            password="smtp-secret",
            use_tls=True,
            use_ssl=False,
            timeout=None,
        )
        self.assertIs(mock_email_message.call_args.kwargs["connection"], connection)

    @override_settings(
        REPORT_EMAIL_TO="to1@example.com",
        EMAIL_BACKEND="django.core.mail.backends.console.EmailBackend",
        EMAIL_HOST="smtp.gmail.com",
        EMAIL_PORT=587,
        EMAIL_USE_TLS=True,
        EMAIL_HOST_USER="",
        EMAIL_HOST_PASSWORD="",
        DEFAULT_FROM_EMAIL="sender@example.com",
    )
    def test_send_report_by_email_rejects_console_backend_without_smtp_credentials(self):
        result = app_report.send_report_by_email(self._sample_report())

        self.assertFalse(result["ok"])
        self.assertIn("mode console", result["detail"])

    @override_settings(
        REPORT_EMAIL_TO="to1@example.com",
        EMAIL_BACKEND="django.core.mail.backends.smtp.EmailBackend",
        EMAIL_HOST="smtp.gmail.com",
        EMAIL_PORT=587,
        EMAIL_USE_TLS=True,
        EMAIL_HOST_USER="sender@example.com",
        EMAIL_HOST_PASSWORD="smtp-secret",
        DEFAULT_FROM_EMAIL="sender@example.com",
    )
    @patch("App_PADESCE.reporting.app_report._get_report_logo_path", return_value=None)
    @patch("App_PADESCE.reporting.app_report.EmailMessage")
    @patch("App_PADESCE.reporting.app_report.build_report_email_html", return_value="<p>rapport</p>")
    def test_send_report_by_email_returns_error_when_smtp_send_fails(
        self,
        _mock_html,
        mock_email_message,
        _mock_logo,
    ):
        mock_email_message.return_value.send.side_effect = RuntimeError("SMTP auth failed")

        with patch.object(app_report, "HAS_DOCX", False):
            result = app_report.send_report_by_email(self._sample_report())

        self.assertFalse(result["ok"])
        self.assertIn("SMTP auth failed", result["detail"])
