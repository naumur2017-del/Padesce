import base64
import csv
import io
import unicodedata
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model as _get_user_model
from django.contrib.auth.models import Group
from django.core.paginator import Paginator
from django.db import OperationalError, connection, transaction
from django.db.models import Avg, Count, Q, Sum
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook

from App_PADESCE.appels.models import Appel, AppelPasFormeII, is_call_attempted_status
from App_PADESCE.apprenants.models import Apprenant, SmsLog
from App_PADESCE.core.access import require_analysis_access, require_superadmin_access
from App_PADESCE.environnement.models import EnqueteEnvironnement
from App_PADESCE.formations.models import (
    Beneficiaire,
    Classe,
    Formation,
    Lieu,
    Prestataire,
    Prestation,
)
from App_PADESCE.presences.models import Presence
from App_PADESCE.reporting.app_report import (
    build_application_report,
    export_application_report_anomalies_excel,
    export_application_report_csv,
    export_application_report_excel,
    export_application_report_word,
    get_report_email_recipients,
    normalize_report_call_scope,
    parse_report_dates,
    send_report_by_email,
)
from App_PADESCE.reporting import daily_digest_jobs
from App_PADESCE.reporting.forms import ConsolidationUploadForm
from App_PADESCE.reporting.manuals import (
    get_reporting_manual_path,
    load_reporting_manual_markdown,
    render_reporting_manual_html,
)
from App_PADESCE.reporting.models import (
    ConcordanceRecord,
    ConsolidationRecord,
    PendingLearnerContactImport,
    PendingLearnerContactRecord,
)
from App_PADESCE.satisfaction_apprenants.models import SatisfactionApprenant
from App_PADESCE.satisfaction_formateurs.models import SatisfactionFormateur


def _application_report_public_url(request) -> str:
    params = {
        "scope": "apprenant",
        "section": "rapport",
    }
    for key in ("start", "end", "call_scope", "classe", "cga_anomaly_page"):
        value = str(request.GET.get(key) or request.POST.get(key) or "").strip()
        if value:
            params[key] = value
    return f"{reverse('public_space')}?{urlencode(params)}"


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize_header(value: str) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize("NFKD", str(value))
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_only = ascii_only.lower().replace("&", "and").replace("’", "'")
    ascii_only = "".join(ch if ch.isalnum() or ch.isspace() else " " for ch in ascii_only)
    return " ".join(ascii_only.split())


MAX_CONSO_COLS = 40  # Augmenté pour être sûr de ne pas couper la colonne Classe ID

# Mapping plus tolérant pour "Classe ID"
CONSOLIDATION_HEADER_MAP = {
    "n": "numero",
    "nom et prenom 0 name first name": "nom_complet",
    "nom et prenom 0 name and first name": "nom_complet",
    "beneficiaires": "beneficiaire",
    "genre h0f 0 gender m0f": "genre",
    "age": "age",
    "fonction d c e m b": "fonction",
    "qualification chiffre 0 1 2 etc": "qualification",
    "nb d annees d experience chiffre 0 1 2 etc": "nb_annees_experience",
    "ville de residence de l appprenant": "ville_residence",
    "prestataire": "prestataire",
    "type de formation declaree": "intitule_formation_solicitee",
    "formation padesce": "intitule_formation_dispensee",
    "fenetre": "fenetre",
    "ville de la formation": "ville_formation",
    "arrondissement": "arrondissement",
    "departement": "departement",
    "region": "region",
    "lieux": "lieu_formation",
    "precision sur le lieu 0 quartier de formation": "precision_lieu",
    "coordonnees gps du lieu de formation longitude": "longitude",
    "coordonnees gps du lieu de formation latitude": "latitude",
    "1er no tel 0 tel no apprenant": "telephone1",
    "2e no tel 0 tel no apprenant si disponible": "telephone2",
    "cohorte": "cohorte",
    "tel formateur 0 point focal sur place": "tel_formateur",
    "code": "code",
    "cout unitaire subvention mcdc ttc": "cout_unitaire_subvention",
    "montant total subvention mcdc ttc": "montant_total_subvention",
    "statut de la prestation": "statut_prestation",
    # Variantes pour Classe ID – très tolérant
    "classe id": "classe_id",
    "class id": "classe_id",
    "classeid": "classe_id",
    "classid": "classe_id",
    "id classe": "classe_id",
    "classe": "classe_id",  # fallback si "ID" est absent
    "classe_id": "classe_id",
    "id de classe": "classe_id",
    "numero classe": "classe_id",
    "code classe": "classe_id",
}


SESSION_KEY_CONSO = "consolidation_upload"

APPRENANT_COLUMNS = [
    ("numero", "Numero"),
    ("code", "Code"),
    ("nom_complet", "Nom complet"),
    ("beneficiaire", "Beneficiaire"),
    ("genre", "Genre"),
    ("age", "Age"),
    ("fonction", "Fonction"),
    ("qualification", "Qualification"),
    ("nb_annees_experience", "Nb annees experience"),
    ("ville_residence", "Ville residence"),
    ("prestataire", "Prestataire"),
    ("intitule_formation_solicitee", "Intitule formation sollicitee"),
    ("intitule_formation_dispensee", "Intitule formation dispensee"),
    ("fenetre", "Fenetre"),
    ("ville_formation", "Ville formation"),
    ("arrondissement", "Arrondissement"),
    ("departement", "Departement"),
    ("region", "Region"),
    ("lieu_formation", "Lieu formation"),
    ("precision_lieu", "Precision lieu"),
    ("longitude", "Longitude"),
    ("latitude", "Latitude"),
    ("telephone1", "Telephone 1"),
    ("telephone2", "Telephone 2"),
    ("cohorte", "Cohorte"),
    ("tel_formateur", "Tel formateur"),
    ("classe_code", "Classe"),
    ("formation_nom", "Formation"),
    ("code_ville", "Code ville"),
    ("appartenance_beneficiaire", "Appartenance beneficiaire"),
    ("actif", "Actif"),
]

APPEL_TERMINE_COLUMNS = [
    ("code", "Code"),
    ("nom", "Nom"),
    ("prestataire", "Prestataire"),
    ("beneficiaire", "Beneficiaire"),
    ("lieu", "Lieu"),
    ("classe_label", "Classe (label)"),
    ("classe_code", "Classe (code)"),
    ("telephone1", "Telephone 1"),
    ("telephone2", "Telephone 2"),
    ("taux_presence", "Taux presence"),
    ("type_formation_declaree", "Type formation declaree"),
    ("formation_padesce", "Formation PADESCE"),
    ("fenetre", "Fenetre"),
    ("status", "Statut"),
]


def _apprenant_value(apprenant: Apprenant, key: str):
    if key == "classe_code":
        return apprenant.classe.code if apprenant.classe_id else ""
    if key == "formation_nom":
        return apprenant.formation.nom if apprenant.formation_id else ""
    return getattr(apprenant, key, "")


def _appel_value(appel: Appel, key: str):
    if key == "classe_code":
        return appel.classe.code if appel.classe_id else ""
    return getattr(appel, key, "")


def _excel_safe(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    if isinstance(value, (int, float, Decimal)):
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _ensure_prestataire(name: str) -> Prestataire | None:
    if not name:
        return None
    raw = str(name).strip()
    if not raw:
        return None
    code = raw[:50]
    obj, _ = Prestataire.objects.get_or_create(code=code, defaults={"raison_sociale": raw})
    if obj.raison_sociale != raw:
        obj.raison_sociale = raw
        obj.save(update_fields=["raison_sociale"])
    return obj


def _reset_consolidation_tables():
    """
    Empty the tables that are rebuilt from consolidation imports using raw SQL to avoid
    instantiating large querysets and keep the request from timing out.
    """
    tables = (
        SatisfactionApprenant,
        SatisfactionFormateur,
        Presence,
        SmsLog,
        Apprenant,
        Classe,
        Prestation,
        Formation,
        Prestataire,
        Beneficiaire,
        Lieu,
    )
    with connection.cursor() as cursor:
        needs_toggle = connection.vendor in {"sqlite", "mysql"}
        if needs_toggle:
            if connection.vendor == "sqlite":
                cursor.execute("PRAGMA foreign_keys = OFF")
            else:
                cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        for model in tables:
            table_name = connection.ops.quote_name(model._meta.db_table)
            cursor.execute(f"DELETE FROM {table_name}")
        if needs_toggle:
            if connection.vendor == "sqlite":
                cursor.execute("PRAGMA foreign_keys = ON")
            else:
                cursor.execute("SET FOREIGN_KEY_CHECKS = 1")


def _ensure_formation(intitule: str, fenetre: str = "") -> Formation | None:
    if not intitule:
        return None
    raw = str(intitule).strip()
    if not raw:
        return None
    code = raw[:50]
    obj, _ = Formation.objects.get_or_create(code=code, defaults={"nom": raw, "fenetre": fenetre})
    if obj.nom != raw or (fenetre and obj.fenetre != fenetre):
        obj.nom = raw
        if fenetre:
            obj.fenetre = fenetre
        obj.save(update_fields=["nom", "fenetre"])
    return obj


def _ensure_beneficiaire(
    name: str, region: str = "", departement: str = "", arrondissement: str = "", ville: str = ""
) -> Beneficiaire | None:
    if not name:
        return None
    raw = str(name).strip()
    if not raw:
        return None
    obj, _ = Beneficiaire.objects.get_or_create(
        nom_structure=raw,
        defaults={
            "region": region,
            "departement": departement,
            "arrondissement": arrondissement,
            "ville": ville,
        },
    )
    return obj


def _ensure_lieu(
    nom: str,
    region: str = "",
    departement: str = "",
    arrondissement: str = "",
    ville: str = "",
    longitude: str = "",
    latitude: str = "",
    precision: str = "",
) -> Lieu | None:
    if not nom:
        return None
    raw = str(nom).strip()
    if not raw:
        return None
    code = raw[:50]
    obj, _ = Lieu.objects.get_or_create(
        code=code,
        defaults={
            "nom_lieu": raw,
            "region": region,
            "departement": departement,
            "arrondissement": arrondissement,
            "ville": ville,
            "longitude": longitude,
            "latitude": latitude,
            "precision": precision,
        },
    )
    return obj


def _ensure_prestation(
    prestataire: Prestataire | None,
    formation: Formation | None,
    beneficiaire: Beneficiaire | None,
    code_hint: str = "",
) -> Prestation | None:
    if not prestataire or not formation:
        return None
    code_raw = (code_hint or "").strip()
    code = (code_raw or "null")[:50]
    obj, _ = Prestation.objects.get_or_create(
        code=code,
        defaults={"prestataire": prestataire, "formation": formation, "beneficiaire": beneficiaire},
    )
    return obj


def _ensure_classe(
    prestation: Prestation | None,
    formation: Formation | None,
    fenetre: str = "",
    cohorte: str = "",
    classe_id: str = "",
) -> Classe | None:
    if not prestation or not formation:
        return None

    # Priorité : utiliser classe_id comme code si présent et non vide
    code_raw = (classe_id or "").strip()
    if not code_raw:
        # Fallback si classe_id absent → code court unique
        code_raw = f"CL-{formation.code[:6] or 'XX'}-{fenetre or 'X'}-{cohorte or '1'}"

    code = code_raw[:20]

    # Construction du nom/intitulé lisible
    formation_nom = formation.nom.strip()
    if len(formation_nom) > 120:
        formation_nom = formation_nom[:117] + "..."

    # Format préféré : CLAxxx — Nom de la formation
    classe_nom = f"{code} — {formation_nom}"

    # Si cohorte > 1, on peut l'ajouter pour plus de clarté (optionnel)
    cohorte_int = _to_int(cohorte)
    if cohorte_int and cohorte_int > 1:
        classe_nom += f" (cohorte {cohorte_int})"

    obj, created = Classe.objects.get_or_create(
        code=code,
        defaults={
            "prestation": prestation,
            "formation": formation,
            "intitule_formation": formation_nom,
            "fenetre": fenetre or "",
            "cohorte": cohorte_int if cohorte_int else 1,
        },
    )

    # Mise à jour si déjà existant mais nom différent
    if not created:
        updated = False
        if obj.intitule_formation != formation_nom:
            obj.intitule_formation = formation_nom
            updated = True
        if updated:
            obj.save(update_fields=["intitule_formation"])

    return obj


def _to_int(value):
    try:
        if value is None or value == "":
            return None
        val = str(value).strip().replace(" ", "")
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_decimal(value):
    if value in (None, ""):
        return None
    try:
        cleaned = str(value).replace(" ", "").replace(",", ".")
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _read_consolidation_sheet(file_obj, max_rows: int | None = 60):
    wb = load_workbook(file_obj, data_only=True)
    if "Consolidation" not in wb.sheetnames:
        raise ValueError("Feuille 'Consolidation' introuvable dans le fichier.")
    ws = wb["Consolidation"]
    header = None
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [_normalize_cell(c) for c in row][:MAX_CONSO_COLS]
        if header is None:
            header = cells
            continue
        if not any(cells):
            continue
        rows.append(cells)
        if max_rows and len(rows) >= max_rows:
            break
    return header or [], rows


def _rows_to_records(header, rows):
    header_norm = [_normalize_header(h) for h in header]

    # Mapping index → field
    header_map = {}
    for idx, h in enumerate(header_norm):
        field = CONSOLIDATION_HEADER_MAP.get(h)
        if field:
            header_map[idx] = field

    records = []
    related_payload = []

    for row in rows:
        cells = [_normalize_cell(c) for c in row]
        if not any(cells):
            continue

        data = {}
        for idx, field in header_map.items():
            if idx < len(cells):
                data[field] = cells[idx]

        # Debug rapide (à commenter après test)
        # if "classe_id" not in data or not data["classe_id"]:
        #     print("Ligne sans classe_id →", data.get("nom_complet", "inconnu"))

        records.append(
            ConsolidationRecord(
                numero=data.get("numero", ""),
                nom_complet=data.get("nom_complet", ""),
                beneficiaire=data.get("beneficiaire", ""),
                genre=data.get("genre", ""),
                age=_to_int(data.get("age")),
                fonction=data.get("fonction", ""),
                qualification=data.get("qualification", ""),
                nb_annees_experience=_to_int(data.get("nb_annees_experience")),
                ville_residence=data.get("ville_residence", ""),
                prestataire=data.get("prestataire", ""),
                intitule_formation_solicitee=data.get("intitule_formation_solicitee", ""),
                intitule_formation_dispensee=data.get("intitule_formation_dispensee", ""),
                fenetre=data.get("fenetre", ""),
                ville_formation=data.get("ville_formation", ""),
                arrondissement=data.get("arrondissement", ""),
                departement=data.get("departement", ""),
                region=data.get("region", ""),
                lieu_formation=data.get("lieu_formation", ""),
                precision_lieu=data.get("precision_lieu", ""),
                longitude=data.get("longitude", ""),
                latitude=data.get("latitude", ""),
                telephone1=data.get("telephone1", ""),
                telephone2=data.get("telephone2", ""),
                cohorte=data.get("cohorte", ""),
                tel_formateur=data.get("tel_formateur", ""),
                code=data.get("code", ""),
                cout_unitaire_subvention=_to_decimal(data.get("cout_unitaire_subvention")),
                montant_total_subvention=_to_decimal(data.get("montant_total_subvention")),
                statut_prestation=data.get("statut_prestation", ""),
            )
        )
        related_payload.append(data)

    return records, related_payload


def _extract_unique_classe_ids(payload: list[dict]) -> list[str]:
    classes = {(item.get("classe_id") or "").strip() for item in payload}
    return sorted(code for code in classes if code)


def _save_related_from_payload(payload: list[dict]):
    seen_benef = {}
    seen_prest = {}
    seen_form = {}
    seen_lieu = {}
    seen_classe = {}
    created_apprenants = 0

    for item in payload:
        ben_name = item.get("beneficiaire", "").strip()
        ben_key = ben_name.lower()
        prest_name = item.get("prestataire", "").strip()
        prest_key = prest_name.lower()
        intitule = (
            item.get("intitule_formation_dispensee")
            or item.get("intitule_formation_solicitee")
            or ""
        )
        intitule_key = intitule.lower()
        fenetre = item.get("fenetre", "") or ""
        lieu_nom = item.get("lieu_formation", "").strip()
        lieu_key = lieu_nom.lower()
        classe_id = (item.get("classe_id") or "").strip()
        cohorte_raw = item.get("cohorte", "")

        region = item.get("region", "").strip()
        departement = item.get("departement", "").strip()
        arrondissement = item.get("arrondissement", "").strip()
        ville = item.get("ville_formation", "").strip()

        beneficiaire = seen_benef.get(ben_key) or _ensure_beneficiaire(
            ben_name,
            region=region,
            departement=departement,
            arrondissement=arrondissement,
            ville=ville,
        )
        if beneficiaire:
            seen_benef[ben_key] = beneficiaire

        prestataire = seen_prest.get(prest_key) or _ensure_prestataire(prest_name)
        if prestataire:
            seen_prest[prest_key] = prestataire

        formation = seen_form.get(intitule_key) or _ensure_formation(intitule, fenetre=fenetre)
        if formation:
            seen_form[intitule_key] = formation

        lieu = seen_lieu.get(lieu_key) or _ensure_lieu(
            lieu_nom,
            region=region,
            departement=departement,
            arrondissement=arrondissement,
            ville=ville,
            longitude=item.get("longitude", ""),
            latitude=item.get("latitude", ""),
            precision=item.get("precision_lieu", ""),
        )
        if lieu:
            seen_lieu[lieu_key] = lieu

        prestation = _ensure_prestation(
            prestataire, formation, beneficiaire, code_hint=item.get("code", "")
        )

        # Clé unique pour la classe : on priorise classe_id si présent
        classe_key = (
            classe_id.lower()
            if classe_id
            else f"{prestation.id if prestation else 'noprest'}-{fenetre}-{cohorte_raw}".lower()
        )

        classe = seen_classe.get(classe_key)
        if not classe:
            classe = _ensure_classe(
                prestation,
                formation,
                fenetre=fenetre,
                cohorte=cohorte_raw,
                classe_id=classe_id,
            )
            if classe:
                seen_classe[classe_key] = classe

        if classe and formation:
            code_appr = (item.get("code") or "").strip()
            if not code_appr:
                code_appr = f"AP-{slugify(item.get('nom_complet', '')[:12])}-{classe.code[:8]}"
            tel1 = (item.get("telephone1") or "").strip()
            tel2 = (item.get("telephone2") or "").strip()
            defaults = {
                "nom_complet": item.get("nom_complet", ""),
                "genre": item.get("genre", ""),
                "age": _to_int(item.get("age")),
                "fonction": item.get("fonction", ""),
                "qualification": item.get("qualification", ""),
                "nb_annees_experience": _to_int(item.get("nb_annees_experience")) or 0,
                "fenetre": fenetre,
                "telephone1": tel1 or None,
                "telephone2": tel2 or None,
                "ville_residence": item.get("ville_residence", ""),
                "region": region,
                "departement": departement,
                "arrondissement": arrondissement,
                "code_ville": item.get("ville_formation", ""),
                "appartenance_beneficiaire": True,
            }
            try:
                existing = None
                if tel1:
                    existing = Apprenant.objects.filter(
                        formation=formation, telephone1=tel1
                    ).first()
                if existing:
                    for field, val in defaults.items():
                        setattr(existing, field, val)
                    existing.classe = classe
                    existing.formation = formation
                    existing.save()
                else:
                    obj, created = Apprenant.objects.get_or_create(
                        code=code_appr,
                        defaults={**defaults, "classe": classe, "formation": formation},
                    )
                    if created:
                        created_apprenants += 1
            except Exception:
                continue  # on saute la ligne en cas de conflit unique

    return created_apprenants


# Fonction d'analyse des headers (a conserver ou a reactiver pour debug)
def _analyze_headers(headers):
    normalized_headers = [_normalize_header(h or "") for h in headers]
    mapped_fields = set()
    for h in normalized_headers:
        mapped = CONSOLIDATION_HEADER_MAP.get(h)
        if mapped:
            mapped_fields.add(mapped)
    expected = {
        v
        for v in CONSOLIDATION_HEADER_MAP.values()
        if not v.startswith("cout") and not v.startswith("montant") and not v.startswith("statut")
    }
    missing = sorted(expected - mapped_fields)
    extras = [
        headers[idx]
        for idx, h in enumerate(normalized_headers)
        if h not in CONSOLIDATION_HEADER_MAP and h
    ]
    return {
        "mapped": sorted(mapped_fields),
        "missing": missing,
        "extras": [e for e in extras if e],
    }


def _read_consolidation_sheet(file_obj, max_rows: int | None = 60):
    wb = load_workbook(file_obj, data_only=True)
    if "Consolidation" not in wb.sheetnames:
        raise ValueError("Feuille 'Consolidation' introuvable dans le fichier.")
    ws = wb["Consolidation"]
    header = None
    rows = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Stop after the expected columns, including "Classe ID".
        cells = [_normalize_cell(c) for c in row][:MAX_CONSO_COLS]
        if header is None:
            header = cells
            continue
        if not any(cells):
            continue
        rows.append(cells)
        if max_rows and len(rows) >= max_rows:
            break
    return header or [], rows


def _rows_to_records(header, rows):
    header_norm = [_normalize_header(h) for h in header[:MAX_CONSO_COLS]]
    header_map = {
        idx: CONSOLIDATION_HEADER_MAP[h]
        for idx, h in enumerate(header_norm)
        if h in CONSOLIDATION_HEADER_MAP
    }
    records = []
    related_payload = []
    for row in rows:
        cells = [_normalize_cell(c) for c in row][:29]
        if not any(cells):
            continue
        data = {}
        for idx, field in header_map.items():
            if idx < len(cells):
                data[field] = cells[idx]
        records.append(
            ConsolidationRecord(
                numero=data.get("numero", ""),
                nom_complet=data.get("nom_complet", ""),
                beneficiaire=data.get("beneficiaire", ""),
                genre=data.get("genre", ""),
                age=_to_int(data.get("age")),
                fonction=data.get("fonction", ""),
                qualification=data.get("qualification", ""),
                nb_annees_experience=_to_int(data.get("nb_annees_experience")),
                ville_residence=data.get("ville_residence", ""),
                prestataire=data.get("prestataire", ""),
                intitule_formation_solicitee=data.get("intitule_formation_solicitee", ""),
                intitule_formation_dispensee=data.get("intitule_formation_dispensee", ""),
                fenetre=data.get("fenetre", ""),
                ville_formation=data.get("ville_formation", ""),
                arrondissement=data.get("arrondissement", ""),
                departement=data.get("departement", ""),
                region=data.get("region", ""),
                lieu_formation=data.get("lieu_formation", ""),
                precision_lieu=data.get("precision_lieu", ""),
                longitude=data.get("longitude", ""),
                latitude=data.get("latitude", ""),
                telephone1=data.get("telephone1", ""),
                telephone2=data.get("telephone2", ""),
                cohorte=data.get("cohorte", ""),
                tel_formateur=data.get("tel_formateur", ""),
                code=data.get("code", ""),
                cout_unitaire_subvention=_to_decimal(data.get("cout_unitaire_subvention")),
                montant_total_subvention=_to_decimal(data.get("montant_total_subvention")),
                statut_prestation=data.get("statut_prestation", ""),
            )
        )
        related_payload.append(data)
    return records, related_payload


def _save_related_from_payload(payload: list[dict]):
    # Deduplicate creations for speed.
    seen_benef = {}
    seen_prest = {}
    seen_form = {}
    seen_lieu = {}
    seen_classe = {}
    created_apprenants = 0

    for item in payload:
        ben_name = item.get("beneficiaire", "").strip()
        ben_key = ben_name.lower()
        prest_name = item.get("prestataire", "").strip()
        prest_key = prest_name.lower()
        intitule = (
            item.get("intitule_formation_dispensee")
            or item.get("intitule_formation_solicitee")
            or ""
        )
        intitule_key = intitule.lower()
        fenetre = item.get("fenetre", "") or ""
        lieu_nom = item.get("lieu_formation", "").strip()
        lieu_key = lieu_nom.lower()
        classe_id = (item.get("classe_id") or "").strip()
        classe_id_key = classe_id.lower()
        cohorte_raw = item.get("cohorte", "")

        region = item.get("region", "").strip()
        departement = item.get("departement", "").strip()
        arrondissement = item.get("arrondissement", "").strip()
        ville = item.get("ville_formation", "").strip()

        beneficiaire = seen_benef.get(ben_key) or _ensure_beneficiaire(
            ben_name,
            region=region,
            departement=departement,
            arrondissement=arrondissement,
            ville=ville,
        )
        if beneficiaire:
            seen_benef[ben_key] = beneficiaire

        prestataire = seen_prest.get(prest_key) or _ensure_prestataire(prest_name)
        if prestataire:
            seen_prest[prest_key] = prestataire

        formation = seen_form.get(intitule_key) or _ensure_formation(intitule, fenetre=fenetre)
        if formation:
            seen_form[intitule_key] = formation

        lieu = seen_lieu.get(lieu_key) or _ensure_lieu(
            lieu_nom,
            region=region,
            departement=departement,
            arrondissement=arrondissement,
            ville=ville,
            longitude=item.get("longitude", ""),
            latitude=item.get("latitude", ""),
            precision=item.get("precision_lieu", ""),
        )
        if lieu:
            seen_lieu[lieu_key] = lieu

        prestation = _ensure_prestation(
            prestataire, formation, beneficiaire, code_hint=item.get("code", "")
        )
        classe_key = (
            classe_id_key
            or f"{prestation.id if prestation else 'noprest'}-{fenetre}-{cohorte_raw}".lower()
        )
        classe = seen_classe.get(classe_key)
        if not classe:
            classe = _ensure_classe(
                prestation,
                formation,
                fenetre=fenetre,
                cohorte=cohorte_raw,
                classe_id=classe_id,
            )
            if classe:
                seen_classe[classe_key] = classe

        if classe and formation:
            code_appr = (item.get("code") or "").strip()
            if not code_appr:
                code_appr = f"AP-{slugify(item.get('nom_complet', ''))[:6]}-{classe.code[:6]}"
            tel1 = (item.get("telephone1") or "").strip()
            tel2 = (item.get("telephone2") or "").strip()
            defaults = {
                "nom_complet": item.get("nom_complet", ""),
                "genre": item.get("genre", ""),
                "age": _to_int(item.get("age")),
                "fonction": item.get("fonction", ""),
                "qualification": item.get("qualification", ""),
                "nb_annees_experience": _to_int(item.get("nb_annees_experience")) or 0,
                "fenetre": fenetre,
                "telephone1": tel1 or None,
                "telephone2": tel2 or None,
                "ville_residence": item.get("ville_residence", ""),
                "region": region,
                "departement": departement,
                "arrondissement": arrondissement,
                "code_ville": item.get("ville_formation", ""),
                "appartenance_beneficiaire": True,
            }
            try:
                # Si un apprenant existe deja avec le meme tel1 dans la formation, on le met a jour.
                existing = None
                if tel1:
                    existing = Apprenant.objects.filter(
                        formation=formation, telephone1=tel1
                    ).first()
                if existing:
                    for field, val in defaults.items():
                        setattr(existing, field, val)
                    existing.classe = classe
                    existing.formation = formation
                    existing.save()
                else:
                    obj, created = Apprenant.objects.get_or_create(
                        code=code_appr,
                        defaults={**defaults, "classe": classe, "formation": formation},
                    )
                    if created:
                        created_apprenants += 1
            except Exception:
                # En cas de conflit unique, on ignore la ligne pour ne pas casser l'import.
                continue
    return created_apprenants


@require_superadmin_access
def consolidation_view(request):
    form = ConsolidationUploadForm(request.POST or None, request.FILES or None)
    headers = []
    preview_rows = []
    analysis = {"mapped": [], "missing": [], "extras": []}
    errors = []
    file_meta = request.session.get(SESSION_KEY_CONSO, {}).get("meta", {})
    save_requested = bool(request.POST.get("save"))
    prestataire_filter = (request.GET.get("prestataire") or "").strip()
    beneficiaire_filter = (request.GET.get("beneficiaire") or "").strip()

    prestataires = (
        Apprenant.objects.exclude(prestataire__isnull=True)
        .exclude(prestataire__exact="")
        .values_list("prestataire", flat=True)
        .distinct()
        .order_by("prestataire")
    )
    beneficiaires = (
        Apprenant.objects.exclude(beneficiaire__isnull=True)
        .exclude(beneficiaire__exact="")
        .values_list("beneficiaire", flat=True)
        .distinct()
        .order_by("beneficiaire")
    )
    apprenants_qs = Apprenant.objects.select_related("classe", "formation")
    if prestataire_filter:
        apprenants_qs = apprenants_qs.filter(prestataire=prestataire_filter)
    if beneficiaire_filter:
        apprenants_qs = apprenants_qs.filter(beneficiaire=beneficiaire_filter)
    apprenants = list(apprenants_qs.order_by("nom_complet"))
    apprenant_rows = [
        [_apprenant_value(apprenant, key) for key, _ in APPRENANT_COLUMNS]
        for apprenant in apprenants
    ]
    appel_prestataire_filters = [
        p.strip() for p in request.GET.getlist("appel_prestataire") if p.strip()
    ]
    appel_beneficiaire_filter = (request.GET.get("appel_beneficiaire") or "").strip()
    appel_filter_mode = (request.GET.get("appel_filter_mode") or "and").lower()
    if appel_filter_mode not in {"and", "or"}:
        appel_filter_mode = "and"
    appel_status_filter = (request.GET.get("appel_status") or "termine").strip()
    valid_statuses = {code for code, _ in Appel.STATUS_CHOICES}
    if appel_status_filter not in valid_statuses:
        appel_status_filter = "termine"

    appel_prestataires = (
        Appel.objects.filter(status=appel_status_filter)
        .exclude(prestataire__isnull=True)
        .exclude(prestataire__exact="")
        .values_list("prestataire", flat=True)
        .distinct()
        .order_by("prestataire")
    )
    appel_beneficiaires = (
        Appel.objects.filter(status=appel_status_filter)
        .exclude(beneficiaire__isnull=True)
        .exclude(beneficiaire__exact="")
        .values_list("beneficiaire", flat=True)
        .distinct()
        .order_by("beneficiaire")
    )

    appels_termine_qs = Appel.objects.filter(status=appel_status_filter).select_related("classe")
    if appel_prestataire_filters or appel_beneficiaire_filter:
        q = Q()
        if appel_filter_mode == "or":
            if appel_prestataire_filters:
                q |= Q(prestataire__in=appel_prestataire_filters)
            if appel_beneficiaire_filter:
                q |= Q(beneficiaire=appel_beneficiaire_filter)
        else:
            if appel_prestataire_filters:
                q &= Q(prestataire__in=appel_prestataire_filters)
            if appel_beneficiaire_filter:
                q &= Q(beneficiaire=appel_beneficiaire_filter)
        appels_termine_qs = appels_termine_qs.filter(q)
    appels_termine = list(appels_termine_qs.order_by("nom"))
    appels_termine_rows = [
        [_appel_value(appel, key) for key, _ in APPEL_TERMINE_COLUMNS] for appel in appels_termine
    ]
    prestataire_appels_termine = (
        appels_termine_qs.values("prestataire")
        .annotate(total=Count("id"))
        .order_by("-total", "prestataire")
    )
    appel_querystring = urlencode(
        {
            "appel_prestataire": appel_prestataire_filters,
            "appel_beneficiaire": appel_beneficiaire_filter,
            "appel_filter_mode": appel_filter_mode,
            "appel_status": appel_status_filter,
        },
        doseq=True,
    )
    extract_requested = bool(request.POST.get("extract_classes"))
    unique_classe_ids: list[str] = []

    if request.method == "POST" and form.is_valid():
        fichier = form.cleaned_data.get("fichier")
        try:
            content = None
            if fichier:
                content = fichier.read()
                file_meta = {
                    "name": getattr(fichier, "name", ""),
                    "size": getattr(fichier, "size", 0),
                }
                request.session[SESSION_KEY_CONSO] = {
                    "meta": file_meta,
                    "b64": base64.b64encode(content).decode("ascii"),
                }
                request.session.modified = True
            elif save_requested and request.session.get(SESSION_KEY_CONSO):
                cached = request.session.get(SESSION_KEY_CONSO, {})
                b64 = cached.get("b64")
                if b64:
                    content = base64.b64decode(b64)
                    file_meta = cached.get("meta", {})
            if not content:
                raise ValueError("Veuillez charger un fichier consolide avant de valider.")

            buffer_preview = io.BytesIO(content)
            headers, preview_rows = _read_consolidation_sheet(buffer_preview, max_rows=60)
            analysis = _analyze_headers(headers)
            if save_requested or extract_requested:
                buffer_full = io.BytesIO(content)
                full_headers, all_rows = _read_consolidation_sheet(buffer_full, max_rows=None)
                records, payload = _rows_to_records(full_headers, all_rows)
                if not records:
                    raise ValueError("Aucune ligne valide a enregistrer.")
                unique_classe_ids = _extract_unique_classe_ids(payload)
                if save_requested:
                    try:
                        # Always wipe before inserting, even if the insert later fails, to behave like a seed/replace.  # noqa: E501
                        _reset_consolidation_tables()
                        ConsolidationRecord.objects.all().delete()
                        with transaction.atomic():
                            ConsolidationRecord.objects.bulk_create(records, ignore_conflicts=False)
                            _save_related_from_payload(payload)
                        messages.success(
                            request,
                            f"{len(records)} lignes consolidees enregistrees (remplacement complet).",  # noqa: E501
                        )
                    except OperationalError:
                        errors.append(
                            "Base de donnees occupee (database locked). Reessayez dans un instant."
                        )
        except Exception as exc:  # pragma: no cover - runtime feedback
            errors.append(str(exc))
            preview_rows = []

    return render(
        request,
        "reporting/consolidation.html",
        {
            "form": form,
            "headers": headers,
            "preview_rows": preview_rows,
            "analysis": analysis,
            "errors": errors,
            "file_meta": file_meta,
            "apprenant_columns": APPRENANT_COLUMNS,
            "apprenant_rows": apprenant_rows,
            "apprenant_count": len(apprenants),
            "prestataires": prestataires,
            "beneficiaires": beneficiaires,
            "selected_prestataire": prestataire_filter,
            "selected_beneficiaire": beneficiaire_filter,
            "unique_classe_ids": unique_classe_ids,
            "appels_termine_columns": APPEL_TERMINE_COLUMNS,
            "appels_termine_rows": appels_termine_rows,
            "appels_termine_count": len(appels_termine),
            "prestataire_appels_termine": prestataire_appels_termine,
            "appel_prestataires": appel_prestataires,
            "appel_beneficiaires": appel_beneficiaires,
            "selected_appel_prestataire": appel_prestataire_filters,
            "selected_appel_beneficiaire": appel_beneficiaire_filter,
            "selected_appel_filter_mode": appel_filter_mode,
            "selected_appel_status": appel_status_filter,
            "appel_status_choices": Appel.STATUS_CHOICES,
            "appel_querystring": appel_querystring,
        },
    )


@require_analysis_access
def export_consolidation_apprenants_excel(request):
    prestataire_filter = (request.GET.get("prestataire") or "").strip()
    beneficiaire_filter = (request.GET.get("beneficiaire") or "").strip()

    apprenants_qs = Apprenant.objects.select_related("classe", "formation")
    if prestataire_filter:
        apprenants_qs = apprenants_qs.filter(prestataire=prestataire_filter)
    if beneficiaire_filter:
        apprenants_qs = apprenants_qs.filter(beneficiaire=beneficiaire_filter)

    wb = Workbook()
    ws = wb.active
    ws.title = "Apprenants"
    ws.append([label for _, label in APPRENANT_COLUMNS])
    for apprenant in apprenants_qs.order_by("nom_complet"):
        ws.append([_excel_safe(_apprenant_value(apprenant, key)) for key, _ in APPRENANT_COLUMNS])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=apprenants_consolidation.xlsx"
    wb.save(response)
    return response


@require_analysis_access
def export_appels_termines_excel(request):
    appel_prestataire_filters = [
        p.strip() for p in request.GET.getlist("appel_prestataire") if p.strip()
    ]
    appel_beneficiaire_filter = (request.GET.get("appel_beneficiaire") or "").strip()
    appel_filter_mode = (request.GET.get("appel_filter_mode") or "and").lower()
    if appel_filter_mode not in {"and", "or"}:
        appel_filter_mode = "and"
    appel_status_filter = (request.GET.get("appel_status") or "termine").strip()
    valid_statuses = {code for code, _ in Appel.STATUS_CHOICES}
    if appel_status_filter not in valid_statuses:
        appel_status_filter = "termine"

    appels_qs = Appel.objects.filter(status=appel_status_filter).select_related("classe")
    if appel_prestataire_filters or appel_beneficiaire_filter:
        q = Q()
        if appel_filter_mode == "or":
            if appel_prestataire_filters:
                q |= Q(prestataire__in=appel_prestataire_filters)
            if appel_beneficiaire_filter:
                q |= Q(beneficiaire=appel_beneficiaire_filter)
        else:
            if appel_prestataire_filters:
                q &= Q(prestataire__in=appel_prestataire_filters)
            if appel_beneficiaire_filter:
                q &= Q(beneficiaire=appel_beneficiaire_filter)
        appels_qs = appels_qs.filter(q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Appels termines"
    ws.append([label for _, label in APPEL_TERMINE_COLUMNS])
    for appel in appels_qs.order_by("nom"):
        ws.append([_excel_safe(_appel_value(appel, key)) for key, _ in APPEL_TERMINE_COLUMNS])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=appels_termines.xlsx"
    wb.save(response)
    return response


@require_analysis_access
def export_appels_termines_csv(request):
    appel_prestataire_filters = [
        p.strip() for p in request.GET.getlist("appel_prestataire") if p.strip()
    ]
    appel_beneficiaire_filter = (request.GET.get("appel_beneficiaire") or "").strip()
    appel_filter_mode = (request.GET.get("appel_filter_mode") or "and").lower()
    if appel_filter_mode not in {"and", "or"}:
        appel_filter_mode = "and"
    appel_status_filter = (request.GET.get("appel_status") or "termine").strip()
    valid_statuses = {code for code, _ in Appel.STATUS_CHOICES}
    if appel_status_filter not in valid_statuses:
        appel_status_filter = "termine"

    appels_qs = Appel.objects.filter(status=appel_status_filter).select_related("classe")
    if appel_prestataire_filters or appel_beneficiaire_filter:
        q = Q()
        if appel_filter_mode == "or":
            if appel_prestataire_filters:
                q |= Q(prestataire__in=appel_prestataire_filters)
            if appel_beneficiaire_filter:
                q |= Q(beneficiaire=appel_beneficiaire_filter)
        else:
            if appel_prestataire_filters:
                q &= Q(prestataire__in=appel_prestataire_filters)
            if appel_beneficiaire_filter:
                q &= Q(beneficiaire=appel_beneficiaire_filter)
        appels_qs = appels_qs.filter(q)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=appels_termines.csv"
    writer = csv.writer(response)
    writer.writerow([label for _, label in APPEL_TERMINE_COLUMNS])
    for appel in appels_qs.order_by("nom"):
        writer.writerow([_excel_safe(_appel_value(appel, key)) for key, _ in APPEL_TERMINE_COLUMNS])
    return response


def safe_rate(num: float, den: float) -> float:
    return round((num / den) * 100, 2) if den else 0.0


@require_analysis_access
def reporting_home(request):
    nb_classes = Classe.objects.count()
    nb_apprenants = Apprenant.objects.count()
    nb_formateurs = SatisfactionFormateur.objects.values("formateur").distinct().count()
    nb_enquetes_presence = Presence.objects.count()
    nb_sat_apprenants = SatisfactionApprenant.objects.count()
    nb_sat_formateurs = SatisfactionFormateur.objects.count()
    nb_env = EnqueteEnvironnement.objects.count()

    presence_rates = (
        Presence.objects.values("classe__code")
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")[:10]
    )
    sat_appr_moy = (
        SatisfactionApprenant.objects.values("classe__code")
        .annotate(moy=Avg("q9_satisfaction_globale"))
        .order_by("-moy")[:10]
    )
    sat_form_moy = (
        SatisfactionFormateur.objects.values("classe__code")
        .annotate(moy=Avg("q9_satisfaction_globale_prestataire"))
        .order_by("-moy")[:10]
    )

    # Synthèses globales
    total_pres = Presence.objects.count()
    total_pr = Presence.objects.filter(presence="PR").count()
    taux_presence_global = safe_rate(total_pr, total_pres)

    # RES00-04
    sat_appr_agg = SatisfactionApprenant.objects.aggregate(
        total_q9=Sum("q9_satisfaction_globale"), count=Count("id")
    )
    sat_appr_sum = sat_appr_agg["total_q9"] or 0
    sat_appr_count = sat_appr_agg["count"] or 0
    taux_sat_appr_global = safe_rate(sat_appr_sum, sat_appr_count * 5)
    # RES00-05
    sat_form_agg = SatisfactionFormateur.objects.aggregate(
        total_q9=Sum("q9_satisfaction_globale_prestataire"), count=Count("id")
    )
    sat_form_sum = sat_form_agg["total_q9"] or 0
    sat_form_count = sat_form_agg["count"] or 0
    taux_sat_form_global = safe_rate(sat_form_sum, sat_form_count * 5)

    # Environnement : moyenne des booleens principaux (8 points indicatifs)
    env_bool_fields = [
        "tables",
        "chaises",
        "ecran",
        "videoprojecteur",
        "ventilation",
        "eclairage",
        "salle_propre",
        "salle_securisee",
    ]
    env_counts = EnqueteEnvironnement.objects.aggregate(
        total=Count("id"), **{f: Sum(f) for f in env_bool_fields}
    )
    env_score = safe_rate(
        sum((env_counts.get(f) or 0) for f in env_bool_fields),
        (env_counts.get("total") or 0) * len(env_bool_fields),
    )

    # Taux presence par axes
    def with_rate(qs):
        return [{**r, "taux": safe_rate(r.get("pr") or 0, r.get("total") or 0)} for r in qs]

    taux_presence_prestataire = with_rate(
        Presence.objects.values("classe__prestation__prestataire__raison_sociale")
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")
    )
    taux_presence_prestation = with_rate(
        Presence.objects.values("classe__prestation__code")
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")
    )
    taux_presence_beneficiaire = with_rate(
        Presence.objects.values("classe__prestation__beneficiaire__nom_structure")
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")
    )
    taux_presence_formation = with_rate(
        Presence.objects.values("classe__formation__nom")
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")
    )
    taux_presence_formation_harmo = with_rate(
        Presence.objects.values("classe__formation__nom_harmonise")
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")
    )

    # Satisfaction globale par axes (apprenants)
    sat_appr_prestataire = (
        SatisfactionApprenant.objects.values("classe__prestation__prestataire__raison_sociale")
        .annotate(moy=Avg("q9_satisfaction_globale"))
        .order_by("-moy")
    )
    sat_appr_prestation = (
        SatisfactionApprenant.objects.values("classe__prestation__code")
        .annotate(moy=Avg("q9_satisfaction_globale"))
        .order_by("-moy")
    )
    sat_appr_benef = (
        SatisfactionApprenant.objects.values("classe__prestation__beneficiaire__nom_structure")
        .annotate(moy=Avg("q9_satisfaction_globale"))
        .order_by("-moy")
    )
    sat_appr_formation = (
        SatisfactionApprenant.objects.values("classe__formation__nom")
        .annotate(moy=Avg("q9_satisfaction_globale"))
        .order_by("-moy")
    )
    sat_appr_formation_harmo = (
        SatisfactionApprenant.objects.values("classe__formation__nom_harmonise")
        .annotate(moy=Avg("q9_satisfaction_globale"))
        .order_by("-moy")
    )

    # Satisfaction globale par axes (formateurs)
    sat_form_prestataire = (
        SatisfactionFormateur.objects.values("classe__prestation__prestataire__raison_sociale")
        .annotate(moy=Avg("q9_satisfaction_globale_prestataire"))
        .order_by("-moy")
    )
    sat_form_prestation = (
        SatisfactionFormateur.objects.values("classe__prestation__code")
        .annotate(moy=Avg("q9_satisfaction_globale_prestataire"))
        .order_by("-moy")
    )
    sat_form_benef = (
        SatisfactionFormateur.objects.values("classe__prestation__beneficiaire__nom_structure")
        .annotate(moy=Avg("q9_satisfaction_globale_prestataire"))
        .order_by("-moy")
    )
    sat_form_formation = (
        SatisfactionFormateur.objects.values("classe__formation__nom")
        .annotate(moy=Avg("q9_satisfaction_globale_prestataire"))
        .order_by("-moy")
    )
    sat_form_formation_harmo = (
        SatisfactionFormateur.objects.values("classe__formation__nom_harmonise")
        .annotate(moy=Avg("q9_satisfaction_globale_prestataire"))
        .order_by("-moy")
    )

    repart_apprenants_ville = (
        Apprenant.objects.values("ville_residence").annotate(total=Count("id")).order_by("-total")
    )
    repart_apprenants_region = (
        Apprenant.objects.values("region").annotate(total=Count("id")).order_by("-total")
    )
    repart_apprenants_formation = (
        Apprenant.objects.values("formation__nom").annotate(total=Count("id")).order_by("-total")
    )
    repart_apprenants_formation_harmo = (
        Apprenant.objects.values("formation__nom_harmonise")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    repart_apprenants_benef = (
        Apprenant.objects.values("classe__prestation__beneficiaire__nom_structure")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    repart_apprenants_prestataire = (
        Apprenant.objects.values("classe__prestation__prestataire__raison_sociale")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # Effectifs / femmes / appartenance par prestation
    prestations_effectifs = (
        Prestation.objects.annotate(
            appr_total=Count("classes__apprenants", distinct=True),
            appr_femmes=Count(
                "classes__apprenants",
                filter=Q(classes__apprenants__genre__iexact="f"),
                distinct=True,
            ),
            appr_appart=Count(
                "classes__apprenants",
                filter=Q(classes__apprenants__appartenance_beneficiaire=True),
                distinct=True,
            ),
        )
        .values(
            "code",
            "effectif_a_former",
            "femmes",
            "appr_total",
            "appr_femmes",
            "appr_appart",
            "prestataire__raison_sociale",
            "beneficiaire__nom_structure",
        )
        .order_by("code")
    )

    # Durées par prestation
    prestations_durees = Prestation.objects.values(
        "code", "prestataire__raison_sociale", "duree_prevue_heures", "duree_reelle_heures"
    ).order_by("code")

    # Environnement par lieu
    env_fields = [
        "tables",
        "chaises",
        "ecran",
        "videoprojecteur",
        "ventilation",
        "eclairage",
        "salle_propre",
        "salle_securisee",
    ]
    env_qs = (
        EnqueteEnvironnement.objects.values("classe__lieu__nom_lieu", "classe__lieu__region")
        .annotate(
            total=Count("id"),
            **{f: Sum(f) for f in env_fields},
        )
        .order_by("-total")
    )
    env_par_lieu = []
    for row in env_qs:
        total = row.get("total") or 0
        somme = sum((row.get(f) or 0) for f in env_fields)
        score = safe_rate(somme, total * len(env_fields))
        env_par_lieu.append(
            {
                "lieu": row.get("classe__lieu__nom_lieu"),
                "region": row.get("classe__lieu__region"),
                "total": total,
                "score": score,
            }
        )

    # Logic for RES02-03, RES02-04, RES02-05
    prestations_effectifs_list = []
    for p in prestations_effectifs:
        eff_ok = (p["appr_total"] or 0) >= (p["effectif_a_former"] or 0)
        femmes_ok = (p["appr_femmes"] or 0) >= (p["femmes"] or 0)
        p["respect_effectif"] = eff_ok
        p["respect_femmes"] = femmes_ok
        p["taux_appartenance"] = safe_rate(p["appr_appart"] or 0, p["appr_total"] or 0)
        prestations_effectifs_list.append(p)

    # Logic for RES03-01
    from App_PADESCE.formations.models import Lieu

    carte_lieux = []
    for l in Lieu.objects.filter(actif=True):  # noqa: E741
        if l.latitude and l.longitude:
            carte_lieux.append({"nom": l.nom_lieu, "lat": l.latitude, "lng": l.longitude})

    charts = [
        {
            "code": "RES00-01",
            "title": "Taux de présence global",
            "value": f"{taux_presence_global} %",
        },
        {"code": "RES00-02", "title": "Synthèse du suivi contractuel", "value": "N/A"},
        {
            "code": "RES00-03",
            "title": "Synthèse de l'évaluation de l'environnement (8 points)",
            "value": f"{env_score} %",
        },
        {
            "code": "RES00-04",
            "title": "Taux de satisfaction global apprenants",
            "value": f"{taux_sat_appr_global} %",
        },
        {
            "code": "RES00-05",
            "title": "Taux de satisfaction global formateurs",
            "value": f"{taux_sat_form_global} %",
        },
    ]

    context = {
        "nb_classes": nb_classes,
        "nb_apprenants": nb_apprenants,
        "nb_formateurs": nb_formateurs,
        "nb_enquetes_presence": nb_enquetes_presence,
        "nb_sat_apprenants": nb_sat_apprenants,
        "nb_sat_formateurs": nb_sat_formateurs,
        "nb_env": nb_env,
        "presence_rates": presence_rates,
        "sat_appr_moy": sat_appr_moy,
        "sat_form_moy": sat_form_moy,
        "formations": Formation.objects.all().order_by("nom")[:20],
        "charts": charts,
        "carte_lieux": carte_lieux,
        "taux_presence_prestataire": taux_presence_prestataire,
        "taux_presence_prestation": taux_presence_prestation,
        "taux_presence_beneficiaire": taux_presence_beneficiaire,
        "taux_presence_formation": taux_presence_formation,
        "taux_presence_formation_harmo": taux_presence_formation_harmo,
        "sat_appr_prestataire": sat_appr_prestataire,
        "sat_appr_prestation": sat_appr_prestation,
        "sat_appr_benef": sat_appr_benef,
        "sat_appr_formation": sat_appr_formation,
        "sat_appr_formation_harmo": sat_appr_formation_harmo,
        "sat_form_prestataire": sat_form_prestataire,
        "sat_form_prestation": sat_form_prestation,
        "sat_form_benef": sat_form_benef,
        "sat_form_formation": sat_form_formation,
        "sat_form_formation_harmo": sat_form_formation_harmo,
        "repart_apprenants_ville": repart_apprenants_ville,
        "repart_apprenants_region": repart_apprenants_region,
        "repart_apprenants_formation": repart_apprenants_formation,
        "repart_apprenants_formation_harmo": repart_apprenants_formation_harmo,
        "repart_apprenants_benef": repart_apprenants_benef,
        "repart_apprenants_prestataire": repart_apprenants_prestataire,
        "prestations_effectifs": prestations_effectifs_list,
        "prestations_durees": prestations_durees,
        "env_par_lieu": env_par_lieu,
    }
    return render(request, "reporting/index.html", context)


def _table_presence_rates(field: str):
    return (
        Presence.objects.values(field)
        .annotate(total=Count("id"), pr=Count("id", filter=Q(presence="PR")))
        .order_by("-total")
    )


def _table_sat_avg(model, field: str, display: str):
    return model.objects.values(display).annotate(moy=Avg(field)).order_by("-moy")


def get_table_data(code: str) -> dict:
    code = code.lower()
    if code == "presence-classe":
        qs = _table_presence_rates("classe__code")[:10]
        return {
            "title": "Top presence (classe)",
            "headers": ["Classe", "PR", "Total"],
            "rows": [[r["classe__code"], r["pr"], r["total"]] for r in qs],
        }
    if code == "sat-appr-q9":
        qs = _table_sat_avg(SatisfactionApprenant, "q9_satisfaction_globale", "classe__code")[:10]
        return {
            "title": "Sat. apprenants (Q9)",
            "headers": ["Classe", "Moyenne"],
            "rows": [[r["classe__code"], round(r["moy"] or 0, 2)] for r in qs],
        }
    if code == "sat-form-q9":
        qs = _table_sat_avg(
            SatisfactionFormateur, "q9_satisfaction_globale_prestataire", "classe__code"
        )[:10]
        return {
            "title": "Sat. formateurs (Q9)",
            "headers": ["Classe", "Moyenne"],
            "rows": [[r["classe__code"], round(r["moy"] or 0, 2)] for r in qs],
        }
    if code == "presence-prestataire":
        qs = _table_presence_rates("classe__prestation__prestataire__raison_sociale")
        return {
            "title": "Presence par prestataire",
            "headers": ["Prestataire", "PR", "Total", "Taux %"],
            "rows": [
                [
                    r["classe__prestation__prestataire__raison_sociale"],
                    r["pr"],
                    r["total"],
                    safe_rate(r["pr"], r["total"]),
                ]
                for r in qs
            ],
        }
    if code == "presence-prestation":
        qs = _table_presence_rates("classe__prestation__code")
        return {
            "title": "Presence par prestation",
            "headers": ["Prestation", "PR", "Total", "Taux %"],
            "rows": [
                [r["classe__prestation__code"], r["pr"], r["total"], safe_rate(r["pr"], r["total"])]
                for r in qs
            ],
        }
    if code == "presence-beneficiaire":
        qs = _table_presence_rates("classe__prestation__beneficiaire__nom_structure")
        return {
            "title": "Presence par beneficiaire",
            "headers": ["Beneficiaire", "PR", "Total", "Taux %"],
            "rows": [
                [
                    r["classe__prestation__beneficiaire__nom_structure"],
                    r["pr"],
                    r["total"],
                    safe_rate(r["pr"], r["total"]),
                ]
                for r in qs
            ],
        }
    if code == "presence-formation":
        qs = _table_presence_rates("classe__formation__nom")
        return {
            "title": "Presence par formation",
            "headers": ["Formation", "PR", "Total", "Taux %"],
            "rows": [
                [r["classe__formation__nom"], r["pr"], r["total"], safe_rate(r["pr"], r["total"])]
                for r in qs
            ],
        }
    if code == "presence-formation-harmo":
        qs = _table_presence_rates("classe__formation__nom_harmonise")
        return {
            "title": "Presence par formation harmonisee",
            "headers": ["Formation harmo.", "PR", "Total", "Taux %"],
            "rows": [
                [
                    r["classe__formation__nom_harmonise"],
                    r["pr"],
                    r["total"],
                    safe_rate(r["pr"], r["total"]),
                ]
                for r in qs
            ],
        }
    if code == "sat-appr-prestataire":
        qs = _table_sat_avg(
            SatisfactionApprenant,
            "q9_satisfaction_globale",
            "classe__prestation__prestataire__raison_sociale",
        )
        return {
            "title": "Satisfaction apprenants par axes",
            "headers": ["Groupe", "Moyenne"],
            "rows": [
                [r["classe__prestation__prestataire__raison_sociale"], round(r["moy"] or 0, 2)]
                for r in qs
            ],
        }
    if code == "sat-form-prestataire":
        qs = _table_sat_avg(
            SatisfactionFormateur,
            "q9_satisfaction_globale_prestataire",
            "classe__prestation__prestataire__raison_sociale",
        )
        return {
            "title": "Satisfaction formateurs par axes",
            "headers": ["Groupe", "Moyenne"],
            "rows": [
                [r["classe__prestation__prestataire__raison_sociale"], round(r["moy"] or 0, 2)]
                for r in qs
            ],
        }
    if code == "prestations-effectifs":
        qs = (
            Prestation.objects.annotate(
                appr_total=Count("classes__apprenants", distinct=True),
                appr_femmes=Count(
                    "classes__apprenants",
                    filter=Q(classes__apprenants__genre__iexact="f"),
                    distinct=True,
                ),
                appr_appart=Count(
                    "classes__apprenants",
                    filter=Q(classes__apprenants__appartenance_beneficiaire=True),
                    distinct=True,
                ),
            )
            .values(
                "code", "effectif_a_former", "femmes", "appr_total", "appr_femmes", "appr_appart"
            )
            .order_by("code")
        )
        rows = []
        for r in qs:
            eff_ok = (r["appr_total"] or 0) >= (r["effectif_a_former"] or 0)
            femmes_ok = (r["appr_femmes"] or 0) >= (r["femmes"] or 0)
            appart_rate = safe_rate(r["appr_appart"] or 0, r["appr_total"] or 0)
            rows.append(
                [
                    r["code"],
                    r["effectif_a_former"],
                    r["appr_total"],
                    "OK" if eff_ok else "NOK",
                    r["femmes"],
                    r["appr_femmes"],
                    "OK" if femmes_ok else "NOK",
                    r["appr_appart"],
                    f"{appart_rate} %",
                ]
            )
        return {
            "title": "Effectifs / Femmes / Appartenance par prestation",
            "headers": [
                "Prestation",
                "Eff. prevu",
                "Eff. reel",
                "Resp. Eff",
                "Fem. prevues",
                "Fem. reelles",
                "Resp. Fem",
                "Appart.",
                "Taux App.",
            ],
            "rows": rows,
        }
    if code == "prestations-durees":
        qs = Prestation.objects.values(
            "code", "duree_prevue_heures", "duree_reelle_heures"
        ).order_by("code")
        return {
            "title": "Durees par prestation",
            "headers": ["Prestation", "Duree prevue (h)", "Duree reelle (h)"],
            "rows": [[r["code"], r["duree_prevue_heures"], r["duree_reelle_heures"]] for r in qs],
        }
    if code == "repart-ville":
        qs = (
            Apprenant.objects.values("ville_residence")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        return {
            "title": "Repartition apprenants (villes)",
            "headers": ["Ville", "Total"],
            "rows": [[r["ville_residence"], r["total"]] for r in qs],
        }
    if code == "repart-region":
        qs = Apprenant.objects.values("region").annotate(total=Count("id")).order_by("-total")
        return {
            "title": "Repartition apprenants (regions)",
            "headers": ["Region", "Total"],
            "rows": [[r["region"], r["total"]] for r in qs],
        }
    if code == "repart-formation":
        qs = (
            Apprenant.objects.values("formation__nom")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        return {
            "title": "Repartition formations",
            "headers": ["Formation", "Total"],
            "rows": [[r["formation__nom"], r["total"]] for r in qs],
        }
    if code == "repart-formation-harmo":
        qs = (
            Apprenant.objects.values("formation__nom_harmonise")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        return {
            "title": "Repartition formations harmonisees",
            "headers": ["Formation harmonisee", "Total"],
            "rows": [[r["formation__nom_harmonise"], r["total"]] for r in qs],
        }
    if code == "repart-beneficiaire":
        qs = (
            Apprenant.objects.values("classe__prestation__beneficiaire__nom_structure")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        return {
            "title": "Repartition beneficiaires",
            "headers": ["Beneficiaire", "Total"],
            "rows": [
                [r["classe__prestation__beneficiaire__nom_structure"], r["total"]] for r in qs
            ],
        }
    if code == "repart-prestataire":
        qs = (
            Apprenant.objects.values("classe__prestation__prestataire__raison_sociale")
            .annotate(total=Count("id"))
            .order_by("-total")
        )
        return {
            "title": "Repartition prestataires",
            "headers": ["Prestataire", "Total"],
            "rows": [
                [r["classe__prestation__prestataire__raison_sociale"], r["total"]] for r in qs
            ],
        }
    if code == "env-lieu":
        env_fields = [
            "tables",
            "chaises",
            "ecran",
            "videoprojecteur",
            "ventilation",
            "eclairage",
            "salle_propre",
            "salle_securisee",
        ]
        env_qs = (
            EnqueteEnvironnement.objects.values("classe__lieu__nom_lieu", "classe__lieu__region")
            .annotate(total=Count("id"), **{f: Sum(f) for f in env_fields})
            .order_by("-total")
        )
        rows = []
        for row in env_qs:
            total = row.get("total") or 0
            somme = sum((row.get(f) or 0) for f in env_fields)
            rows.append(
                [
                    row.get("classe__lieu__nom_lieu"),
                    row.get("classe__lieu__region"),
                    total,
                    safe_rate(somme, total * len(env_fields)),
                ]
            )
        return {
            "title": "Environnement par lieu",
            "headers": ["Lieu", "Region", "Nb enquetes", "Score (8 pts)"],
            "rows": rows,
        }
    raise Http404("Table inconnue")


@require_analysis_access
def export_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=reporting_summary.csv"
    writer = csv.writer(response)
    writer.writerow(["type", "label", "valeur"])
    writer.writerow(["nb_classes", "Classes", Classe.objects.count()])
    writer.writerow(["nb_apprenants", "Apprenants", Apprenant.objects.count()])
    writer.writerow(["nb_presences", "Enquetes presence", Presence.objects.count()])
    writer.writerow(["nb_sat_appr", "Sat apprenants", SatisfactionApprenant.objects.count()])
    writer.writerow(["nb_sat_form", "Sat formateurs", SatisfactionFormateur.objects.count()])
    writer.writerow(["nb_env", "Environnement", EnqueteEnvironnement.objects.count()])
    return response


@require_analysis_access
def export_excel(request):
    # Simplified: reuse CSV content with .xls extension to keep it lightweight here.
    response = export_csv(request)
    response["Content-Disposition"] = "attachment; filename=reporting_summary.xls"
    return response


@require_analysis_access
def reporting_embed(request, code: str):
    response = render(request, "reporting/embed.html", {"code": code.upper()})
    response["X-Frame-Options"] = "ALLOWALL"
    response = render(request, "reporting/embed.html", {"code": code.upper()})
    response["X-Frame-Options"] = "ALLOWALL"
    return response


@require_analysis_access
def reporting_embed_table(request, code: str):
    payload = get_table_data(code)
    response = render(request, "reporting/embed_table.html", payload)
    response["X-Frame-Options"] = "ALLOWALL"
    return response


# ---------------------------------------------------------------------------
# Section 7 – Vues Rapport Application
# ---------------------------------------------------------------------------


def _report_call_scope_from_request(request, source: str = "GET") -> str:
    data = request.POST if source == "POST" else request.GET
    return normalize_report_call_scope(data.get("call_scope"))


def _build_application_report_context(request) -> dict:
    start_date, end_date = parse_report_dates(request.GET.get("start"), request.GET.get("end"))
    selected_class_code = (request.GET.get("classe") or "").strip()
    call_scope = _report_call_scope_from_request(request)
    report = build_application_report(
        start_date,
        end_date,
        selected_class_code=selected_class_code,
        call_scope=call_scope,
    )

    user_model = _get_user_model()
    download_params = {
        "start": start_date.isoformat(),
        "end": end_date.isoformat(),
        "call_scope": call_scope,
    }
    if selected_class_code:
        download_params["classe"] = selected_class_code
    cga_anomaly_page_obj = None
    cga_anomaly_query = urlencode(download_params)
    if call_scope == "cga":
        cga_anomaly_paginator = Paginator(report.get("anomalies", {}).get("anomaly_rows", []), 5)
        cga_anomaly_page_obj = cga_anomaly_paginator.get_page(request.GET.get("cga_anomaly_page"))
    context = {
        "report": report,
        "start_value": start_date.isoformat(),
        "end_value": end_date.isoformat(),
        "selected_class_code": selected_class_code,
        "call_scope": call_scope,
        "report_scope_tabs": [
            {"key": "padesce", "label": "PADESCE / Formateurs", "active": call_scope == "padesce"},
            {"key": "cga", "label": "CGA", "active": call_scope == "cga"},
        ],
        "class_options": report["anomalies"]["class_options"],
        "cga_anomaly_page_obj": cga_anomaly_page_obj,
        "cga_anomaly_query": cga_anomaly_query,
        "download_query": urlencode(download_params),
        "manual_send_enabled": request.user.is_superuser,
        "user_groups_total": Group.objects.count(),
        "user_managers_total": user_model.objects.filter(
            groups__name__in=["manager_padesce", "manager_cga"]
        )
        .distinct()
        .count(),
        "mail_recipients": get_report_email_recipients(),
    }
    return context


@require_analysis_access
def application_report_view(request):
    return redirect(_application_report_public_url(request))


@require_analysis_access
def application_report_export_excel(request):
    start_date, end_date = parse_report_dates(request.GET.get("start"), request.GET.get("end"))
    selected_class_code = (request.GET.get("classe") or "").strip()
    call_scope = _report_call_scope_from_request(request)
    report = build_application_report(
        start_date,
        end_date,
        selected_class_code=selected_class_code,
        call_scope=call_scope,
    )
    payload = export_application_report_excel(report)
    filename = f"rapport_application_{call_scope}_{start_date}_{end_date}.xlsx"
    response = HttpResponse(
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_analysis_access
def application_report_export_csv_view(request):
    start_date, end_date = parse_report_dates(request.GET.get("start"), request.GET.get("end"))
    selected_class_code = (request.GET.get("classe") or "").strip()
    call_scope = _report_call_scope_from_request(request)
    report = build_application_report(
        start_date,
        end_date,
        selected_class_code=selected_class_code,
        call_scope=call_scope,
    )
    filename = f"rapport_application_{call_scope}_{start_date}_{end_date}.csv"
    response = HttpResponse(
        export_application_report_csv(report), content_type="text/csv; charset=utf-8"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_analysis_access
def application_report_export_word_view(request):
    start_date, end_date = parse_report_dates(request.GET.get("start"), request.GET.get("end"))
    selected_class_code = (request.GET.get("classe") or "").strip()
    call_scope = _report_call_scope_from_request(request)
    report = build_application_report(
        start_date,
        end_date,
        selected_class_code=selected_class_code,
        call_scope=call_scope,
    )
    payload = export_application_report_word(report)
    filename = f"rapport_application_{call_scope}_{start_date}_{end_date}.docx"
    response = HttpResponse(
        payload,
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_analysis_access
def application_report_export_anomalies_excel_view(request):
    start_date, end_date = parse_report_dates(request.GET.get("start"), request.GET.get("end"))
    selected_class_code = (request.GET.get("classe") or "").strip()
    call_scope = _report_call_scope_from_request(request)
    report = build_application_report(
        start_date,
        end_date,
        selected_class_code=selected_class_code,
        call_scope=call_scope,
    )
    payload = export_application_report_anomalies_excel(report)
    class_suffix = f"_{slugify(selected_class_code)}" if selected_class_code else ""
    filename = f"rapport_anomalies_{call_scope}_{start_date}_{end_date}{class_suffix}.xlsx"
    response = HttpResponse(
        payload,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@require_superadmin_access
def application_report_send_mail_view(request):
    if request.method != "POST":
        return redirect(reverse("public_space") + "?scope=apprenant&section=rapport")
    start_date, end_date = parse_report_dates(request.POST.get("start"), request.POST.get("end"))
    call_scope = _report_call_scope_from_request(request, source="POST")
    report = build_application_report(start_date, end_date, call_scope=call_scope)
    result = send_report_by_email(report)
    if result["ok"]:
        messages.success(request, result["detail"])
    else:
        messages.warning(request, result["detail"])
    return redirect(
        reverse("public_space")
        + "?"
        + urlencode(
            {
                "scope": "apprenant",
                "section": "rapport",
                "start": start_date.isoformat(),
                "end": end_date.isoformat(),
                "call_scope": call_scope,
            }
        )
    )


@csrf_exempt
@require_POST
def reporting_daily_digest_trigger_view(request):
    token_error = _validate_reporting_trigger_token(request)
    if token_error:
        return token_error

    start_date, end_date = parse_report_dates(request.POST.get("start"), request.POST.get("end"))
    job_id = daily_digest_jobs.start_daily_digest(
        start_date,
        end_date,
        backup_job_id=request.POST.get("job_id", ""),
        backup_error=request.POST.get("backup_error", ""),
        recipients=request.POST.get("recipients") or None,
    )
    return JsonResponse(
        {
            "job_id": job_id,
            "status": "pending",
            "message": "Digest quotidien demarre.",
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
        },
        status=202,
    )


def _validate_reporting_trigger_token(request):
    expected_token = str(
        getattr(settings, "REPORT_TRIGGER_TOKEN", "")
        or getattr(settings, "BACKUP_TRIGGER_TOKEN", "")
        or ""
    ).strip()
    if not expected_token:
        return JsonResponse({"error": "REPORT_TRIGGER_TOKEN non configuré."}, status=503)

    provided_token = (
        request.headers.get("X-Report-Token", "")
        or request.headers.get("X-Backup-Token", "")
        or request.POST.get("token", "")
    )
    provided_token = str(provided_token or "").strip()
    if not provided_token or provided_token != expected_token:
        return JsonResponse({"error": "Token invalide ou manquant."}, status=403)
    return None


def reporting_daily_digest_status_view(request, job_id):
    token_error = _validate_reporting_trigger_token(request)
    if token_error:
        return token_error

    job = daily_digest_jobs.get_job(job_id)
    if not job:
        return JsonResponse({"error": "Job introuvable."}, status=404)
    return JsonResponse(job)


@require_analysis_access
def reporting_manual_view(request):
    try:
        manual_path = get_reporting_manual_path()
        markdown_text = load_reporting_manual_markdown()
    except FileNotFoundError as exc:
        raise Http404("Le manuel de reporting est introuvable.") from exc

    context = {
        "manual_filename": manual_path.name,
        "manual_html": mark_safe(render_reporting_manual_html(markdown_text)),
        "manual_download_url": reverse("reporting_manual_download"),
    }
    return render(request, "reporting/documentation.html", context)


@require_analysis_access
def reporting_manual_download_view(request):
    try:
        manual_path = get_reporting_manual_path()
        payload = manual_path.read_bytes()
    except FileNotFoundError as exc:
        raise Http404("Le manuel de reporting est introuvable.") from exc

    response = HttpResponse(payload, content_type="text/markdown; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{manual_path.name}"'
    return response


def public_reporting_manual_view(request):
    """Public access to reporting documentation (no login required)."""
    anchor = request.GET.get("section", "")
    try:
        manual_path = get_reporting_manual_path()
        markdown_text = load_reporting_manual_markdown()
    except FileNotFoundError as exc:
        raise Http404("Le manuel de reporting est introuvable.") from exc

    context = {
        "manual_filename": manual_path.name,
        "manual_html": mark_safe(render_reporting_manual_html(markdown_text)),
        "manual_download_url": reverse("reporting_manual_download"),
        "anchor": anchor,
        "is_public": True,
    }
    return render(request, "reporting/documentation_public.html", context)


def _concordance_header_key(value):
    return _normalize_header(value).replace(" ", "_")


CONCORDANCE_FEUIL2_DISPLAY_HEADERS = (
    "NBRE",
    "PRESTA ID",
    "PRESTATAIRE",
    "BENEFICIAIRE",
    "FENETRE",
    "NBRE PERSONNES FORMEES SELON FICHE DE PRESENCE RAPPORT PRESTATAIRE - T",
    "NBRE PERSONNES FORMEES SELON FICHE DE PRESENCE RAPPORT PRESTATAIRE - H",
    "NBRE PERSONNES FORMEES SELON FICHE DE PRESENCE RAPPORT PRESTATAIRE - F",
    "TAUX_CONCORDANCE",
    "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - H",
    "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - F",
    "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - T",
)


def _concordance_display_headers(records):
    """Keep Feuil2's display order stable across SQLite and PostgreSQL JSONB."""
    discovered = []
    for record in records:
        for key in record.payload:
            if key not in discovered:
                discovered.append(key)

    is_feuil2_layout = set(CONCORDANCE_FEUIL2_DISPLAY_HEADERS).issubset(discovered)
    if is_feuil2_layout:
        return list(CONCORDANCE_FEUIL2_DISPLAY_HEADERS), True
    return discovered[:13], False


def _pending_contact_rows_from_file(uploaded_file):
    """Read a flexible contact workbook while preserving its source column order."""
    filename = str(getattr(uploaded_file, "name", "")).lower()
    if filename.endswith(".csv"):
        try:
            text_data = uploaded_file.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("Le fichier CSV doit être encodé en UTF-8.") from exc
        try:
            dialect = csv.Sniffer().sniff(text_data[:4096], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        sheet_rows = list(csv.reader(io.StringIO(text_data), dialect=dialect))
    elif filename.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
    else:
        raise ValueError("Le fichier doit être au format XLSX, XLSM ou CSV.")

    header_index = next(
        (
            index
            for index, row in enumerate(sheet_rows)
            if any(_normalize_cell(value) for value in row)
        ),
        None,
    )
    if header_index is None:
        raise ValueError("Le fichier est vide.")

    source_headers = sheet_rows[header_index]
    data_rows = sheet_rows[header_index + 1 :]
    column_count = max([len(source_headers), *(len(row) for row in data_rows)] or [0])
    if column_count > 100:
        raise ValueError("Le fichier dépasse la limite de 100 colonnes.")

    headers = []
    occurrences = {}
    for index in range(column_count):
        base_header = (
            _normalize_cell(source_headers[index])
            if index < len(source_headers)
            else ""
        ) or f"Colonne {index + 1}"
        occurrence = occurrences.get(base_header, 0) + 1
        occurrences[base_header] = occurrence
        headers.append(base_header if occurrence == 1 else f"{base_header} ({occurrence})")

    payloads = []
    for row in data_rows:
        payload = {
            header: _normalize_cell(row[index]) if index < len(row) else ""
            for index, header in enumerate(headers)
        }
        if any(payload.values()):
            payloads.append(payload)
    return headers, payloads


def _concordance_rows_from_file(uploaded_file):
    """Read a concordance workbook, favouring Feuil2 when it is available."""
    filename = str(getattr(uploaded_file, "name", "")).lower()
    if filename.endswith(".csv"):
        text_data = uploaded_file.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text_data))
        sheet_rows = list(reader)
    elif filename.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        worksheet = workbook["Feuil2"] if "Feuil2" in workbook.sheetnames else workbook.active
        sheet_rows = list(worksheet.iter_rows(values_only=True))
    else:
        raise ValueError("Le fichier doit être au format XLSX, XLSM ou CSV.")
    if not sheet_rows:
        raise ValueError("Le fichier est vide.")
    # Le fichier de rapprochement comporte deux lignes d'en-tête dans Feuil2.
    # On les fusionne et rend les colonnes H/F/T distinctes et filtrables.
    if filename.endswith((".xlsx", ".xlsm")) and "Feuil2" in workbook.sheetnames and len(sheet_rows) > 2:
        top_headers = [_normalize_cell(value) for value in sheet_rows[0]]
        bottom_headers = [_normalize_cell(value) for value in sheet_rows[1]]
        inherited = ""
        header_specs = []
        for index, value in enumerate(top_headers):
            # Feuil2 contient une première colonne A entièrement vide : elle ne
            # fait pas partie du tableau de rapprochement et ne doit surtout pas
            # décaler la dernière colonne « T ».
            if not value and not (bottom_headers[index] if index < len(bottom_headers) else "") and not any(
                _normalize_cell(row[index]) for row in sheet_rows[2:] if index < len(row)
            ):
                continue
            if value:
                inherited = value
            child = bottom_headers[index] if index < len(bottom_headers) else ""
            # Fenêtre est une colonne autonome dans Feuil2, même si Excel la
            # place sous la cellule visuellement fusionnée « Bénéficiaire ».
            if _concordance_header_key(child) == "fenetre":
                header = child
            elif child and inherited and child != inherited:
                header = f"{inherited} - {child}"
            elif inherited == "NBRE PERSONNES FORMEES SELON FICHE DE PRESENCE RAPPORT PRESTATAIRE":
                # La première colonne de ce bloc est le total (T), puis H et F.
                header = f"{inherited} - T"
            else:
                header = child or inherited or f"Colonne {index + 1}"
            header_specs.append((index, header))
        headers = [header for _index, header in header_specs]
        seen_headers = {}
        unique_headers = []
        for header in headers:
            occurrence = seen_headers.get(header, 0)
            seen_headers[header] = occurrence + 1
            unique_headers.append(header if occurrence == 0 else f"{header} ({occurrence + 1})")
        headers = unique_headers
        header_specs = [(header_specs[index][0], header) for index, header in enumerate(headers)]
        data_rows = sheet_rows[2:]
    else:
        header_specs = [
            (index, _normalize_cell(value) or f"Colonne {index + 1}")
            for index, value in enumerate(sheet_rows[0])
        ]
        headers = [header for _index, header in header_specs]
        data_rows = sheet_rows[1:]
    if not any(headers):
        raise ValueError("La première ligne doit contenir les en-têtes.")
    normalized = [_concordance_header_key(value) for value in headers]
    rows = []
    for row in data_rows:
        payload = {
            header: _normalize_cell(row[index])
            for index, header in header_specs
            if header and index < len(row) and _normalize_cell(row[index])
        }
        if not payload:
            continue
        lookup = {
            _concordance_header_key(header): _normalize_cell(row[index])
            for index, header in header_specs
            if index < len(row)
        }
        genre = lookup.get("genre") or lookup.get("sexe") or "Non renseigné"
        fenetre = lookup.get("fenetre") or lookup.get("fenetre_appel") or "Non renseignée"
        rows.append(ConcordanceRecord(genre=genre, fenetre=fenetre, payload=payload))
    return headers, rows


def _concordance_payload_value(payload, *aliases):
    normalized_aliases = {_concordance_header_key(alias) for alias in aliases}
    for key, value in payload.items():
        if _concordance_header_key(key) in normalized_aliases:
            return str(value or "").strip()
    return ""


def _synthesis_payload_identity(payload):
    """Read the source identifiers used by RC and the freely-shaped R import."""
    return {
        "presta_id": _concordance_payload_value(
            payload,
            "PRESTA ID", "PRESTAID", "PRESTATION ID", "PRESTATION",
            "CODE PRESTATION", "CODE PRESTA",
        ),
        "prestataire": _concordance_payload_value(
            payload,
            "PRESTATAIRE", "PRESTATAIRES", "NOM PRESTATAIRE",
            "NOM DU PRESTATAIRE", "STRUCTURE PRESTATAIRE",
        ),
        "beneficiaire": _concordance_payload_value(
            payload,
            "BENEFICIAIRE", "BENEFICIAIRES", "NOM BENEFICIAIRE",
            "NOM DU BENEFICIAIRE", "STRUCTURE BENEFICIAIRE",
        ),
        "fenetre": _concordance_payload_value(
            payload, "FENETRE", "FENETRE APPEL", "FENETRE D APPEL", "COHORTE"
        ),
    }


def _build_pas_forme_ii_campaign():
    """Return every Pas Formés II prestation and its 10% call threshold status."""
    thresholded = {}
    aggregates = AppelPasFormeII.objects.filter(is_active=True).values("prestation_id").annotate(
        total=Count("id"), appeles=Count("id", filter=Q(formulaire_rempli_at__isnull=False))
    )
    for item in aggregates:
        total, appeles = int(item["total"] or 0), int(item["appeles"] or 0)
        # A prestation reaches the call-reconciliation threshold once at
        # least 10% of its learners have completed the form (minimum one).
        thresholded[item["prestation_id"]] = bool(
            total and appeles >= max(1, (total * 10 + 99) // 100)
        )

    calls = list(AppelPasFormeII.objects.filter(
        is_active=True
    ).order_by("prestation_id", "fenetre", "nom"))
    segments = {}
    for call in calls:
        key = (call.prestation_id or "Non renseigné", call.prestataire or "Non renseigné", call.beneficiaire or "Non renseigné", call.fenetre or "Non renseignée")
        segment = segments.setdefault(key, {
            "presta_id": key[0], "prestataire": key[1], "beneficiaire": key[2], "fenetre": key[3],
            "total": 0, "appeles": 0, "hommes": 0, "femmes": 0, "apprenants": [],
            "formes_total": 0, "formes_hommes": 0, "formes_femmes": 0,
            "formes_fenetre_2": 0, "formes_fenetre_3": 0, "pas_formes_total": 0,
            "seuil_atteint": thresholded.get(call.prestation_id, False),
        })
        segment["total"] += 1
        # A completed form proves the person was called even if an old record
        # still carries the waiting status. Otherwise keep only started calls.
        if not (
            is_call_attempted_status(call.status) or call.formulaire_rempli_at
        ):
            continue
        genre = (call.genre or "Non renseigné").strip()
        genre_key = _gender_key(genre)
        declared_sessions = call.nombre_seances_declare
        planned_sessions = call.total_seances
        # The reconciliation counts only a call when the form has actually
        # been completed and its number of attended sessions is provided.
        if call.formulaire_rempli_at and declared_sessions is not None:
            segment["appeles"] += 1
            if genre_key == "men":
                segment["hommes"] += 1
            elif genre_key == "women":
                segment["femmes"] += 1
        audio_url = ""
        if call.audio_file and call.audio_file.name:
            try:
                audio_url = call.audio_file.url
            except Exception:
                pass
        attendance_rate = None
        training_status = "Non renseigné"
        if declared_sessions is not None and planned_sessions is not None and planned_sessions > 0:
            if declared_sessions == 0:
                # A declared count of 0 does not reliably mean "attended
                # nothing" (blank inputs default to 0); treat it as
                # indeterminate instead of asserting "Pas formé".
                training_status = "Non déterminé"
            else:
                attendance_rate = declared_sessions / planned_sessions * 100
                training_status = (
                    "Formé" if 75 <= attendance_rate <= 120 else "Pas formé"
                )
        if training_status == "Formé":
            segment["formes_total"] += 1
            if genre_key == "men":
                segment["formes_hommes"] += 1
            elif genre_key == "women":
                segment["formes_femmes"] += 1
            window = _campaign_window_key(key[3])
            if window == "Fenêtre 2":
                segment["formes_fenetre_2"] += 1
            elif window == "Fenêtre 3":
                segment["formes_fenetre_3"] += 1
        elif training_status == "Pas formé":
            segment["pas_formes_total"] += 1
        if training_status in {"Non déterminé", "Non renseigné"}:
            # Neither an explicit 0 nor a missing declaration is a real,
            # reviewable answer. Both stay in the underlying call record
            # (DB) but are left out of the "Détail des apprenants appelés"
            # list shown to the enquêteur.
            continue
        segment["apprenants"].append({
            "nom": call.nom,
            "code": call.reference_code,
            "telephone": call.telephone,
            "membre_structure": call.membre_structure,
            "genre": genre,
            "fenetre": key[3],
            "presta_id": key[0],
            "seances_declarees": declared_sessions,
            "seances_prevues": planned_sessions,
            "taux_presence": attendance_rate,
            "statut_formation": training_status,
            "audio_url": audio_url,
            "audio_disponible": bool(audio_url),
        })

    # Keep the prestations still awaiting their first call.  They are present in
    # the RA source tab and must consequently also be visible in the synthesis.
    rows = list(segments.values())
    for row in rows:
        denom = row["formes_total"] + row["pas_formes_total"]
        row["taux_formation"] = (row["formes_total"] / denom * 100) if denom else None
        row["decision"] = row["total"] if (row["taux_formation"] is not None and row["taux_formation"] > 75) else "_"
    summary = {
        "prestations": sum(1 for atteint in thresholded.values() if atteint),
        "appels_effectues": 0,
        "fenetre_2": 0, "fenetre_3": 0, "hommes": 0, "femmes": 0,
    }
    for row in rows:
        # The H/F recap intentionally has no "non renseigné" column.
        # Count only the same H and F people that are displayed in it.
        gendered_calls = row["hommes"] + row["femmes"]
        summary["appels_effectues"] += row["appeles"]
        if _campaign_window_key(row["fenetre"]) == "Fenêtre 2":
            summary["fenetre_2"] += gendered_calls
        elif _campaign_window_key(row["fenetre"]) == "Fenêtre 3":
            summary["fenetre_3"] += gendered_calls
        summary["hommes"] += row["hommes"]
        summary["femmes"] += row["femmes"]
    return sorted(rows, key=lambda row: (row["presta_id"], row["fenetre"])), summary


def _format_concordance_value(header, value):
    """Match Feuil2's displayed number formats without changing stored data."""
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    try:
        number = float(raw_value.replace(",", "."))
    except (TypeError, ValueError):
        return raw_value

    if _concordance_header_key(header) == "taux_concordance":
        # Feuil2 formats this value as 0.0%.
        return f"{number:.1%}"
    if str(header).startswith("NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE"):
        # The calculated H/F/T values in Feuil2 use the Excel number format 0.
        return f"{number:.0f}"
    return raw_value


def _campaign_window_key(value):
    """Return the reporting window used by the H/F recap table."""
    normalized = _concordance_header_key(value).replace("_", "")
    if normalized in {"2", "fenetre2", "f2"}:
        return "Fenêtre 2"
    if normalized in {"3", "fenetre3", "f3"}:
        return "Fenêtre 3"
    return ""


def _gender_key(value):
    normalized = _concordance_header_key(value)
    if normalized in {"homme", "masculin", "h", "m"}:
        return "men"
    if normalized in {"femme", "feminin", "féminin", "f"}:
        return "women"
    return ""


def _formed_people_summary(concordance_window_summary):
    """Reshape the concordance H/F/T recap (NOMBRE FORME TOTAL) into stat cards."""
    by_window = {row["window"]: row for row in concordance_window_summary}
    total_row = by_window.get("Total", {"men": 0, "women": 0, "total": 0})
    return {
        "total": total_row["total"],
        "fenetre_2": by_window.get("Fenêtre 2", {}).get("total", 0),
        "fenetre_3": by_window.get("Fenêtre 3", {}).get("total", 0),
        "hommes": total_row["men"],
        "femmes": total_row["women"],
    }


def _window_gender_summary(rows, *, gender_key, window_key, value_key=lambda _row: 1):
    """Build the Fenêtre 2 / Fenêtre 3 H/F/T recap used on each tab."""
    totals = {
        "Fenêtre 2": {"men": 0, "women": 0},
        "Fenêtre 3": {"men": 0, "women": 0},
    }
    for row in rows:
        window = _campaign_window_key(window_key(row))
        gender = _gender_key(gender_key(row))
        if not window or not gender:
            continue
        totals[window][gender] += value_key(row) or 0

    summary_rows = []
    for window, values in totals.items():
        summary_rows.append({
            "window": window,
            "men": values["men"],
            "women": values["women"],
            "total": values["men"] + values["women"],
        })
    summary_rows.append({
        "window": "Total",
        "men": sum(row["men"] for row in summary_rows),
        "women": sum(row["women"] for row in summary_rows),
        "total": sum(row["total"] for row in summary_rows),
    })
    return summary_rows


def _combine_gender_summaries(*summaries):
    """Add the H/F recap figures already calculated for other tabs."""
    windows = ("Fenêtre 2", "Fenêtre 3")
    totals = {window: {"men": 0, "women": 0} for window in windows}
    for summary in summaries:
        for row in summary:
            window = row.get("window")
            if window not in totals:
                continue
            totals[window]["men"] += row.get("men", 0) or 0
            totals[window]["women"] += row.get("women", 0) or 0

    result = [
        {
            "window": window,
            "men": values["men"],
            "women": values["women"],
            "total": values["men"] + values["women"],
        }
        for window, values in totals.items()
    ]
    result.append({
        "window": "Total",
        "men": sum(row["men"] for row in result),
        "women": sum(row["women"] for row in result),
        "total": sum(row["total"] for row in result),
    })
    return result


def _number_from_concordance_payload(payload, *aliases):
    """Read a workbook numeric cell without failing on French decimal formatting."""
    raw_value = _concordance_payload_value(payload, *aliases).replace(" ", "").replace(",", ".")
    try:
        return float(raw_value)
    except (TypeError, ValueError):
        return 0


def _concordance_window_summary(rows, headers):
    """Sum the final H/F/T values produced by the concordance workbook."""
    totals = {
        "Fenêtre 2": {"men": 0, "women": 0, "total": 0},
        "Fenêtre 3": {"men": 0, "women": 0, "total": 0},
    }
    for row in rows:
        window = _campaign_window_key(row.fenetre)
        if not window:
            continue
        men_value = _concordance_payload_value(
            row.payload,
            "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - H",
        )
        women_value = _concordance_payload_value(
            row.payload,
            "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - F",
        )
        if men_value or women_value:
            # Prefer explicit labels: JSONB does not preserve the Excel column order.
            men = int(round(_number_from_concordance_payload({"value": men_value}, "value")))
            women = int(round(_number_from_concordance_payload({"value": women_value}, "value")))
        else:
            # Some Feuil2 exports have merged headers without usable labels.
            # Use the common displayed order because JSONB can reorder payload keys.
            values = [row.payload.get(header, "") for header in headers]
            if len(values) < 3:
                continue
            # The detail grid displays each calculated H/F cell as a whole person.
            men = int(round(_number_from_concordance_payload({"value": values[-3]}, "value")))
            women = int(round(_number_from_concordance_payload({"value": values[-2]}, "value")))
        totals[window]["men"] += men
        totals[window]["women"] += women
        totals[window]["total"] += men + women

    summary_rows = [{"window": window, **values} for window, values in totals.items()]
    summary_rows.append({
        "window": "Total",
        "men": sum(row["men"] for row in summary_rows),
        "women": sum(row["women"] for row in summary_rows),
        "total": sum(row["total"] for row in summary_rows),
    })
    return summary_rows


def _concordance_final_gender_values(row, headers):
    """Return the final H/F values for one concordance prestation."""
    men_value = _concordance_payload_value(
        row.payload,
        "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - H",
    )
    women_value = _concordance_payload_value(
        row.payload,
        "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - F",
    )
    if men_value or women_value:
        return (
            int(round(_number_from_concordance_payload({"value": men_value}, "value"))),
            int(round(_number_from_concordance_payload({"value": women_value}, "value"))),
        )

    values = [row.payload.get(header, "") for header in headers]
    if len(values) < 3:
        return 0, 0
    return (
        int(round(_number_from_concordance_payload({"value": values[-3]}, "value"))),
        int(round(_number_from_concordance_payload({"value": values[-2]}, "value"))),
    )


def _build_synthesis_reconciliation_rows(concordance, headers, campaign_rows, pending_contact_import):
    """Consolidate RC, RA and R sources for the real-time synthesis table."""
    rows = []

    concordance_segments = {}
    for record in concordance:
        identity = _synthesis_payload_identity(record.payload)
        presta_id = identity["presta_id"]
        if not presta_id or presta_id.upper() == "TOTAL GENERAL":
            continue
        prestataire = identity["prestataire"]
        beneficiaire = identity["beneficiaire"]
        fenetre = record.fenetre or identity["fenetre"]
        key = (presta_id, prestataire, beneficiaire, fenetre)
        segment = concordance_segments.setdefault(
            key,
            {
                "presta_id": presta_id,
                "prestataire": prestataire,
                "beneficiaire": beneficiaire,
                "fenetre": fenetre,
                "hommes": 0,
                "femmes": 0,
                "appeles": 0,
                "total": 0,
                "methode": "RC",
            },
        )
        hommes, femmes = _concordance_final_gender_values(record, headers)
        source_total_value = _concordance_payload_value(
            record.payload,
            "NBRE PERSONNES FORMEES SELON FICHE DE PRESENCE RAPPORT PRESTATAIRE - T",
        )
        final_total_value = _concordance_payload_value(
            record.payload,
            "NOMBRE FORME TOTAL AVEC TAUX DE CONCORDANCE - T",
        )
        # RC displays the reconciled people called against the original
        # presence-list total, as in the concordance tab.
        total_value = source_total_value or final_total_value
        total = int(round(_number_from_concordance_payload({"value": total_value}, "value")))
        total = total if total_value else hommes + femmes
        segment["hommes"] += hommes
        segment["femmes"] += femmes
        segment["appeles"] += hommes + femmes
        segment["total"] += total
    rows.extend(concordance_segments.values())

    rows.extend(
        {
            "presta_id": row["presta_id"],
            "prestataire": row["prestataire"],
            "beneficiaire": row["beneficiaire"],
            "fenetre": row["fenetre"],
            "hommes": row["hommes"],
            "femmes": row["femmes"],
            "appeles": row["appeles"],
            "total": row["total"],
            "methode": "RA",
        }
        for row in campaign_rows
    )

    if pending_contact_import:
        pending_segments = {}
        for record in pending_contact_import.records.all():
            identity = _synthesis_payload_identity(record.payload)
            presta_id = identity["presta_id"] or "Non renseigné"
            prestataire = identity["prestataire"]
            beneficiaire = identity["beneficiaire"]
            fenetre = identity["fenetre"]
            key = (presta_id, prestataire, beneficiaire, fenetre)
            segment = pending_segments.setdefault(
                key,
                {
                    "presta_id": presta_id,
                    "prestataire": prestataire,
                    "beneficiaire": beneficiaire,
                    "fenetre": fenetre,
                    "hommes": 0,
                    "femmes": 0,
                    "appeles": 0,
                    "total": 0,
                    "methode": "R",
                },
            )
            segment["total"] += 1
        rows.extend(pending_segments.values())

    method_order = {"RC": 0, "RA": 1, "R": 2}
    return sorted(
        rows,
        key=lambda row: (
            method_order[row["methode"]],
            row["presta_id"],
            row["fenetre"],
        ),
    )


def _dated_export_filename(base_name, extension):
    """Append today's date to an export filename, e.g. base_2026-08-04.csv."""
    return f"{base_name}_{timezone.now().strftime('%Y-%m-%d')}.{extension}"


def _csv_export_response(base_name, headers, rows):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{_dated_export_filename(base_name, "csv")}"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _excel_export_response(base_name, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_dated_export_filename(base_name, "xlsx")}"'
    return response


def _concordance_export_rows(request):
    """Full filtered RC dataset (headers + row values), unlike the 100-row preview."""
    concordance = list(ConcordanceRecord.objects.order_by("id"))
    selected_fenetre = (request.GET.get("fenetre") or "").strip()
    selected_prestataire = (request.GET.get("prestataire") or "").strip()
    selected_beneficiaire = (request.GET.get("beneficiaire") or "").strip()
    selected_presta_id = (request.GET.get("presta_id") or "").strip()
    search_query = (request.GET.get("q") or "").strip().lower()

    def source_value(row, key):
        return _concordance_payload_value(row.payload, key)

    filtered = []
    for row in concordance:
        prestataire = source_value(row, "PRESTATAIRE")
        beneficiaire = source_value(row, "BENEFICIAIRE")
        presta_id = source_value(row, "PRESTA ID")
        searchable = " ".join([row.genre, row.fenetre, prestataire, beneficiaire, presta_id, *map(str, row.payload.values())]).lower()
        if selected_fenetre and row.fenetre != selected_fenetre:
            continue
        if selected_prestataire and prestataire != selected_prestataire:
            continue
        if selected_beneficiaire and beneficiaire != selected_beneficiaire:
            continue
        if selected_presta_id and presta_id != selected_presta_id:
            continue
        if search_query and search_query not in searchable:
            continue
        filtered.append(row)

    headers, _ = _concordance_display_headers(concordance)
    rows = [
        [_format_concordance_value(header, row.payload.get(header, "")) for header in headers]
        for row in filtered
    ]
    return list(headers), rows


def _campaign_export_rows():
    """Full RA dataset (one row per prestation), matching the 'Prestations et seuil des appels' table."""
    rows, _ = _build_pas_forme_ii_campaign()
    headers = [
        "PRESTAID", "Prestataire", "Bénéficiaire", "Fenêtre",
        "Appelés H", "Appelés F", "Appelés", "Total",
        "Formés total", "Formés H", "Formés F", "Formés fenêtre 2", "Formés fenêtre 3",
        "Pas formés total", "Taux de formation (%)", "Décision", "Seuil atteint",
    ]
    data = [
        [
            row["presta_id"], row["prestataire"], row["beneficiaire"], row["fenetre"],
            row["hommes"], row["femmes"], row["appeles"], row["total"],
            row["formes_total"], row["formes_hommes"], row["formes_femmes"],
            row["formes_fenetre_2"], row["formes_fenetre_3"], row["pas_formes_total"],
            round(row["taux_formation"], 1) if row["taux_formation"] is not None else "",
            row["decision"] if row["decision"] != "_" else "",
            "Oui" if row["seuil_atteint"] else "Non",
        ]
        for row in rows
    ]
    return headers, data


def _synthesis_export_rows():
    """Full 'Réconciliation par prestation' dataset (RC/RA/R consolidated)."""
    concordance = list(ConcordanceRecord.objects.order_by("id"))
    headers, _ = _concordance_display_headers(concordance)
    not_formed_campaign_rows, _ = _build_pas_forme_ii_campaign()
    pending_contact_import = PendingLearnerContactImport.objects.first()
    rows = _build_synthesis_reconciliation_rows(
        concordance, headers, not_formed_campaign_rows, pending_contact_import
    )
    export_headers = [
        "PRESTAID", "Prestataire", "Bénéficiaire", "Fenêtre",
        "Appelés H", "Appelés F", "Appelés", "Total", "Méthode",
    ]
    data = [
        [
            row["presta_id"], row["prestataire"] or "—", row["beneficiaire"] or "—", row["fenetre"] or "—",
            row["hommes"], row["femmes"], row["appeles"], row["total"], row["methode"],
        ]
        for row in rows
    ]
    return export_headers, data


def _pending_contacts_export_rows():
    """Full pending-contacts dataset, unlike the 100-row preview."""
    pending_contact_import = PendingLearnerContactImport.objects.first()
    if not pending_contact_import:
        return ["#"], []
    headers = ["#"] + list(pending_contact_import.headers)
    rows = [
        [record.row_number] + [record.payload.get(header, "") for header in pending_contact_import.headers]
        for record in pending_contact_import.records.all()
    ]
    return headers, rows


def concordance_export_rc_csv(request):
    headers, rows = _concordance_export_rows(request)
    return _csv_export_response("reconciliation_concordance", headers, rows)


def concordance_export_rc_excel(request):
    headers, rows = _concordance_export_rows(request)
    return _excel_export_response("reconciliation_concordance", headers, rows)


def concordance_export_ra_csv(request):
    headers, rows = _campaign_export_rows()
    return _csv_export_response("reconciliation_appels", headers, rows)


def concordance_export_ra_excel(request):
    headers, rows = _campaign_export_rows()
    return _excel_export_response("reconciliation_appels", headers, rows)


def concordance_export_synthesis_csv(request):
    headers, rows = _synthesis_export_rows()
    return _csv_export_response("synthese_reconciliation", headers, rows)


def concordance_export_synthesis_excel(request):
    headers, rows = _synthesis_export_rows()
    return _excel_export_response("synthese_reconciliation", headers, rows)


def concordance_export_contacts_csv(request):
    headers, rows = _pending_contacts_export_rows()
    return _csv_export_response("contacts_en_attente", headers, rows)


def concordance_export_contacts_excel(request):
    headers, rows = _pending_contacts_export_rows()
    return _excel_export_response("contacts_en_attente", headers, rows)


def concordance_campaigns_view(request):
    can_manage_concordance_import = bool(
        request.user.is_authenticated and request.user.is_superuser
    )
    if request.method == "POST":
        action = request.POST.get("action")
        protected_actions = {
            "upload", "clear", "pending_contacts_upload", "pending_contacts_clear",
        }
        if action in protected_actions and not can_manage_concordance_import:
            messages.error(
                request,
                "Seul un administrateur peut importer, remplacer ou supprimer ces données.",
            )
            if action.startswith("pending_contacts"):
                return redirect(f"{reverse('concordance_campaigns')}?tab=pending-contacts")
            return redirect("concordance_campaigns")
        if action == "pending_contacts_clear":
            PendingLearnerContactImport.objects.all().delete()
            messages.success(request, "Les contacts importés ont été supprimés.")
        elif action == "pending_contacts_upload":
            uploaded_file = request.FILES.get("fichier")
            if not uploaded_file:
                messages.error(request, "Sélectionnez un fichier avant de lancer l’import.")
            else:
                try:
                    headers, payloads = _pending_contact_rows_from_file(uploaded_file)
                    if not payloads:
                        raise ValueError("Aucune ligne de contact n’a été trouvée.")
                    source_filename = str(uploaded_file.name).replace("\\", "/").rsplit("/", 1)[-1]
                    with transaction.atomic():
                        PendingLearnerContactImport.objects.all().delete()
                        import_batch = PendingLearnerContactImport.objects.create(
                            source_filename=source_filename[:255],
                            headers=headers,
                        )
                        PendingLearnerContactRecord.objects.bulk_create(
                            [
                                PendingLearnerContactRecord(
                                    import_batch=import_batch,
                                    row_number=index,
                                    payload=payload,
                                )
                                for index, payload in enumerate(payloads, start=1)
                            ],
                            batch_size=500,
                        )
                    messages.success(
                        request,
                        f"{len(payloads)} contacts importés et ancien fichier remplacé.",
                    )
                except Exception as exc:
                    messages.error(request, str(exc))
        elif action == "clear":
            ConcordanceRecord.objects.all().delete()
            messages.success(request, "Les travaux de concordance ont été supprimés.")
        elif action == "upload":
            uploaded_file = request.FILES.get("fichier")
            if not uploaded_file:
                messages.error(request, "Sélectionnez un fichier avant de lancer l’import.")
            else:
                try:
                    headers, records = _concordance_rows_from_file(uploaded_file)
                    if not records:
                        raise ValueError("Aucune ligne de données n’a été trouvée.")
                    with transaction.atomic():
                        ConcordanceRecord.objects.all().delete()
                        ConcordanceRecord.objects.bulk_create(records, batch_size=500)
                    messages.success(request, f"{len(records)} lignes de concordance importées.")
                except Exception as exc:
                    messages.error(request, str(exc))
        if action in {"pending_contacts_clear", "pending_contacts_upload"}:
            return redirect(f"{reverse('concordance_campaigns')}?tab=pending-contacts")
        return redirect("concordance_campaigns")

    # L'ordre d'import est celui de Feuil2 : en particulier, la ligne TOTAL
    # GENERAL doit rester à la fin du tableau.
    concordance = list(ConcordanceRecord.objects.order_by("id"))
    selected_fenetre = (request.GET.get("fenetre") or "").strip()
    selected_prestataire = (request.GET.get("prestataire") or "").strip()
    selected_beneficiaire = (request.GET.get("beneficiaire") or "").strip()
    selected_presta_id = (request.GET.get("presta_id") or "").strip()
    search_query = (request.GET.get("q") or "").strip().lower()

    def source_value(row, key):
        return _concordance_payload_value(row.payload, key)

    filter_options = {
        "fenetres": sorted({row.fenetre for row in concordance if row.fenetre}),
        "prestataires": sorted({source_value(row, "PRESTATAIRE") for row in concordance if source_value(row, "PRESTATAIRE")}),
        "beneficiaires": sorted({source_value(row, "BENEFICIAIRE") for row in concordance if source_value(row, "BENEFICIAIRE")}),
        "presta_ids": sorted({source_value(row, "PRESTA ID") for row in concordance if source_value(row, "PRESTA ID")}),
    }
    filtered_concordance = []
    for row in concordance:
        prestataire = source_value(row, "PRESTATAIRE")
        beneficiaire = source_value(row, "BENEFICIAIRE")
        presta_id = source_value(row, "PRESTA ID")
        searchable = " ".join([row.genre, row.fenetre, prestataire, beneficiaire, presta_id, *map(str, row.payload.values())]).lower()
        if selected_fenetre and row.fenetre != selected_fenetre:
            continue
        if selected_prestataire and prestataire != selected_prestataire:
            continue
        if selected_beneficiaire and beneficiaire != selected_beneficiaire:
            continue
        if selected_presta_id and presta_id != selected_presta_id:
            continue
        if search_query and search_query not in searchable:
            continue
        filtered_concordance.append(row)
    headers, is_feuil2_layout = _concordance_display_headers(concordance)
    concordance_rows = [
        {
            "genre": row.genre,
            "fenetre": row.fenetre,
            "values": [
                _format_concordance_value(header, row.payload.get(header, ""))
                for header in headers
            ],
        }
        for row in filtered_concordance[:100]
    ]

    campaign = {}
    for call in AppelPasFormeII.objects.filter(is_active=True).values(
        "genre", "fenetre", "status", "formulaire_rempli_at",
        "nombre_seances_declare", "total_seances",
    ):
        # This tab reports people who have actually been called.  Do not load
        # records still waiting for their first call into its totals.
        if not (
            is_call_attempted_status(call["status"])
            or call["formulaire_rempli_at"]
        ):
            continue
        genre = call["genre"] or "Non renseigné"
        fenetre = call["fenetre"] or "Non renseignée"
        key = (genre, fenetre)
        bucket = campaign.setdefault(
            key, {"genre": genre, "fenetre": fenetre, "total": 0, "tentes": 0, "reussis": 0, "formes": 0}
        )
        bucket["total"] += 1
        bucket["tentes"] += 1
        if call["formulaire_rempli_at"] or call["status"] in {
            "appel_reussi", "formulaire_rempli", "formulaire_avec_audio", "termine"
        }:
            bucket["reussis"] += 1
        declared_sessions = call["nombre_seances_declare"]
        planned_sessions = call["total_seances"]
        if (
            declared_sessions is not None
            and planned_sessions is not None
            and planned_sessions > 0
            and declared_sessions != 0
            and 75 <= declared_sessions / planned_sessions * 100 <= 120
        ):
            bucket["formes"] += 1
    campaign_rows = sorted(campaign.values(), key=lambda row: (row["fenetre"], row["genre"]))
    not_formed_campaign_rows, not_formed_campaign_summary = _build_pas_forme_ii_campaign()
    formed_people_summary = _formed_people_summary(_concordance_window_summary(concordance, headers))
    concordance_counts = {}
    for row in concordance:
        key = (row.genre, row.fenetre)
        concordance_counts[key] = concordance_counts.get(key, 0) + 1
    synthesis = []
    for key in sorted(set(campaign) | set(concordance_counts), key=lambda value: (value[1], value[0])):
        campaign_row = campaign.get(key, {})
        synthesis.append({
            "genre": key[0], "fenetre": key[1], "concordance": concordance_counts.get(key, 0),
            "appels": campaign_row.get("total", 0), "tentes": campaign_row.get("tentes", 0), "reussis": campaign_row.get("reussis", 0),
        })
    pending_contact_import = PendingLearnerContactImport.objects.first()
    pending_contact_headers = pending_contact_import.headers if pending_contact_import else []
    pending_contact_count = (
        pending_contact_import.records.count() if pending_contact_import else 0
    )
    pending_contact_rows = []
    if pending_contact_import:
        pending_contact_rows = [
            {
                "row_number": record.row_number,
                "values": [record.payload.get(header, "") for header in pending_contact_headers],
            }
            for record in pending_contact_import.records.all()[:100]
        ]
    synthesis_reconciliation_rows = _build_synthesis_reconciliation_rows(
        concordance,
        headers,
        not_formed_campaign_rows,
        pending_contact_import,
    )
    synthesis_reconciliation_summary = {
        method: sum(
            1 for row in synthesis_reconciliation_rows if row["methode"] == method
        )
        for method in ("RC", "RA", "R")
    }
    synthesis_reconciliation_summary["total"] = len(synthesis_reconciliation_rows)

    concordance_gender_summary = _concordance_window_summary(concordance, headers)
    campaign_gender_summary = _window_gender_summary(
        campaign_rows,
        gender_key=lambda row: row["genre"],
        window_key=lambda row: row["fenetre"],
        value_key=lambda row: row["formes"],
    )
    campaign_formed_people_summary = _formed_people_summary(campaign_gender_summary)

    return render(request, "reporting/concordance_campaigns.html", {
        "concordance_count": len(concordance), "filtered_concordance_count": len(filtered_concordance),
        "headers": headers, "concordance_rows": concordance_rows, "is_feuil2_layout": is_feuil2_layout, "filter_options": filter_options,
        "selected_fenetre": selected_fenetre, "selected_prestataire": selected_prestataire,
        "selected_beneficiaire": selected_beneficiaire, "selected_presta_id": selected_presta_id, "search_query": search_query,
        "campaign_rows": campaign_rows, "synthesis_rows": synthesis,
        "not_formed_campaign_rows": not_formed_campaign_rows,
        "not_formed_campaign_summary": not_formed_campaign_summary,
        "formed_people_summary": formed_people_summary,
        "synthesis_reconciliation_rows": synthesis_reconciliation_rows,
        "synthesis_reconciliation_summary": synthesis_reconciliation_summary,
        "pending_contact_import": pending_contact_import,
        "pending_contact_count": pending_contact_count,
        "pending_contact_headers": pending_contact_headers,
        "pending_contact_rows": pending_contact_rows,
        "can_manage_concordance_import": can_manage_concordance_import,
        # This recap is global; filters apply only to the detailed import grid.
        "concordance_gender_summary": concordance_gender_summary,
        "campaign_gender_summary": campaign_gender_summary,
        "campaign_formed_people_summary": campaign_formed_people_summary,
        # The consolidated recap combines trained people from both
        # reconciliation sources: concordance (RC) and call campaigns (RA).
        "synthesis_gender_summary": _combine_gender_summaries(
            concordance_gender_summary, campaign_gender_summary
        ),
    })
