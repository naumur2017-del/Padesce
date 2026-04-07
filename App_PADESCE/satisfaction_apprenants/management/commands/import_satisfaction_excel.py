from __future__ import annotations

import hashlib
import re
import unicodedata
import warnings
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from App_PADESCE.appels.models import APPEL_ANSWER_QUESTION_FIELDS, Appel, AppelAnswers
from App_PADESCE.apprenants.models import Apprenant
from App_PADESCE.formations.models import (
    Beneficiaire,
    Classe,
    Formation,
    Inspecteur,
    Lieu,
    Prestataire,
    Prestation,
)
from App_PADESCE.reporting.network_excel import build_padesce_source_index, normalize_network_lookup
from App_PADESCE.satisfaction_apprenants.models import SatisfactionApprenant

WORKSHEET_NAME = "Appels termines"
QUESTION_INDEXES = {
    "q1_clarte_exposes": 5,
    "q2_interaction_formateur": 6,
    "q3_maitrise_contenu": 7,
    "q4_salle_adequate": 8,
    "q5_materiel_disponible": 9,
    "q6_organisation_temps": 10,
    "q7_utilite_formation": 11,
    "q8_adequation_besoins": 12,
    "q9_satisfaction_globale": 13,
}


@dataclass
class WorksheetRow:
    row_number: int
    apprenant: str
    prestataire: str
    beneficiaire: str
    classe_label: str
    commentaire: str
    recommandations: str
    answers: dict[str, int | None]

    @property
    def has_any_score(self) -> bool:
        return any(value is not None for value in self.answers.values())

    @property
    def has_any_payload(self) -> bool:
        return self.has_any_score or bool(self.commentaire) or bool(self.recommandations)

    @property
    def has_complete_scores(self) -> bool:
        return all(self.answers.get(field) is not None for field in APPEL_ANSWER_QUESTION_FIELDS)


def _normalize_text(value) -> str:
    text = str(value or "").replace("\u2060", " ")
    text = " ".join(text.strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    without_marks = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    cleaned = re.sub(r"[^a-z0-9 ]+", " ", without_marks)
    return " ".join(cleaned.split())


def _tokenize_name(value) -> list[str]:
    return [token for token in _normalize_text(value).split() if token]


def _note_value(value) -> tuple[int | None, bool]:
    if value in (None, ""):
        return None, False
    try:
        numeric = Decimal(str(value).strip())
    except (InvalidOperation, AttributeError):
        return None, False
    rounded = int(numeric.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if not 1 <= rounded <= 5:
        return None, False
    return rounded, numeric != rounded


def _same_structure(left: str, right: str) -> bool:
    left_value = _normalize_text(left)
    right_value = _normalize_text(right)
    return bool(left_value and right_value and left_value == right_value)


def _candidate_score(name: str, candidate_name: str) -> tuple[int, int, float]:
    source_tokens = _tokenize_name(name)
    candidate_tokens = _tokenize_name(candidate_name)
    exact_overlap = len(set(source_tokens) & set(candidate_tokens))
    prefix_hits = sum(
        1
        for token in source_tokens
        if any(
            candidate.startswith(token) or token.startswith(candidate)
            for candidate in candidate_tokens
        )
    )
    sequence = SequenceMatcher(None, _normalize_text(name), _normalize_text(candidate_name)).ratio()
    return prefix_hits, exact_overlap, sequence


def _unique_name_match(candidates: list[dict], row: WorksheetRow) -> tuple[dict | None, str]:
    if not row.classe_label or not row.apprenant:
        return None, ""

    preferred_candidates = [
        record
        for record in candidates
        if (not row.prestataire or _same_structure(record.get("prestataire", ""), row.prestataire))
        and (
            not row.beneficiaire
            or _same_structure(record.get("beneficiaire", ""), row.beneficiaire)
        )
    ]
    pools = (
        [preferred_candidates, candidates]
        if preferred_candidates and preferred_candidates != candidates
        else [candidates]
    )
    row_name = _normalize_text(row.apprenant)
    row_tokens = _tokenize_name(row.apprenant)

    for pool in pools:
        exact = [
            record for record in pool if _normalize_text(record.get("nom_individu", "")) == row_name
        ]
        if len(exact) == 1:
            return exact[0], "exact_name"

        orderless = [
            record
            for record in pool
            if sorted(_tokenize_name(record.get("nom_individu", ""))) == sorted(row_tokens)
        ]
        if len(orderless) == 1:
            return orderless[0], "token_orderless"

        prefix = []
        for record in pool:
            candidate_tokens = _tokenize_name(record.get("nom_individu", ""))
            if not row_tokens or len(row_tokens) > len(candidate_tokens):
                continue
            if all(
                any(
                    candidate.startswith(token) or token.startswith(candidate)
                    for candidate in candidate_tokens
                )
                for token in row_tokens
            ):
                prefix.append(record)
        if len(prefix) == 1:
            return prefix[0], "token_prefix"

        scored = [
            (*_candidate_score(row.apprenant, record.get("nom_individu", "")), record)
            for record in pool
        ]
        scored.sort(key=lambda item: (item[0], item[1], item[2]), reverse=True)
        if not scored:
            continue
        best_prefix, best_overlap, best_ratio, best_record = scored[0]
        second_prefix, second_overlap, second_ratio = (
            scored[1][:3] if len(scored) > 1 else (-1, -1, 0.0)
        )
        prefix_margin = best_prefix - second_prefix
        overlap_margin = best_overlap - second_overlap
        ratio_margin = best_ratio - second_ratio
        confident = (
            (
                best_prefix >= max(2, len(row_tokens) - 1)
                and best_ratio >= 0.55
                and (prefix_margin > 0 or ratio_margin >= 0.08)
            )
            or (
                best_overlap >= 2
                and best_ratio >= 0.5
                and (overlap_margin > 0 or ratio_margin >= 0.08)
            )
            or (len(row_tokens) <= 1 and best_ratio >= 0.8 and ratio_margin >= 0.12)
            or (best_ratio >= 0.92 and ratio_margin >= 0.12)
        )
        if confident:
            return best_record, "heuristic_name"

    return None, ""


def _synthetic_code(row: WorksheetRow, prefix: str) -> str:
    signature = " | ".join(
        [
            _normalize_text(row.apprenant),
            _normalize_text(row.prestataire),
            _normalize_text(row.beneficiaire),
            _normalize_text(row.classe_label),
        ]
    )
    digest = hashlib.sha1(signature.encode("utf-8")).hexdigest()[:12].upper()
    return f"{prefix}-{digest}"


def _find_source_record(
    row: WorksheetRow, source_by_class: dict[str, list[dict]]
) -> tuple[dict | None, str]:
    class_key = normalize_network_lookup(row.classe_label)
    if not class_key:
        return None, "missing_class"
    candidates = list(source_by_class.get(class_key, []))
    if not candidates:
        return None, "missing_class"
    record, method = _unique_name_match(candidates, row)
    return record, method or "unresolved_name"


def _find_existing_appel(
    row: WorksheetRow, source_record: dict | None, synthetic_prefix: str
) -> tuple[Appel | None, str]:
    source_code = str((source_record or {}).get("code") or "").strip()
    if source_code:
        appel = Appel.objects.filter(code__iexact=source_code).first()
        if appel:
            return appel, "source_code"

    synthetic_code = _synthetic_code(row, synthetic_prefix)
    appel = Appel.objects.filter(code__iexact=synthetic_code).first()
    if appel:
        return appel, "synthetic_code"

    classe_label = str((source_record or {}).get("classe_id") or row.classe_label or "").strip()
    if not classe_label:
        return None, ""

    candidates = list(
        Appel.objects.filter(classe_label__iexact=classe_label).only(
            "id",
            "code",
            "nom",
            "prestataire",
            "beneficiaire",
            "classe_label",
        )
    )
    record, method = _unique_name_match(
        [
            {
                "code": candidate.code,
                "nom_individu": candidate.nom,
                "prestataire": candidate.prestataire,
                "beneficiaire": candidate.beneficiaire,
            }
            for candidate in candidates
        ],
        row,
    )
    if not record:
        return None, ""
    matched_code = str(record.get("code") or "").strip()
    return (
        next((candidate for candidate in candidates if candidate.code == matched_code), None),
        method,
    )


def _clean_text(value) -> str:
    return str(value or "").strip()


def _entity_code(prefix: str, *parts: str, max_length: int = 20) -> str:
    signature = " | ".join(_normalize_text(part) for part in parts if _normalize_text(part))
    digest_length = max(6, min(12, max_length - len(prefix) - 1))
    digest = hashlib.sha1(signature.encode("utf-8")).hexdigest()[:digest_length].upper()
    return f"{prefix}-{digest}"


def _cohorte_value(value) -> int:
    text = _clean_text(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    try:
        return max(1, int(digits or "1"))
    except ValueError:
        return 1


def _statut_actif(value) -> bool:
    status = normalize_network_lookup(value)
    if not status:
        return True
    return status not in {"inactif", "inactive", "abandon", "abandonne", "sorti"}


def _local_classe_status(value) -> str:
    status = normalize_network_lookup(value)
    if status in {"termine", "terminee", "arrete"}:
        return "termine"
    if status in {"en cours", "encours", "demarre", "demarree"}:
        return "en_cours"
    return "non_demarre"


def _reference_cache() -> dict[str, dict]:
    return {
        "formations": {},
        "prestataires": {},
        "beneficiaires": {},
        "lieux": {},
        "prestations": {},
        "classes": {},
        "inspecteurs": {},
        "apprenants": {},
    }


def _sync_source_models(
    source_record: dict | None, cache: dict[str, dict] | None = None
) -> dict[str, object | None]:
    if not source_record:
        return {
            "formation": None,
            "prestataire": None,
            "beneficiaire": None,
            "lieu": None,
            "prestation": None,
            "classe": None,
            "inspecteur": None,
            "apprenant": None,
        }

    cache = cache or _reference_cache()

    formation_name = _clean_text(source_record.get("formation") or "")
    formation_key = _normalize_text(formation_name)
    formation = cache["formations"].get(formation_key)
    if formation is None and formation_name:
        formation_code = _entity_code("FOR", formation_name)
        formation = Formation.objects.filter(code=formation_code).first()
        if formation is None:
            formation = Formation.objects.filter(nom__iexact=formation_name).first()
        if formation is None:
            formation = Formation(code=formation_code)
        formation.nom = formation_name
        formation.nom_harmonise = formation_name
        formation.fenetre = _clean_text(source_record.get("fenetre") or formation.fenetre)
        formation.actif = True
        if formation.pk is None:
            formation.statut = _local_classe_status(source_record.get("statut_prestation"))
        formation.save()
        cache["formations"][formation_key] = formation

    prestataire_name = _clean_text(source_record.get("prestataire") or "")
    prestataire_key = _normalize_text(prestataire_name)
    prestataire = cache["prestataires"].get(prestataire_key)
    if prestataire is None and prestataire_name:
        prestataire_code = _entity_code("PST", prestataire_name, max_length=50)
        prestataire = Prestataire.objects.filter(code=prestataire_code).first()
        if prestataire is None:
            prestataire = Prestataire.objects.filter(
                raison_sociale__iexact=prestataire_name
            ).first()
        if prestataire is None:
            prestataire = Prestataire(code=prestataire_code)
        prestataire.raison_sociale = prestataire_name
        prestataire.actif = True
        prestataire.save()
        cache["prestataires"][prestataire_key] = prestataire

    beneficiaire_name = _clean_text(source_record.get("beneficiaire") or "")
    beneficiaire_key = _normalize_text(beneficiaire_name)
    beneficiaire = cache["beneficiaires"].get(beneficiaire_key)
    if beneficiaire is None and beneficiaire_name:
        beneficiaire = Beneficiaire.objects.filter(nom_structure__iexact=beneficiaire_name).first()
        if beneficiaire is None:
            beneficiaire = Beneficiaire()
        beneficiaire.nom_structure = beneficiaire_name
        beneficiaire.region = _clean_text(source_record.get("region") or "")
        beneficiaire.ville = _clean_text(source_record.get("ville") or "")
        beneficiaire.actif = True
        beneficiaire.save()
        cache["beneficiaires"][beneficiaire_key] = beneficiaire

    lieu_name = _clean_text(source_record.get("lieu") or "")
    lieu_key = _normalize_text(lieu_name)
    lieu = cache["lieux"].get(lieu_key)
    if lieu is None and lieu_name:
        lieu_code = _entity_code("LIEU", lieu_name, source_record.get("ville"), max_length=50)
        lieu = Lieu.objects.filter(code=lieu_code).first()
        if lieu is None:
            lieu = Lieu.objects.filter(nom_lieu__iexact=lieu_name).first()
        if lieu is None:
            lieu = Lieu(code=lieu_code)
        lieu.nom_lieu = lieu_name
        lieu.region = _clean_text(source_record.get("region") or "")
        lieu.ville = _clean_text(source_record.get("ville") or "")
        lieu.actif = True
        lieu.save()
        cache["lieux"][lieu_key] = lieu

    prestation = None
    prestation_code = _clean_text(source_record.get("prestation_id") or "")
    if prestation_code and formation and prestataire:
        prestation = cache["prestations"].get(prestation_code)
        if prestation is None:
            prestation = Prestation.objects.filter(code=prestation_code).first()
        if prestation is None:
            prestation = Prestation(code=prestation_code)
        prestation.prestataire = prestataire
        prestation.formation = formation
        prestation.beneficiaire = beneficiaire
        prestation.actif = True
        prestation.save()
        cache["prestations"][prestation_code] = prestation

    classe = None
    classe_code = _clean_text(source_record.get("classe_id") or "")
    if classe_code and prestation and formation:
        classe = cache["classes"].get(classe_code)
        if classe is None:
            classe = Classe.objects.filter(code=classe_code).first()
        if classe is None:
            classe = Classe(code=classe_code)
        classe.prestation = prestation
        classe.lieu = lieu
        classe.formation = formation
        classe.intitule_formation = (
            formation_name or getattr(classe, "intitule_formation", "") or classe_code
        )
        classe.fenetre = _clean_text(source_record.get("fenetre") or "")
        classe.cohorte = _cohorte_value(source_record.get("cohorte"))
        classe.statut = _local_classe_status(source_record.get("statut_prestation"))
        classe.actif = True
        classe.save()
        cache["classes"][classe_code] = classe

    inspecteur = None
    inspecteur_code = _clean_text(source_record.get("inspecteur_id") or "")
    if inspecteur_code:
        inspecteur = cache["inspecteurs"].get(inspecteur_code)
        if inspecteur is None:
            inspecteur, _created = Inspecteur.objects.update_or_create(
                code=inspecteur_code,
                defaults={
                    "nom_complet": _clean_text(
                        source_record.get("inspecteur_label") or inspecteur_code
                    ),
                    "actif": True,
                },
            )
            cache["inspecteurs"][inspecteur_code] = inspecteur

    apprenant = None
    apprenant_code = _clean_text(source_record.get("apprenant_id") or "")
    apprenant_name = _clean_text(source_record.get("nom_individu") or "")
    if apprenant_code and classe and formation:
        apprenant = cache["apprenants"].get(apprenant_code)
        if apprenant is None:
            apprenant = Apprenant.objects.filter(code=apprenant_code).first()
        if apprenant is None and apprenant_name:
            apprenant = Apprenant.objects.filter(
                classe=classe, nom_complet__iexact=apprenant_name
            ).first()
        if apprenant is None:
            apprenant = Apprenant(code=apprenant_code, classe=classe, formation=formation)
        apprenant.code = apprenant_code
        apprenant.numero = _clean_text(source_record.get("numero") or "")
        apprenant.classe = classe
        apprenant.formation = formation
        apprenant.nom_complet = apprenant_name or apprenant.nom_complet
        apprenant.beneficiaire = beneficiaire_name
        apprenant.genre = _clean_text(source_record.get("sexe") or "")
        apprenant.fenetre = _clean_text(source_record.get("fenetre") or "")
        apprenant.prestataire = prestataire_name
        apprenant.intitule_formation_solicitee = formation_name
        apprenant.intitule_formation_dispensee = formation_name
        apprenant.ville_formation = _clean_text(source_record.get("ville") or "")
        apprenant.cohorte = _clean_text(source_record.get("cohorte") or "")
        apprenant.region = _clean_text(source_record.get("region") or "")
        apprenant.lieu_formation = lieu_name
        apprenant.telephone1 = None
        apprenant.telephone2 = None
        apprenant.actif = _statut_actif(source_record.get("statut_apprenant"))
        apprenant.save()
        cache["apprenants"][apprenant_code] = apprenant

    return {
        "formation": formation,
        "prestataire": prestataire,
        "beneficiaire": beneficiaire,
        "lieu": lieu,
        "prestation": prestation,
        "classe": classe,
        "inspecteur": inspecteur,
        "apprenant": apprenant,
    }


def _worksheet_row(values, row_number: int) -> WorksheetRow | None:
    if not any(value not in (None, "") for value in values):
        return None
    answers: dict[str, int | None] = {}
    for field, index in QUESTION_INDEXES.items():
        note, _rounded = _note_value(values[index] if index < len(values) else None)
        answers[field] = note
    return WorksheetRow(
        row_number=row_number,
        apprenant=_clean_text(values[1] if len(values) > 1 else ""),
        prestataire=_clean_text(values[2] if len(values) > 2 else ""),
        beneficiaire=_clean_text(values[3] if len(values) > 3 else ""),
        classe_label=_clean_text(values[4] if len(values) > 4 else ""),
        commentaire=_clean_text(values[14] if len(values) > 14 else ""),
        recommandations=_clean_text(values[15] if len(values) > 15 else ""),
        answers=answers,
    )


class Command(BaseCommand):
    help = (
        "Importe la feuille Appels termines d'un fichier Excel de satisfaction, "
        "met a jour les appels existants et synchronise les reponses."
    )

    def add_arguments(self, parser):
        parser.add_argument("file_path", help="Chemin du fichier Excel a importer.")
        parser.add_argument(
            "--sheet",
            default=WORKSHEET_NAME,
            help=f"Nom de la feuille source. Defaut: {WORKSHEET_NAME}.",
        )
        parser.add_argument(
            "--synthetic-prefix", default="SATXLS", help="Prefixe utilise pour les codes generes."
        )
        parser.add_argument(
            "--dry-run", action="store_true", help="Analyse le fichier sans modifier la base."
        )

    def handle(self, *args, **options):
        file_path = Path(options["file_path"]).expanduser()
        worksheet_name = str(options["sheet"] or WORKSHEET_NAME).strip() or WORKSHEET_NAME
        synthetic_prefix = str(options["synthetic_prefix"] or "SATXLS").strip().upper() or "SATXLS"
        dry_run = bool(options["dry_run"])

        if not file_path.exists():
            raise CommandError(f"Fichier introuvable: {file_path}")

        try:
            source_bundle = build_padesce_source_index()
        except Exception as exc:
            raise CommandError(f"Impossible de charger la source reseau: {exc}") from exc

        warnings.filterwarnings("ignore", category=UserWarning)
        workbook = load_workbook(file_path, data_only=True)
        if worksheet_name not in workbook.sheetnames:
            raise CommandError(f"Feuille introuvable: {worksheet_name}")

        source_by_class: dict[str, list[dict]] = {}
        for record in source_bundle.get("records", {}).values():
            class_key = normalize_network_lookup(record.get("classe_id", ""))
            if class_key:
                source_by_class.setdefault(class_key, []).append(record)

        now = timezone.localtime()
        summary = {
            "rows": 0,
            "rounded_scores": 0,
            "source_matched": 0,
            "source_missing": 0,
            "created_appels": 0,
            "updated_appels": 0,
            "created_answers": 0,
            "updated_answers": 0,
            "deleted_answers": 0,
            "created_surveys": 0,
            "updated_surveys": 0,
            "deleted_surveys": 0,
            "synthetic_codes": 0,
        }
        source_methods: dict[str, int] = {}
        unresolved_rows: list[str] = []
        reference_cache = _reference_cache()

        worksheet = workbook[worksheet_name]
        row_iter = worksheet.iter_rows(min_row=2, values_only=True)

        with transaction.atomic():
            for excel_row_number, raw_values in enumerate(row_iter, start=2):
                row = _worksheet_row(raw_values, excel_row_number)
                if row is None:
                    continue
                summary["rows"] += 1
                for field, index in QUESTION_INDEXES.items():
                    _note, rounded = _note_value(
                        raw_values[index] if index < len(raw_values) else None
                    )
                    if rounded:
                        summary["rounded_scores"] += 1

                source_record, source_method = _find_source_record(row, source_by_class)
                source_methods[source_method] = source_methods.get(source_method, 0) + 1
                if source_record:
                    summary["source_matched"] += 1
                else:
                    summary["source_missing"] += 1
                    unresolved_rows.append(
                        f"Ligne {row.row_number}: {row.apprenant or '-'} | {row.classe_label or 'classe absente'}"  # noqa: E501
                    )

                appel, matched_by = _find_existing_appel(row, source_record, synthetic_prefix)
                source_links = _sync_source_models(source_record, reference_cache)
                target_code = _clean_text((source_record or {}).get("code") or "")
                if not target_code and appel is not None:
                    target_code = _clean_text(appel.code)
                if not target_code:
                    target_code = _synthetic_code(row, synthetic_prefix)
                    summary["synthetic_codes"] += 1
                target_name = _clean_text(
                    (source_record or {}).get("nom_individu") or row.apprenant
                )
                target_class = _clean_text(
                    (source_record or {}).get("classe_id") or row.classe_label
                )
                target_prestataire = _clean_text(
                    (source_record or {}).get("prestataire") or row.prestataire
                )
                target_beneficiaire = _clean_text(
                    (source_record or {}).get("beneficiaire") or row.beneficiaire
                )

                appel_defaults = {
                    "code": target_code,
                    "nom": target_name,
                    "prestataire": target_prestataire,
                    "beneficiaire": target_beneficiaire,
                    "lieu": _clean_text((source_record or {}).get("lieu") or ""),
                    "classe_label": target_class,
                    "fenetre": _clean_text((source_record or {}).get("fenetre") or ""),
                    "formation_padesce": _clean_text((source_record or {}).get("formation") or ""),
                    "classe": source_links.get("classe"),
                    "status": "termine",
                    "is_active": True,
                }

                if appel is None:
                    appel = Appel.objects.create(**appel_defaults)
                    summary["created_appels"] += 1
                else:
                    for field, value in appel_defaults.items():
                        setattr(appel, field, value)
                    appel.save()
                    summary["updated_appels"] += 1

                if not row.has_any_payload:
                    deleted_answers, _ = AppelAnswers.objects.filter(appel=appel).delete()
                    if deleted_answers:
                        summary["deleted_answers"] += 1
                    deleted_surveys, _ = SatisfactionApprenant.objects.filter(appel=appel).delete()
                    if deleted_surveys:
                        summary["deleted_surveys"] += 1
                    continue

                answer_defaults = {
                    **{field: row.answers.get(field) for field in APPEL_ANSWER_QUESTION_FIELDS},
                    "commentaire": row.commentaire,
                    "recommandations": row.recommandations,
                    "modified_by": None,
                    "modified_at": now,
                }
                answers, created_answers = AppelAnswers.objects.update_or_create(
                    appel=appel,
                    defaults=answer_defaults,
                )
                summary["created_answers" if created_answers else "updated_answers"] += 1

                if row.has_complete_scores:
                    survey_defaults = {
                        "classe": source_links.get("classe"),
                        "apprenant": source_links.get("apprenant"),
                        "inspecteur": source_links.get("inspecteur"),
                        "enqueteur": None,
                        "date": now.date(),
                        "heure": now.time().replace(microsecond=0),
                        "commentaire": row.commentaire,
                        "recommandations": row.recommandations,
                        "transcription": "",
                        **{field: row.answers[field] for field in APPEL_ANSWER_QUESTION_FIELDS},
                    }
                    survey, created_survey = SatisfactionApprenant.objects.update_or_create(
                        appel=appel,
                        defaults=survey_defaults,
                    )
                    del survey
                    summary["created_surveys" if created_survey else "updated_surveys"] += 1
                else:
                    deleted_surveys, _ = SatisfactionApprenant.objects.filter(appel=appel).delete()
                    if deleted_surveys:
                        summary["deleted_surveys"] += 1

            if dry_run:
                transaction.set_rollback(True)

        workbook.close()

        source_breakdown = ", ".join(
            f"{label}={count}"
            for label, count in sorted(source_methods.items(), key=lambda item: item[0])
            if count
        )
        mode_label = "Simulation terminee" if dry_run else "Import termine"
        self.stdout.write(
            self.style.SUCCESS(
                (
                    f"{mode_label}. "
                    f"{summary['rows']} ligne(s) traitee(s), "
                    f"{summary['source_matched']} rattachee(s) au reseau, "
                    f"{summary['created_appels']} appel(s) cree(s), {summary['updated_appels']} mis a jour, "  # noqa: E501
                    f"{summary['created_answers']} reponse(s) creee(s), {summary['updated_answers']} mise(s) a jour, "  # noqa: E501
                    f"{summary['deleted_answers']} reponse(s) supprimee(s), "
                    f"{summary['created_surveys']} enquete(s) complete(s) creee(s), "
                    f"{summary['updated_surveys']} mise(s) a jour, {summary['deleted_surveys']} supprimee(s), "  # noqa: E501
                    f"{summary['synthetic_codes']} code(s) synthetique(s), "
                    f"{summary['rounded_scores']} note(s) decimale(s) arrondie(s)."
                )
            )
        )
        if source_breakdown:
            self.stdout.write(f"Resolution source: {source_breakdown}")
        if unresolved_rows:
            preview = "\n".join(unresolved_rows[:10])
            self.stdout.write(
                self.style.WARNING(
                    f"{len(unresolved_rows)} ligne(s) sans rattachement reseau strict.\n{preview}"
                )
            )
