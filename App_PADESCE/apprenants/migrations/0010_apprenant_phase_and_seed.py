from datetime import date
from pathlib import Path
import unicodedata

from django.db import migrations, models
import django.db.models.deletion


def _normalize(value: str) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).upper()


def _load_cutoff_terminated_prestations(base_dir: Path) -> set[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return set()

    candidates = [
        base_dir / "data" / "network_excel_bundle" / "network-fichier-consolide-cutoff.xlsm",
        base_dir / "docs" / "data" / "network_excel_cache" / "network-fichier-consolide-cutoff.xlsm",
    ]
    workbook_path = next((path for path in candidates if path.exists()), None)
    if workbook_path is None:
        return set()

    wb = load_workbook(workbook_path, read_only=True, data_only=True, keep_vba=True)
    if "Prestations" not in wb.sheetnames:
        return set()

    ws = wb["Prestations"]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers_norm = [_normalize(h) for h in headers]

    try:
        idx_id = headers_norm.index("ID PRESTATION")
    except ValueError:
        return set()

    status_headers = {"STATUT DE LA PRESTATION", "STATUT AVEC TAUX", "STATUT"}
    idx_status = next((i for i, h in enumerate(headers_norm) if h in status_headers), None)
    if idx_status is None:
        return set()

    terminated_keywords = ("TERMIN", "ARRET")
    presta_codes: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code = str(row[idx_id] or "").strip().upper()
        raw_status = _normalize(row[idx_status] or "")
        if not raw_code:
            continue
        if not any(keyword in raw_status for keyword in terminated_keywords):
            continue
        presta_codes.add(raw_code)
    return presta_codes


def _seed_and_assign_phases(apps, schema_editor):
    Phase = apps.get_model("formations", "Phase")
    Prestation = apps.get_model("formations", "Prestation")
    Classe = apps.get_model("formations", "Classe")
    Formateur = apps.get_model("formations", "Formateur")
    Apprenant = apps.get_model("apprenants", "Apprenant")
    AppelFormateur = apps.get_model("appels", "AppelFormateur")

    phase1, _ = Phase.objects.get_or_create(
        id_phase=1,
        defaults={"date_debut": date(2025, 9, 24), "date_fin": date(2026, 2, 28)},
    )
    phase2, _ = Phase.objects.get_or_create(
        id_phase=2,
        defaults={"date_debut": date(2026, 3, 1), "date_fin": None},
    )

    base_dir = Path(__file__).resolve().parents[4]
    phase1_presta_codes = _load_cutoff_terminated_prestations(base_dir)

    if phase1_presta_codes:
        phase1_presta_ids = set(
            Prestation.objects.filter(code__in=phase1_presta_codes).values_list("id", flat=True)
        )
    else:
        phase1_presta_ids = set()

    if phase1_presta_ids:
        Prestation.objects.filter(id__in=phase1_presta_ids).update(phase=phase1)
    Prestation.objects.exclude(id__in=phase1_presta_ids).update(phase=phase2)

    phase1_class_ids = set(
        Classe.objects.filter(prestation_id__in=phase1_presta_ids).values_list("id", flat=True)
    )
    if phase1_class_ids:
        Classe.objects.filter(id__in=phase1_class_ids).update(phase=phase1)
    Classe.objects.exclude(id__in=phase1_class_ids).update(phase=phase2)

    if phase1_class_ids:
        Apprenant.objects.filter(classe_id__in=phase1_class_ids).update(phase=phase1)
    Apprenant.objects.exclude(classe_id__in=phase1_class_ids).update(phase=phase2)

    cutoff_formateur_phones = set(
        AppelFormateur.objects.exclude(telephone__isnull=True)
        .exclude(telephone__exact="")
        .values_list("telephone", flat=True)
    )
    phase1_formateur_ids = set(
        Formateur.objects.filter(telephone__in=cutoff_formateur_phones).values_list("id", flat=True)
    )

    if phase1_formateur_ids:
        Formateur.objects.filter(id__in=phase1_formateur_ids).update(phase=phase1)

    # Règle métier demandée: le formateur restant (valeur test) doit rester sans phase.
    Formateur.objects.exclude(id__in=phase1_formateur_ids).exclude(code="FORMA001").update(phase=phase2)
    Formateur.objects.filter(code="FORMA001").update(phase=None)


class Migration(migrations.Migration):

    dependencies = [
        ("apprenants", "0009_data_presence_controls"),
        ("formations", "0009_phase_and_phase_fk"),
    ]

    operations = [
        migrations.AddField(
            model_name="apprenant",
            name="phase",
            field=models.ForeignKey(
                blank=True,
                null=True,
                on_delete=django.db.models.deletion.SET_NULL,
                related_name="apprenants",
                to="formations.phase",
            ),
        ),
        migrations.RunPython(_seed_and_assign_phases, migrations.RunPython.noop),
    ]
