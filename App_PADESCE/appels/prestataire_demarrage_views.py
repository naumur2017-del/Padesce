import csv
import datetime
import io
import re
import unicodedata

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from App_PADESCE.appels.models import (
    CALL_FORM_STATUSES,
    CALL_SUCCESS_STATUSES,
    AppelPrestataireDemarrage,
    infer_padesce_status,
)
from App_PADESCE.appels.pagination import build_pagination_tokens
from App_PADESCE.appels.views import _bind_audio_state, _cleanup_stale_locks, _safe_audio_url
from App_PADESCE.core.phase_scope import PHASE_SCOPE_V1_COMBINED, normalize_phase_scope
from App_PADESCE.formations.models import Prestataire


def _normalize_header(value):
    if value is None:
        return ""
    text = " ".join(str(value).strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_lookup(value):
    text = re.sub(r"[^a-z0-9]+", " ", _normalize_header(value))
    return " ".join(text.split())


def _as_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_phone(value):
    digits = re.sub(r"\D", "", _as_text(value))
    if digits.startswith("00237"):
        digits = digits[5:]
    elif digits.startswith("237") and len(digits) > 9:
        digits = digits[3:]
    return digits[-9:] if len(digits) > 9 else digits


def _reference_for_item(item: dict) -> str:
    code = item.get("prestataire_code") or ""
    phone = item.get("telephone") or "sans-tel"
    name_key = re.sub(r"[^a-z0-9]+", "-", _normalize_lookup(item.get("nom_prestataire"))).strip("-")
    return f"{code or name_key or 'prestataire'}-{phone}"[:120]


def _resolve_prestataire(item: dict) -> tuple[Prestataire | None, str]:
    code = _as_text(item.get("prestataire_code"))
    if code:
        prestataire = Prestataire.objects.filter(code__iexact=code).first()
        if prestataire:
            return prestataire, "code"

    name_key = _normalize_lookup(item.get("nom_prestataire"))
    short_key = _normalize_lookup(item.get("nom_simplifie"))
    if not name_key and not short_key:
        return None, ""

    for prestataire in Prestataire.objects.filter(actif=True).only("id", "raison_sociale", "code"):
        db_name = _normalize_lookup(prestataire.raison_sociale)
        db_code = _normalize_lookup(prestataire.code)
        if name_key and name_key == db_name:
            return prestataire, "nom_exact"
        if short_key and short_key in {db_name, db_code}:
            return prestataire, "nom_simplifie"
        if name_key and db_name and (name_key in db_name or db_name in name_key):
            return prestataire, "nom_partiel"
    return None, ""


def _parse_prestataire_demarrage_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            idx = header_map.get(_normalize_header(key))
            if idx is not None and idx < len(row):
                return row[idx] or ""
        return ""

    payload = []
    for row in rows:
        nom = get(row, "Nom du prestataire", "Prestataire", "Nom")
        code = get(row, "", "Code", "Code prestataire")
        if not nom and not code:
            continue
        item = {
            "numero": get(row, "#", "Numero", "N"),
            "prestataire_code": _as_text(code),
            "nom_prestataire": _as_text(nom),
            "nom_simplifie": _as_text(get(row, "Nom simplifié", "Nom simplifie", "Sigle")),
            "telephone": _normalize_phone(get(row, "Numéro de téléphone", "Telephone")),
        }
        if not item["telephone"]:
            continue
        try:
            item["numero"] = int(str(item["numero"]).strip()) if item["numero"] != "" else None
        except (TypeError, ValueError):
            item["numero"] = None
        item["reference_code"] = _reference_for_item(item)
        prestataire, match_method = _resolve_prestataire(item)
        item["prestataire"] = prestataire
        item["match_method"] = match_method
        payload.append(item)
    return payload


def _has_form_data(row: AppelPrestataireDemarrage) -> bool:
    return bool(
        row.prestation_debutee
        or row.date_debut_prestation
        or row.motif_non_demarrage
        or row.commentaire
        or row.faux_numero
        or row.bon_numero
        or row.satisfaction_completed_at
    )


def _has_audio(row: AppelPrestataireDemarrage) -> bool:
    return bool(getattr(row.audio_file, "name", ""))


def _sync_status(row: AppelPrestataireDemarrage, *, save=True) -> str:
    new_status = infer_padesce_status(
        row.status, has_form=_has_form_data(row), has_audio=_has_audio(row)
    )
    if save and new_status != row.status:
        row.status = new_status
        row.save(update_fields=["status", "updated_at"])
    else:
        row.status = new_status
    return new_status


def _build_stats(queryset):
    stats = queryset.aggregate(
        total=Count("id"),
        lies_bd=Count("id", filter=Q(prestataire__isnull=False)),
        appels_tentes=Count("id", filter=~Q(status="en_attente")),
        appels_reussis=Count("id", filter=Q(status__in=CALL_SUCCESS_STATUSES)),
        formulaires_remplis=Count("id", filter=Q(status__in=CALL_FORM_STATUSES)),
        audios=Count("id", filter=Q(audio_file__isnull=False) & ~Q(audio_file="")),
    )
    total = int(stats.get("total") or 0)
    completed = int(stats.get("formulaires_remplis") or 0)
    stats["completion_rate"] = round((completed / total) * 100, 1) if total else 0
    stats["completion_label"] = f"{completed} / {total}" if total else "0 / 0"
    return stats


def _filtered_queryset(request):
    qs = AppelPrestataireDemarrage.objects.filter(is_active=True).select_related("prestataire")
    status = request.GET.get("status", "")
    linked = request.GET.get("linked", "")
    agent = request.GET.get("agent", "")
    q = request.GET.get("q", "").strip()
    if status:
        qs = qs.filter(status=status)
    if linked == "1":
        qs = qs.filter(prestataire__isnull=False)
    elif linked == "0":
        qs = qs.filter(prestataire__isnull=True)
    if agent:
        qs = qs.filter(locked_by__username__iexact=agent)
    if q:
        qs = qs.filter(
            Q(nom_prestataire__icontains=q)
            | Q(nom_simplifie__icontains=q)
            | Q(prestataire_code__icontains=q)
            | Q(telephone__icontains=q)
            | Q(prestataire__raison_sociale__icontains=q)
        )
    filters = {
        "phase_scope": normalize_phase_scope(
            request.GET.get("phase_scope"), default=PHASE_SCOPE_V1_COMBINED
        ),
        "status": status,
        "linked": linked,
        "agent": agent,
        "q": q,
    }
    return qs, filters


@login_required
@transaction.atomic
def prestataire_demarrage_index(request):
    if request.method == "POST" and request.FILES.get("file"):
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut importer des fichiers d'appels.")
            return redirect(request.path_info)
        mode = request.POST.get("update_mode", "replace")
        try:
            payload = _parse_prestataire_demarrage_excel(io.BytesIO(request.FILES["file"].read()))
        except Exception as exc:
            messages.error(request, f"Impossible de lire le fichier : {exc}")
            return redirect(request.path_info)
        if mode == "replace":
            AppelPrestataireDemarrage.objects.update(is_active=False)
        created = 0
        updated = 0
        skipped_duplicates = 0
        seen_references = set()
        for item in payload:
            reference = item["reference_code"]
            if reference in seen_references:
                skipped_duplicates += 1
                continue
            seen_references.add(reference)
            row = AppelPrestataireDemarrage.objects.filter(reference_code=reference).first()
            if row:
                if mode != "replace" and row.is_active:
                    skipped_duplicates += 1
                    continue
                for key, value in item.items():
                    setattr(row, key, value)
                row.is_active = True
                row.save()
                updated += 1
            else:
                AppelPrestataireDemarrage.objects.create(**item, is_active=True)
                created += 1
        messages.success(
            request,
            (
                f"Fichier importe. {created} nouveau(x) appel(s), "
                f"{updated} appel(s) mis a jour, {skipped_duplicates} doublon(s) ignore(s). "
                "Les lignes sans numero de telephone ne sont pas importees."
            ),
        )
        return redirect(request.path_info)

    qs, filters = _filtered_queryset(request)
    stats = _build_stats(qs)
    qs = qs.order_by("status", "nom_prestataire")
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    rows = _bind_audio_state(list(page_obj.object_list))
    page_obj.object_list = rows
    params = request.GET.copy()
    params.pop("page", None)
    return render(
        request,
        "appels/prestataire_demarrage.html",
        {
            "rows": rows,
            "page_obj": page_obj,
            "paginator": paginator,
            "pagination_tokens": build_pagination_tokens(page_obj),
            "querystring_no_page": params.urlencode(),
            "filters": filters,
            "stats": stats,
            "appels_count": qs.count(),
            "motif_choices": AppelPrestataireDemarrage.MOTIF_CHOICES,
        },
    )


@login_required
def prestataire_demarrage_export_filtered_csv(request):
    qs, _filters = _filtered_queryset(request)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="appels-prestataires-demarrage.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(
        [
            "Code",
            "Nom prestataire",
            "Nom simplifie",
            "Telephone",
            "Prestataire BD",
            "Methode liaison",
            "Statut",
            "Prestation debutee",
            "Date debut",
            "Motif non demarrage",
            "Commentaire",
            "Faux numero",
            "Bon numero",
            "Audio URL",
        ]
    )
    for row in qs.order_by("nom_prestataire"):
        writer.writerow(
            [
                row.prestataire_code,
                row.nom_prestataire,
                row.nom_simplifie,
                row.telephone,
                row.prestataire.raison_sociale if row.prestataire else "",
                row.match_method,
                row.get_status_display(),
                row.prestation_debutee,
                row.date_debut_prestation.isoformat() if row.date_debut_prestation else "",
                row.get_motif_non_demarrage_display() if row.motif_non_demarrage else "",
                row.commentaire,
                row.faux_numero,
                row.bon_numero,
                _safe_audio_url(row),
            ]
        )
    return response


@login_required
@require_POST
def prestataire_demarrage_action(request, pk: int):
    _cleanup_stale_locks()
    row = get_object_or_404(AppelPrestataireDemarrage, pk=pk)
    action = request.POST.get("action", "")
    now = timezone.now()
    if action in {"start", "resume"}:
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = now
    elif action == "pause":
        row.status = "pause"
    elif action == "rappeler":
        row.status = "a_rappeler"
        row.locked_by = request.user
        row.locked_at = now
        rappel_at = request.POST.get("rappel_at")
        if rappel_at:
            try:
                row.rappel_at = timezone.make_aware(datetime.datetime.fromisoformat(rappel_at))
            except (TypeError, ValueError):
                row.rappel_at = None
    else:
        return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)
    row.save(update_fields=["status", "locked_by", "locked_at", "rappel_at", "updated_at"])
    return JsonResponse(
        {
            "ok": True,
            "status": row.status,
            "status_label": row.get_status_display(),
            "locked_by": row.locked_by.username if row.locked_by else "",
            "rappel_at": row.rappel_at.isoformat() if row.rappel_at else "",
        }
    )


def _clean_yes_no(value):
    value = str(value or "").strip().upper()
    return value if value in {"OUI", "NON"} else ""


def _parse_date(value):
    value = str(value or "").strip()
    if not value:
        return None
    try:
        return datetime.date.fromisoformat(value)
    except (TypeError, ValueError):
        return None


@login_required
@require_POST
def prestataire_demarrage_finalize(request, pk: int):
    row = get_object_or_404(AppelPrestataireDemarrage, pk=pk)
    action = request.POST.get("action", "terminer")
    if action == "rappeler":
        row.status = "a_rappeler"
        rappel_at = request.POST.get("rappel_at")
        if rappel_at:
            try:
                row.rappel_at = timezone.make_aware(datetime.datetime.fromisoformat(rappel_at))
            except (TypeError, ValueError):
                row.rappel_at = None
        row.save(update_fields=["status", "rappel_at", "updated_at"])
        return JsonResponse(
            {"ok": True, "status": row.status, "status_label": row.get_status_display()}
        )

    row.prestation_debutee = _clean_yes_no(request.POST.get("prestation_debutee"))
    if row.prestation_debutee == "OUI":
        row.date_debut_prestation = _parse_date(request.POST.get("date_debut_prestation"))
        row.motif_non_demarrage = ""
    elif row.prestation_debutee == "NON":
        row.date_debut_prestation = None
        motif = str(request.POST.get("motif_non_demarrage") or "").strip()
        valid_motifs = {key for key, _label in AppelPrestataireDemarrage.MOTIF_CHOICES}
        row.motif_non_demarrage = motif if motif in valid_motifs else ""
    else:
        row.date_debut_prestation = None
        row.motif_non_demarrage = ""
    row.commentaire = str(request.POST.get("commentaire") or "").strip()
    row.faux_numero = _clean_yes_no(request.POST.get("faux_numero"))
    row.bon_numero = _normalize_phone(request.POST.get("bon_numero"))
    row.satisfaction_completed_at = timezone.now()
    file_obj = request.FILES.get("audio")
    if not file_obj and not _has_audio(row):
        return JsonResponse(
            {"ok": False, "error": "L'audio est obligatoire pour finaliser cet appel."},
            status=400,
        )
    if file_obj:
        row.audio_file = file_obj
    row.status = "formulaire_avec_audio"
    row.rappel_at = None
    row.save()
    _sync_status(row)
    return JsonResponse(
        {
            "ok": True,
            "status": row.status,
            "status_label": row.get_status_display(),
            "audio_saved": bool(file_obj),
            "audio_url": _safe_audio_url(row),
        }
    )
