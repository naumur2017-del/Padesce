import io
import re
from collections import defaultdict

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from App_PADESCE.appels.models import AppelPasFormeII
from App_PADESCE.appels.pagination import build_pagination_tokens
from App_PADESCE.appels.thresholds import (
    PAS_FORME_II_THRESHOLD_PERCENT,
    pas_forme_ii_threshold_reached,
    pas_forme_ii_threshold_target,
)


def _header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[-9:] if len(digits) > 9 else digits


def _number(value):
    try:
        return int(value) if value not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _boolean(value):
    return _header(value) in {"1", "oui", "yes", "true", "vrai", "x"}


def _parse(file_obj):
    workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    sheet_name = next(
        (name for name in workbook.sheetnames if _header(name) == "feuil1"),
        None,
    )
    if sheet_name is None:
        raise ValueError("la feuille 'Feuil1' est introuvable")
    ws = workbook[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None) or []
    lookup = {_header(v): i for i, v in enumerate(headers)}

    missing = [
        label
        for key, label in (
            ("prestation id", "PRESTATION ID"),
            ("apprenants", "APPRENANTS"),
        )
        if key not in lookup
    ]
    if missing:
        raise ValueError(f"colonne(s) obligatoire(s) absente(s) : {', '.join(missing)}")

    def get(row, key):
        i = lookup.get(key)
        return row[i] if i is not None and i < len(row) else ""

    result = []
    for row in rows:
        prestation = str(get(row, "prestation id") or "").strip()
        nom = str(get(row, "apprenants") or "").strip()
        if not prestation or not nom:
            continue
        phone = _phone(get(row, "numero"))
        nom_key = re.sub(r"[^a-z0-9]+", "-", nom.lower()).strip("-")
        ref = f"{prestation}-{phone or 'sans-tel'}-{nom_key}"[:160]
        result.append(
            {
                "reference_code": ref,
                "prestation_id": prestation,
                "nom": nom,
                "telephone": phone,
                "beneficiaire": str(get(row, "beneficiaires") or "").strip(),
                "prestataire": str(get(row, "prestataires") or "").strip(),
                "genre": str(get(row, "genre") or "").strip(),
                "fenetre": str(get(row, "fenetre") or "").strip(),
                "absent_dans_consolide": _boolean(get(row, "absent dans consolide")),
                "total_presence": _number(get(row, "total presence")),
                "total_seances": _number(get(row, "total seance")),
                "seuil_75": _number(get(row, "seuil 75%")),
                "nombre_seances_source": _number(get(row, "nombre seance")),
                "forme_final": str(get(row, "forme final") or "").strip(),
            }
        )
    return result


_SOURCE_FIELDS = (
    "prestation_id",
    "nom",
    "telephone",
    "beneficiaire",
    "prestataire",
    "genre",
    "fenetre",
    "absent_dans_consolide",
    "total_presence",
    "total_seances",
    "seuil_75",
    "nombre_seances_source",
    "forme_final",
)


def _comparison_sync(payload):
    """Synchronize the active list without deleting call or form history."""
    incoming_by_reference = {}
    duplicates = 0
    for item in payload:
        reference = item["reference_code"]
        if reference in incoming_by_reference:
            duplicates += 1
            continue
        incoming_by_reference[reference] = item

    existing_by_reference = {row.reference_code: row for row in AppelPasFormeII.objects.all()}
    now = timezone.now()
    to_create = []
    to_reactivate = []
    unchanged = 0

    for reference, item in incoming_by_reference.items():
        row = existing_by_reference.get(reference)
        if row is None:
            to_create.append(AppelPasFormeII(**item, is_active=True))
        elif row.is_active:
            unchanged += 1
        else:
            for field in _SOURCE_FIELDS:
                setattr(row, field, item[field])
            row.is_active = True
            row.updated_at = now
            to_reactivate.append(row)

    incoming_references = set(incoming_by_reference)
    deactivate_ids = [
        row.pk
        for reference, row in existing_by_reference.items()
        if row.is_active and reference not in incoming_references
    ]
    for start in range(0, len(deactivate_ids), 500):
        AppelPasFormeII.objects.filter(pk__in=deactivate_ids[start : start + 500]).update(
            is_active=False,
            updated_at=now,
        )

    if to_reactivate:
        AppelPasFormeII.objects.bulk_update(
            to_reactivate,
            [*_SOURCE_FIELDS, "is_active", "updated_at"],
            batch_size=500,
        )
    if to_create:
        AppelPasFormeII.objects.bulk_create(to_create, batch_size=500)

    return {
        "created": len(to_create),
        "reactivated": len(to_reactivate),
        "deactivated": len(deactivate_ids),
        "unchanged": unchanged,
        "duplicates": duplicates,
    }


def _thresholds():
    rows = (
        AppelPasFormeII.objects.filter(is_active=True)
        .values("prestation_id")
        .annotate(
            total=Count("id"),
            completed=Count("id", filter=Q(formulaire_rempli_at__isnull=False)),
        )
    )
    return [
        {
            **row,
            "target": pas_forme_ii_threshold_target(row["total"]),
            "pct": round(100 * row["completed"] / row["total"], 1) if row["total"] else 0,
            "reached": pas_forme_ii_threshold_reached(row["total"], row["completed"]),
        }
        for row in rows
    ]


def _prestation_threshold_state(prestation_id: str, *, lock: bool = False) -> dict:
    queryset = AppelPasFormeII.objects.filter(is_active=True, prestation_id=prestation_id)
    if lock:
        list(queryset.select_for_update().order_by("pk").values_list("pk", flat=True))
    counts = queryset.aggregate(
        total=Count("id"),
        completed=Count("id", filter=Q(formulaire_rempli_at__isnull=False)),
    )
    total = int(counts["total"] or 0)
    completed = int(counts["completed"] or 0)
    return {
        "total": total,
        "completed": completed,
        "target": pas_forme_ii_threshold_target(total),
        "reached": pas_forme_ii_threshold_reached(total, completed),
    }


def _contact_blocked_error(state: dict) -> str:
    return (
        f"Quota de {PAS_FORME_II_THRESHOLD_PERCENT} % atteint pour cette prestation "
        f"({state['completed']} formulaire(s) sur {state['total']}). "
        "Aucun nouvel appel ne peut être effectué."
    )


def _contact_blocked_response(request, state: dict, *, wants_json: bool):
    error = _contact_blocked_error(state)
    if wants_json:
        return JsonResponse(
            {"ok": False, "error": error, "threshold_reached": True, **state},
            status=409,
        )
    messages.error(request, error)
    return redirect("pas_forme_ii_index")


def _filtered_rows_queryset(request):
    rows = AppelPasFormeII.objects.filter(is_active=True)
    filters = {
        key: request.GET.get(key, "").strip()
        for key in (
            "status",
            "prestation",
            "prestataire",
            "beneficiaire",
            "genre",
            "fenetre",
            "q",
        )
    }
    if filters["status"]:
        rows = rows.filter(status=filters["status"])
    if filters["prestation"]:
        rows = rows.filter(prestation_id=filters["prestation"])
    if filters["prestataire"]:
        rows = rows.filter(prestataire__icontains=filters["prestataire"])
    if filters["beneficiaire"]:
        rows = rows.filter(beneficiaire__icontains=filters["beneficiaire"])
    if filters["genre"]:
        rows = rows.filter(genre__iexact=filters["genre"])
    if filters["fenetre"]:
        rows = rows.filter(fenetre__iexact=filters["fenetre"])
    if filters["q"]:
        rows = rows.filter(
            Q(nom__icontains=filters["q"]) | Q(telephone__icontains=filters["q"])
        )
    return rows, filters


@login_required
@transaction.atomic
def index(request):
    if request.method == "POST" and request.FILES.get("file"):
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut importer.")
            return redirect(request.path_info)
        try:
            payload = _parse(io.BytesIO(request.FILES["file"].read()))
        except Exception as exc:
            messages.error(request, f"Impossible de lire le fichier : {exc}")
            return redirect(request.path_info)

        action_name = request.POST.get("import_action", "add")
        if action_name not in {"add", "update_segments", "compare_sync"}:
            messages.error(request, "Action d'import inconnue.")
            return redirect(request.path_info)
        if not payload:
            label = "Comparaison annulée" if action_name == "compare_sync" else "Import annulé"
            messages.warning(
                request,
                f"{label} : aucune ligne exploitable dans Feuil1.",
            )
            return redirect(request.path_info)

        if action_name == "compare_sync":
            result = _comparison_sync(payload)
            messages.success(
                request,
                "Comparaison terminée : "
                f"{result['unchanged']} inchangés, "
                f"{result['reactivated']} réactivés, "
                f"{result['deactivated']} désactivés, "
                f"{result['created']} ajoutés, "
                f"{result['duplicates']} doublons du fichier ignorés.",
            )
            return redirect(request.path_info)

        existing_by_reference = {
            row.reference_code: row for row in AppelPasFormeII.objects.all()
        }
        seen_references = set()
        created = updated = reactivated = duplicates = not_found = 0
        for item in payload:
            reference = item["reference_code"]
            if reference in seen_references:
                duplicates += 1
                continue
            seen_references.add(reference)
            row = existing_by_reference.get(reference)
            if action_name == "update_segments":
                if not row or not row.is_active:
                    not_found += 1
                    continue
                if row.genre != item["genre"] or row.fenetre != item["fenetre"]:
                    row.genre = item["genre"]
                    row.fenetre = item["fenetre"]
                    row.save(update_fields=["genre", "fenetre", "updated_at"])
                    updated += 1
                continue

            if row:
                if row.is_active:
                    duplicates += 1
                    continue
                for field in _SOURCE_FIELDS:
                    setattr(row, field, item[field])
                row.is_active = True
                row.save(update_fields=[*_SOURCE_FIELDS, "is_active", "updated_at"])
                reactivated += 1
                continue
            row = AppelPasFormeII.objects.create(**item)
            existing_by_reference[reference] = row
            created += 1

        if action_name == "update_segments":
            messages.success(
                request,
                f"Mise à jour terminée : {updated} genre/fenêtre actualisés, "
                f"{not_found} lignes non trouvées, {duplicates} doublons du fichier ignorés.",
            )
        else:
            messages.success(
                request,
                f"Import terminé : {created} ajoutés, {reactivated} réactivés, "
                f"{duplicates} doublons ignorés.",
            )
        return redirect(request.path_info)

    rows, filters = _filtered_rows_queryset(request)
    rows = rows.select_related("locked_by", "modified_by")

    all_rows = AppelPasFormeII.objects.filter(is_active=True)
    filters.update(
        {
            "prestations": list(
                all_rows.values_list("prestation_id", flat=True)
                .distinct()
                .order_by("prestation_id")
            ),
            "prestataires": list(
                all_rows.values_list("prestataire", flat=True)
                .distinct()
                .order_by("prestataire")
            ),
            "beneficiaires": list(
                all_rows.values_list("beneficiaire", flat=True)
                .distinct()
                .order_by("beneficiaire")
            ),
            "genres": list(
                all_rows.exclude(genre="")
                .values_list("genre", flat=True)
                .distinct()
                .order_by("genre")
            ),
            "fenetres": list(
                all_rows.exclude(fenetre="")
                .values_list("fenetre", flat=True)
                .distinct()
                .order_by("fenetre")
            ),
        }
    )
    campaign_segments = list(
        all_rows.values("genre", "fenetre")
        .annotate(
            total=Count("id"),
            completed=Count("id", filter=Q(formulaire_rempli_at__isnull=False)),
            successful=Count(
                "id",
                filter=Q(
                    status__in=[
                        "appel_reussi",
                        "formulaire_rempli",
                        "formulaire_avec_audio",
                    ]
                ),
            ),
        )
        .order_by("fenetre", "genre")
    )
    page_obj = Paginator(rows.order_by("prestation_id", "nom"), 50).get_page(
        request.GET.get("page", 1)
    )
    threshold_map = {item["prestation_id"]: item for item in _thresholds()}
    params = request.GET.copy()
    params.pop("page", None)
    page_rows = list(page_obj.object_list)
    for row in page_rows:
        row.can_edit_form = request.user.is_superuser or row.locked_by_id == request.user.id
        row.contact_blocked = bool(
            threshold_map.get(row.prestation_id, {}).get("reached")
        )
    return render(
        request,
        "appels/pas_forme_ii.html",
        {
            "rows": page_rows,
            "page_obj": page_obj,
            "pagination_tokens": build_pagination_tokens(page_obj),
            "querystring_no_page": params.urlencode(),
            "filters": filters,
            "thresholds": list(threshold_map.values()),
            "threshold_map": threshold_map,
            "campaign_segments": campaign_segments,
            "status_choices": AppelPasFormeII.STATUS_CHOICES,
            "pas_forme_ii_threshold_percent": PAS_FORME_II_THRESHOLD_PERCENT,
        },
    )


@login_required
def export_xlsx(request):
    if not request.user.is_superuser:
        messages.error(request, "Seul un superadmin peut exporter ce tableau.")
        return redirect("pas_forme_ii_index")

    rows, _filters = _filtered_rows_queryset(request)
    rows = rows.select_related("locked_by", "modified_by").order_by("prestation_id", "nom")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Appels Pas Formes II"
    headers = [
        "Apprenant",
        "Genre",
        "Fenetre",
        "Telephone",
        "Prestation",
        "Beneficiaire",
        "Prestataire",
        "Presences",
        "Seances prevues",
        "Seances source",
        "Statut",
        "Membre structure",
        "Seances declarees",
        "Pas forme du tout",
        "Faux nom",
        "Vrai nom",
        "Formulaire rempli le",
        "Audio disponible",
        "Audio URL",
        "Appele par",
        "Modifie par",
    ]
    worksheet.append(headers)

    for row in rows:
        worksheet.append(
            [
                row.nom,
                row.genre,
                row.fenetre,
                row.telephone,
                row.prestation_id,
                row.beneficiaire,
                row.prestataire,
                row.total_presence,
                row.total_seances,
                row.nombre_seances_source,
                row.get_status_display(),
                row.membre_structure,
                row.nombre_seances_declare,
                "Oui" if row.pas_forme_du_tout else "Non",
                "Oui" if row.faux_nom else "Non",
                row.vrai_nom,
                (
                    timezone.localtime(row.formulaire_rempli_at).strftime("%Y-%m-%d %H:%M:%S")
                    if row.formulaire_rempli_at
                    else ""
                ),
                "Oui" if row.audio_file else "Non",
                row.audio_file.url if row.audio_file else "",
                row.locked_by.username if row.locked_by else "",
                row.modified_by.username if row.modified_by else "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="appels-pas-formes-ii.xlsx"'
    workbook.save(response)
    return response


@login_required
@require_POST
@transaction.atomic
def save_form(request, pk):
    row = get_object_or_404(AppelPasFormeII, pk=pk, is_active=True)
    wants_json = request.headers.get("x-requested-with") == "XMLHttpRequest"
    was_completed = row.formulaire_rempli_at is not None

    if was_completed and not (
        request.user.is_superuser or row.locked_by_id == request.user.id
    ):
        error = (
            "Seul l'agent ayant réalisé l'appel ou un administrateur peut "
            "modifier ce formulaire."
        )
        if wants_json:
            return JsonResponse({"ok": False, "error": error}, status=403)
        messages.error(request, error)
        return redirect("pas_forme_ii_index")

    if not was_completed:
        threshold_state = _prestation_threshold_state(row.prestation_id, lock=True)
        if threshold_state["reached"]:
            return _contact_blocked_response(
                request,
                threshold_state,
                wants_json=wants_json,
            )

    if request.POST.get("action") == "rappeler":
        row.status = "a_rappeler"
        row.rappel_at = request.POST.get("rappel_at") or None
        row.locked_by = request.user
        row.save(update_fields=["status", "rappel_at", "locked_by", "updated_at"])
        if wants_json:
            return JsonResponse({"ok": True, "status": row.status})
        messages.success(request, "Rappel enregistré.")
        return redirect("pas_forme_ii_index")

    membre_structure = str(request.POST.get("q2", "")).strip().upper()
    pas_forme_du_tout = request.POST.get("pas_forme_du_tout") == "on"
    nombre_seances = (
        0
        if pas_forme_du_tout
        else _number(request.POST.get("nombre_seances_declare"))
    )
    if membre_structure not in {"OUI", "NON"} or nombre_seances is None or nombre_seances < 0:
        error = "Répondez par Oui ou Non et indiquez un nombre de séances valide."
        if wants_json:
            return JsonResponse({"ok": False, "error": error}, status=400)
        messages.error(request, error)
        return redirect("pas_forme_ii_index")

    faux_nom = request.POST.get("faux_nom") == "on"
    vrai_nom = str(request.POST.get("vrai_nom", "")).strip()
    if faux_nom and not vrai_nom:
        error = "Indiquez le vrai nom lorsque le nom déclaré est faux."
        if wants_json:
            return JsonResponse({"ok": False, "error": error}, status=400)
        messages.error(request, error)
        return redirect("pas_forme_ii_index")

    row.nombre_seances_declare = nombre_seances
    row.membre_structure = membre_structure
    row.est_forme = request.POST.get("est_forme") == "on"
    row.pas_forme_du_tout = pas_forme_du_tout
    row.faux_nom = faux_nom
    row.vrai_nom = vrai_nom if faux_nom else ""
    row.commentaire = str(request.POST.get("commentaire", "")).strip()
    row.formulaire_rempli_at = row.formulaire_rempli_at or timezone.now()
    row.status = "formulaire_avec_audio" if request.FILES.get("audio") else "formulaire_rempli"
    if not row.locked_by_id:
        row.locked_by = request.user
    if was_completed:
        row.modified_by = request.user
        row.modified_at = timezone.now()
    row.locked_at = None
    row.rappel_at = None
    if request.FILES.get("audio"):
        row.audio_file = request.FILES["audio"]
    row.save()

    if wants_json:
        return JsonResponse(
            {
                "ok": True,
                "status": row.status,
                "status_label": row.get_status_display(),
                "membre_structure": row.membre_structure,
                "nombre_seances_declare": row.nombre_seances_declare,
                "pas_forme_du_tout": row.pas_forme_du_tout,
                "audio_url": row.audio_file.url if row.audio_file else "",
            }
        )
    messages.success(request, "Formulaire enregistré.")
    return redirect("pas_forme_ii_index")


def _can_edit(request, row):
    return request.user.is_superuser or (
        row.locked_by_id and row.locked_by_id == request.user.id
    )


@login_required
@require_POST
@transaction.atomic
def update_form(request, pk):
    row = get_object_or_404(AppelPasFormeII, pk=pk)
    wants_json = request.headers.get("x-requested-with") == "XMLHttpRequest"

    if not _can_edit(request, row):
        error = "Seul l'agent ayant traité cette fiche ou un administrateur peut la modifier."
        if wants_json:
            return JsonResponse({"ok": False, "error": error}, status=403)
        messages.error(request, error)
        return redirect("pas_forme_ii_index")

    telephone = _phone(request.POST.get("telephone", row.telephone))
    prestation_id = str(request.POST.get("prestation_id", "")).strip()
    membre_structure = str(request.POST.get("q2", "")).strip().upper()
    pas_forme_du_tout = request.POST.get("pas_forme_du_tout") == "on"
    nombre_seances = (
        0
        if pas_forme_du_tout
        else _number(request.POST.get("nombre_seances_declare"))
    )
    status = str(request.POST.get("status", "")).strip()
    faux_nom = request.POST.get("faux_nom") == "on"
    vrai_nom = str(request.POST.get("vrai_nom", "")).strip()
    remove_audio = request.POST.get("remove_audio") == "on"

    if not prestation_id:
        error = "La prestation est requise."
    elif membre_structure not in {"", "OUI", "NON"}:
        error = "Répondez par Oui ou Non pour l'appartenance à la structure."
    elif nombre_seances is not None and nombre_seances < 0:
        error = "Le nombre de séances déclaré est invalide."
    elif faux_nom and not vrai_nom:
        error = "Indiquez le vrai nom lorsque le nom déclaré est faux."
    elif status and status not in dict(AppelPasFormeII.STATUS_CHOICES):
        error = "Statut invalide."
    else:
        error = None

    if error:
        if wants_json:
            return JsonResponse({"ok": False, "error": error}, status=400)
        messages.error(request, error)
        return redirect("pas_forme_ii_index")

    row.telephone = telephone
    row.prestation_id = prestation_id
    row.membre_structure = membre_structure
    if nombre_seances is not None:
        row.nombre_seances_declare = nombre_seances
    row.pas_forme_du_tout = pas_forme_du_tout
    row.faux_nom = faux_nom
    row.vrai_nom = vrai_nom if faux_nom else ""
    if status:
        row.status = status
    if request.FILES.get("audio"):
        row.audio_file = request.FILES["audio"]
    elif remove_audio and row.audio_file:
        row.audio_file.delete(save=False)
        row.audio_file = None
    row.modified_by = request.user
    row.modified_at = timezone.now()
    row.save()

    if wants_json:
        return JsonResponse(
            {
                "ok": True,
                "telephone": row.telephone,
                "prestation_id": row.prestation_id,
                "membre_structure": row.membre_structure,
                "nombre_seances_declare": row.nombre_seances_declare,
                "pas_forme_du_tout": row.pas_forme_du_tout,
                "faux_nom": row.faux_nom,
                "vrai_nom": row.vrai_nom,
                "status": row.status,
                "status_label": row.get_status_display(),
                "audio_url": row.audio_file.url if row.audio_file else "",
            }
        )
    messages.success(request, "Fiche mise à jour.")
    return redirect("pas_forme_ii_index")


@login_required
@require_POST
@transaction.atomic
def action(request, pk):
    row = get_object_or_404(AppelPasFormeII, pk=pk, is_active=True)
    action_name = request.POST.get("action")
    wants_json = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if action_name not in {"start", "resume", "pause", "reussi"}:
        if wants_json:
            return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)
        return redirect("pas_forme_ii_index")
    if action_name != "pause":
        threshold_state = _prestation_threshold_state(row.prestation_id, lock=True)
        if threshold_state["reached"]:
            return _contact_blocked_response(
                request,
                threshold_state,
                wants_json=wants_json,
            )
    if action_name in {"start", "resume"}:
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = timezone.now()
    elif action_name == "pause":
        row.status = "pause"
    elif action_name == "reussi":
        row.status = "appel_reussi"
    row.save(update_fields=["status", "locked_by", "locked_at", "updated_at"])
    if wants_json:
        return JsonResponse({"ok": True, "status": row.status})
    return redirect("pas_forme_ii_index")
