from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

from django.conf import settings
from django.db.models import Count, Max

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - safe fallback if dependency missing
    load_workbook = None


FORMATEUR_NAMES_XLSX = "fichier_concatene (1) 1 (1).xlsx"
FORMATEUR_NAME_COLUMN = "Nom du formateur"
FORMATEUR_PHONE_COLUMN = "Telephone formateur1"
FORMATEUR_DATA_SHEET = "Donnees"
FORMATEUR_NAME_FALLBACK = "Non renseigné"


def _normalize_phone(value: str) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _extract_phone_numbers(value: str) -> list[str]:
    return re.findall(r"\d{8,15}", str(value or ""))


def _normalized_header(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _find_header_index(header_map: dict[str, int], target: str, fallback: int) -> int:
    normalized_target = _normalized_header(target)
    direct = header_map.get(normalized_target)
    if direct is not None:
        return direct
    for header_name, index in header_map.items():
        if normalized_target in header_name:
            return index
    return fallback


def _xlsx_path() -> Path:
    return Path(settings.BASE_DIR) / "data" / FORMATEUR_NAMES_XLSX


def _cache_token() -> tuple[str, float]:
    path = _xlsx_path()
    if not path.exists():
        return (str(path), 0.0)
    return (str(path), path.stat().st_mtime)


@lru_cache(maxsize=2)
def _load_phone_to_name(_token: tuple[str, float]) -> dict[str, str]:
    path = Path(_token[0])
    if not path.exists() or load_workbook is None:
        return {}

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        # Never break public pages if the workbook is missing/corrupted
        # (e.g. LFS pointer, partial deploy, unsupported file).
        return {}
    try:
        ws = wb[FORMATEUR_DATA_SHEET] if FORMATEUR_DATA_SHEET in wb.sheetnames else wb.active
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        header_row = next(rows, ())
        header_map: dict[str, int] = {}
        for index, value in enumerate(header_row):
            normalized = _normalized_header(value)
            if normalized:
                header_map[normalized] = index

        name_index = _find_header_index(header_map, FORMATEUR_NAME_COLUMN, 41)  # AP
        phone_index = _find_header_index(header_map, FORMATEUR_PHONE_COLUMN, 42)  # AQ

        phone_to_name: dict[str, str] = {}
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                raw_name = row[name_index] if name_index < len(row) else ""
                raw_phone = row[phone_index] if phone_index < len(row) else ""
                name = str(raw_name or "").strip()
                if not name:
                    continue
                for phone in _extract_phone_numbers(str(raw_phone or "")):
                    normalized = _normalize_phone(phone)
                    if len(normalized) < 8:
                        continue
                    phone_to_name.setdefault(normalized, name)
            return phone_to_name
        except Exception:
            return {}
    finally:
        wb.close()


def resolve_formateur_name_from_values(*values: str) -> str:
    phone_to_name = _load_phone_to_name(_cache_token())
    if not phone_to_name:
        return ""
    for raw in values:
        for phone in _extract_phone_numbers(str(raw or "")):
            normalized = _normalize_phone(phone)
            if not normalized:
                continue
            found = phone_to_name.get(normalized)
            if found:
                return found
    return ""


def _normalize_resolved_name(value: str) -> str:
    name = str(value or "").strip()
    return name or FORMATEUR_NAME_FALLBACK


def _db_formateur_cache_token() -> tuple[int, float]:
    from App_PADESCE.formations.models import Formateur

    stats = Formateur.objects.filter(actif=True).aggregate(
        max_updated=Max("updated_at"), total=Count("id")
    )
    max_updated = stats.get("max_updated")
    ts = max_updated.timestamp() if max_updated else 0.0
    return int(stats.get("total") or 0), ts


@lru_cache(maxsize=2)
def _load_db_phone_to_name(_token: tuple[int, float]) -> dict[str, str]:
    from App_PADESCE.formations.models import Formateur

    payload: dict[str, str] = {}
    for formateur in Formateur.objects.filter(actif=True).only("nom", "telephone").iterator():
        phone = _normalize_phone(str(getattr(formateur, "telephone", "") or ""))
        if not phone or phone in payload:
            continue
        payload[phone] = _normalize_resolved_name(getattr(formateur, "nom", ""))
    return payload


def resolve_formateur_db_name_from_values(*values: str) -> str:
    phone_candidates: list[str] = []
    for raw in values:
        phone_candidates.extend(_extract_phone_numbers(str(raw or "")))

    db_map = _load_db_phone_to_name(_db_formateur_cache_token())
    if phone_candidates:
        for phone in phone_candidates:
            normalized = _normalize_phone(phone)
            if not normalized:
                continue
            found = db_map.get(normalized)
            if found:
                return found

    return FORMATEUR_NAME_FALLBACK


def sync_formateur_names_from_excel(*, fallback: str = FORMATEUR_NAME_FALLBACK) -> int:
    from App_PADESCE.formations.models import Formateur

    updated = 0
    for formateur in Formateur.objects.filter(actif=True).only(
        "id", "nom", "telephone", "nom_complet"
    ):
        resolved = resolve_formateur_name_from_values(
            str(getattr(formateur, "telephone", "") or ""),
            str(getattr(formateur, "nom_complet", "") or ""),
        )
        target_name = _normalize_resolved_name(resolved if resolved else fallback)
        if str(getattr(formateur, "nom", "") or "").strip() == target_name:
            continue
        formateur.nom = target_name
        formateur.save(update_fields=["nom", "updated_at"])
        updated += 1
    return updated
