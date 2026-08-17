import csv
import datetime
import io
import subprocess
import tempfile
import unicodedata
import zipfile
from pathlib import Path

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.crypto import constant_time_compare
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET, require_POST

from App_PADESCE.appels.cga_report import (
    build_cga_calls_report_workbook,
    get_cga_calls_report_filename,
)
from App_PADESCE.appels.models import CALL_COMPLETED_STATUSES, AppelCGA, AppelCGAAudio
from App_PADESCE.appels.pagination import build_pagination_tokens
from App_PADESCE.appels.views import (
    _bind_audio_state,
    _build_progress_metrics,
    _cga_duplicate_key,
    _cleanup_stale_locks,
    _deactivate_duplicate_rows,
    _has_audio_file,
    _safe_audio_url,
)

IMPORT_BATCH_SIZE = 2000
PAGE_SIZE_DEFAULT = 100
PAGE_SIZE_MAX = 500
CGA_PUBLIC_API_PAGE_SIZE_DEFAULT = 100
CGA_PUBLIC_API_PAGE_SIZE_MAX = 500
CGA_IMPORT_FIELDS = (
    "numero",
    "raison_sociale",
    "sigle",
    "activite_principale",
    "regime",
    "cri",
    "centre_de_rattachement",
    "ville",
    "telephone",
)
CGA_APPEND_UPDATE_FIELDS = [*CGA_IMPORT_FIELDS, "is_active"]
ONECCA_SKIPPED_SHEETS = {"yaoude", "yaounde"}
ONECCA_SECTION_LABELS = {
    "Sec I - EC Libéraux": "EC liberaux",
    "Sec II - Sociétés d'EC": "Societes d'EC",
    "Sec III - EC Salariés": "EC salaries",
    "Sec IV - EC Stagiaires": "EC stagiaires",
}
ONECCA_CITY_PATTERNS = (
    ("Yaounde", ("yaounde", "yde")),
    ("Douala", ("douala", "dla")),
    ("Bamenda", ("bamenda", "bda")),
    ("Buea", ("buea",)),
    ("Limbe", ("limbe", "lbe")),
    ("Garoua", ("garoua", "g'ra")),
    ("Maroua", ("maroua",)),
    ("Bafoussam", ("bafoussam",)),
    ("Ngaoundere", ("ngaoundere", "n'dere", "ndere")),
    ("Bertoua", ("bertoua",)),
    ("Ebolowa", ("ebolowa",)),
)

_build_pagination_tokens = build_pagination_tokens


def _normalize_cga_source(value):
    source = str(value or "").strip().lower()
    valid_sources = {choice[0] for choice in AppelCGA.SOURCE_CHOICES}
    if source in valid_sources:
        return source
    # Some browsers/users can accidentally append copied page-title text to
    # the query value (for example ``source=suivi Connexion | PADESCE``).
    # Keep the intended source when the first token is valid instead of
    # silently falling back to the Entreprise tab.
    first_token = source.split(maxsplit=1)[0] if source else ""
    if first_token in valid_sources:
        return first_token
    return AppelCGA.SOURCE_ENTREPRISE


def _get_active_cga_source(request):
    if request.method == "POST":
        return _normalize_cga_source(request.POST.get("source"))
    return _normalize_cga_source(request.GET.get("source"))


def _cga_source_label(source):
    return dict(AppelCGA.SOURCE_CHOICES).get(source, "Entreprise")


def _normalize_header(value):
    if value is None:
        return ""
    text = " ".join(str(value).strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _as_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _truncate(value, max_length):
    text = _as_text(value)
    return text[:max_length]


def _join_nonempty(*values, separator=" / "):
    return separator.join(_as_text(value) for value in values if _as_text(value))


def _iter_cga_excel_rows(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError("Sheet1 introuvable dans le fichier.")
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return
    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            idx = header_map.get(_normalize_header(key))
            if idx is None or idx >= len(row):
                continue
            return row[idx]
        return None

    for row_index, row in enumerate(rows, start=2):
        numero_raw = get(row, "N°", "No", "N")
        niu = _as_text(get(row, "NIU"))
        raison = _as_text(get(row, "RAISON_SOCIALE"))
        if not raison:
            continue
        numero = None
        try:
            if numero_raw not in (None, ""):
                numero = int(float(numero_raw))
        except Exception:
            numero = None
        if not niu:
            fallback = numero if numero is not None else row_index
            niu = f"AUTO-{fallback}-{row_index}"
        yield {
            "numero": numero,
            "raison_sociale": raison,
            "sigle": _as_text(get(row, "SIGLE")),
            "niu": niu,
            "activite_principale": _as_text(get(row, "ACTIVITE_PRINCIPALE")),
            "regime": _as_text(get(row, "REGIME")),
            "cri": _as_text(get(row, "CRI")),
            "centre_de_rattachement": _as_text(get(row, "CENTRE_DE_RATTACHEMENT")),
            "ville": _as_text(get(row, "VILLE")),
            "telephone": _as_text(get(row, "TELEPHONE")),
        }


def _find_onecca_header_row(ws):
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [_normalize_header(value) for value in row]
        if any("noms" in value and "prenoms" in value for value in normalized):
            return row_index, row
    raise ValueError(f"Entete ONECCA introuvable dans l'onglet {ws.title}.")


def _infer_onecca_city(*values):
    text = _normalize_header(" ".join(_as_text(value) for value in values if _as_text(value)))
    for city, patterns in ONECCA_CITY_PATTERNS:
        if any(pattern in text for pattern in patterns):
            return city
    return ""


def _onecca_source_code(inscription, sheet_title, row_index):
    raw_code = _as_text(inscription)
    if not raw_code:
        raw_code = f"{slugify(sheet_title) or 'cabinet'}-{row_index}"
    return _truncate(f"ONECCA-{raw_code}", 64)


def _iter_onecca_excel_rows(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    yielded = 0
    for ws in wb.worksheets:
        if _normalize_header(ws.title) in ONECCA_SKIPPED_SHEETS:
            continue
        header_row_index, header = _find_onecca_header_row(ws)
        header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}
        section_label = ONECCA_SECTION_LABELS.get(ws.title, ws.title)

        def get(row, *keys):
            for key in keys:
                idx = header_map.get(_normalize_header(key))
                if idx is None or idx >= len(row):
                    continue
                return row[idx]
            return None

        for row_index, row in enumerate(
            ws.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            nom = _as_text(get(row, "Noms & Prénoms", "Noms & Prenoms"))
            if not nom:
                continue
            numero = None
            numero_raw = get(row, "N°", "No", "N")
            try:
                if numero_raw not in (None, ""):
                    numero = int(float(numero_raw))
            except Exception:
                numero = None
            inscription = _as_text(get(row, "Inscription N°", "Inscription No", "Inscription N"))
            date_inscription = _as_text(get(row, "Inscription Date"))
            adresse_postale = _as_text(get(row, "Adresse Postale"))
            ligne_1 = _as_text(get(row, "Ligne 1"))
            ligne_2 = _as_text(get(row, "Ligne 2"))
            email = _as_text(get(row, "Adresse E-mail", "Adresse Email"))
            site_cabinet = _as_text(get(row, "Site du Cabinet"))
            yield {
                "numero": numero,
                "raison_sociale": _truncate(nom, 255),
                "sigle": _truncate(inscription, 255),
                "niu": _onecca_source_code(inscription, ws.title, row_index),
                "activite_principale": _truncate(email, 255),
                "regime": _truncate(section_label, 120),
                "cri": _truncate(date_inscription, 120),
                "centre_de_rattachement": _truncate(
                    _join_nonempty(adresse_postale, site_cabinet, separator=" - "),
                    255,
                ),
                "ville": _truncate(
                    _infer_onecca_city(adresse_postale, site_cabinet),
                    120,
                ),
                "telephone": _truncate(_join_nonempty(ligne_1, ligne_2), 120),
            }
            yielded += 1
    if yielded == 0:
        raise ValueError("Aucune ligne Cabinet lisible dans le fichier ONECCA.")


def _iter_cga_source_rows(file_obj, source):
    if source == AppelCGA.SOURCE_CABINET:
        return _iter_onecca_excel_rows(file_obj)
    return _iter_cga_excel_rows(file_obj)


def _parse_bool_flag(value):
    if value is None:
        return None
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _make_cga_audio_playable(file_obj):
    """Convert browser WebM recordings to MP3 so every browser can play them."""
    filename = getattr(file_obj, "name", "recording.webm")
    if Path(filename).suffix.lower() != ".webm":
        return file_obj
    try:
        with tempfile.TemporaryDirectory(prefix="cga-audio-") as directory:
            source_path = Path(directory) / "recording.webm"
            target_path = Path(directory) / "recording.mp3"
            with source_path.open("wb") as target:
                for chunk in file_obj.chunks():
                    target.write(chunk)
            subprocess.run(
                ["ffmpeg", "-y", "-v", "error", "-i", str(source_path), "-vn", "-q:a", "4", str(target_path)],
                check=True,
                timeout=120,
            )
            return ContentFile(target_path.read_bytes(), name=f"{Path(filename).stem}.mp3")
    except (OSError, subprocess.SubprocessError):
        return file_obj


def _expected_cga_public_api_key() -> str:
    return str(
        getattr(settings, "CGA_PUBLIC_API_KEY", "") or getattr(settings, "EXPORT_API_KEY", "") or ""
    ).strip()


def _provided_cga_public_api_key(request) -> str:
    bearer = str(request.headers.get("Authorization", "") or "").strip()
    if bearer.lower().startswith("bearer "):
        token = bearer[7:].strip()
        if token:
            return token
    for header_name in ("X-CGA-Api-Key", "X-Export-Api-Key"):
        token = str(request.headers.get(header_name, "") or "").strip()
        if token:
            return token
    return ""


def _cga_public_api_auth_error(request):
    expected = _expected_cga_public_api_key()
    if not expected:
        return JsonResponse({"error": "CGA_PUBLIC_API_KEY non configuree."}, status=503)
    provided = _provided_cga_public_api_key(request)
    if not provided or not constant_time_compare(provided, expected):
        return JsonResponse({"error": "Cle API manquante ou invalide."}, status=403)
    return None


def _parse_cga_public_page_size(raw_value):
    if raw_value in (None, ""):
        return CGA_PUBLIC_API_PAGE_SIZE_DEFAULT
    try:
        page_size = int(raw_value)
    except (TypeError, ValueError):
        raise ValueError("page_size doit etre un entier positif.")
    if page_size < 1:
        raise ValueError("page_size doit etre superieur ou egal a 1.")
    return min(page_size, CGA_PUBLIC_API_PAGE_SIZE_MAX)


def _serialize_cga_public_row(row):
    return {
        "id": row.pk,
        "source": row.source,
        "numero": row.numero,
        "niu": row.niu,
        "raison_sociale": row.raison_sociale,
        "sigle": row.sigle,
        "activite_principale": row.activite_principale,
        "regime": row.regime,
        "cri": row.cri,
        "centre_de_rattachement": row.centre_de_rattachement,
        "ville": row.ville,
        "telephone": row.telephone,
        "status": row.status,
        "status_label": row.get_status_display(),
        "interet": row.interet,
        "mauvais_numero": row.mauvais_numero,
        "indisponible": row.indisponible,
        "resultat_summary": row.resultat_summary,
        "rappel_at": row.rappel_at.isoformat() if row.rappel_at else "",
        "created_at": row.created_at.isoformat() if getattr(row, "created_at", None) else "",
        "updated_at": row.updated_at.isoformat() if getattr(row, "updated_at", None) else "",
    }


def _build_filtered_cga_queryset(request):
    active_source = _get_active_cga_source(request)
    qs = (
        AppelCGA.objects.filter(is_active=True, source=active_source)
        .select_related("locked_by")
        .prefetch_related("audio_history")
    )
    campaign_month = (request.GET.get("campaign_month") or "").strip()
    if active_source == AppelCGA.SOURCE_SUIVI and campaign_month:
        try:
            qs = qs.filter(campaign_month=datetime.date.fromisoformat(campaign_month))
        except ValueError:
            campaign_month = ""
    status_filter = (request.GET.get("status") or "").strip()
    resultat_filter = (request.GET.get("resultat") or "").strip()
    regime_filter = (request.GET.get("regime") or "").strip()
    cri_filter = (request.GET.get("cri") or "").strip()
    centre_filter = (request.GET.get("centre") or "").strip()
    ville_filter = (request.GET.get("ville") or "").strip()
    agent_filter = (request.GET.get("agent") or "").strip()
    date_from_str = (request.GET.get("date_from") or "").strip()
    date_to_str = (request.GET.get("date_to") or "").strip()
    search = (request.GET.get("q") or "").strip()

    if status_filter:
        if status_filter == "completed":
            qs = qs.filter(status__in=CALL_COMPLETED_STATUSES)
        else:
            qs = qs.filter(status=status_filter)
    if resultat_filter == "interesse":
        qs = qs.filter(interet="OUI")
    elif resultat_filter == "pas_interesse":
        qs = qs.filter(interet="NON")
    elif resultat_filter == "faux_numero":
        qs = qs.filter(mauvais_numero="OUI")
    elif resultat_filter == "indisponible":
        qs = qs.filter(indisponible="OUI")
    if regime_filter:
        qs = qs.filter(regime__iexact=regime_filter)
    if cri_filter:
        qs = qs.filter(cri__iexact=cri_filter)
    if centre_filter:
        qs = qs.filter(centre_de_rattachement__iexact=centre_filter)
    if ville_filter:
        qs = qs.filter(ville__iexact=ville_filter)
    if agent_filter:
        qs = qs.filter(locked_by__username__iexact=agent_filter)
    if search:
        qs = qs.filter(
            Q(raison_sociale__icontains=search)
            | Q(sigle__icontains=search)
            | Q(niu__icontains=search)
            | Q(telephone__icontains=search)
        )
    if date_from_str:
        try:
            qs = qs.filter(created_at__gte=datetime.datetime.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(created_at__lte=datetime.datetime.fromisoformat(date_to_str))
        except ValueError:
            pass

    filters = {
        "source": active_source,
        "campaign_month": campaign_month,
        "status": status_filter,
        "resultat": resultat_filter,
        "regime": regime_filter,
        "cri": cri_filter,
        "centre": centre_filter,
        "ville": ville_filter,
        "agent": agent_filter,
        "q": search,
        "date_from": date_from_str,
        "date_to": date_to_str,
        "regimes": sorted(
            {v.strip() for v in qs.exclude(regime="").values_list("regime", flat=True) if v}
        ),
        "cris": sorted({v.strip() for v in qs.exclude(cri="").values_list("cri", flat=True) if v}),
        "centres": sorted(
            {
                v.strip()
                for v in qs.exclude(centre_de_rattachement="").values_list(
                    "centre_de_rattachement", flat=True
                )
                if v
            }
        ),
        "villes": sorted(
            {v.strip() for v in qs.exclude(ville="").values_list("ville", flat=True) if v}
        ),
        "agents": sorted(
            {
                u.strip()
                for u in qs.exclude(locked_by__isnull=True).values_list(
                    "locked_by__username", flat=True
                )
                if u
            }
        ),
    }
    return qs, filters


def _flush_cga_batch(batch, *, ignore_conflicts=False):
    if not batch:
        return 0
    AppelCGA.objects.bulk_create(
        batch, batch_size=IMPORT_BATCH_SIZE, ignore_conflicts=ignore_conflicts
    )
    n = len(batch)
    batch.clear()
    return n


def _apply_cga_import_values(row, item):
    for field in CGA_IMPORT_FIELDS:
        setattr(row, field, item[field])


def _sync_cga_append_batch(
    batch_items, *, source=AppelCGA.SOURCE_ENTREPRISE, campaign_month=None
):
    if not batch_items:
        return 0, 0

    # NIU is repeated by the monthly follow-up snapshots, so it is no longer a
    # database-wide unique field.  Limit the lookup to the imported source.
    existing_by_niu = {
        row.niu: row
        for row in AppelCGA.objects.filter(
            source=source,
            campaign_month=campaign_month,
            niu__in=[item["niu"].strip() for item in batch_items],
        )
    }
    rows_to_create = []
    rows_to_update = []

    for item in batch_items:
        niu = item["niu"].strip()
        row = existing_by_niu.get(niu)
        if row is None:
            rows_to_create.append(
                AppelCGA(
                    **item, source=source, campaign_month=campaign_month, is_active=True
                )
            )
            continue
        _apply_cga_import_values(row, item)
        row.is_active = True
        rows_to_update.append(row)

    if rows_to_create:
        AppelCGA.objects.bulk_create(rows_to_create, batch_size=IMPORT_BATCH_SIZE)
    if rows_to_update:
        AppelCGA.objects.bulk_update(
            rows_to_update,
            CGA_APPEND_UPDATE_FIELDS,
            batch_size=IMPORT_BATCH_SIZE,
        )
    return len(rows_to_create), len(rows_to_update)


def _parse_campaign_month(value):
    """Return the first day of the requested month, or raise ValueError."""
    try:
        raw_value = str(value or "").strip()
        parsed = datetime.datetime.strptime(raw_value, "%Y-%m").date()
    except ValueError:
        raise ValueError("Choisissez un mois valide pour la campagne.")
    return parsed.replace(day=1)


def _create_monthly_cga_campaign(month):
    """Create an immutable calling list from the active Entreprise source."""
    source_rows = AppelCGA.objects.filter(
        source=AppelCGA.SOURCE_ENTREPRISE, is_active=True
    ).order_by("id")
    if not source_rows.exists():
        return 0

    fields = (
        "numero", "raison_sociale", "sigle", "niu", "activite_principale", "regime",
        "cri", "centre_de_rattachement", "ville", "telephone",
    )
    batch = []
    created = 0
    with transaction.atomic():
        if AppelCGA.objects.filter(source=AppelCGA.SOURCE_SUIVI, campaign_month=month).exists():
            raise ValueError("Cette campagne mensuelle existe deja et ses donnees sont conservees.")
        for row in source_rows.iterator(chunk_size=IMPORT_BATCH_SIZE):
            values = {field: getattr(row, field) for field in fields}
            batch.append(
                AppelCGA(
                    **values,
                    source=AppelCGA.SOURCE_SUIVI,
                    campaign_month=month,
                    is_active=True,
                )
            )
            if len(batch) >= IMPORT_BATCH_SIZE:
                AppelCGA.objects.bulk_create(batch, batch_size=IMPORT_BATCH_SIZE)
                created += len(batch)
                batch.clear()
        if batch:
            AppelCGA.objects.bulk_create(batch, batch_size=IMPORT_BATCH_SIZE)
            created += len(batch)
    return created


@login_required
def cga_export_xlsx(request):
    source = _get_active_cga_source(request)
    campaign_month = None
    if source == AppelCGA.SOURCE_SUIVI and request.GET.get("campaign_month"):
        try:
            campaign_month = datetime.date.fromisoformat(request.GET["campaign_month"])
        except ValueError:
            messages.error(request, "Mois de campagne invalide pour l'export.")
            return redirect(f"{request.path_info}?source={source}")
    payload = build_cga_calls_report_workbook(source=source, campaign_month=campaign_month)
    response = HttpResponse(
        payload, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{get_cga_calls_report_filename(source=source)}"'
    )
    return response


@never_cache
@require_GET
def cga_public_interested_api(request):
    auth_error = _cga_public_api_auth_error(request)
    if auth_error is not None:
        return auth_error

    qs, _filters = _build_filtered_cga_queryset(request)
    qs = qs.filter(interet="OUI").order_by("-updated_at", "-id")

    updated_since_raw = (request.GET.get("updated_since") or "").strip()
    updated_since = None
    if updated_since_raw:
        updated_since = parse_datetime(updated_since_raw)
        if updated_since is None:
            return JsonResponse(
                {"error": "updated_since invalide. Utilisez un datetime ISO 8601."},
                status=400,
            )
        if timezone.is_naive(updated_since):
            updated_since = timezone.make_aware(updated_since, timezone.get_current_timezone())
        qs = qs.filter(updated_at__gte=updated_since)

    try:
        page_size = _parse_cga_public_page_size(request.GET.get("page_size"))
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)

    paginator = Paginator(qs, page_size)
    page_obj = paginator.get_page(request.GET.get("page"))

    payload = {
        "source": "cga_interested_calls",
        "exported_at": timezone.now().isoformat(),
        "total": paginator.count,
        "page": page_obj.number,
        "page_size": page_size,
        "pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
        "next_page": page_obj.next_page_number() if page_obj.has_next() else None,
        "previous_page": page_obj.previous_page_number() if page_obj.has_previous() else None,
        "updated_since": updated_since.isoformat() if updated_since else "",
        "appels": [_serialize_cga_public_row(row) for row in page_obj.object_list],
    }
    response = JsonResponse(payload)
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Vary"] = "Authorization, X-CGA-Api-Key, X-Export-Api-Key"
    return response


@login_required
def cga_export_filtered_csv(request):
    qs, _ = _build_filtered_cga_queryset(request)
    rows = qs.order_by("status", "raison_sociale")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="cga-filtres.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(
        [
            "Numero",
            "Raison sociale",
            "Sigle",
            "NIU",
            "Activite principale",
            "Regime",
            "CRI",
            "Centre de rattachement",
            "Ville",
            "Telephone",
            "Statut",
            "Agent",
            "Rappel at",
            "Audio URL",
            "Interet",
            "Mauvais numero",
            "Indisponible",
            "Created at",
            "Updated at",
            "Source",
        ]
    )
    for row in rows.iterator(chunk_size=2000):
        writer.writerow(
            [
                row.numero or "",
                row.raison_sociale,
                row.sigle,
                row.niu,
                row.activite_principale,
                row.regime,
                row.cri,
                row.centre_de_rattachement,
                row.ville,
                row.telephone,
                row.get_status_display(),
                row.locked_by.username if row.locked_by else "",
                row.rappel_at.isoformat() if row.rappel_at else "",
                _safe_audio_url(row),
                row.interet,
                row.mauvais_numero,
                row.indisponible,
                row.created_at.isoformat() if getattr(row, "created_at", None) else "",
                row.updated_at.isoformat() if getattr(row, "updated_at", None) else "",
                row.get_source_display(),
            ]
        )
    return response


@never_cache
@login_required
def cga_index(request):
    if request.method == "POST" and request.POST.get("action") == "create_monthly_campaign":
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut creer une campagne mensuelle CGA.")
            return redirect(f"{request.path_info}?source={AppelCGA.SOURCE_SUIVI}")
        try:
            month = _parse_campaign_month(request.POST.get("campaign_month"))
            created = _create_monthly_cga_campaign(month)
            messages.success(
                request,
                f"Campagne Suivi CGA {month:%m/%Y} creee : {created} appel(s) a traiter.",
            )
            return redirect(f"{request.path_info}?source=suivi&campaign_month={month.isoformat()}")
        except (ValueError, IntegrityError) as exc:
            messages.error(request, str(exc))
            return redirect(f"{request.path_info}?source={AppelCGA.SOURCE_SUIVI}")

    if request.method == "POST" and request.FILES.get("file"):
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut importer des fichiers CGA.")
            return redirect(request.path_info)
        source = _get_active_cga_source(request)
        source_label = _cga_source_label(source)
        mode = request.POST.get("update_mode", "replace")
        campaign_month = None
        if source == AppelCGA.SOURCE_SUIVI:
            try:
                campaign_month = _parse_campaign_month(request.POST.get("campaign_month"))
            except ValueError as exc:
                messages.error(request, str(exc))
                return redirect(f"{request.path_info}?source={source}")
        file_obj = io.BytesIO(request.FILES["file"].read())
        try:
            if mode == "replace":
                replacement_qs = AppelCGA.objects.filter(source=source)
                if campaign_month:
                    replacement_qs = replacement_qs.filter(campaign_month=campaign_month)
                replacement_qs.delete()
                batch = []
                seen = set()
                created = 0
                skipped = 0
                for item in _iter_cga_source_rows(file_obj, source):
                    niu = item["niu"].strip()
                    if niu in seen:
                        skipped += 1
                        continue
                    seen.add(niu)
                    batch.append(
                        AppelCGA(
                            **item,
                            source=source,
                            campaign_month=campaign_month,
                            is_active=True,
                        )
                    )
                    if len(batch) >= IMPORT_BATCH_SIZE:
                        created += _flush_cga_batch(batch)
                created += _flush_cga_batch(batch)
                deduped = _deactivate_duplicate_rows(
                    AppelCGA.objects.filter(
                        is_active=True, source=source, campaign_month=campaign_month
                    ),
                    _cga_duplicate_key,
                )
                messages.success(
                    request,
                    (
                        f"Import CGA {source_label} termine. {created} ligne(s) chargee(s), "
                        f"{skipped} doublon(s) ignores, {deduped} doublon(s) desactive(s)."
                    ),
                )
            elif mode == "append":
                seen_file = set()
                created = 0
                skipped = 0
                updated = 0
                batch = []
                for item in _iter_cga_source_rows(file_obj, source):
                    niu = item["niu"].strip()
                    if niu in seen_file:
                        skipped += 1
                        continue
                    seen_file.add(niu)
                    batch.append(item)
                    if len(batch) >= IMPORT_BATCH_SIZE:
                        batch_created, batch_updated = _sync_cga_append_batch(
                            list(batch), source=source, campaign_month=campaign_month
                        )
                        created += batch_created
                        updated += batch_updated
                        batch.clear()
                if batch:
                    batch_created, batch_updated = _sync_cga_append_batch(
                        list(batch), source=source, campaign_month=campaign_month
                    )
                    created += batch_created
                    updated += batch_updated
                deduped = _deactivate_duplicate_rows(
                    AppelCGA.objects.filter(
                        is_active=True, source=source, campaign_month=campaign_month
                    ),
                    _cga_duplicate_key,
                )
                messages.success(
                    request,
                    (
                        f"Import CGA {source_label} termine. {created} nouvelle(s) ligne(s), "
                        f"{updated} mise(s) a jour, {skipped} doublon(s) dans le fichier, "
                        f"{deduped} doublon(s) desactive(s)."
                    ),
                )
            else:
                messages.error(request, "Mode d'import CGA inconnu.")
        except Exception as exc:
            messages.error(request, f"Impossible de lire le fichier CGA : {exc}")
        redirect_url = f"{request.path_info}?source={source}"
        if campaign_month:
            redirect_url += f"&campaign_month={campaign_month.isoformat()}"
        return redirect(redirect_url)

    qs, filters = _build_filtered_cga_queryset(request)
    active_source = filters["source"]
    active_source_label = _cga_source_label(active_source)
    total_count = qs.count()
    stats = _build_progress_metrics(qs)

    try:
        page_size = int(request.GET.get("page_size") or PAGE_SIZE_DEFAULT)
    except ValueError:
        page_size = PAGE_SIZE_DEFAULT
    page_size = max(25, min(page_size, PAGE_SIZE_MAX))

    paginator = Paginator(qs.order_by("status", "raison_sociale"), page_size)
    page_obj = paginator.get_page(request.GET.get("page"))
    rows = _bind_audio_state(list(page_obj.object_list))
    page_obj.object_list = rows
    pagination_tokens = build_pagination_tokens(page_obj)
    params = request.GET.copy()
    params["source"] = active_source
    params.pop("page", None)
    querystring_no_page = params.urlencode()
    source_counts = {
        row["source"]: row["total"]
        for row in AppelCGA.objects.filter(is_active=True)
        .values("source")
        .annotate(total=Count("id"))
    }
    source_tabs = []
    for source_value, source_label in AppelCGA.SOURCE_CHOICES:
        tab_params = request.GET.copy()
        tab_params["source"] = source_value
        tab_params.pop("page", None)
        source_tabs.append(
            {
                "value": source_value,
                "label": source_label,
                "active": source_value == active_source,
                "count": source_counts.get(source_value, 0),
                "url": f"?{tab_params.urlencode()}",
            }
        )
    reset_query = "" if active_source == AppelCGA.SOURCE_ENTREPRISE else f"?source={active_source}"
    export_query = f"?source={active_source}"
    campaign_months = []
    if active_source == AppelCGA.SOURCE_SUIVI:
        campaign_months = list(
            AppelCGA.objects.filter(source=AppelCGA.SOURCE_SUIVI, campaign_month__isnull=False)
            .values_list("campaign_month", flat=True)
            .distinct()
            .order_by("-campaign_month")
        )
        if filters["campaign_month"]:
            export_query += f"&campaign_month={filters['campaign_month']}"
    import_hint = (
        "Importer le Tableau ONECCA 2023 pour charger les cabinets par section."
        if active_source == AppelCGA.SOURCE_CABINET
        else (
            "Importer le fichier CGA (Feuil1) pour charger les colonnes N, raison sociale, "
            "sigle, NIU, activite, regime, CRI, centre, ville et telephone."
        )
    )

    response = render(
        request,
        "appels/cga.html",
        {
            "rows": rows,
            "rows_count": total_count,
            "filters": filters,
            "page_obj": page_obj,
            "page_size": page_size,
            "pagination_tokens": pagination_tokens,
            "querystring_no_page": querystring_no_page,
            "stats": stats,
            "active_source": active_source,
            "active_source_label": active_source_label,
            "source_tabs": source_tabs,
            "reset_query": reset_query,
            "export_query": export_query,
            "import_hint": import_hint,
            "campaign_months": campaign_months,
        },
    )
    response["Cache-Control"] = "private, no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    response["X-PADESCE-CGA-UI-Version"] = "argumentaire-v2"
    return response


@login_required
@require_POST
def cga_action(request, pk: int):
    # Clean up stale locks that may be blocking access
    _cleanup_stale_locks()

    row = get_object_or_404(AppelCGA, pk=pk)
    action = request.POST.get("action")
    rappel_at = request.POST.get("rappel_at")
    _parse_bool_flag(request.POST.get("deja_forme"))

    now = timezone.now()

    # Nouveaux champs resultats
    interet_val = request.POST.get("interet", "")
    mauvais_val = request.POST.get("mauvais_numero", "NON")
    indisponible_val = request.POST.get("indisponible", "NON")

    if action == "start":
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = now
    elif action == "pause":
        row.status = "pause"
    elif action == "resume":
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = now
    elif action == "rappeler":
        row.status = "a_rappeler"
        row.locked_by = request.user
        row.locked_at = now
        # The follow-up call modal can save its outcome while scheduling a
        # visit/reminder.  Keep the legacy reminder flow unchanged when no
        # outcome fields were submitted.
        if "interet" in request.POST:
            row.interet = interet_val
        if "mauvais_numero" in request.POST:
            row.mauvais_numero = mauvais_val
        if "indisponible" in request.POST:
            row.indisponible = indisponible_val
        if rappel_at:
            try:
                parsed_rappel_at = datetime.datetime.fromisoformat(rappel_at)
                if timezone.is_naive(parsed_rappel_at):
                    parsed_rappel_at = timezone.make_aware(
                        parsed_rappel_at, timezone.get_current_timezone()
                    )
                row.rappel_at = parsed_rappel_at
            except ValueError:
                row.rappel_at = None
    elif action == "terminer":
        row.interet = interet_val
        row.mauvais_numero = mauvais_val
        row.indisponible = indisponible_val
        if row.indisponible == "OUI":
            row.status = "a_rappeler"
        else:
            row.status = "termine"
    else:
        return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)

    row.save(
        update_fields=[
            "status",
            "locked_by",
            "locked_at",
            "rappel_at",
            "updated_at",
            "interet",
            "mauvais_numero",
            "indisponible",
        ]
    )
    return JsonResponse(
        {
            "ok": True,
            "status": row.status,
            "status_label": row.get_status_display(),
            "locked_by": row.locked_by.username if row.locked_by else "",
            "rappel_at": row.rappel_at.isoformat() if row.rappel_at else "",
            "resultat_summary": row.resultat_summary,
        }
    )


@login_required
@require_POST
def cga_upload_audio(request, pk: int):
    row = get_object_or_404(AppelCGA, pk=pk)
    file_obj = request.FILES.get("audio")
    if not file_obj:
        return JsonResponse({"ok": False, "error": "Aucun fichier audio."}, status=400)
    recording = AppelCGAAudio.objects.create(
        appel=row,
        audio_file=_make_cga_audio_playable(file_obj),
        uploaded_by=request.user,
    )
    # Keep the latest recording in the existing player while the complete list
    # remains available through audio_history. Never delete an earlier file.
    row.audio_file.name = recording.audio_file.name
    row.save(update_fields=["audio_file", "updated_at"])
    return JsonResponse(
        {
            "ok": True,
            "audio_saved": True,
            "audio_url": _safe_audio_url(row),
            "audio_count": row.audio_history.count(),
        }
    )


@login_required
@require_POST
def download_cga_audios(request):
    ids = request.POST.getlist("ids")
    if not ids:
        return JsonResponse({"ok": False, "error": "Aucun appel selectionne."}, status=400)
    try:
        ids = [int(val) for val in ids]
    except ValueError:
        return JsonResponse({"ok": False, "error": "Identifiants invalides."}, status=400)

    rows = list(
        AppelCGA.objects.filter(pk__in=ids, audio_file__isnull=False).order_by("raison_sociale")
    )
    if not rows:
        return JsonResponse(
            {"ok": False, "error": "Aucun audio disponible pour la selection."}, status=404
        )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        written = 0
        for row in rows:
            if not _has_audio_file(row):
                continue
            try:
                with row.audio_file.open("rb") as audio:
                    suffix = Path(row.audio_file.name).suffix or ".mp3"
                    safe_name = (
                        f"{slugify(row.niu) or 'niu'}-"
                        f"{slugify(row.raison_sociale) or 'cga'}{suffix}"
                    )
                    archive.writestr(safe_name, audio.read())
                    written += 1
            except Exception:
                continue
        if written == 0:
            return JsonResponse({"ok": False, "error": "Pas d'audio recuperable."}, status=404)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="cga-audios.zip"'
    return response


@login_required
@require_POST
def cga_transcription_detail(request, pk: int):
    row = get_object_or_404(AppelCGA, pk=pk)
    try:
        obj, generated = _ensure_cga_transcription(row)
        return JsonResponse(
            {"ok": True, "generated": generated, "transcription": _transcription_to_payload(obj)}
        )
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


@login_required
def cga_transcription_download(request, pk: int):
    row = get_object_or_404(AppelCGA, pk=pk)
    try:
        obj, _generated = _ensure_cga_transcription(row)
    except Exception as exc:
        return HttpResponse(str(exc), status=400, content_type="text/plain; charset=utf-8")
    response = HttpResponse(obj.transcription_text or "", content_type="text/plain; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="transcription-cga-{slugify(row.niu) or row.pk}.txt"'
    )
    return response


@login_required
@require_POST
def start_filtered_cga_transcription(request):
    qs, _ = _build_filtered_cga_queryset(request)
    jobs = _collect_jobs_from_cga(qs)
    ok, payload, status_code = _start_transcription_task(
        CGA_FILTERED_TRANSCRIPTION_TASK_KEY,
        jobs,
        "Transcription du tableau filtre CGA",
    )
    return JsonResponse(payload, status=status_code)


@login_required
def filtered_cga_transcription_status(request):
    return JsonResponse(
        {"ok": True, "status": _transcription_status_response(CGA_FILTERED_TRANSCRIPTION_TASK_KEY)}
    )


@login_required
@require_POST
def stop_filtered_cga_transcription(request):
    ok, payload, status_code = _stop_transcription_task(CGA_FILTERED_TRANSCRIPTION_TASK_KEY)
    return JsonResponse(payload, status=status_code)
