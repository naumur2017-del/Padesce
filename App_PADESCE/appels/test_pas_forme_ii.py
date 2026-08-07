import io
import tempfile

import openpyxl
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from App_PADESCE.appels.models import AppelPasFormeII


def _import_file(rows, headers=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Feuil1"
    sheet.append(
        headers
        or [
            "PRESTATION ID",
            "APPRENANTS",
            "NUMERO",
            "BENEFICIAIRES",
            "PRESTATAIRES",
            "GENRE",
            "FENETRE",
            "ABSENT DANS CONSOLIDE",
            "TOTAL PRESENCE",
            "TOTAL SEANCE",
            "SEUIL 75%",
            "NOMBRE SEANCE",
            "FORME FINAL",
        ]
    )
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        "pas-formes-ii.xlsx",
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@override_settings(
    MEDIA_ROOT=tempfile.mkdtemp(prefix="padesce-pas-forme-ii-import-tests-"),
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    },
)
class AppelPasFormeIIImportTests(TestCase):
    def setUp(self):
        self.admin = get_user_model().objects.create_superuser(
            username="pas-forme-ii-admin",
            password="test123",
            email="admin@example.com",
        )
        self.client.force_login(self.admin)

    @staticmethod
    def _row(
        prestation="PRESTA-001",
        nom="Apprenant Test",
        telephone="690000001",
        genre="F",
        fenetre="2",
        absent="NON",
    ):
        return [
            prestation,
            nom,
            telephone,
            "Bénéficiaire",
            "Prestataire",
            genre,
            fenetre,
            absent,
            4,
            6,
            5,
            4,
            "OUI",
        ]

    def test_export_xlsx_returns_filtered_detailed_table(self):
        AppelPasFormeII.objects.create(
            reference_code="PRESTA-EXPORT-1-690000011-apprenante-export",
            prestation_id="PRESTA-EXPORT-1",
            nom="Apprenante Export",
            telephone="690000011",
            beneficiaire="Beneficiaire Export",
            prestataire="Prestataire Export",
            genre="F",
            fenetre="3",
            total_presence=12,
            total_seances=12,
            nombre_seances_source=8,
            membre_structure="OUI",
            nombre_seances_declare=8,
            status="formulaire_rempli",
            formulaire_rempli_at=timezone.now(),
        )
        AppelPasFormeII.objects.create(
            reference_code="PRESTA-EXPORT-2-690000012-apprenant-ignore",
            prestation_id="PRESTA-EXPORT-2",
            nom="Apprenant Ignore",
            telephone="690000012",
            genre="M",
        )

        response = self.client.get(
            reverse("pas_forme_ii_export_xlsx"),
            {"genre": "F"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(len(rows), 2)
        exported = dict(zip(rows[0], rows[1]))
        self.assertEqual(exported["Apprenant"], "Apprenante Export")
        self.assertEqual(exported["Genre"], "F")
        self.assertEqual(exported["Prestation"], "PRESTA-EXPORT-1")
        self.assertEqual(exported["Statut"], "Formulaire Rempli")

    def test_add_import_ignores_file_duplicates_and_parses_non_as_false(self):
        row = self._row()
        response = self.client.post(
            reverse("pas_forme_ii_index"),
            {"import_action": "add", "file": _import_file([row, row])},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(AppelPasFormeII.objects.count(), 1)
        imported = AppelPasFormeII.objects.get()
        self.assertFalse(imported.absent_dans_consolide)
        self.assertEqual(imported.genre, "F")
        self.assertEqual(imported.fenetre, "2")
        self.assertContains(response, "1 ajoutés")
        self.assertContains(response, "1 doublons ignorés")

    def test_add_import_reactivates_without_losing_call_history(self):
        existing = AppelPasFormeII.objects.create(
            reference_code="PRESTA-001-690000001-apprenant-test",
            prestation_id="PRESTA-001",
            nom="Apprenant Test",
            telephone="690000001",
            is_active=False,
            commentaire="Historique conservé",
            formulaire_rempli_at=timezone.now(),
        )

        self.client.post(
            reverse("pas_forme_ii_index"),
            {"import_action": "add", "file": _import_file([self._row(genre="M", fenetre="3")])},
        )

        existing.refresh_from_db()
        self.assertTrue(existing.is_active)
        self.assertEqual(existing.genre, "M")
        self.assertEqual(existing.fenetre, "3")
        self.assertEqual(existing.commentaire, "Historique conservé")
        self.assertIsNotNone(existing.formulaire_rempli_at)

    def test_update_segments_changes_only_genre_and_window(self):
        existing = AppelPasFormeII.objects.create(
            reference_code="PRESTA-001-690000001-apprenant-test",
            prestation_id="PRESTA-001",
            nom="Apprenant Test",
            telephone="690000001",
            genre="Ancien genre",
            fenetre="Ancienne fenêtre",
            commentaire="Ne pas modifier",
            status="formulaire_rempli",
            formulaire_rempli_at=timezone.now(),
        )

        response = self.client.post(
            reverse("pas_forme_ii_index"),
            {
                "import_action": "update_segments",
                "file": _import_file([self._row(genre="F", fenetre="3")]),
            },
            follow=True,
        )

        existing.refresh_from_db()
        self.assertEqual(existing.genre, "F")
        self.assertEqual(existing.fenetre, "3")
        self.assertEqual(existing.commentaire, "Ne pas modifier")
        self.assertEqual(existing.status, "formulaire_rempli")
        self.assertIsNotNone(existing.formulaire_rempli_at)
        self.assertContains(response, "1 genre/fenêtre actualisés")

    def test_comparison_preserves_matches_deactivates_missing_and_adds_new(self):
        matching = AppelPasFormeII.objects.create(
            reference_code="PRESTA-001-690000001-apprenant-test",
            prestation_id="PRESTA-001",
            nom="Apprenant Test",
            telephone="690000001",
            genre="Genre plateforme",
            commentaire="Historique à préserver",
            status="formulaire_rempli",
        )
        matching_updated_at = matching.updated_at
        missing = AppelPasFormeII.objects.create(
            reference_code="PRESTA-002-690000002-apprenant-absent",
            prestation_id="PRESTA-002",
            nom="Apprenant Absent",
            telephone="690000002",
        )

        response = self.client.post(
            reverse("pas_forme_ii_index"),
            {
                "import_action": "compare_sync",
                "file": _import_file(
                    [
                        self._row(genre="Genre fichier"),
                        self._row(
                            prestation="PRESTA-003",
                            nom="Nouvel Apprenant",
                            telephone="690000003",
                        ),
                    ]
                ),
            },
            follow=True,
        )

        matching.refresh_from_db()
        missing.refresh_from_db()
        self.assertTrue(matching.is_active)
        self.assertEqual(matching.genre, "Genre plateforme")
        self.assertEqual(matching.commentaire, "Historique à préserver")
        self.assertEqual(matching.updated_at, matching_updated_at)
        self.assertFalse(missing.is_active)
        self.assertTrue(
            AppelPasFormeII.objects.get(
                reference_code="PRESTA-003-690000003-nouvel-apprenant"
            ).is_active
        )
        self.assertContains(response, "1 inchangés")
        self.assertContains(response, "1 désactivés")
        self.assertContains(response, "1 ajoutés")

    def test_empty_or_invalid_file_does_not_change_existing_rows(self):
        existing = AppelPasFormeII.objects.create(
            reference_code="PRESTA-PROTECTED",
            prestation_id="PRESTA-PROTECTED",
            nom="Apprenant Protégé",
        )

        empty = self.client.post(
            reverse("pas_forme_ii_index"),
            {"import_action": "compare_sync", "file": _import_file([])},
            follow=True,
        )
        invalid = self.client.post(
            reverse("pas_forme_ii_index"),
            {
                "import_action": "add",
                "file": _import_file([], headers=["COLONNE INCONNUE"]),
            },
            follow=True,
        )

        existing.refresh_from_db()
        self.assertTrue(existing.is_active)
        self.assertContains(empty, "Comparaison annulée")
        self.assertContains(invalid, "colonne(s) obligatoire(s) absente(s)")

    def test_page_uses_ten_percent_threshold_with_ceiling(self):
        calls = []
        for prestation_id, total in (("PRESTA-TEN", 10), ("PRESTA-ELEVEN", 11)):
            for index in range(total):
                calls.append(
                    AppelPasFormeII(
                        reference_code=f"{prestation_id}-{index}",
                        prestation_id=prestation_id,
                        nom=f"Apprenant {prestation_id} {index}",
                        formulaire_rempli_at=timezone.now() if index == 0 else None,
                    )
                )
        AppelPasFormeII.objects.bulk_create(calls)

        response = self.client.get(reverse("pas_forme_ii_index"))
        thresholds = {row["prestation_id"]: row for row in response.context["thresholds"]}

        self.assertEqual(response.context["pas_forme_ii_threshold_percent"], 10)
        self.assertEqual(thresholds["PRESTA-TEN"]["target"], 1)
        self.assertEqual(thresholds["PRESTA-ELEVEN"]["target"], 2)
        self.assertTrue(thresholds["PRESTA-TEN"]["reached"])
        self.assertFalse(thresholds["PRESTA-ELEVEN"]["reached"])
        self.assertContains(response, "Seuil de formulaires remplis par prestation — 10 %")


@override_settings(
    MEDIA_ROOT=tempfile.mkdtemp(prefix="padesce-pas-forme-ii-tests-"),
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    },
)
class AppelPasFormeIISaveTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="pas-forme-ii-agent",
            password="test123",
        )
        self.client.force_login(self.user)
        self.row = AppelPasFormeII.objects.create(
            reference_code="PFII-001",
            prestation_id="PRESTA-001",
            nom="Apprenant Test",
            telephone="690000001",
            beneficiaire="Structure Test",
            prestataire="Prestataire Test",
            total_presence=4,
            total_seances=6,
            nombre_seances_source=4,
        )

    def _reach_prestation_threshold(self):
        AppelPasFormeII.objects.bulk_create(
            [
                AppelPasFormeII(
                    reference_code=f"PFII-QUOTA-{index}",
                    prestation_id=self.row.prestation_id,
                    nom=f"Apprenant quota {index}",
                    telephone=f"6900001{index:02d}",
                    formulaire_rempli_at=timezone.now() if index == 0 else None,
                )
                for index in range(9)
            ]
        )

    def test_completed_form_is_saved_counted_and_exposes_audio(self):
        response = self.client.post(
            reverse("pas_forme_ii_save_form", args=[self.row.pk]),
            {
                "action": "terminer",
                "q2": "OUI",
                "nombre_seances_declare": "5",
                "audio": SimpleUploadedFile(
                    "appel.webm",
                    b"test-audio-content",
                    content_type="audio/webm",
                ),
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        self.row.refresh_from_db()
        self.assertEqual(self.row.membre_structure, "OUI")
        self.assertEqual(self.row.nombre_seances_declare, 5)
        self.assertEqual(self.row.status, "formulaire_avec_audio")
        self.assertIsNotNone(self.row.formulaire_rempli_at)
        self.assertTrue(self.row.audio_file.name)

        page = self.client.get(reverse("pas_forme_ii_index"))
        self.assertEqual(page.context["thresholds"][0]["completed"], 1)
        self.assertNotContains(page, 'class="btn open"')
        self.assertContains(page, "Modifier")
        self.assertContains(page, 'data-q2="OUI"')
        self.assertContains(page, 'data-declared="5"')
        self.assertContains(page, "<audio", html=False)
        self.assertContains(page, "fetch(form.dataset.submitUrl")
        self.assertNotContains(page, "fetch(form.action")

    def test_not_trained_at_all_checkbox_is_saved_and_shown(self):
        response = self.client.post(
            reverse("pas_forme_ii_save_form", args=[self.row.pk]),
            {
                "action": "terminer",
                "q2": "NON",
                "pas_forme_du_tout": "on",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["pas_forme_du_tout"])
        self.row.refresh_from_db()
        self.assertTrue(self.row.pas_forme_du_tout)
        self.assertEqual(self.row.nombre_seances_declare, 0)

        page = self.client.get(reverse("pas_forme_ii_index"))
        self.assertContains(page, "N’a pas été formé du tout")
        self.assertContains(page, 'data-pas-forme-du-tout="true"')
        self.assertContains(page, 'id="edit-pas-forme-du-tout"')

    def test_not_trained_at_all_checkbox_can_be_corrected(self):
        self.row.locked_by = self.user
        self.row.pas_forme_du_tout = True
        self.row.nombre_seances_declare = 0
        self.row.save(
            update_fields=[
                "locked_by",
                "pas_forme_du_tout",
                "nombre_seances_declare",
                "updated_at",
            ]
        )
        url = reverse("pas_forme_ii_update", args=[self.row.pk])

        response = self.client.post(
            url,
            {
                "telephone": self.row.telephone,
                "prestation_id": self.row.prestation_id,
                "q2": "OUI",
                "nombre_seances_declare": "4",
                "status": "formulaire_rempli",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.json()["pas_forme_du_tout"])
        self.row.refresh_from_db()
        self.assertFalse(self.row.pas_forme_du_tout)
        self.assertEqual(self.row.nombre_seances_declare, 4)

    def test_invalid_form_is_rejected_without_being_counted(self):
        response = self.client.post(
            reverse("pas_forme_ii_save_form", args=[self.row.pk]),
            {"action": "terminer", "q2": "", "nombre_seances_declare": ""},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["ok"])
        self.row.refresh_from_db()
        self.assertIsNone(self.row.formulaire_rempli_at)
        self.assertEqual(self.row.status, "en_attente")

    def test_false_name_requires_and_saves_the_real_name(self):
        url = reverse("pas_forme_ii_save_form", args=[self.row.pk])
        incomplete = self.client.post(
            url,
            {"q2": "OUI", "nombre_seances_declare": "3", "faux_nom": "on"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(incomplete.status_code, 400)

        response = self.client.post(
            url,
            {
                "q2": "OUI",
                "nombre_seances_declare": "3",
                "faux_nom": "on",
                "vrai_nom": "Vrai Apprenant",
            },
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        self.row.refresh_from_db()
        self.assertTrue(self.row.faux_nom)
        self.assertEqual(self.row.vrai_nom, "Vrai Apprenant")

    def test_start_action_is_persisted_for_ajax(self):
        response = self.client.post(
            reverse("pas_forme_ii_action", args=[self.row.pk]),
            {"action": "start"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["ok"])
        self.row.refresh_from_db()
        self.assertEqual(self.row.status, "en_cours")
        self.assertEqual(self.row.locked_by, self.user)

    def test_ten_percent_quota_blocks_new_contacts_on_page_and_backend(self):
        self._reach_prestation_threshold()

        page = self.client.get(reverse("pas_forme_ii_index"))
        self.assertContains(page, "Quota 10 % atteint — contact bloqué")

        action_response = self.client.post(
            reverse("pas_forme_ii_action", args=[self.row.pk]),
            {"action": "start"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        form_response = self.client.post(
            reverse("pas_forme_ii_save_form", args=[self.row.pk]),
            {"q2": "OUI", "nombre_seances_declare": "4"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(action_response.status_code, 409)
        self.assertTrue(action_response.json()["threshold_reached"])
        self.assertEqual(form_response.status_code, 409)
        self.assertTrue(form_response.json()["threshold_reached"])
        self.row.refresh_from_db()
        self.assertEqual(self.row.status, "en_attente")
        self.assertIsNone(self.row.formulaire_rempli_at)

    def test_pause_remains_available_after_quota_is_reached(self):
        self._reach_prestation_threshold()
        self.row.status = "en_cours"
        self.row.locked_by = self.user
        self.row.save(update_fields=["status", "locked_by", "updated_at"])

        response = self.client.post(
            reverse("pas_forme_ii_action", args=[self.row.pk]),
            {"action": "pause"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.row.refresh_from_db()
        self.assertEqual(self.row.status, "pause")

    def test_completed_form_can_only_be_modified_by_call_agent_or_admin(self):
        self.row.locked_by = self.user
        self.row.formulaire_rempli_at = timezone.now()
        self.row.status = "formulaire_rempli"
        self.row.save()
        other_user = get_user_model().objects.create_user(
            username="other-agent", password="test123"
        )
        url = reverse("pas_forme_ii_save_form", args=[self.row.pk])

        self.client.force_login(other_user)
        denied = self.client.post(
            url,
            {"q2": "OUI", "nombre_seances_declare": "2"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(denied.status_code, 403)

        self.client.force_login(self.user)
        updated = self.client.post(
            url,
            {"q2": "NON", "nombre_seances_declare": "2"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(updated.status_code, 200)
        self.row.refresh_from_db()
        self.assertEqual(self.row.modified_by, self.user)
        self.assertIsNotNone(self.row.modified_at)
        self.assertEqual(self.row.locked_by, self.user)

        admin = get_user_model().objects.create_superuser(
            username="pfii-admin", password="test123", email="admin@example.com"
        )
        self.client.force_login(admin)
        updated_by_admin = self.client.post(
            url,
            {"q2": "OUI", "nombre_seances_declare": "4"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(updated_by_admin.status_code, 200)
        self.row.refresh_from_db()
        self.assertEqual(self.row.modified_by, admin)
        self.assertEqual(self.row.locked_by, self.user)

    def test_call_with_reminder_status_can_be_started(self):
        self.row.status = "a_rappeler"
        self.row.save(update_fields=["status", "updated_at"])

        response = self.client.post(
            reverse("pas_forme_ii_action", args=[self.row.pk]),
            {"action": "start"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.row.refresh_from_db()
        self.assertEqual(self.row.status, "en_cours")
