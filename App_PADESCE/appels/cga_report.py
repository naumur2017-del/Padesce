from __future__ import annotations

import io
from collections import Counter
from datetime import datetime

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from App_PADESCE.appels.models import AppelCGA

CGA_EXPORT_HEADERS = (
    "N°",
    "RAISON_SOCIALE",
    "SIGLE",
    "NIU",
    "ACTIVITE_PRINCIPALE",
    "REGIME",
    "CRI",
    "CENTRE_DE_RATTACHEMENT",
    "VILLE",
    "TELEPHONE",
    "STATUT",
    "RAPPEL_AT",
    "LOCKED_BY",
    "LOCKED_AT",
    "AUDIO_FILE",
    "INTERET",
    "MAUVAIS_NUMERO",
    "INDISPONIBLE",
    "CREATED_AT",
    "UPDATED_AT",
    "SOURCE",
    "CAMPAGNE_MOIS",
)


def _safe_audio_url(row: AppelCGA) -> str:
    audio = getattr(row, "audio_file", None)
    if not audio or not getattr(audio, "name", ""):
        return ""
    try:
        return str(audio.url)
    except Exception:
        return str(audio.name)


def _row_to_values(row: AppelCGA) -> list[str | int]:
    return [
        row.numero or "",
        row.raison_sociale,
        row.sigle,
        row.niu,
        row.activite_principale,
        row.regime,
        row.cri,
        row.centre_de_rattachement,
        row.ville,
        row.telephone,
        row.status,
        row.rappel_at.isoformat() if row.rappel_at else "",
        row.locked_by.username if row.locked_by else "",
        row.locked_at.isoformat() if row.locked_at else "",
        _safe_audio_url(row),
        row.interet,
        row.mauvais_numero,
        row.indisponible,
        row.created_at.isoformat() if getattr(row, "created_at", None) else "",
        row.updated_at.isoformat() if getattr(row, "updated_at", None) else "",
        row.get_source_display(),
        row.campaign_month.isoformat() if row.campaign_month else "",
    ]


def build_cga_calls_report_workbook(
    source: str | None = None, campaign_month=None
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Synthese"
    details_sheet = workbook.create_sheet("Appels CGA")

    qs = AppelCGA.objects.select_related("locked_by")
    if source:
        qs = qs.filter(source=source)
    if campaign_month:
        qs = qs.filter(campaign_month=campaign_month)
    rows = list(qs.order_by("raison_sociale", "niu"))
    status_counter = Counter(row.status for row in rows if row.status)
    interest_counter = Counter(row.interet for row in rows if row.interet)
    audio_count = sum(1 for row in rows if getattr(getattr(row, "audio_file", None), "name", ""))

    summary_sheet["A1"] = "Rapport appels CGA"
    summary_sheet["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    summary_sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary_sheet["A3"] = "Genere le"
    summary_sheet["B3"] = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    summary_sheet["A4"] = "Source"
    summary_sheet["B4"] = (
        dict(AppelCGA.SOURCE_CHOICES).get(source, "Toutes") if source else "Toutes"
    )
    summary_sheet["A5"] = "Total appels"
    summary_sheet["B5"] = len(rows)
    summary_sheet["A6"] = "Appels avec audio"
    summary_sheet["B6"] = audio_count
    if campaign_month:
        summary_sheet["A7"] = "Campagne"
        summary_sheet["B7"] = campaign_month.strftime("%m/%Y")

    summary_sheet["A8"] = "Statut"
    summary_sheet["B8"] = "Nombre"
    summary_sheet["D8"] = "Interet"
    summary_sheet["E8"] = "Nombre"
    for cell in ("A8", "B8", "D8", "E8"):
        summary_sheet[cell].font = Font(bold=True)

    status_row = 9
    for status, count in sorted(status_counter.items()):
        summary_sheet[f"A{status_row}"] = status
        summary_sheet[f"B{status_row}"] = count
        status_row += 1

    interest_row = 9
    for interest, count in sorted(interest_counter.items()):
        summary_sheet[f"D{interest_row}"] = interest
        summary_sheet[f"E{interest_row}"] = count
        interest_row += 1

    details_sheet.append(list(CGA_EXPORT_HEADERS))
    for cell in details_sheet[1]:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = PatternFill("solid", fgColor="D9E2F3")

    for row in rows:
        details_sheet.append(_row_to_values(row))

    for sheet in (summary_sheet, details_sheet):
        for column in sheet.columns:
            values = [len(str(cell.value or "")) for cell in column]
            sheet.column_dimensions[column[0].column_letter].width = min(max(values) + 2, 40)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def get_cga_calls_report_filename(
    report_date: datetime | None = None,
    source: str | None = None,
) -> str:
    current = report_date or timezone.localtime()
    suffix = f"_{source}" if source else ""
    return f"rapport_appels_cga{suffix}_{current:%Y%m%d}.xlsx"
