import csv
import datetime
import io
import re
import unicodedata

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from App_PADESCE.appels.models import (
    CALL_FORM_STATUSES,
    CALL_SUCCESS_STATUSES,
    AppelPasForme,
    infer_padesce_status,
)
from App_PADESCE.appels.pagination import build_pagination_tokens
from App_PADESCE.appels.views import _bind_audio_state, _cleanup_stale_locks, _safe_audio_url
from App_PADESCE.core.phase_scope import PHASE_SCOPE_V1_COMBINED, normalize_phase_scope
from App_PADESCE.formations.models import Beneficiaire, Prestataire, Prestation


def _normalize_header(value):
    if value is None:
        return ""
    text = " ".join(str(value).strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _as_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_phone(value):
    digits = re.sub(r"\D", "", _as_text(value))
    if digits.startswith("00237"):
        digits = digits[5:]
    elif digits.startswith("237") and len(digits) > 9:
        digits = digits[3:]
    return digits[-9:] if len(digits) > 9 else digits


def _reference_for_item(item: dict) -> str:
    phone = item.get("telephone") or "sans-tel"
    prestation = item.get("prestation_id") or "sans-prestation"
    nom_key = re.sub(r"[^a-z0-9]+", "-", _normalize_header(item.get("nom"))).strip("-")
    return f"{prestation}-{phone}-{nom_key or 'apprenant'}"[:120]


def _parse_pas_forme_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            idx = header_map.get(_normalize_header(key))
            if idx is not None and idx < len(row):
                return row[idx] or ""
        return ""

    payload = []
    for row in rows:
        nom = get(row, "APPRENANT ABSENT CHEZ NAUMUR", "APPRENANT", "NOM")
        if not nom:
            continue
        item = {
            "numero": get(row, "NUMERO"),
            "nom": _as_text(nom),
            "telephone": _normalize_phone(get(row, "NUMERO DE TELEPHONE", "TELEPHONE")),
            "est_forme": _as_text(get(row, "EST FORME")),
            "prestation_id": _as_text(get(row, "PRESTATION ID", "PRESTATION")),
            "prestataire": _as_text(get(row, "NOM DU PRESTATAIRE", "PRESTATAIRE")),
            "beneficiaire": _as_text(get(row, "NOM DU BENEFICIAIRE", "BENEFICIAIRE")),
        }
        if not item["telephone"]:
            continue
        try:
            item["numero"] = int(str(item["numero"]).strip()) if item["numero"] != "" else None
        except (TypeError, ValueError):
            item["numero"] = None
        item["reference_code"] = _reference_for_item(item)
        payload.append(item)
    return payload


def _has_form_data(row: AppelPasForme) -> bool:
    return bool(
        row.connait_structure
        or row.beneficiaire_corrige
        or row.connait_prestataire
        or row.prestataire_corrige
        or row.membre_structure
        or row.a_assiste_formation
        or row.connait_theme
        or row.nombre_seances is not None
        or row.faux_numero
        or row.bon_numero
        or row.faux_nom
        or row.vrai_nom
        or row.satisfaction_completed_at
    )


def _has_audio(row: AppelPasForme) -> bool:
    return bool(getattr(row.audio_file, "name", ""))


def _sync_pas_forme_status(row: AppelPasForme, *, save=True) -> str:
    new_status = infer_padesce_status(
        row.status,
        has_form=_has_form_data(row),
        has_audio=_has_audio(row),
    )
    if save and new_status != row.status:
        row.status = new_status
        row.save(update_fields=["status", "updated_at"])
    else:
        row.status = new_status
    return new_status


def _build_stats(queryset):
    stats = queryset.aggregate(
        total=Count("id"),
        appels_tentes=Count("id", filter=~Q(status="en_attente")),
        appels_reussis=Count("id", filter=Q(status__in=CALL_SUCCESS_STATUSES)),
        formulaires_remplis=Count("id", filter=Q(status__in=CALL_FORM_STATUSES)),
        audios=Count("id", filter=Q(audio_file__isnull=False) & ~Q(audio_file="")),
    )
    total = int(stats.get("total") or 0)
    completed = int(stats.get("formulaires_remplis") or 0)
    stats["completion_rate"] = round((completed / total) * 100, 1) if total else 0
    stats["completion_label"] = f"{completed} / {total}" if total else "0 / 0"
    return stats


def _filtered_queryset(request):
    qs = AppelPasForme.objects.filter(is_active=True)
    status = request.GET.get("status", "")
    prestataire = request.GET.get("prestataire", "")
    beneficiaire = request.GET.get("beneficiaire", "")
    prestation = request.GET.get("prestation", "")
    agent = request.GET.get("agent", "")
    q = request.GET.get("q", "").strip()
    if status:
        qs = qs.filter(status=status)
    if prestataire:
        qs = qs.filter(prestataire__icontains=prestataire)
    if beneficiaire:
        qs = qs.filter(beneficiaire__icontains=beneficiaire)
    if prestation:
        qs = qs.filter(prestation_id__icontains=prestation)
    if agent:
        qs = qs.filter(locked_by__username__iexact=agent)
    if q:
        qs = qs.filter(
            Q(nom__icontains=q)
            | Q(telephone__icontains=q)
            | Q(prestation_id__icontains=q)
            | Q(prestataire__icontains=q)
            | Q(beneficiaire__icontains=q)
        )
    dropdowns = qs.values("prestataire", "beneficiaire", "prestation_id").distinct()
    filters = {
        "phase_scope": normalize_phase_scope(
            request.GET.get("phase_scope"), default=PHASE_SCOPE_V1_COMBINED
        ),
        "status": status,
        "prestataire": prestataire,
        "beneficiaire": beneficiaire,
        "prestation": prestation,
        "agent": agent,
        "q": q,
        "prestataires": sorted({r["prestataire"].strip() for r in dropdowns if r["prestataire"]}),
        "beneficiaires": sorted(
            {r["beneficiaire"].strip() for r in dropdowns if r["beneficiaire"]}
        ),
        "prestations": sorted(
            {r["prestation_id"].strip() for r in dropdowns if r["prestation_id"]}
        ),
    }
    return qs, filters


@login_required
@transaction.atomic
def pas_forme_index(request):
    if request.method == "POST" and request.FILES.get("file"):
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut importer des fichiers d'appels.")
            return redirect(request.path_info)
        mode = request.POST.get("update_mode", "replace")
        try:
            payload = _parse_pas_forme_excel(io.BytesIO(request.FILES["file"].read()))
        except Exception as exc:
            messages.error(request, f"Impossible de lire le fichier : {exc}")
            return redirect(request.path_info)
        if mode == "replace":
            AppelPasForme.objects.update(is_active=False)
        created = 0
        updated = 0
        skipped_duplicates = 0
        seen_references = set()
        for item in payload:
            reference = item["reference_code"]
            if reference in seen_references:
                skipped_duplicates += 1
                continue
            seen_references.add(reference)
            row = AppelPasForme.objects.filter(reference_code=reference).first()
            if row:
                if mode != "replace" and row.is_active:
                    skipped_duplicates += 1
                    continue
                for key, value in item.items():
                    setattr(row, key, value)
                row.is_active = True
                row.save()
                updated += 1
            else:
                AppelPasForme.objects.create(**item, is_active=True)
                created += 1
        messages.success(
            request,
            (
                f"Fichier importe. {created} nouveau(x) appel(s), "
                f"{updated} appel(s) mis a jour, {skipped_duplicates} doublon(s) ignore(s). "
                "Les lignes sans numero de telephone ne sont pas importees."
            ),
        )
        return redirect(request.path_info)

    qs, filters = _filtered_queryset(request)
    stats = _build_stats(qs)
    qs = qs.select_related("locked_by").order_by("status", "nom")
    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    rows = _bind_audio_state(list(page_obj.object_list))
    page_obj.object_list = rows
    params = request.GET.copy()
    params.pop("page", None)
    return render(
        request,
        "appels/pas_forme.html",
        {
            "rows": rows,
            "page_obj": page_obj,
            "paginator": paginator,
            "pagination_tokens": build_pagination_tokens(page_obj),
            "querystring_no_page": params.urlencode(),
            "filters": filters,
            "stats": stats,
            "appels_count": qs.count(),
            "beneficiaire_options": Beneficiaire.objects.filter(actif=True).order_by(
                "nom_structure"
            ),
            "prestataire_options": Prestataire.objects.filter(actif=True).order_by(
                "raison_sociale"
            ),
            "prestation_options": Prestation.objects.filter(actif=True)
            .select_related("prestataire", "beneficiaire")
            .order_by("code"),
        },
    )


@login_required
def pas_forme_export_filtered_csv(request):
    qs, _filters = _filtered_queryset(request)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="appels-pas-forme.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(
        [
            "Nom",
            "Telephone",
            "Prestation",
            "Prestataire",
            "Beneficiaire",
            "Statut",
            "Connait structure",
            "Beneficiaire corrige",
            "Connait prestataire",
            "Prestataire corrige",
            "Membre structure",
            "A assiste formation PADESCE",
            "Connait theme",
            "Nombre seances",
            "Faux numero",
            "Bon numero",
            "Faux nom",
            "Vrai nom",
            "Audio URL",
        ]
    )
    for row in qs.order_by("nom"):
        writer.writerow(
            [
                row.nom,
                row.telephone,
                row.prestation_id,
                row.prestataire,
                row.beneficiaire,
                row.get_status_display(),
                row.connait_structure,
                row.beneficiaire_corrige,
                row.connait_prestataire,
                row.prestataire_corrige,
                row.membre_structure,
                row.a_assiste_formation,
                row.connait_theme,
                row.nombre_seances if row.nombre_seances is not None else "",
                row.faux_numero,
                row.bon_numero,
                row.faux_nom,
                row.vrai_nom,
                _safe_audio_url(row),
            ]
        )
    return response


@login_required
@require_POST
@transaction.atomic
def pas_forme_manual_add(request):
    if not request.user.is_superuser:
        messages.error(request, "Seul un superadmin peut ajouter une ligne manuellement.")
        return redirect("pas_forme_index")

    nom = _as_text(request.POST.get("nom"))
    telephone = _normalize_phone(request.POST.get("telephone"))
    prestation_id = str(request.POST.get("prestation_id") or "").strip()
    if not nom or not telephone or not prestation_id:
        messages.error(request, "Renseignez le nom, le telephone et la prestation.")
        return redirect("pas_forme_index")

    prestation = (
        Prestation.objects.filter(pk=prestation_id, actif=True)
        .select_related("prestataire", "beneficiaire")
        .first()
    )
    if not prestation:
        messages.error(request, "Prestation introuvable ou inactive.")
        return redirect("pas_forme_index")

    item = {
        "numero": None,
        "nom": nom,
        "telephone": telephone,
        "est_forme": "",
        "prestation_id": prestation.code,
        "prestataire": prestation.prestataire.raison_sociale if prestation.prestataire else "",
        "beneficiaire": prestation.beneficiaire.nom_structure if prestation.beneficiaire else "",
    }
    item["reference_code"] = _reference_for_item(item)
    row = AppelPasForme.objects.filter(reference_code=item["reference_code"]).first()
    if row:
        for key, value in item.items():
            setattr(row, key, value)
        row.is_active = True
        row.save()
        messages.success(request, "Ligne existante mise a jour et reactivee.")
    else:
        AppelPasForme.objects.create(**item, is_active=True)
        messages.success(request, "Ligne ajoutee dans Appels Pas Forme.")
    return redirect("pas_forme_index")


@login_required
@require_POST
def pas_forme_action(request, pk: int):
    _cleanup_stale_locks()
    row = get_object_or_404(AppelPasForme, pk=pk)
    action = request.POST.get("action", "")
    now = timezone.now()
    if action in {"start", "resume"}:
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = now
    elif action == "pause":
        row.status = "pause"
    elif action == "rappeler":
        row.status = "a_rappeler"
        row.locked_by = request.user
        row.locked_at = now
        rappel_at = request.POST.get("rappel_at")
        if rappel_at:
            try:
                row.rappel_at = timezone.make_aware(datetime.datetime.fromisoformat(rappel_at))
            except (TypeError, ValueError):
                row.rappel_at = None
    else:
        return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)
    row.save(update_fields=["status", "locked_by", "locked_at", "rappel_at", "updated_at"])
    return JsonResponse(
        {
            "ok": True,
            "status": row.status,
            "status_label": row.get_status_display(),
            "locked_by": row.locked_by.username if row.locked_by else "",
            "rappel_at": row.rappel_at.isoformat() if row.rappel_at else "",
        }
    )


def _clean_yes_no(value):
    value = str(value or "").strip().upper()
    return value if value in {"OUI", "NON"} else ""


@login_required
@require_POST
def pas_forme_finalize(request, pk: int):
    row = get_object_or_404(AppelPasForme, pk=pk)
    action = request.POST.get("action", "terminer")
    if action == "rappeler":
        row.status = "a_rappeler"
        rappel_at = request.POST.get("rappel_at")
        if rappel_at:
            try:
                row.rappel_at = timezone.make_aware(datetime.datetime.fromisoformat(rappel_at))
            except (TypeError, ValueError):
                row.rappel_at = None
        row.save(update_fields=["status", "rappel_at", "updated_at"])
        return JsonResponse(
            {"ok": True, "status": row.status, "status_label": row.get_status_display()}
        )

    row.connait_structure = _clean_yes_no(request.POST.get("q1"))
    row.beneficiaire_corrige = (
        str(request.POST.get("beneficiaire_corrige") or "").strip()
        if row.connait_structure == "NON"
        else ""
    )
    row.connait_prestataire = _clean_yes_no(request.POST.get("q_prestataire"))
    row.prestataire_corrige = (
        str(request.POST.get("prestataire_corrige") or "").strip()
        if row.connait_prestataire == "NON"
        else ""
    )
    row.membre_structure = _clean_yes_no(request.POST.get("q2"))
    row.a_assiste_formation = _clean_yes_no(request.POST.get("q3"))
    row.connait_theme = _clean_yes_no(request.POST.get("q4"))
    try:
        row.nombre_seances = int(request.POST.get("q5") or 0)
    except (TypeError, ValueError):
        row.nombre_seances = None
    row.faux_numero = _clean_yes_no(request.POST.get("faux_numero"))
    row.bon_numero = _normalize_phone(request.POST.get("bon_numero"))
    row.faux_nom = _clean_yes_no(request.POST.get("faux_nom"))
    row.vrai_nom = str(request.POST.get("vrai_nom") or "").strip()
    row.satisfaction_completed_at = timezone.now()
    file_obj = request.FILES.get("audio")
    if file_obj:
        row.audio_file = file_obj
    row.status = "formulaire_avec_audio" if file_obj else "formulaire_rempli"
    row.rappel_at = None
    row.save()
    _sync_pas_forme_status(row)
    return JsonResponse(
        {
            "ok": True,
            "status": row.status,
            "status_label": row.get_status_display(),
            "audio_saved": bool(file_obj),
            "audio_url": _safe_audio_url(row),
        }
    )
