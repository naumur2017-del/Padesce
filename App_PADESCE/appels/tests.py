import os
import shutil
import tempfile
import io
import datetime
from unittest.mock import patch

import openpyxl
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.paginator import Paginator
from django.test import RequestFactory, SimpleTestCase, TestCase
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone

from App_PADESCE.appels.cga_views import _build_pagination_tokens
from App_PADESCE.appels.consolidation_views import (
    _build_filter_options,
    _create_appels_from_candidates,
    _filtered_candidates,
    consolidation_pending_appels,
)
from App_PADESCE.appels.models import (
    Appel,
    AppelAnswers,
    AppelCGA,
    AppelImportArchive,
    AppelPrestataireDemarrage,
)
from App_PADESCE.satisfaction_apprenants.models import SatisfactionApprenant


class ConsolidationFilterTests(SimpleTestCase):
    def test_filter_options_only_expose_related_values(self):
        candidates = [
            {
                "code": "AB0E",
                "nom": "Fanta Adamou",
                "classe_label": "CLA001",
                "prestataire": "Prestataire A",
                "beneficiaire": "Beneficiaire A",
                "formation_padesce": "Formation A",
                "cohorte": "1",
                "fenetre": "2",
                "statut_prestation": "TERMINE",
            },
            {
                "code": "AB0F",
                "nom": "Hadjaratou Abbou",
                "classe_label": "CLA001",
                "prestataire": "Prestataire A",
                "beneficiaire": "Beneficiaire A",
                "formation_padesce": "Formation A",
                "cohorte": "1",
                "fenetre": "2",
                "statut_prestation": "TERMINE",
            },
            {
                "code": "XY10",
                "nom": "Autre Nom",
                "classe_label": "CLA002",
                "prestataire": "Prestataire B",
                "beneficiaire": "Beneficiaire B",
                "formation_padesce": "Formation B",
                "cohorte": "3",
                "fenetre": "3",
                "statut_prestation": "ARRETE",
            },
        ]
        filters = {
            "q": "",
            "classe": "CLA001",
            "prestataire": "",
            "beneficiaire": "",
            "formation": "",
            "cohorte": "",
            "fenetre": "",
            "statut_prestation": "TERMINE",
        }

        filtered = _filtered_candidates(candidates, filters)
        options = _build_filter_options(candidates, filters)

        self.assertEqual(len(filtered), 2)
        self.assertEqual([item["value"] for item in options["prestataire"]], ["Prestataire A"])
        self.assertEqual([item["value"] for item in options["beneficiaire"]], ["Beneficiaire A"])
        self.assertEqual([item["value"] for item in options["cohorte"]], ["1"])


class CgaPaginationTests(SimpleTestCase):
    def test_build_pagination_tokens_shows_first_pages_gap_and_last_page(self):
        page_obj = Paginator(range(100), 10).get_page(1)

        self.assertEqual(_build_pagination_tokens(page_obj), [1, 2, 3, None, 10])

    def test_build_pagination_tokens_keeps_neighbors_for_middle_page(self):
        page_obj = Paginator(range(100), 10).get_page(5)

        self.assertEqual(_build_pagination_tokens(page_obj), [1, 2, 3, 4, 5, 6, None, 10])


class ConsolidationImportTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.user = get_user_model().objects.create_user(
            username="admin",
            password="test123",
            is_staff=True,
            is_superuser=True,
        )

    def test_create_appels_from_candidates_skips_existing_codes(self):
        Appel.objects.create(code="AB0E", nom="Deja present")

        created, skipped = _create_appels_from_candidates(
            [
                {
                    "code": "AB0E",
                    "nom": "Fanta Adamou",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "lieu": "Mindif",
                    "classe_label": "CLA001",
                    "fenetre": "3",
                    "telephone1": "694666763",
                    "telephone2": "",
                    "type_formation_declaree": "Transformation",
                    "formation_padesce": "Transformation des produits laitiers",
                },
                {
                    "code": "NEW1",
                    "nom": "Nouvel Apprenant",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "lieu": "Mindif",
                    "classe_label": "CLA001",
                    "fenetre": "3",
                    "telephone1": "699999999",
                    "telephone2": "",
                    "type_formation_declaree": "Transformation",
                    "formation_padesce": "Transformation des produits laitiers",
                },
            ]
        )

        self.assertEqual(created, 1)
        self.assertEqual(skipped, 1)
        self.assertTrue(Appel.objects.filter(code="NEW1").exists())
        self.assertEqual(AppelImportArchive.objects.filter(source_code="NEW1").count(), 1)

    @patch("App_PADESCE.appels.consolidation_views.messages.success")
    @patch("App_PADESCE.appels.consolidation_views.build_consolidation_call_candidates")
    def test_consolidation_pending_appels_batch_action_uses_current_filter(
        self, mock_build_candidates, mock_messages_success
    ):
        mock_build_candidates.return_value = {
            "source": {"name": "fichier consolide.xlsm", "modified_label": "24/03/2026 a 12:00"},
            "sheet_name": "Consolidation",
            "count": 2,
            "source_rows": 2,
            "duplicate_codes": [],
            "records": [
                {
                    "code": "CLA001A",
                    "nom": "Alpha",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "formation_padesce": "Formation A",
                    "type_formation_declaree": "Formation A",
                    "classe_label": "CLA001",
                    "fenetre": "2",
                    "cohorte": "1",
                    "telephone1": "690000001",
                    "telephone2": "",
                    "lieu": "Lieu A",
                    "statut_prestation": "TERMINE",
                },
                {
                    "code": "CLA002A",
                    "nom": "Beta",
                    "prestataire": "Prestataire B",
                    "beneficiaire": "Beneficiaire B",
                    "formation_padesce": "Formation B",
                    "type_formation_declaree": "Formation B",
                    "classe_label": "CLA002",
                    "fenetre": "3",
                    "cohorte": "2",
                    "telephone1": "690000002",
                    "telephone2": "",
                    "lieu": "Lieu B",
                    "statut_prestation": "TERMINE",
                },
            ],
        }

        request = self.factory.post(
            "/appels/consolidation/a-charger/",
            data={
                "action": "create_batch",
                "batch_size": "1",
                "return_query": "classe=CLA001&statut_prestation=TERMINE",
            },
        )
        request.user = self.user

        response = consolidation_pending_appels(request)

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Appel.objects.filter(code="CLA001A").exists())
        self.assertFalse(Appel.objects.filter(code="CLA002A").exists())
        self.assertTrue(mock_messages_success.called)


class CgaImportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="cga-admin",
            password="test123",
            is_staff=True,
            is_superuser=True,
        )
        self.client.force_login(self.user)

    def _build_cga_upload(self, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(
            [
                "N°",
                "RAISON_SOCIALE",
                "SIGLE",
                "NIU",
                "ACTIVITE_PRINCIPALE",
                "REGIME",
                "CRI",
                "CENTRE_DE_RATTACHEMENT",
                "VILLE",
                "TELEPHONE",
            ]
        )
        for row in rows:
            ws.append(
                [
                    row.get("numero"),
                    row.get("raison_sociale"),
                    row.get("sigle", ""),
                    row.get("niu", ""),
                    row.get("activite_principale", ""),
                    row.get("regime", ""),
                    row.get("cri", ""),
                    row.get("centre_de_rattachement", ""),
                    row.get("ville", ""),
                    row.get("telephone", ""),
                ]
            )
        stream = io.BytesIO()
        wb.save(stream)
        return SimpleUploadedFile(
            "cga-import.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _build_onecca_upload(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sec I - EC Libéraux"
        headers = [
            "N°",
            "Noms & Prénoms",
            "Inscription N°",
            "Inscription Date",
            "Adresse Postale",
            "Ligne 1",
            "Ligne 2",
            "Adresse E-mail",
            "Site du Cabinet",
        ]
        ws.append(headers)
        ws.append(
            [
                "1",
                "ABEGE Patrick AGI",
                "182 ECP",
                "31/07/2013",
                "BP: 745 Bda",
                "679 62 69 13",
                "652 91 13 07",
                "abegepat20@example.com",
                "Commercial avenue Bamenda",
            ]
        )
        yaoude = wb.create_sheet("YAOUDE")
        yaoude.append(["SECTION DES EXPERTS-COMPTABLES LIBERAUX"])
        yaoude.append(headers)
        yaoude.append(
            [
                "1",
                "ABEGE Patrick AGI",
                "182 ECP",
                "31/07/2013",
                "BP: 745 Bda",
                "679 62 69 13",
                "",
                "duplicate@example.com",
                "Duplicate subset row",
            ]
        )
        societes = wb.create_sheet("Sec II - Sociétés d'EC")
        societes.append(headers)
        societes.append(
            [
                "1",
                "ACN & CO",
                "48 SEC",
                "04/05/2017",
                "BP: 183 Buéa",
                "676 54 87 77",
                "",
                "contact@example.com",
                "MAHAN House Molyko - Buea",
            ]
        )
        stream = io.BytesIO()
        wb.save(stream)
        return SimpleUploadedFile(
            "tableau-onecca.xlsx",
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_append_import_updates_existing_rows_and_creates_missing_ones(self):
        locked_at = timezone.now()
        existing = AppelCGA.objects.create(
            numero=1,
            raison_sociale="Ancienne raison",
            sigle="OLD",
            niu="NIU-001",
            activite_principale="Ancienne activite",
            regime="Ancien regime",
            cri="Ancien cri",
            centre_de_rattachement="Ancien centre",
            ville="Ancienne ville",
            telephone="690000000",
            is_active=False,
            status="pause",
            locked_by=self.user,
            locked_at=locked_at,
        )

        response = self.client.post(
            reverse("cga_index"),
            {
                "update_mode": "append",
                "file": self._build_cga_upload(
                    [
                        {
                            "numero": 12,
                            "raison_sociale": "Nouvelle raison",
                            "sigle": "NEW",
                            "niu": "NIU-001",
                            "activite_principale": "Nouvelle activite",
                            "regime": "Regime A",
                            "cri": "CRI A",
                            "centre_de_rattachement": "Centre A",
                            "ville": "Garoua",
                            "telephone": "699000001",
                        },
                        {
                            "numero": 13,
                            "raison_sociale": "Entreprise B",
                            "sigle": "ENTB",
                            "niu": "NIU-002",
                            "activite_principale": "Commerce",
                            "regime": "Regime B",
                            "cri": "CRI B",
                            "centre_de_rattachement": "Centre B",
                            "ville": "Maroua",
                            "telephone": "699000002",
                        },
                        {
                            "numero": 14,
                            "raison_sociale": "Entreprise B duplicate",
                            "sigle": "ENTB",
                            "niu": "NIU-002",
                            "activite_principale": "Commerce",
                            "regime": "Regime B",
                            "cri": "CRI B",
                            "centre_de_rattachement": "Centre B",
                            "ville": "Maroua",
                            "telephone": "699000002",
                        },
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 302)

        existing.refresh_from_db()
        self.assertEqual(existing.numero, 12)
        self.assertEqual(existing.raison_sociale, "Nouvelle raison")
        self.assertEqual(existing.sigle, "NEW")
        self.assertEqual(existing.activite_principale, "Nouvelle activite")
        self.assertEqual(existing.telephone, "699000001")
        self.assertTrue(existing.is_active)
        self.assertEqual(existing.status, "pause")
        self.assertEqual(existing.locked_by, self.user)
        self.assertEqual(existing.locked_at, locked_at)

        created = AppelCGA.objects.get(niu="NIU-002")
        self.assertEqual(created.raison_sociale, "Entreprise B")
        self.assertEqual(
            AppelCGA.objects.filter(source=AppelCGA.SOURCE_ENTREPRISE).count(),
            2,
        )

    @patch("App_PADESCE.appels.cga_views._sync_cga_append_batch")
    def test_append_import_splits_file_into_batches_of_2000(self, mock_sync_cga_append_batch):
        mock_sync_cga_append_batch.side_effect = lambda batch, **kwargs: (len(batch), 0)

        rows = [
            {
                "numero": index + 1,
                "raison_sociale": f"Entreprise {index + 1}",
                "sigle": f"SIG{index + 1}",
                "niu": f"NIU-{index + 1:04d}",
                "activite_principale": "Commerce",
                "regime": "Regime",
                "cri": "CRI",
                "centre_de_rattachement": "Centre",
                "ville": "Ville",
                "telephone": f"699{index + 1:06d}",
            }
            for index in range(2000)
        ]
        rows.append(
            {
                "numero": 2001,
                "raison_sociale": "Duplicate after first batch",
                "sigle": "DUP",
                "niu": "NIU-0001",
                "activite_principale": "Commerce",
                "regime": "Regime",
                "cri": "CRI",
                "centre_de_rattachement": "Centre",
                "ville": "Ville",
                "telephone": "699999998",
            }
        )
        rows.append(
            {
                "numero": 2002,
                "raison_sociale": "Entreprise 2001",
                "sigle": "SIG2001",
                "niu": "NIU-2001",
                "activite_principale": "Commerce",
                "regime": "Regime",
                "cri": "CRI",
                "centre_de_rattachement": "Centre",
                "ville": "Ville",
                "telephone": "699999999",
            }
        )

        response = self.client.post(
            reverse("cga_index"),
            {
                "update_mode": "append",
                "file": self._build_cga_upload(rows),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(mock_sync_cga_append_batch.call_count, 2)
        self.assertEqual(len(mock_sync_cga_append_batch.call_args_list[0].args[0]), 2000)
        self.assertEqual(len(mock_sync_cga_append_batch.call_args_list[1].args[0]), 1)
        self.assertEqual(
            mock_sync_cga_append_batch.call_args_list[0].kwargs["source"],
            AppelCGA.SOURCE_ENTREPRISE,
        )

    def test_onecca_import_loads_cabinet_source_and_skips_yaoude_subset(self):
        response = self.client.post(
            reverse("cga_index"),
            {
                "source": AppelCGA.SOURCE_CABINET,
                "update_mode": "append",
                "file": self._build_onecca_upload(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            response.headers["Location"],
            f"/cga/?source={AppelCGA.SOURCE_CABINET}",
        )
        cabinets = AppelCGA.objects.filter(source=AppelCGA.SOURCE_CABINET).order_by("niu")
        self.assertEqual(cabinets.count(), 2)
        first = AppelCGA.objects.get(source=AppelCGA.SOURCE_CABINET, niu="ONECCA-182 ECP")
        self.assertEqual(first.raison_sociale, "ABEGE Patrick AGI")
        self.assertEqual(first.sigle, "182 ECP")
        self.assertEqual(first.regime, "EC liberaux")
        self.assertEqual(first.ville, "Bamenda")
        self.assertIn("679 62 69 13", first.telephone)

    def test_replace_import_only_clears_selected_source(self):
        AppelCGA.objects.create(
            source=AppelCGA.SOURCE_ENTREPRISE,
            raison_sociale="Entreprise conservee",
            niu="NIU-KEEP",
        )
        AppelCGA.objects.create(
            source=AppelCGA.SOURCE_CABINET,
            raison_sociale="Ancien cabinet",
            niu="ONECCA-OLD",
        )

        response = self.client.post(
            reverse("cga_index"),
            {
                "source": AppelCGA.SOURCE_CABINET,
                "update_mode": "replace",
                "file": self._build_onecca_upload(),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(AppelCGA.objects.filter(niu="NIU-KEEP").exists())
        self.assertFalse(AppelCGA.objects.filter(niu="ONECCA-OLD").exists())
        self.assertEqual(AppelCGA.objects.filter(source=AppelCGA.SOURCE_CABINET).count(), 2)

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_cga_index_switches_between_entreprise_and_cabinet_rows(self):
        AppelCGA.objects.create(
            source=AppelCGA.SOURCE_ENTREPRISE,
            raison_sociale="Entreprise Alpha",
            niu="NIU-ALPHA",
        )
        AppelCGA.objects.create(
            source=AppelCGA.SOURCE_CABINET,
            raison_sociale="Cabinet Beta",
            niu="ONECCA-BETA",
        )

        entreprise_response = self.client.get(reverse("cga_index"))
        cabinet_response = self.client.get(
            reverse("cga_index"),
            {"source": AppelCGA.SOURCE_CABINET},
        )

        self.assertContains(entreprise_response, "Entreprise Alpha")
        self.assertNotContains(entreprise_response, "Cabinet Beta")
        self.assertContains(cabinet_response, "Cabinet Beta")
        self.assertNotContains(cabinet_response, "Entreprise Alpha")
        self.assertEqual(cabinet_response.context["active_source"], AppelCGA.SOURCE_CABINET)


class CgaPublicApiTests(TestCase):
    @override_settings(CGA_PUBLIC_API_KEY="cga-public-secret")
    def test_public_interested_api_requires_api_key(self):
        response = self.client.get(reverse("cga_public_interested_api"))

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response.json()["error"], "Cle API manquante ou invalide.")

    @override_settings(CGA_PUBLIC_API_KEY="cga-public-secret")
    def test_public_interested_api_returns_only_active_interested_rows(self):
        old_row = AppelCGA.objects.create(
            raison_sociale="Entreprise Alpha",
            niu="NIU-CGA-OLD",
            telephone="699000100",
            status="termine",
            interet="OUI",
            is_active=True,
        )
        fresh_row = AppelCGA.objects.create(
            raison_sociale="Entreprise Beta",
            niu="NIU-CGA-NEW",
            telephone="699000101",
            status="termine",
            interet="OUI",
            is_active=True,
            ville="Garoua",
        )
        AppelCGA.objects.create(
            raison_sociale="Entreprise Gamma",
            niu="NIU-CGA-NO",
            telephone="699000102",
            status="termine",
            interet="NON",
            is_active=True,
        )
        AppelCGA.objects.create(
            raison_sociale="Entreprise Delta",
            niu="NIU-CGA-INACTIVE",
            telephone="699000103",
            status="termine",
            interet="OUI",
            is_active=False,
        )

        old_updated_at = timezone.now() - datetime.timedelta(days=2)
        fresh_updated_at = timezone.now() - datetime.timedelta(minutes=5)
        AppelCGA.objects.filter(pk=old_row.pk).update(updated_at=old_updated_at)
        AppelCGA.objects.filter(pk=fresh_row.pk).update(updated_at=fresh_updated_at)

        response = self.client.get(
            reverse("cga_public_interested_api"),
            {
                "updated_since": (timezone.now() - datetime.timedelta(hours=1)).isoformat(),
            },
            HTTP_X_CGA_API_KEY="cga-public-secret",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["page"], 1)
        self.assertEqual(payload["page_size"], 100)
        self.assertEqual([item["niu"] for item in payload["appels"]], ["NIU-CGA-NEW"])
        self.assertEqual(payload["appels"][0]["telephone"], "699000101")
        self.assertEqual(payload["appels"][0]["ville"], "Garoua")
        self.assertEqual(payload["appels"][0]["status"], "termine")

    @override_settings(CGA_PUBLIC_API_KEY="cga-public-secret")
    def test_public_interested_api_accepts_bearer_token(self):
        AppelCGA.objects.create(
            raison_sociale="Entreprise Bearer",
            niu="NIU-CGA-BEARER",
            telephone="699000104",
            status="termine",
            interet="OUI",
            is_active=True,
        )

        response = self.client.get(
            reverse("cga_public_interested_api"),
            HTTP_AUTHORIZATION="Bearer cga-public-secret",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["total"], 1)

    @override_settings(CGA_PUBLIC_API_KEY="cga-public-secret")
    def test_public_interested_api_disables_shared_cache(self):
        response = self.client.get(
            reverse("cga_public_interested_api"),
            HTTP_X_CGA_API_KEY="cga-public-secret",
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn("no-store", response["Cache-Control"])
        self.assertIn("Authorization", response["Vary"])
        self.assertIn("X-CGA-Api-Key", response["Vary"])

    @override_settings(CGA_PUBLIC_API_KEY="cga-public-secret")
    def test_public_interested_api_rejects_invalid_updated_since(self):
        response = self.client.get(
            reverse("cga_public_interested_api"),
            {"updated_since": "not-a-date"},
            HTTP_X_CGA_API_KEY="cga-public-secret",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json()["error"], "updated_since invalide. Utilisez un datetime ISO 8601."
        )


class AppelsIndexFilterTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="appels-viewer",
            password="test123",
        )
        self.modifier = get_user_model().objects.create_user(
            username="answer-editor",
            password="test123",
        )
        self.client.force_login(self.user)

    def test_appels_index_filters_formulaire_and_modified_by(self):
        first = Appel.objects.create(
            code="APP001", nom="Alpha One", locked_by=self.user, status="formulaire_rempli"
        )
        second = Appel.objects.create(
            code="APP002", nom="Beta Two", locked_by=self.user, status="en_cours"
        )
        AppelAnswers.objects.create(
            appel=first,
            q1_clarte_exposes=4,
            q2_interaction_formateur=4,
            q3_maitrise_contenu=4,
            q4_salle_adequate=4,
            q5_materiel_disponible=4,
            q6_organisation_temps=4,
            q7_utilite_formation=4,
            q8_adequation_besoins=4,
            q9_satisfaction_globale=4,
            commentaire="RAS",
            recommandations="RAS",
            modified_by=self.modifier,
        )
        AppelAnswers.objects.create(
            appel=second,
            commentaire="Brouillon",
            recommandations="Brouillon",
            modified_by=self.modifier,
        )

        response = self.client.get(
            reverse("appels_index"),
            {"formulaire": "rempli", "modified_by": self.modifier.username, "q": "APP001"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "APP001")
        self.assertNotContains(response, "APP002")

    def test_appels_index_treats_single_answer_as_filled_form(self):
        appel = Appel.objects.create(
            code="APP003", nom="Gamma Three", locked_by=self.user, status="appel_tente"
        )
        AppelAnswers.objects.create(
            appel=appel,
            q1_clarte_exposes=5,
            commentaire="Premier point repondu",
            recommandations="RAS",
            modified_by=self.modifier,
        )

        response = self.client.get(reverse("appels_index"), {"formulaire": "rempli", "q": "APP003"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "APP003")

    def test_appels_index_completed_filter_includes_all_finalized_statuses(self):
        Appel.objects.create(
            code="APP005", nom="Echo Five", locked_by=self.user, status="appel_reussi"
        )
        Appel.objects.create(
            code="APP006", nom="Foxtrot Six", locked_by=self.user, status="formulaire_rempli"
        )
        Appel.objects.create(
            code="APP007", nom="Golf Seven", locked_by=self.user, status="formulaire_avec_audio"
        )
        Appel.objects.create(
            code="APP008", nom="Hotel Eight", locked_by=self.user, status="a_rappeler"
        )

        response = self.client.get(reverse("appels_index"), {"status": "completed"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "APP005")
        self.assertContains(response, "APP006")
        self.assertContains(response, "APP007")
        self.assertNotContains(response, "APP008")

    def test_appel_answers_detail_syncs_status_after_single_answer(self):
        appel = Appel.objects.create(
            code="APP004", nom="Delta Four", locked_by=self.user, status="appel_tente"
        )

        response = self.client.post(
            reverse("appel_answers_detail", args=[appel.pk]),
            {
                "q1": "4",
                "commentaire": "Reponse partielle",
                "recommandations": "RAS",
            },
        )

        self.assertEqual(response.status_code, 200)
        appel.refresh_from_db()
        self.assertEqual(appel.status, "formulaire_rempli")

    def test_appels_index_hides_missing_audio_files(self):
        Appel.objects.create(
            code="APP404",
            nom="Audio Missing",
            status="termine",
            audio_file="padesce/2026/03/18/missing-audio.webm",
        )

        response = self.client.get(reverse("appels_index"), {"q": "APP404"})

        self.assertEqual(response.status_code, 200)
        row = next(item for item in response.context["appels"] if item.code == "APP404")
        self.assertFalse(row.has_audio_file)
        self.assertEqual(row.audio_file_url, "")
        self.assertContains(response, "Audio introuvable")
        self.assertNotContains(response, "missing-audio.webm")

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_hides_reached_classes_and_ignores_learners_without_phone(
        self, mock_build_source_index
    ):
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "classes": {
                "cla001": {
                    "classe_id": "CLA001",
                    "prestation_id": "PRESTA001",
                    "statut_prestation": "TERMINE",
                },
                "cla002": {
                    "classe_id": "CLA002",
                    "prestation_id": "PRESTA001",
                    "statut_prestation": "TERMINE",
                },
            },
            "prestations": {
                "presta001": {"prestation_id": "PRESTA001", "statut_prestation": "TERMINE"},
            },
            "records": {
                "a1": {"classe_id": "CLA001", "telephone1": "690000001", "telephone2": ""},
                "a2": {"classe_id": "CLA001", "telephone1": "690000002", "telephone2": ""},
                "a3": {"classe_id": "CLA001", "telephone1": "", "telephone2": ""},
                "b1": {"classe_id": "CLA002", "telephone1": "690000010", "telephone2": ""},
            },
        }

        Appel.objects.create(
            code="CLA001-A",
            nom="Alpha",
            classe_label="CLA001",
            telephone1="690000001",
            status="formulaire_rempli",
        )
        Appel.objects.create(
            code="CLA001-B",
            nom="Bravo",
            classe_label="CLA001",
            telephone1="690000002",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA001-C",
            nom="Charlie",
            classe_label="CLA001",
            telephone1="",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA002-A",
            nom="Delta",
            classe_label="CLA002",
            telephone1="690000010",
            status="en_attente",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        mock_build_source_index.assert_any_call(source_key="cutoff")
        visible_classes = {row.classe_label for row in response.context["appels"]}
        self.assertEqual(visible_classes, {"CLA002"})
        self.assertEqual(
            [item["value"] for item in response.context["filters"]["classes_enriched"]],
            ["CLA002"],
        )
        classe_progress = {item["classe"]: item for item in response.context["classe_progress"]}
        self.assertEqual(classe_progress["CLA001"]["total"], 2)
        self.assertTrue(classe_progress["CLA001"]["reached"])

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_uses_saved_forms_for_25_percent_threshold(self, mock_build_source_index):
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "classes": {
                "cla100": {
                    "classe_id": "CLA100",
                    "prestation_id": "PRESTA100",
                    "statut_prestation": "TERMINE",
                },
            },
            "prestations": {
                "presta100": {"prestation_id": "PRESTA100", "statut_prestation": "TERMINE"},
            },
            "records": {
                "a1": {"classe_id": "CLA100", "telephone1": "690100001", "telephone2": ""},
                "a2": {"classe_id": "CLA100", "telephone1": "690100002", "telephone2": ""},
                "a3": {"classe_id": "CLA100", "telephone1": "690100003", "telephone2": ""},
                "a4": {"classe_id": "CLA100", "telephone1": "690100004", "telephone2": ""},
            },
        }

        counted_row = Appel.objects.create(
            code="CLA100-A",
            nom="Alpha",
            classe_label="CLA100",
            telephone1="690100001",
            status="appel_reussi",
        )
        Appel.objects.create(
            code="CLA100-B",
            nom="Bravo",
            classe_label="CLA100",
            telephone1="690100002",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA100-C",
            nom="Charlie",
            classe_label="CLA100",
            telephone1="690100003",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA100-D",
            nom="Delta",
            classe_label="CLA100",
            telephone1="690100004",
            status="en_attente",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual({row.classe_label for row in response.context["appels"]}, {"CLA100"})
        classe_progress = {item["classe"]: item for item in response.context["classe_progress"]}
        self.assertEqual(classe_progress["CLA100"]["termines"], 0)
        self.assertFalse(classe_progress["CLA100"]["reached"])

        counted_row.status = "formulaire_rempli"
        counted_row.save(update_fields=["status"])

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["filters"]["classes_enriched"]), [])
        self.assertEqual(list(response.context["appels"]), [])

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_counts_termine_status_for_25_percent_threshold(
        self, mock_build_source_index
    ):
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "classes": {
                "cla200": {
                    "classe_id": "CLA200",
                    "prestation_id": "PRESTA200",
                    "statut_prestation": "TERMINE",
                },
            },
            "prestations": {
                "presta200": {"prestation_id": "PRESTA200", "statut_prestation": "TERMINE"},
            },
            "records": {
                "a1": {"classe_id": "CLA200", "telephone1": "690200001", "telephone2": ""},
                "a2": {"classe_id": "CLA200", "telephone1": "690200002", "telephone2": ""},
                "a3": {"classe_id": "CLA200", "telephone1": "690200003", "telephone2": ""},
                "a4": {"classe_id": "CLA200", "telephone1": "690200004", "telephone2": ""},
            },
        }

        Appel.objects.create(
            code="CLA200-A",
            nom="Alpha",
            classe_label="CLA200",
            telephone1="690200001",
            status="termine",
        )
        Appel.objects.create(
            code="CLA200-B",
            nom="Bravo",
            classe_label="CLA200",
            telephone1="690200002",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA200-C",
            nom="Charlie",
            classe_label="CLA200",
            telephone1="690200003",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA200-D",
            nom="Delta",
            classe_label="CLA200",
            telephone1="690200004",
            status="en_attente",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["filters"]["classes_enriched"]), [])
        self.assertEqual(list(response.context["appels"]), [])
        classe_progress = {item["classe"]: item for item in response.context["classe_progress"]}
        self.assertEqual(classe_progress["CLA200"]["termines"], 1)
        self.assertTrue(classe_progress["CLA200"]["reached"])

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_maps_completed_rows_without_class_label_from_source_code(
        self, mock_build_source_index
    ):
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "classes": {
                "cla210": {
                    "classe_id": "CLA210",
                    "prestation_id": "PRESTA210",
                    "statut_prestation": "TERMINE",
                },
            },
            "prestations": {
                "presta210": {"prestation_id": "PRESTA210", "statut_prestation": "TERMINE"},
            },
            "records": {
                "app210a": {
                    "code": "APP210A",
                    "classe_id": "CLA210",
                    "telephone1": "690210001",
                    "telephone2": "",
                },
            },
        }

        Appel.objects.create(
            code="APP210A",
            nom="Alpha",
            classe_label="",
            telephone1="690210001",
            status="termine",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context["appels"]), [])
        self.assertEqual(response.context["hidden_class_summary"]["hidden_class_count"], 1)
        classe_progress = {item["classe"]: item for item in response.context["classe_progress"]}
        self.assertEqual(classe_progress["CLA210"]["termines"], 1)
        self.assertTrue(classe_progress["CLA210"]["reached"])

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_exposes_goal_summary_for_70_of_75_prestations(
        self, mock_build_source_index
    ):
        classes = {}
        prestations = {}
        records = {}
        for index in range(1, 76):
            prestation_id = f"PRESTA{index:03d}"
            classe_id = f"CLA{index:03d}"
            classes[classe_id.lower()] = {
                "classe_id": classe_id,
                "prestation_id": prestation_id,
                "statut_prestation": "TERMINE",
            }
            prestations[prestation_id.lower()] = {
                "prestation_id": prestation_id,
                "statut_prestation": "TERMINE",
            }
            records[f"r{index:03d}"] = {
                "classe_id": classe_id,
                "telephone1": f"690{index:06d}",
                "telephone2": "",
            }
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "counts": {"prestations": 75, "classes": 75, "apprenants": 75},
            "classes": classes,
            "prestations": prestations,
            "records": records,
        }

        Appel.objects.create(
            code="CLA001-A",
            nom="Alpha",
            classe_label="CLA001",
            telephone1="690000001",
            status="termine",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        summary = response.context["analysis_goal_summary"]
        self.assertEqual(summary["total_prestations_count"], 75)
        self.assertEqual(summary["analysis_prestations_count"], 1)
        self.assertEqual(summary["minimum_target"], 70)
        self.assertEqual(summary["remaining_to_minimum_target"], 69)
        self.assertEqual(summary["actionable_prestations_count"], 74)
        self.assertEqual(summary["blocked_prestations_count"], 0)
        self.assertEqual(summary["max_reachable_prestations_count"], 75)
        self.assertEqual(summary["minimum_target_gap"], 0)

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_recommends_last_class_to_finish_prestation(self, mock_build_source_index):
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "classes": {
                "cla010": {
                    "classe_id": "CLA010",
                    "prestation_id": "PRESTA010",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "cla011": {
                    "classe_id": "CLA011",
                    "prestation_id": "PRESTA010",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "cla012": {
                    "classe_id": "CLA012",
                    "prestation_id": "PRESTA010",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "cla020": {
                    "classe_id": "CLA020",
                    "prestation_id": "PRESTA020",
                    "prestataire": "Prestataire B",
                    "beneficiaire": "Beneficiaire B",
                    "statut_prestation": "TERMINE",
                },
            },
            "prestations": {
                "presta010": {
                    "prestation_id": "PRESTA010",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "presta020": {
                    "prestation_id": "PRESTA020",
                    "prestataire": "Prestataire B",
                    "beneficiaire": "Beneficiaire B",
                    "statut_prestation": "TERMINE",
                },
            },
            "records": {
                "a1": {"classe_id": "CLA010", "telephone1": "690100001", "telephone2": ""},
                "a2": {"classe_id": "CLA010", "telephone1": "690100002", "telephone2": ""},
                "b1": {"classe_id": "CLA011", "telephone1": "690110001", "telephone2": ""},
                "b2": {"classe_id": "CLA011", "telephone1": "690110002", "telephone2": ""},
                "c1": {"classe_id": "CLA012", "telephone1": "690120001", "telephone2": ""},
                "d1": {"classe_id": "CLA020", "telephone1": "690200001", "telephone2": ""},
            },
        }

        Appel.objects.create(
            code="CLA010-A",
            nom="A",
            classe_label="CLA010",
            telephone1="690100001",
            status="formulaire_rempli",
        )
        Appel.objects.create(
            code="CLA010-B",
            nom="B",
            classe_label="CLA010",
            telephone1="690100002",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA011-A",
            nom="C",
            classe_label="CLA011",
            telephone1="690110001",
            status="formulaire_rempli",
        )
        Appel.objects.create(
            code="CLA011-B",
            nom="D",
            classe_label="CLA011",
            telephone1="690110002",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA012-A",
            nom="E",
            classe_label="CLA012",
            telephone1="690120001",
            status="en_attente",
        )
        Appel.objects.create(
            code="CLA020-A",
            nom="F",
            classe_label="CLA020",
            telephone1="690200001",
            status="en_attente",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        recommendations = response.context["recommended_classes"]
        self.assertGreaterEqual(len(recommendations), 1)
        self.assertEqual(recommendations[0]["classe"], "CLA012")
        self.assertEqual(recommendations[0]["priority_label"], "Prestation a finir")

    @patch("App_PADESCE.appels.views.build_padesce_source_index")
    def test_appels_index_counts_terminated_prestation_when_one_class_reaches_analysis_threshold(
        self, mock_build_source_index
    ):
        mock_build_source_index.return_value = {
            "source": {"label": "Fichier consolide", "modified_label": "27/03/2026 a 12:20"},
            "classes": {
                "cla100": {
                    "classe_id": "CLA100",
                    "prestation_id": "PRESTA100",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "cla101": {
                    "classe_id": "CLA101",
                    "prestation_id": "PRESTA100",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "cla200": {
                    "classe_id": "CLA200",
                    "prestation_id": "PRESTA200",
                    "prestataire": "Prestataire B",
                    "beneficiaire": "Beneficiaire B",
                    "statut_prestation": "EN COURS",
                },
            },
            "prestations": {
                "presta100": {
                    "prestation_id": "PRESTA100",
                    "prestataire": "Prestataire A",
                    "beneficiaire": "Beneficiaire A",
                    "statut_prestation": "TERMINE",
                },
                "presta200": {
                    "prestation_id": "PRESTA200",
                    "prestataire": "Prestataire B",
                    "beneficiaire": "Beneficiaire B",
                    "statut_prestation": "EN COURS",
                },
            },
            "records": {
                "a1": {"classe_id": "CLA100", "telephone1": "690001100"},
                "a2": {"classe_id": "CLA101", "telephone1": "690001101"},
                "b1": {"classe_id": "CLA200", "telephone1": "690002200"},
            },
        }

        Appel.objects.create(
            code="APP-100-1",
            nom="Alpha",
            classe_label="CLA100",
            telephone1="690001100",
            status="termine",
        )

        response = self.client.get(reverse("appels_index"))

        self.assertEqual(response.status_code, 200)
        summary = response.context["hidden_class_summary"]
        self.assertEqual(summary["analysis_prestations_count"], 1)
        self.assertEqual(summary["analysis_prestations_total_count"], 1)
        self.assertEqual(summary["analysis_prestations_ratio"], "1/1")


class AppelActionFlagTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="appels-agent",
            password="test123",
        )
        self.client.force_login(self.user)

    def test_appel_action_persists_not_formed_flag(self):
        appel = Appel.objects.create(
            code="APP-FLAG-001",
            nom="Apprenant Non Forme",
            status="en_attente",
            is_active=True,
        )

        response = self.client.post(
            reverse("appel_action", args=[appel.pk]),
            {"action": "terminer", "flag_pas_forme": "1"},
        )

        self.assertEqual(response.status_code, 200)
        appel.refresh_from_db()
        self.assertTrue(appel.flag_pas_forme)

    def test_appel_action_start_marks_call_in_progress(self):
        appel = Appel.objects.create(
            code="APP-FLAG-002",
            nom="Apprenant En Cours",
            status="en_attente",
            is_active=True,
        )

        response = self.client.post(
            reverse("appel_action", args=[appel.pk]),
            {"action": "start"},
        )

        self.assertEqual(response.status_code, 200)
        appel.refresh_from_db()
        self.assertEqual(appel.status, "en_cours")


class AppelFinalizeSaveTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="appels-finalize-agent",
            password="test123",
        )
        self.client.force_login(self.user)

    def test_finalize_saves_form_when_questions_are_submitted_without_form_modified_flag(self):
        Appel.objects.create(
            code="APP-FINAL-002",
            nom="Autre Apprenant",
            classe_label="CLA-FINAL",
            telephone1="690000001",
            status="en_attente",
            is_active=True,
        )
        appel = Appel.objects.create(
            code="APP-FINAL-001",
            nom="Apprenant Test",
            classe_label="CLA-FINAL",
            telephone1="690000002",
            status="en_cours",
            is_active=True,
        )

        response = self.client.post(
            reverse("appel_finalize", args=[appel.pk]),
            {
                "action": "terminer",
                "q1": "4",
                "q9": "5",
                "commentaire": "Formulaire PADESCE bien renseigne",
                "recommandations": "Continuer comme cela",
                "form_modified": "0",
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        appel.refresh_from_db()
        answers = appel.answers

        self.assertEqual(appel.status, "formulaire_rempli")
        self.assertEqual(payload["status"], "formulaire_rempli")
        self.assertFalse(payload["satisfaction_saved"])
        self.assertEqual(answers.q1_clarte_exposes, 4)
        self.assertEqual(answers.q9_satisfaction_globale, 5)
        self.assertEqual(answers.commentaire, "Formulaire PADESCE bien renseigne")
        self.assertEqual(answers.recommandations, "Continuer comme cela")
        self.assertFalse(SatisfactionApprenant.objects.filter(appel=appel).exists())
        self.assertEqual(payload["class_progress"]["termines"], 1)
        self.assertEqual(payload["class_progress"]["target"], 1)
        self.assertTrue(payload["class_progress"]["reached"])

    def test_finalize_without_questionnaire_answers_marks_call_successful(self):
        appel = Appel.objects.create(
            code="APP-FINAL-003",
            nom="Apprenant Sans Formulaire",
            classe_label="CLA-FINAL-2",
            telephone1="690000003",
            status="en_cours",
            is_active=True,
        )

        response = self.client.post(
            reverse("appel_finalize", args=[appel.pk]),
            {
                "action": "terminer",
                "commentaire": "Pas de notes disponibles",
                "recommandations": "",
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()

        appel.refresh_from_db()

        self.assertEqual(appel.status, "appel_reussi")
        self.assertEqual(payload["status"], "appel_reussi")
        self.assertFalse(payload["satisfaction_saved"])
        self.assertFalse(SatisfactionApprenant.objects.filter(appel=appel).exists())


class CgaAudioAndTemplateTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username="cga-agent",
            password="test123",
        )
        self.client.force_login(self.user)

    def test_cga_index_hides_transcription_controls(self):
        with open("templates/appels/cga.html", "r", encoding="utf-8") as template_file:
            template_source = template_file.read()

        self.assertNotIn("Transcrire le tableau filtre", template_source)
        self.assertNotIn('class="btn-small js-transcription"', template_source)
        self.assertNotIn("Transcription locale", template_source)

    def test_cga_upload_audio_replaces_previous_file(self):
        temp_media_root = tempfile.mkdtemp(prefix="cga-audio-test-")
        try:
            with override_settings(MEDIA_ROOT=temp_media_root):
                row = AppelCGA.objects.create(
                    raison_sociale="Entreprise Beta",
                    niu="NIU-CGA-002",
                    telephone="690000222",
                    status="en_cours",
                    is_active=True,
                )

                first_audio = SimpleUploadedFile(
                    "premier.mp3",
                    b"first-audio",
                    content_type="audio/mpeg",
                )
                first_response = self.client.post(
                    reverse("cga_upload_audio", args=[row.pk]),
                    {"audio": first_audio},
                )
                self.assertEqual(first_response.status_code, 200)
                self.assertTrue(first_response.json()["audio_saved"])

                row.refresh_from_db()
                first_name = row.audio_file.name
                self.assertTrue(os.path.exists(os.path.join(temp_media_root, first_name)))

                second_audio = SimpleUploadedFile(
                    "second.webm",
                    b"second-audio",
                    content_type="audio/webm",
                )
                second_response = self.client.post(
                    reverse("cga_upload_audio", args=[row.pk]),
                    {"audio": second_audio},
                )

                self.assertEqual(second_response.status_code, 200)
                row.refresh_from_db()
                second_name = row.audio_file.name

                self.assertNotEqual(second_name, first_name)
                self.assertFalse(os.path.exists(os.path.join(temp_media_root, first_name)))
                self.assertTrue(os.path.exists(os.path.join(temp_media_root, second_name)))
                self.assertTrue(second_response.json()["audio_url"])
        finally:
            shutil.rmtree(temp_media_root, ignore_errors=True)


class PrestataireDemarrageBulkDeactivateTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.superadmin = User.objects.create_superuser(
            username="demarrage-superadmin",
            email="demarrage-superadmin@example.com",
            password="test-pass-123",
        )
        self.operator = User.objects.create_user(
            username="demarrage-operatrice",
            password="test-pass-123",
        )
        self.row_a = AppelPrestataireDemarrage.objects.create(
            reference_code="PREST-A-690000001",
            nom_prestataire="Prestataire A",
            telephone="690000001",
        )
        self.row_b = AppelPrestataireDemarrage.objects.create(
            reference_code="PREST-B-690000002",
            nom_prestataire="Prestataire B",
            telephone="690000002",
        )

    def test_superadmin_can_deactivate_selected_rows(self):
        self.client.force_login(self.superadmin)

        response = self.client.post(
            reverse("prestataire_demarrage_bulk_deactivate"),
            {"selected_ids": [str(self.row_a.pk)]},
        )

        self.assertEqual(response.status_code, 302)
        self.row_a.refresh_from_db()
        self.row_b.refresh_from_db()
        self.assertFalse(self.row_a.is_active)
        self.assertTrue(self.row_b.is_active)

    def test_superadmin_can_reactivate_selected_rows(self):
        self.row_a.is_active = False
        self.row_a.save(update_fields=["is_active", "updated_at"])
        self.client.force_login(self.superadmin)

        response = self.client.post(
            reverse("prestataire_demarrage_bulk_reactivate"),
            {"selected_ids": [str(self.row_a.pk)]},
        )

        self.assertEqual(response.status_code, 302)
        self.row_a.refresh_from_db()
        self.assertTrue(self.row_a.is_active)

    def test_operator_cannot_deactivate_selected_rows(self):
        self.client.force_login(self.operator)

        response = self.client.post(
            reverse("prestataire_demarrage_bulk_deactivate"),
            {"selected_ids": [str(self.row_a.pk)]},
        )

        self.assertEqual(response.status_code, 302)
        self.row_a.refresh_from_db()
        self.assertTrue(self.row_a.is_active)

    def test_operator_cannot_reactivate_selected_rows(self):
        self.row_a.is_active = False
        self.row_a.save(update_fields=["is_active", "updated_at"])
        self.client.force_login(self.operator)

        response = self.client.post(
            reverse("prestataire_demarrage_bulk_reactivate"),
            {"selected_ids": [str(self.row_a.pk)]},
        )

        self.assertEqual(response.status_code, 302)
        self.row_a.refresh_from_db()
        self.assertFalse(self.row_a.is_active)

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_operator_list_hides_deactivated_rows(self):
        self.row_a.is_active = False
        self.row_a.save(update_fields=["is_active", "updated_at"])
        self.client.force_login(self.operator)

        response = self.client.get(
            reverse("prestataire_demarrage_index"),
            {"active_state": "inactive"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Prestataire A")
        self.assertContains(response, "Prestataire B")

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_superadmin_can_view_deactivated_rows(self):
        self.row_a.is_active = False
        self.row_a.save(update_fields=["is_active", "updated_at"])
        self.client.force_login(self.superadmin)

        response = self.client.get(
            reverse("prestataire_demarrage_index"),
            {"active_state": "inactive"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Prestataire A")
        self.assertContains(response, "Desactivee")
        self.assertContains(response, "Reactiver la selection")
        self.assertNotContains(response, "Prestataire B")

    def test_inactive_row_rejects_direct_call_actions(self):
        self.row_a.is_active = False
        self.row_a.save(update_fields=["is_active", "updated_at"])
        self.client.force_login(self.operator)

        response = self.client.post(
            reverse("prestataire_demarrage_action", args=[self.row_a.pk]),
            {"action": "start"},
            HTTP_ACCEPT="application/json",
        )

        self.assertEqual(response.status_code, 404)
