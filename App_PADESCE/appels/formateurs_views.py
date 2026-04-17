# ruff: noqa: E501
import csv
import datetime
import io
import re
import unicodedata
import zipfile
from pathlib import Path
from types import SimpleNamespace

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from App_PADESCE.appels.formateur_names import resolve_formateur_db_name_from_values
from App_PADESCE.appels.models import (
    CALL_COMPLETED_STATUSES,
    CALL_FORM_STATUSES,
    CALL_SUCCESS_STATUSES,
    AppelFormateur,
    _short_slug,
    sync_formateur_status,
)
from App_PADESCE.appels.views import (
    _bind_audio_state,
    _cleanup_stale_locks,
    _deactivate_duplicate_rows,
    _formateur_duplicate_key,
    _has_audio_file,
    _safe_audio_url,
)
from App_PADESCE.formations.models import Classe
from App_PADESCE.satisfaction_formateurs.models import SatisfactionFormateur

FORMATEUR_THRESHOLD_PERCENT = 50
FORMATEURS_FILTERED_TRANSCRIPTION_TASK_KEY = "formateurs_filtered_transcription"

FORMATEUR_SATISFACTION_HEADER_FIELDS = (
    (
        "q1_prerequis_apprenants",
        "Q1 - Niveau des prerequis des apprenants",
        "Les apprenants avaient-ils les prerequis necessaires pour suivre cette formation ?",
    ),
    (
        "q2_interaction_apprenants",
        "Q2 - Niveau d'interaction des apprenants",
        "Les apprenants ont-ils ete suffisamment interactifs, posant des questions "
        "et participant activement ?",
    ),
    (
        "q3_competences_acquises",
        "Q3 - Competences acquises par les apprenants",
        "Estimez-vous que les apprenants ont acquis les competences cibles "
        "a l'issue de la formation ?",
    ),
)


def _transcription_to_payload(obj) -> dict:
    return {
        "text": str(getattr(obj, "transcription_text", "") or ""),
        "engine": str(getattr(obj, "engine", "local") or "local"),
    }


def _ensure_formateur_transcription(row):
    return (
        SimpleNamespace(
            transcription_text=str(getattr(row, "transcription_text", "") or ""),
            engine="local",
        ),
        False,
    )


def _collect_jobs_from_formateurs(queryset) -> list[dict]:
    jobs = []
    for row in queryset:
        if not getattr(row, "audio_file", None):
            continue
        jobs.append({"id": row.pk})
    return jobs


def _transcription_status_response(_task_key: str) -> dict:
    return {
        "state": "idle",
        "message": "Transcription groupée indisponible sur cette version.",
        "total": 0,
        "processed": 0,
    }


def _start_transcription_task(task_key: str, jobs: list[dict], _label: str):
    return True, {"ok": True, "status": _transcription_status_response(task_key)}, 200


def _stop_transcription_task(task_key: str):
    return True, {"ok": True, "status": _transcription_status_response(task_key)}, 200


def _build_formateur_progress_metrics(queryset):
    from django.db.models import Q

    completed_q = Q(status__in=CALL_SUCCESS_STATUSES)
    threshold_q = Q(status__in=CALL_FORM_STATUSES)
    stats = queryset.aggregate(
        total=Count("id"),
        termines=Count("id", filter=threshold_q),
        appels_tentes=Count("id", filter=~Q(status="en_attente")),
        appels_reussis=Count("id", filter=completed_q),
        formulaires_remplis=Count("id", filter=Q(status__in=CALL_FORM_STATUSES)),
        formulaires_avec_audio=Count("id", filter=Q(status="formulaire_avec_audio")),
        audios=Count("id", filter=Q(audio_file__isnull=False) | Q(status="formulaire_avec_audio")),
    )
    total = int(stats.get("total") or 0)
    termines = int(stats.get("termines") or 0)
    threshold_target = max(1, int(total * FORMATEUR_THRESHOLD_PERCENT / 100)) if total else 0
    threshold_reached = total > 0 and termines >= threshold_target
    threshold_remaining = max(threshold_target - termines, 0)
    stats.update(
        {
            "remaining": max(total - termines, 0),
            "completion_rate": round((termines / total) * 100, 1) if total else 0.0,
            "completion_label": f"{termines} / {total}" if total else "0 / 0",
            "threshold_target": threshold_target,
            "threshold_remaining": threshold_remaining,
            "threshold_reached": threshold_reached,
            "threshold_message": (
                f"Seuil de {FORMATEUR_THRESHOLD_PERCENT}% atteint. "
                f"Vous pouvez passer a autre chose."
                if threshold_reached
                else (
                    f"Encore {threshold_remaining} appel(s) pour atteindre "
                    f"{FORMATEUR_THRESHOLD_PERCENT}%."
                    if total
                    else "Aucun appel dans ce filtre."
                )
            ),
            "threshold_label": f"{FORMATEUR_THRESHOLD_PERCENT}%",
        }
    )
    return stats


def _build_formateur_satisfaction_header_metrics(queryset):
    aggregate_map = {}
    for index, (field_name, _title, _question) in enumerate(
        FORMATEUR_SATISFACTION_HEADER_FIELDS,
        start=1,
    ):
        aggregate_map[f"avg_{index}"] = Avg(field_name)
        aggregate_map[f"count_{index}"] = Count("id", filter=Q(**{f"{field_name}__isnull": False}))

    aggregates = queryset.aggregate(**aggregate_map)
    return [
        {
            "field": field_name,
            "title": title,
            "question": question,
            "average": round(aggregates.get(f"avg_{index}") or 0, 2),
            "responses": int(aggregates.get(f"count_{index}") or 0),
        }
        for index, (field_name, title, question) in enumerate(
            FORMATEUR_SATISFACTION_HEADER_FIELDS,
            start=1,
        )
    ]


IMPORT_BATCH_SIZE = 1000
PAGE_SIZE_DEFAULT = 100
PAGE_SIZE_MAX = 500

FRENCH_MONTHS = {
    "janvier": 1,
    "fevrier": 2,
    "fevrier_alt": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "aout_alt": 8,
    "septembre": 9,
    "octobre": 10,
    "novembre": 11,
    "decembre": 12,
    "decembre_alt": 12,
}


def _normalize_header(value):
    if value is None:
        return ""
    text = " ".join(str(value).strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _as_text(value):
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _normalize_text(value):
    text = " ".join(_as_text(value).split())
    normalized = unicodedata.normalize("NFKD", text.lower())
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_phone_number(value):
    digits = re.sub(r"\D", "", _as_text(value))
    if not digits:
        return ""
    if digits.startswith("00237") and len(digits) >= 14:
        digits = digits[5:]
    elif digits.startswith("237") and len(digits) >= 12:
        digits = digits[3:]
    if len(digits) > 9:
        last9 = digits[-9:]
        if len(last9) == 9:
            return last9
    return digits


def _extract_phone_numbers(value):
    raw = _as_text(value)
    if not raw:
        return []
    # Extract 9-digit phone sequences even when embedded in prose like
    # "Promoteur ... Tel: 690 33 00 41/ 678 47 55 41".
    regex_matches = re.findall(r"(?<!\d)(?:\+?237[\s\-./]*)?(?:\d[\s\-./]*){9}(?!\d)", raw)
    numbers = []
    seen = set()
    for chunk in regex_matches:
        number = _normalize_phone_number(chunk)
        if not number or number in seen:
            continue
        seen.add(number)
        numbers.append(number)
    if numbers:
        return numbers

    # Fallback for slash-delimited cells and spreadsheet artifacts such as '$699...$'.
    cleaned = (
        raw.replace("$", " ")
        .replace("\\", "/")
        .replace("|", "/")
        .replace(";", "/")
        .replace(",", "/")
    )
    chunks = re.split(r"[/\n\r]+", cleaned)
    for chunk in chunks:
        number = _normalize_phone_number(chunk)
        if not number or number in seen:
            continue
        seen.add(number)
        numbers.append(number)
    if numbers:
        return numbers
    fallback = _normalize_phone_number(raw)
    return [fallback] if fallback else []


def _clean_phone_source(value):
    numbers = _extract_phone_numbers(value)
    return "/".join(numbers)


def _parse_french_date(value):
    text = _as_text(value)
    if not text:
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    normalized = _normalize_text(text)
    match = re.search(r"(\d{1,2})\s+([a-z]+)\s+(\d{4})", normalized)
    if not match:
        return None
    day = int(match.group(1))
    month = FRENCH_MONTHS.get(match.group(2))
    year = int(match.group(3))
    if not month:
        return None
    try:
        return datetime.date(year, month, day)
    except ValueError:
        return None


def _build_reference_code(item):
    # Extraire les informations pertinentes
    numero_seance = item.get("numero_seance") or "0"
    formateur = item.get("formateur") or item.get("contact_formateur") or ""

    # Formater le nom du formateur (prendre les premiers caractères significatifs)
    formateur_slug = _short_slug(formateur, "formateur", max_len=20)

    # Construire un code simple et lisible avec seulement le numéro et le formateur
    return f"FORM-{numero_seance}-{formateur_slug}"


def _iter_formateur_excel_rows(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    if "Calendrier" not in wb.sheetnames:
        raise ValueError("La feuille Calendrier est introuvable dans le fichier.")
    ws = wb["Calendrier"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return
    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            idx = header_map.get(_normalize_header(key))
            if idx is None or idx >= len(row):
                continue
            return row[idx]
        return None

    for row in rows:
        numero_raw = get(row, "Numero seance", "N seance", "No seance")
        prestataire = _as_text(get(row, "PRESTATAIRE"))
        beneficiaire = _as_text(get(row, "BENEFICIAIRE"))
        formation = _as_text(get(row, "FORMATION"))
        lieu = _as_text(get(row, "LIEU"))
        contact_raw = _as_text(get(row, "CONTACT DU FORMATEUR"))
        contact_clean = _clean_phone_source(contact_raw)
        cohorte = _as_text(get(row, "COHORTE"))
        date_label = _as_text(get(row, "DATE"))
        heure_debut = _as_text(get(row, "Heure de debut", "Heure debut"))
        heure_fin = _as_text(get(row, "Heure de fin"))
        if not any([prestataire, beneficiaire, formation, lieu, contact_raw]):
            continue
        try:
            numero_seance = int(float(numero_raw)) if numero_raw not in (None, "") else None
        except Exception:
            numero_seance = None
        session_date = _parse_french_date(get(row, "DATE"))
        phone_numbers = _extract_phone_numbers(contact_raw)
        if not phone_numbers:
            phone_numbers = [""]
        for phone in phone_numbers:
            item = {
                "numero_seance": numero_seance,
                "prestataire": prestataire,
                "beneficiaire": beneficiaire,
                "formation": formation,
                "lieu": lieu,
                "telephone": phone,
                "cohorte": cohorte.replace("$", "").strip(),
                "date_label": date_label,
                "session_date": session_date,
                "heure_debut": heure_debut,
                "heure_fin": heure_fin,
                "source_contact": contact_clean,
            }
            item["reference_code"] = _build_reference_code(item)
            yield item


# Mots-clés indiquant qu'un formateur n'a pas assuré la formation.
# Recherchés dans commentaires, recommandations et les champs Q4-Q6.
_PAS_FORME_KEYWORDS = [
    "n'a pas formé",
    "na pas formé",
    "n'a pas forme",
    "na pas forme",
    "pas formé",
    "pas forme",
    "n'a pas assuré",
    "n'a pas assure",
    "n'a pas pu former",
    "pas de formation",
    "n'a pas travaillé",
    "n'a pas travaille",
    "non formé",
    "non forme",
]

_PAS_FORME_TEXT_FIELDS = (
    "commentaires",
    "recommandations",
    "q4_gestion_administrative",
    "q5_gestion_financiere",
    "q6_communication",
)


def _pas_forme_q_filter():
    """Retourne un Q combiné pour détecter les formateurs signalés comme non-formateurs."""
    combined = Q()
    for field in _PAS_FORME_TEXT_FIELDS:
        for kw in _PAS_FORME_KEYWORDS:
            combined |= Q(**{f"{field}__icontains": kw})
    return combined


def _build_filtered_formateurs_queryset(request):
    qs = AppelFormateur.objects.filter(is_active=True).select_related("locked_by")
    status_filter = (request.GET.get("status") or "").strip()
    prestataire_filter = (request.GET.get("prestataire") or "").strip()
    beneficiaire_filter = (request.GET.get("beneficiaire") or "").strip()
    formation_filter = (request.GET.get("formation") or "").strip()
    cohorte_filter = (request.GET.get("cohorte") or "").strip()
    agent_filter = (request.GET.get("agent") or "").strip()
    session_date_filter = (request.GET.get("session_date") or "").strip()
    date_from_str = (request.GET.get("date_from") or "").strip()
    date_to_str = (request.GET.get("date_to") or "").strip()
    search = (request.GET.get("q") or "").strip()
    pas_forme_filter = request.GET.get("pas_forme") == "1"

    if status_filter:
        if status_filter == "completed":
            qs = qs.filter(status__in=CALL_COMPLETED_STATUSES)
        else:
            qs = qs.filter(status=status_filter)
    if prestataire_filter:
        qs = qs.filter(prestataire__iexact=prestataire_filter)
    if beneficiaire_filter:
        qs = qs.filter(beneficiaire__iexact=beneficiaire_filter)
    if formation_filter:
        qs = qs.filter(formation__iexact=formation_filter)
    if cohorte_filter:
        qs = qs.filter(cohorte__iexact=cohorte_filter)
    if agent_filter:
        qs = qs.filter(locked_by__username__iexact=agent_filter)
    if session_date_filter:
        try:
            qs = qs.filter(session_date=datetime.date.fromisoformat(session_date_filter))
        except ValueError:
            pass
    if search:
        qs = qs.filter(
            Q(telephone__icontains=search)
            | Q(prestataire__icontains=search)
            | Q(beneficiaire__icontains=search)
            | Q(formation__icontains=search)
            | Q(lieu__icontains=search)
            | Q(reference_code__icontains=search)
            | Q(commentaires__icontains=search)
            | Q(recommandations__icontains=search)
            | Q(q4_gestion_administrative__icontains=search)
            | Q(q5_gestion_financiere__icontains=search)
            | Q(q6_communication__icontains=search)
        )
    if pas_forme_filter:
        qs = qs.filter(_pas_forme_q_filter())
    if date_from_str:
        try:
            qs = qs.filter(created_at__gte=datetime.datetime.fromisoformat(date_from_str))
        except ValueError:
            pass
    if date_to_str:
        try:
            qs = qs.filter(created_at__lte=datetime.datetime.fromisoformat(date_to_str))
        except ValueError:
            pass

    filters = {
        "status": status_filter,
        "prestataire": prestataire_filter,
        "beneficiaire": beneficiaire_filter,
        "formation": formation_filter,
        "cohorte": cohorte_filter,
        "agent": agent_filter,
        "session_date": session_date_filter,
        "date_from": date_from_str,
        "date_to": date_to_str,
        "q": search,
        "pas_forme": pas_forme_filter,
        "prestataires": sorted(
            {
                v.strip()
                for v in qs.exclude(prestataire="").values_list("prestataire", flat=True)
                if v
            }
        ),
        "beneficiaires": sorted(
            {
                v.strip()
                for v in qs.exclude(beneficiaire="").values_list("beneficiaire", flat=True)
                if v
            }
        ),
        "formations": sorted(
            {v.strip() for v in qs.exclude(formation="").values_list("formation", flat=True) if v}
        ),
        "cohortes": sorted(
            {v.strip() for v in qs.exclude(cohorte="").values_list("cohorte", flat=True) if v}
        ),
        "session_dates": [
            value.isoformat()
            for value in qs.exclude(session_date__isnull=True)
            .values_list("session_date", flat=True)
            .distinct()
            .order_by("session_date")
        ],
        "agents": sorted(
            {
                u.strip()
                for u in qs.exclude(locked_by__isnull=True).values_list(
                    "locked_by__username", flat=True
                )
                if u
            }
        ),
    }
    return qs, filters


def _parse_note_1_5(value):
    text = _as_text(value)
    if not text:
        return None
    try:
        note = int(float(text))
    except (TypeError, ValueError):
        return None
    return note if 1 <= note <= 5 else None


def _clean_open_answer(value):
    return _as_text(value)


def _parse_session_time(value):
    text = _as_text(value)
    if not text:
        return None
    normalized = text.replace("h", ":").replace("H", ":").strip()
    for fmt in ("%H:%M", "%H:%M:%S", "%H%M"):
        try:
            return datetime.datetime.strptime(normalized, fmt).time()
        except ValueError:
            continue
    return None


def _resolve_classe_for_formateur_row(row: AppelFormateur):
    qs = Classe.objects.select_related(
        "formateur",
        "prestation__prestataire",
        "prestation__beneficiaire",
    ).filter(formateur__isnull=False)

    if row.prestataire:
        qs = qs.filter(prestation__prestataire__raison_sociale__iexact=row.prestataire)
    if row.beneficiaire:
        qs = qs.filter(prestation__beneficiaire__nom_structure__iexact=row.beneficiaire)
    if row.cohorte and str(row.cohorte).strip().isdigit():
        qs = qs.filter(cohorte=int(str(row.cohorte).strip()))

    target_phone = _normalize_phone_number(row.telephone)
    candidates = list(qs)
    if target_phone:
        phone_matches = [
            item
            for item in candidates
            if _normalize_phone_number(getattr(item.formateur, "telephone", "")) == target_phone
        ]
        if phone_matches:
            return phone_matches[0]

    if candidates:
        return candidates[0]

    if target_phone:
        fallback = [
            item
            for item in Classe.objects.select_related("formateur").filter(formateur__isnull=False)
            if _normalize_phone_number(getattr(item.formateur, "telephone", "")) == target_phone
        ]
        if fallback:
            return fallback[0]
    return None


def _sync_satisfaction_from_formateur_row(row: AppelFormateur, user):
    if not (
        row.q1_prerequis_apprenants
        and row.q2_interaction_apprenants
        and row.q3_competences_acquises
    ):
        return False

    classe = _resolve_classe_for_formateur_row(row)
    if not classe or not classe.formateur:
        return False

    date_value = row.session_date or timezone.localdate()
    heure_value = _parse_session_time(row.heure_debut)

    defaults = {
        "inspecteur": None,
        "enqueteur": user if getattr(user, "is_authenticated", False) else None,
        "heure": heure_value,
        "q1_prerequis_apprenants": row.q1_prerequis_apprenants,
        "q2_interaction_apprenants": row.q2_interaction_apprenants,
        "q3_competences_acquises": row.q3_competences_acquises,
        "q4_gestion_administrative": row.q4_gestion_administrative or "",
        "q5_gestion_financiere": row.q5_gestion_financiere or "",
        "q6_communication": row.q6_communication or "",
        "commentaires": row.commentaires or "",
        "recommandations": row.recommandations or "",
    }

    obj, _created = SatisfactionFormateur.objects.update_or_create(
        classe=classe,
        formateur=classe.formateur,
        date=date_value,
        defaults=defaults,
    )
    if row.audio_file and not obj.audio_appel:
        obj.audio_appel = row.audio_file
        obj.save(update_fields=["audio_appel", "updated_at"])
    return True


FORMATEUR_SCORE_FIELDS = (
    "q1_prerequis_apprenants",
    "q2_interaction_apprenants",
    "q3_competences_acquises",
)

FORMATEUR_TEXT_FIELDS = (
    "q4_gestion_administrative",
    "q5_gestion_financiere",
    "q6_communication",
    "commentaires",
    "recommandations",
)


def _collect_formateur_survey_updates(request, row: AppelFormateur) -> tuple[dict, bool]:
    updates = {}
    survey_posted = False

    for field_name in FORMATEUR_SCORE_FIELDS:
        raw_value = request.POST.get(field_name)
        if raw_value is None:
            updates[field_name] = getattr(row, field_name)
            continue
        survey_posted = True
        updates[field_name] = _parse_note_1_5(raw_value)

    for field_name in FORMATEUR_TEXT_FIELDS:
        raw_value = request.POST.get(field_name)
        if raw_value is None:
            updates[field_name] = getattr(row, field_name)
            continue
        survey_posted = True
        updates[field_name] = _clean_open_answer(raw_value)

    return updates, survey_posted


@login_required
def formateurs_export_xlsx(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calendrier"
    ws.append(
        [
            "Reference",
            "Numero seance",
            "Prestataire",
            "Beneficiaire",
            "Formation",
            "Lieu",
            "Telephone",
            "Source contact",
            "Cohorte",
            "Date label",
            "Date ISO",
            "Heure de debut",
            "Heure de fin",
            "Statut",
            "Rappel at",
            "Locked by",
            "Locked at",
            "Audio file",
            "Q1 prerequis apprenants",
            "Q2 interaction apprenants",
            "Q3 competences acquises",
            "Q4 gestion administrative",
            "Q5 gestion financiere",
            "Q6 communication",
            "Commentaires",
            "Recommandations",
            "Satisfaction completed at",
            "Created at",
            "Updated at",
        ]
    )
    for row in AppelFormateur.objects.select_related("locked_by").order_by(
        "session_date", "numero_seance", "telephone"
    ):
        ws.append(
            [
                row.reference_code,
                row.numero_seance,
                row.prestataire,
                row.beneficiaire,
                row.formation,
                row.lieu,
                row.telephone,
                row.source_contact,
                row.cohorte,
                row.date_label,
                row.session_date.isoformat() if row.session_date else "",
                row.heure_debut,
                row.heure_fin,
                row.status,
                row.rappel_at.isoformat() if row.rappel_at else "",
                row.locked_by.username if row.locked_by else "",
                row.locked_at.isoformat() if row.locked_at else "",
                _safe_audio_url(row),
                row.q1_prerequis_apprenants,
                row.q2_interaction_apprenants,
                row.q3_competences_acquises,
                row.q4_gestion_administrative,
                row.q5_gestion_financiere,
                row.q6_communication,
                row.commentaires,
                row.recommandations,
                row.satisfaction_completed_at.isoformat() if row.satisfaction_completed_at else "",
                row.created_at.isoformat() if getattr(row, "created_at", None) else "",
                row.updated_at.isoformat() if getattr(row, "updated_at", None) else "",
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="appels-formateurs-export.xlsx"'
    wb.save(response)
    return response


@login_required
def formateurs_export_filtered_csv(request):
    qs, _ = _build_filtered_formateurs_queryset(request)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="appels-formateurs-filtres.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(
        [
            "Reference",
            "Numero seance",
            "Prestataire",
            "Beneficiaire",
            "Formation",
            "Lieu",
            "Telephone",
            "Cohorte",
            "Date label",
            "Date ISO",
            "Heure debut",
            "Heure fin",
            "Statut",
            "Agent",
            "Rappel at",
            "Audio URL",
            "Q1 prerequis apprenants",
            "Q2 interaction apprenants",
            "Q3 competences acquises",
            "Q4 gestion administrative",
            "Q5 gestion financiere",
            "Q6 communication",
            "Commentaires",
            "Recommandations",
            "Satisfaction completed at",
            "Created at",
            "Updated at",
        ]
    )
    for row in qs.order_by("status", "session_date", "numero_seance", "telephone").iterator(
        chunk_size=2000
    ):
        writer.writerow(
            [
                row.reference_code,
                row.numero_seance or "",
                row.prestataire,
                row.beneficiaire,
                row.formation,
                row.lieu,
                row.telephone,
                row.cohorte,
                row.date_label,
                row.session_date.isoformat() if row.session_date else "",
                row.heure_debut,
                row.heure_fin,
                row.get_status_display(),
                row.locked_by.username if row.locked_by else "",
                row.rappel_at.isoformat() if row.rappel_at else "",
                _safe_audio_url(row),
                row.q1_prerequis_apprenants,
                row.q2_interaction_apprenants,
                row.q3_competences_acquises,
                row.q4_gestion_administrative,
                row.q5_gestion_financiere,
                row.q6_communication,
                row.commentaires,
                row.recommandations,
                row.satisfaction_completed_at.isoformat() if row.satisfaction_completed_at else "",
                row.created_at.isoformat() if getattr(row, "created_at", None) else "",
                row.updated_at.isoformat() if getattr(row, "updated_at", None) else "",
            ]
        )
    return response


@login_required
def formateurs_export_filtered_xlsx(request):
    """Export XLSX des appels formateurs filtrés selon les filtres appliqués."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    
    qs, _ = _build_filtered_formateurs_queryset(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appels Formateurs Filtrés"
    
    # En-têtes
    headers = [
        "Reference",
        "Numero seance",
        "Prestataire",
        "Beneficiaire",
        "Formation",
        "Lieu",
        "Telephone",
        "Cohorte",
        "Date label",
        "Date ISO",
        "Heure debut",
        "Heure fin",
        "Statut",
        "Agent",
        "Rappel at",
        "Audio URL",
        "Q1 prerequis apprenants",
        "Q2 interaction apprenants",
        "Q3 competences acquises",
        "Q4 gestion administrative",
        "Q5 gestion financiere",
        "Q6 communication",
        "Commentaires",
        "Recommandations",
        "Satisfaction completed at",
        "Created at",
        "Updated at",
    ]
    ws.append(headers)
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Données
    for row in qs.order_by("status", "session_date", "numero_seance", "telephone").iterator(chunk_size=2000):
        ws.append([
            row.reference_code,
            row.numero_seance or "",
            row.prestataire,
            row.beneficiaire,
            row.formation,
            row.lieu,
            row.telephone,
            row.cohorte,
            row.date_label,
            row.session_date.isoformat() if row.session_date else "",
            row.heure_debut,
            row.heure_fin,
            row.get_status_display(),
            row.locked_by.username if row.locked_by else "",
            row.rappel_at.isoformat() if row.rappel_at else "",
            _safe_audio_url(row),
            row.q1_prerequis_apprenants,
            row.q2_interaction_apprenants,
            row.q3_competences_acquises,
            row.q4_gestion_administrative,
            row.q5_gestion_financiere,
            row.q6_communication,
            row.commentaires,
            row.recommandations,
            row.satisfaction_completed_at.isoformat() if row.satisfaction_completed_at else "",
            row.created_at.isoformat() if getattr(row, "created_at", None) else "",
            row.updated_at.isoformat() if getattr(row, "updated_at", None) else "",
        ])
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="appels-formateurs-filtres.xlsx"'
    wb.save(response)
    return response


@login_required
def formateurs_index(request):
    if request.method == "POST" and request.FILES.getlist("files"):
        if not request.user.is_superuser:
            messages.error(request, "Seul un superadmin peut importer des appels formateurs.")
            return redirect(request.path_info)
        mode = request.POST.get("update_mode", "replace")
        uploaded_files = request.FILES.getlist("files")
        try:
            payload = []
            files_without_numbers = []
            rows_without_numbers = 0
            for uploaded in uploaded_files:
                file_obj = io.BytesIO(uploaded.read())
                current_payload = list(_iter_formateur_excel_rows(file_obj))
                payload.extend(current_payload)
                if not current_payload:
                    files_without_numbers.append(uploaded.name)
            raw_count = len(payload)
            if mode == "replace":
                AppelFormateur.objects.update(is_active=False)
                created = 0
                updated = 0
                seen = set()
                duplicate_refs = 0
                for item in payload:
                    ref = item["reference_code"]
                    if not item.get("telephone"):
                        rows_without_numbers += 1
                        continue
                    if ref in seen:
                        duplicate_refs += 1
                        continue
                    seen.add(ref)
                    row = AppelFormateur.objects.filter(reference_code=ref).first()
                    if row:
                        for key, value in item.items():
                            setattr(row, key, value)
                        row.is_active = True
                        row.save()
                        updated += 1
                    else:
                        AppelFormateur.objects.create(**item, is_active=True)
                        created += 1
                messages.success(
                    request,
                    (
                        f"Import formateurs termine. {len(uploaded_files)} fichier(s) traite(s), "
                        f"{created} ligne(s) creee(s), {updated} mise(s) a jour depuis Calendrier."
                    ),
                )
                removed_duplicates = _deactivate_duplicate_rows(
                    AppelFormateur.objects.filter(is_active=True),
                    _formateur_duplicate_key,
                )
            elif mode == "append":
                created = 0
                updated = 0
                seen = set()
                duplicate_refs = 0
                for item in payload:
                    ref = item["reference_code"]
                    if not item.get("telephone"):
                        rows_without_numbers += 1
                        continue
                    if ref in seen:
                        duplicate_refs += 1
                        continue
                    seen.add(ref)
                    row = AppelFormateur.objects.filter(reference_code=ref).first()
                    if row:
                        for key, value in item.items():
                            setattr(row, key, value)
                        row.is_active = True
                        row.save()
                        updated += 1
                    else:
                        AppelFormateur.objects.create(**item, is_active=True)
                        created += 1
                messages.success(
                    request,
                    (
                        f"Import formateurs termine. {len(uploaded_files)} fichier(s) traite(s), "
                        f"{created} nouvelle(s) ligne(s) ajoutee(s), {updated} mise(s) a jour."
                    ),
                )
                removed_duplicates = _deactivate_duplicate_rows(
                    AppelFormateur.objects.filter(is_active=True),
                    _formateur_duplicate_key,
                )
            else:
                messages.error(request, "Mode d'import formateurs inconnu.")
                removed_duplicates = 0
                duplicate_refs = 0

            skipped_count = rows_without_numbers + duplicate_refs
            if skipped_count or removed_duplicates or files_without_numbers:
                warning_parts = []
                if rows_without_numbers:
                    warning_parts.append(
                        f"{rows_without_numbers} ligne(s) sans numero exploitable ignoree(s)"
                    )
                if duplicate_refs:
                    warning_parts.append(f"{duplicate_refs} doublon(s) detecte(s) pendant l'upload")
                if removed_duplicates:
                    warning_parts.append(
                        f"{removed_duplicates} doublon(s) supplementaire(s) "
                        f"desactive(s) apres import"
                    )
                if files_without_numbers:
                    warning_parts.append(
                        (
                            "fichier(s) sans contact exploitable: "
                            + ", ".join(files_without_numbers[:5])
                        )
                    )
                messages.warning(request, "Import formateurs: " + " ; ".join(warning_parts) + ".")
            if raw_count == 0:
                messages.warning(
                    request,
                    "Aucun numero de telephone n'a ete pris dans les fichiers importes. "
                    "Verifiez la colonne CONTACT DU FORMATEUR.",
                )
        except Exception as exc:
            messages.error(request, f"Impossible de lire la feuille Calendrier : {exc}")
        return redirect(request.path_info)

    qs, filters = _build_filtered_formateurs_queryset(request)
    total_count = qs.count()
    stats = _build_formateur_progress_metrics(qs)
    satisfaction_header_metrics = _build_formateur_satisfaction_header_metrics(qs)

    try:
        page_size = int(request.GET.get("page_size") or PAGE_SIZE_DEFAULT)
    except ValueError:
        page_size = PAGE_SIZE_DEFAULT
    page_size = max(25, min(page_size, PAGE_SIZE_MAX))
    paginator = Paginator(
        qs.order_by("status", "session_date", "numero_seance", "telephone"), page_size
    )
    page_obj = paginator.get_page(request.GET.get("page"))
    rows = _bind_audio_state(list(page_obj.object_list))
    _pas_forme_ids = set(
        AppelFormateur.objects.filter(_pas_forme_q_filter()).values_list("id", flat=True)
    )
    for row in rows:
        row.formateur_nom = resolve_formateur_db_name_from_values(
            getattr(row, "telephone", ""),
            getattr(row, "source_contact", ""),
        )
        row.pas_forme_detecte = row.id in _pas_forme_ids
    page_obj.object_list = rows
    params = request.GET.copy()
    params.pop("page", None)
    querystring_no_page = params.urlencode()

    return render(
        request,
        "appels/formateurs.html",
        {
            "rows": rows,
            "rows_count": total_count,
            "filters": filters,
            "page_obj": page_obj,
            "page_size": page_size,
            "querystring_no_page": querystring_no_page,
            "stats": stats,
            "satisfaction_header_metrics": satisfaction_header_metrics,
        },
    )


@login_required
@require_POST
def formateurs_delete_missing_phones(request):
    if not request.user.is_superuser:
        messages.error(request, "Seul un superadmin peut supprimer les lignes sans telephone.")
        return redirect(request.META.get("HTTP_REFERER") or "/appels-formateurs/")
    deleted = (
        AppelFormateur.objects.filter(is_active=True)
        .filter(Q(telephone="") | Q(telephone__isnull=True))
        .update(is_active=False)
    )
    messages.success(request, f"{deleted} ligne(s) sans numero de telephone ont ete desactivees.")
    return redirect(request.META.get("HTTP_REFERER") or "/appels-formateurs/")


@login_required
@require_POST
def formateur_action(request, pk: int):
    # Clean up stale locks that may be blocking access
    _cleanup_stale_locks()

    row = get_object_or_404(AppelFormateur, pk=pk)
    action = request.POST.get("action")
    rappel_at = request.POST.get("rappel_at")

    now = timezone.now()
    survey_saved = False
    synced_to_padesce = False
    survey_updates, survey_posted = _collect_formateur_survey_updates(request, row)
    update_fields = ["updated_at"]

    if action in ("start", "update_form", "terminer", "rappeler") and survey_posted:
        for field_name, value in survey_updates.items():
            setattr(row, field_name, value)
        update_fields.extend([*FORMATEUR_SCORE_FIELDS, *FORMATEUR_TEXT_FIELDS])
        survey_saved = True

    if action in ("terminer", "rappeler"):
        for field_name, value in survey_updates.items():
            setattr(row, field_name, value)
        if any(getattr(row, field_name) is None for field_name in FORMATEUR_SCORE_FIELDS):
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Renseignez d'abord Q1, Q2 et Q3 via Demarrer ou en cliquant "
                    "sur la ligne du formateur.",
                },
                status=400,
            )
        row.satisfaction_completed_at = now
        if "satisfaction_completed_at" not in update_fields:
            update_fields.append("satisfaction_completed_at")
        survey_saved = True

    if action == "start":
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = now
        update_fields.extend(["status", "locked_by", "locked_at"])
    elif action == "pause":
        row.status = "pause"
        update_fields.append("status")
    elif action == "resume":
        row.status = "en_cours"
        row.locked_by = request.user
        row.locked_at = now
        update_fields.extend(["status", "locked_by", "locked_at"])
    elif action == "rappeler":
        row.status = "a_rappeler"
        row.locked_by = request.user
        row.locked_at = now
        update_fields.extend(["status", "locked_by", "locked_at"])
        if rappel_at:
            try:
                row.rappel_at = datetime.datetime.fromisoformat(rappel_at)
            except ValueError:
                row.rappel_at = None
        else:
            row.rappel_at = None
        update_fields.append("rappel_at")
    elif action == "terminer":
        row.status = "termine"
        update_fields.append("status")
    elif action == "update_form":
        pass
    else:
        return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)

    row.save(update_fields=list(dict.fromkeys(update_fields)))
    sync_formateur_status(row)

    if survey_saved and (
        action in ("terminer", "rappeler")
        or (
            action == "update_form"
            and row.status in {"termine", "a_rappeler"}
            and all(getattr(row, field_name) is not None for field_name in FORMATEUR_SCORE_FIELDS)
        )
    ):
        try:
            synced_to_padesce = _sync_satisfaction_from_formateur_row(row, request.user)
        except Exception:
            synced_to_padesce = False

    return JsonResponse(
        {
            "ok": True,
            "status": row.status,
            "status_label": row.get_status_display(),
            "locked_by": row.locked_by.username if row.locked_by else "",
            "rappel_at": row.rappel_at.isoformat() if row.rappel_at else "",
            "survey_saved": survey_saved,
            "synced_to_padesce": synced_to_padesce,
        }
    )


@login_required
@require_POST
def formateur_upload_audio(request, pk: int):
    row = get_object_or_404(AppelFormateur, pk=pk)
    file_obj = request.FILES.get("audio")
    if not file_obj:
        return JsonResponse({"ok": False, "error": "Aucun fichier audio."}, status=400)
    row.audio_file = file_obj
    row.save(update_fields=["audio_file", "updated_at"])
    sync_formateur_status(row)
    return JsonResponse(
        {
            "ok": True,
            "audio_url": _safe_audio_url(row),
            "status": row.status,
            "status_label": row.get_status_display(),
        }
    )


@login_required
@require_POST
def download_formateur_audios(request):
    ids = request.POST.getlist("ids")
    if not ids:
        return JsonResponse({"ok": False, "error": "Aucun appel selectionne."}, status=400)
    try:
        ids = [int(value) for value in ids]
    except ValueError:
        return JsonResponse({"ok": False, "error": "Identifiants invalides."}, status=400)

    rows = list(
        AppelFormateur.objects.filter(pk__in=ids, audio_file__isnull=False).order_by(
            "session_date", "numero_seance", "telephone"
        )
    )
    if not rows:
        return JsonResponse(
            {"ok": False, "error": "Aucun audio disponible pour la selection."}, status=404
        )

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        written = 0
        for row in rows:
            if not _has_audio_file(row):
                continue
            try:
                with row.audio_file.open("rb") as audio:
                    suffix = Path(row.audio_file.name).suffix or ".mp3"
                    safe_name = (
                        f"{slugify(row.reference_code) or 'session'}-"
                        f"{slugify(row.telephone) or 'telephone'}{suffix}"
                    )
                    archive.writestr(safe_name, audio.read())
                    written += 1
            except Exception:
                continue
        if written == 0:
            return JsonResponse({"ok": False, "error": "Pas d'audio recuperable."}, status=404)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="appels-formateurs-audios.zip"'
    return response


@login_required
@require_POST
def formateur_transcription_detail(request, pk: int):
    row = get_object_or_404(AppelFormateur, pk=pk)
    try:
        obj, generated = _ensure_formateur_transcription(row)
        return JsonResponse(
            {"ok": True, "generated": generated, "transcription": _transcription_to_payload(obj)}
        )
    except Exception as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=400)


@login_required
def formateur_transcription_download(request, pk: int):
    row = get_object_or_404(AppelFormateur, pk=pk)
    try:
        obj, _generated = _ensure_formateur_transcription(row)
    except Exception as exc:
        return HttpResponse(str(exc), status=400, content_type="text/plain; charset=utf-8")
    response = HttpResponse(obj.transcription_text or "", content_type="text/plain; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="transcription-formateur-'
        f'{slugify(row.reference_code) or row.pk}.txt"'
    )
    return response


@login_required
@require_POST
def start_filtered_formateurs_transcription(request):
    qs, _ = _build_filtered_formateurs_queryset(request)
    jobs = _collect_jobs_from_formateurs(qs)
    ok, payload, status_code = _start_transcription_task(
        FORMATEURS_FILTERED_TRANSCRIPTION_TASK_KEY,
        jobs,
        "Transcription du tableau filtre formateurs",
    )
    return JsonResponse(payload, status=status_code)


@login_required
def filtered_formateurs_transcription_status(request):
    if "_transcription_status_response" not in globals():
        return JsonResponse(
            {
                "ok": True,
                "status": {
                    "state": "idle",
                    "message": "Transcription groupée indisponible sur cette version.",
                    "total": 0,
                    "processed": 0,
                },
            }
        )
    return JsonResponse(
        {
            "ok": True,
            "status": _transcription_status_response(FORMATEURS_FILTERED_TRANSCRIPTION_TASK_KEY),
        }
    )


@login_required
@require_POST
def stop_filtered_formateurs_transcription(request):
    ok, payload, status_code = _stop_transcription_task(FORMATEURS_FILTERED_TRANSCRIPTION_TASK_KEY)
    return JsonResponse(payload, status=status_code)
