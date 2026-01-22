import datetime
import io
import zipfile
from decimal import Decimal
from pathlib import Path

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Case, DecimalField, ExpressionWrapper, F, When
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from App_PADESCE.appels.models import Appel
from App_PADESCE.formations.models import Classe


def _normalize_header(value):
    return (value or "").strip().lower()


def _parse_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    if "Feuil2" not in wb.sheetnames:
        raise ValueError("Feuil2 introuvable dans le fichier.")
    ws = wb["Feuil2"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    def get(row, key):
        idx = header_map.get(_normalize_header(key))
        if idx is None or idx >= len(row):
            return ""
        return row[idx] or ""

    data = []
    for row in rows:
        nom = get(row, "Nom")
        code = get(row, "Code")
        if not nom or not code:
            continue
        prestataire = get(row, "Prestataire")
        beneficiaire = get(row, "Beneficiaire")
        lieu = get(row, "Lieux")
        classe_label = get(row, "Classe")
        taux_presence = get(row, "Taux de presence") or 0
        tel1 = get(row, "1er No tél 0 Tel No")
        tel2 = get(row, "2e No tél 0 Tel No")
        type_formation = get(row, "Type de formation declarée")
        formation_padesce = get(row, "Formation Padesce")
        try:
            taux_presence = Decimal(str(taux_presence))
        except Exception:
            taux_presence = Decimal("0")
        pct = Decimal("0")
        try:
            pct = taux_presence * 100
        except Exception:
            pct = Decimal("0")
        data.append(
            {
                "code": str(code).strip(),
                "nom": str(nom).strip(),
                "prestataire": str(prestataire).strip(),
                "beneficiaire": str(beneficiaire).strip(),
                "lieu": str(lieu).strip(),
                "classe_label": str(classe_label).strip(),
                "taux_presence": pct,
                "telephone1": str(tel1).strip() if tel1 not in (None, "") else "",
                "telephone2": str(tel2).strip() if tel2 not in (None, "") else "",
                "type_formation_declaree": str(type_formation).strip(),
                "formation_padesce": str(formation_padesce).strip(),
            }
        )
    return data


def _parse_bool_flag(value):
    if value is None:
        return None
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


@login_required
@transaction.atomic
def appels_index(request):
    if request.method == "POST" and request.FILES.get("file"):
        f = request.FILES["file"]
        mode = request.POST.get("update_mode", "replace")
        try:
            payload = _parse_excel(io.BytesIO(f.read()))
        except Exception as exc:
            messages.error(request, f"Impossible de lire le fichier : {exc}")
            return redirect(request.path_info)

        if mode == "replace":
            Appel.objects.all().delete()
            created = 0
            for item in payload:
                classe_obj = None
                if item["classe_label"]:
                    classe_obj = Classe.objects.filter(code=item["classe_label"]).first()
                Appel.objects.create(
                    **item,
                    classe=classe_obj,
                )
                created += 1
            messages.success(request, f"Fichier importe. {created} ligne(s) enregistree(s). Anciennes donnees supprimees.")
        else:
            updated = 0
            for item in payload:
                code = item.get("code")
                if not code:
                    continue
                appel = Appel.objects.filter(code=code.strip()).first()
                if not appel:
                    continue
                appel.type_formation_declaree = str(item.get("type_formation_declaree") or "").strip()
                appel.formation_padesce = str(item.get("formation_padesce") or "").strip()
                appel.save(update_fields=["type_formation_declaree", "formation_padesce", "updated_at"])
                updated += 1
            messages.success(
                request,
                f"Fichier importe. {updated} appel(s) mis a jour avec le type de formation et la formation Padesce."
            )
        return redirect(request.path_info)

    appels_qs = Appel.objects.all()
    status_filter = request.GET.get("status") or ""
    prestataire_filter = request.GET.get("prestataire") or ""
    beneficiaire_filter = request.GET.get("beneficiaire") or ""
    classe_filter = request.GET.get("classe") or ""
    taux_filter = request.GET.get("taux_min", "").strip()
    date_from_str = request.GET.get("date_from", "").strip()
    date_to_str = request.GET.get("date_to", "").strip()
    search = request.GET.get("q", "").strip()

    if status_filter:
        appels_qs = appels_qs.filter(status=status_filter)
    if prestataire_filter:
        appels_qs = appels_qs.filter(prestataire__icontains=prestataire_filter)
    if beneficiaire_filter:
        appels_qs = appels_qs.filter(beneficiaire__icontains=beneficiaire_filter)
    if classe_filter:
        appels_qs = appels_qs.filter(classe_label__icontains=classe_filter)
    if taux_filter:
        try:
            seuil = Decimal(taux_filter)
            appels_qs = appels_qs.filter(taux_presence__gte=seuil)
        except Exception:
            pass
    if search:
        appels_qs = appels_qs.filter(nom__icontains=search)
    if date_from_str:
        try:
            date_from = datetime.datetime.fromisoformat(date_from_str)
            appels_qs = appels_qs.filter(created_at__gte=date_from)
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.datetime.fromisoformat(date_to_str)
            appels_qs = appels_qs.filter(created_at__lte=date_to)
        except ValueError:
            pass

    appels_qs = appels_qs.annotate(
        taux_presence_display=Case(
            When(
                taux_presence__lte=1,
                then=ExpressionWrapper(F("taux_presence") * 100, output_field=DecimalField(max_digits=7, decimal_places=2)),
            ),
            default=F("taux_presence"),
            output_field=DecimalField(max_digits=7, decimal_places=2),
        )
    )

    appels_count = appels_qs.count()
    appels = appels_qs.order_by("status", "nom")
    filters = {
        "status": status_filter,
        "prestataire": prestataire_filter,
        "beneficiaire": beneficiaire_filter,
        "classe": classe_filter,
        "q": search,
        "prestataires": sorted(
            {p.strip() for p in Appel.objects.exclude(prestataire="").values_list("prestataire", flat=True) if p}
        ),
        "beneficiaires": sorted(
            {b.strip() for b in Appel.objects.exclude(beneficiaire="").values_list("beneficiaire", flat=True) if b}
        ),
        "classes": Appel.objects.exclude(classe_label="").values_list("classe_label", flat=True).distinct(),
        "taux_min": taux_filter,
        "date_from": date_from_str,
        "date_to": date_to_str,
    }
    return render(request, "appels/index.html", {"appels": appels, "filters": filters, "appels_count": appels_count})


@login_required
@require_POST
def appel_action(request, pk: int):
    appel = get_object_or_404(Appel, pk=pk)
    action = request.POST.get("action")
    rappel_at = request.POST.get("rappel_at")
    deja_forme_flag = _parse_bool_flag(request.POST.get("deja_forme"))

    if appel.locked_by and appel.locked_by != request.user and appel.status in ("en_cours", "pause"):
        return JsonResponse({"ok": False, "error": "Appel en cours par un autre utilisateur."}, status=409)

    now = timezone.now()
    if action == "start":
        appel.status = "en_cours"
        appel.locked_by = request.user
        appel.locked_at = now
    elif action == "pause":
        appel.status = "pause"
    elif action == "resume":
        appel.status = "en_cours"
        appel.locked_by = request.user
        appel.locked_at = now
    elif action == "rappeler":
        appel.status = "a_rappeler"
        appel.locked_by = request.user
        appel.locked_at = now
        if rappel_at:
            try:
                appel.rappel_at = datetime.datetime.fromisoformat(rappel_at)
            except ValueError:
                appel.rappel_at = None
    elif action == "terminer":
        appel.status = "termine"
    else:
        return JsonResponse({"ok": False, "error": "Action inconnue."}, status=400)

    if deja_forme_flag is not None:
        appel.deja_forme = deja_forme_flag

    appel.save(update_fields=["status", "locked_by", "locked_at", "rappel_at", "deja_forme", "updated_at"])
    return JsonResponse(
        {
            "ok": True,
            "status": appel.status,
            "status_label": appel.get_status_display(),
            "locked_by": appel.locked_by.username if appel.locked_by else "",
            "rappel_at": appel.rappel_at.isoformat() if appel.rappel_at else "",
        }
    )


@login_required
@require_POST
def appel_upload_audio(request, pk: int):
    appel = get_object_or_404(Appel, pk=pk)
    file_obj = request.FILES.get("audio")
    if not file_obj:
        return JsonResponse({"ok": False, "error": "Aucun fichier audio."}, status=400)
    appel.audio_file = file_obj
    appel.save(update_fields=["audio_file", "updated_at"])
    return JsonResponse({"ok": True, "audio_url": appel.audio_file.url if appel.audio_file else ""})


@login_required
@require_POST
def download_appel_audios(request):
    ids = request.POST.getlist("ids")
    if not ids:
        return JsonResponse({"ok": False, "error": "Aucun appel sélectionné."}, status=400)
    try:
        ids = [int(val) for val in ids]
    except ValueError:
        return JsonResponse({"ok": False, "error": "Identifiants invalides."}, status=400)

    appels = list(
        Appel.objects.filter(pk__in=ids, audio_file__isnull=False)
        .order_by("nom")
    )
    if not appels:
        return JsonResponse({"ok": False, "error": "Aucun audio disponible pour les appels sélectionnés."}, status=404)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        written = 0
        for appel in appels:
            if not appel.audio_file:
                continue
            try:
                with appel.audio_file.open("rb") as audio:
                    suffix = Path(appel.audio_file.name).suffix or ".mp3"
                    safe_name = f"{slugify(appel.code) or 'code'}-{slugify(appel.nom) or 'appel'}{suffix}"
                    archive.writestr(safe_name, audio.read())
                    written += 1
            except Exception:
                continue
        if written == 0:
            return JsonResponse({"ok": False, "error": "Pas d'audio récupérable."}, status=404)

    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="appels-audios.zip"'
    return response
