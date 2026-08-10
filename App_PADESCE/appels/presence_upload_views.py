import io
import logging
import re

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.shortcuts import redirect
from django.views.decorators.http import require_POST

from App_PADESCE.appels.models import Appel
from App_PADESCE.apprenants.models import Apprenant
from App_PADESCE.formations.models import Classe, Prestation

logger = logging.getLogger(__name__)


def _normalize_header(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[-9:] if len(digits) > 9 else digits


def _presence_marker(value):
    val = str(value or "").strip().upper()
    if val == "PR":
        return "PR"
    if val == "AB":
        return "AB"
    return ""


def _next_p_app_counter():
    last = (
        Apprenant.objects.filter(code__startswith="P-APP")
        .order_by("-code")
        .values_list("code", flat=True)
        .first()
    )
    if not last:
        return 1
    digits = re.sub(r"\D", "", last.replace("P-APP", "", 1))
    try:
        return int(digits) + 1
    except (ValueError, TypeError):
        return 1


def parse_presence_file(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Le fichier ne contient aucune feuille.")
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None) or []
    lookup = {_normalize_header(v): i for i, v in enumerate(headers)}

    required = [
        ("nom complet", "Nom complet"),
        ("statut", "Statut"),
    ]
    missing = [label for key, label in required if key not in lookup]
    if missing:
        raise ValueError(f"Colonne(s) obligatoire(s) absente(s) : {', '.join(missing)}")

    def get(row, *keys):
        for key in keys:
            idx = lookup.get(_normalize_header(key))
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    presence_cols = []
    for i in range(1, 11):
        key = f"c{i}"
        if key in lookup:
            presence_cols.append((key, i))

    result = []
    for row in rows:
        nom = str(get(row, "nom complet") or "").strip()
        if not nom:
            continue
        statut = str(get(row, "statut") or "").strip()
        apprenant_id = str(get(row, "apprenant id") or "").strip() or None
        telephone = _phone(get(row, "numero de telephone", "telephone", "numero"))
        prestataire = str(get(row, "prestataire") or "").strip()
        beneficiaire = str(get(row, "beneficiaire", "beneficiaires") or "").strip()
        prestation_id = str(get(row, "prestation id") or "").strip()
        classe_code = str(get(row, "classe") or "").strip()

        presences = {}
        for key, num in presence_cols:
            presences[f"c{num}"] = _presence_marker(get(row, key))

        result.append(
            {
                "apprenant_id": apprenant_id,
                "nom_complet": nom,
                "telephone": telephone,
                "prestataire": prestataire,
                "beneficiaire": beneficiaire,
                "prestation_id": prestation_id,
                "classe_code": classe_code,
                "statut": statut,
                "presences": presences,
            }
        )
    return result


PRESENCE_FIELDS = [f"c{i}" for i in range(1, 11)]


def _find_or_create_classe(classe_code, prestation_id):
    if not classe_code:
        return None
    classe_obj = Classe.objects.filter(code=classe_code).first()
    if classe_obj:
        return classe_obj
    if not prestation_id:
        return None
    prestation = Prestation.objects.filter(code=prestation_id).first()
    if not prestation:
        return None
    sid = transaction.savepoint()
    try:
        classe_obj = Classe.objects.create(
            code=classe_code,
            prestation=prestation,
            formation=prestation.formation,
            intitule_formation=prestation.formation.nom if prestation.formation else "",
            actif=True,
        )
        transaction.savepoint_commit(sid)
        logger.info("Classe %s auto-créée via prestation %s", classe_code, prestation_id)
        return classe_obj
    except IntegrityError:
        transaction.savepoint_rollback(sid)
        return Classe.objects.filter(code=classe_code).first()


def _is_non_inscrit(item):
    return "non" in item["statut"].lower() or not item["apprenant_id"]


def _create_apprenant_for_non_inscrit(item, classe_obj, counter):
    code = f"P-APP{counter:04d}"
    tel = item["telephone"] or None

    existing = Apprenant.objects.filter(
        classe=classe_obj,
        nom_complet=item["nom_complet"],
    ).first()
    if existing:
        for field in PRESENCE_FIELDS:
            new_val = item["presences"].get(field, "")
            if new_val:
                setattr(existing, field, new_val)
        update_fields = [f for f in PRESENCE_FIELDS if item["presences"].get(f)]
        if update_fields:
            existing.save(update_fields=update_fields + ["updated_at"])
        return existing.code, "updated"

    sid = transaction.savepoint()
    try:
        Apprenant.objects.create(
            code=code,
            nom_complet=item["nom_complet"],
            classe=classe_obj,
            formation=classe_obj.formation,
            prestataire=item["prestataire"],
            beneficiaire=item["beneficiaire"],
            telephone1=tel,
            actif=True,
            **{k: v for k, v in item["presences"].items() if v},
        )
        transaction.savepoint_commit(sid)
        return code, "created"
    except IntegrityError as exc:
        transaction.savepoint_rollback(sid)
        if "unique_tel1_par_formation" in str(exc) and tel:
            sid2 = transaction.savepoint()
            try:
                Apprenant.objects.create(
                    code=code,
                    nom_complet=item["nom_complet"],
                    classe=classe_obj,
                    formation=classe_obj.formation,
                    prestataire=item["prestataire"],
                    beneficiaire=item["beneficiaire"],
                    telephone1=None,
                    actif=True,
                    **{k: v for k, v in item["presences"].items() if v},
                )
                transaction.savepoint_commit(sid2)
                return code, "created"
            except IntegrityError:
                transaction.savepoint_rollback(sid2)
        logger.warning(
            "Presence upload: impossible de creer %s pour %s: %s",
            code, item["nom_complet"], exc,
        )
        return None, f"{item['nom_complet']}: {exc}"


def _create_appel_for_non_inscrit(item, p_app_code, classe_obj):
    existing = Appel.objects.filter(code=p_app_code).first()
    if existing:
        return

    sid = transaction.savepoint()
    try:
        Appel.objects.create(
            code=p_app_code,
            nom=item["nom_complet"],
            prestataire=item["prestataire"],
            beneficiaire=item["beneficiaire"],
            classe_label=item["classe_code"],
            classe=classe_obj,
            telephone1=item["telephone"] or "",
            is_active=True,
            status="en_attente",
        )
        transaction.savepoint_commit(sid)
    except IntegrityError:
        transaction.savepoint_rollback(sid)


@login_required
@require_POST
@transaction.atomic
def upload_presence_list(request):
    if not request.user.is_superuser:
        messages.error(request, "Seul un superadmin peut importer une liste de présence.")
        return redirect("appels_index")

    uploaded = request.FILES.get("presence_file")
    if not uploaded:
        messages.error(request, "Aucun fichier sélectionné.")
        return redirect("appels_index")

    try:
        payload = parse_presence_file(io.BytesIO(uploaded.read()))
    except Exception as exc:
        messages.error(request, f"Impossible de lire le fichier : {exc}")
        return redirect("appels_index")

    if not payload:
        messages.warning(request, "Aucune ligne exploitable dans le fichier.")
        return redirect("appels_index")

    counter = _next_p_app_counter()
    created_ids = []
    updated_presence = 0
    not_found = 0
    skipped = 0
    appels_created = 0
    errors = []

    classes_cache = {}

    for item in payload:
        classe_code = item["classe_code"]
        prestation_id = item.get("prestation_id", "")
        if classe_code and classe_code not in classes_cache:
            classes_cache[classe_code] = _find_or_create_classe(classe_code, prestation_id)
        classe_obj = classes_cache.get(classe_code)

        if _is_non_inscrit(item):
            if not classe_obj:
                errors.append(f"Classe {classe_code} introuvable pour {item['nom_complet']}")
                skipped += 1
                continue

            result_code, status = _create_apprenant_for_non_inscrit(item, classe_obj, counter)
            if status == "created":
                created_ids.append(result_code)
                _create_appel_for_non_inscrit(item, result_code, classe_obj)
                appels_created += 1
                counter += 1
            elif status == "updated":
                updated_presence += 1
            else:
                errors.append(status)
                skipped += 1
        else:
            apprenant_code = item["apprenant_id"]
            if not apprenant_code:
                skipped += 1
                continue
            apprenant = Apprenant.objects.filter(code=apprenant_code).first()
            if not apprenant:
                if not classe_obj:
                    not_found += 1
                    continue
                tel = item["telephone"] or None
                sid = transaction.savepoint()
                try:
                    apprenant = Apprenant.objects.create(
                        code=apprenant_code,
                        nom_complet=item["nom_complet"],
                        classe=classe_obj,
                        formation=classe_obj.formation,
                        prestataire=item["prestataire"],
                        beneficiaire=item["beneficiaire"],
                        telephone1=tel,
                        actif=True,
                        **{k: v for k, v in item["presences"].items() if v},
                    )
                    transaction.savepoint_commit(sid)
                    updated_presence += 1
                    continue
                except IntegrityError:
                    transaction.savepoint_rollback(sid)
                    apprenant = Apprenant.objects.filter(
                        classe=classe_obj, nom_complet=item["nom_complet"]
                    ).first()
                    if not apprenant:
                        not_found += 1
                        continue

            changed = False
            for field in PRESENCE_FIELDS:
                new_val = item["presences"].get(field, "")
                if new_val and getattr(apprenant, field, "") != new_val:
                    setattr(apprenant, field, new_val)
                    changed = True
            if changed:
                apprenant.save(
                    update_fields=[f for f in PRESENCE_FIELDS if f in item["presences"]] + ["updated_at"]
                )
                updated_presence += 1

    parts = []
    if created_ids:
        first = created_ids[0]
        last = created_ids[-1]
        parts.append(f"{len(created_ids)} non-inscrit(s) créé(s) ({first} à {last})")
    if appels_created:
        parts.append(f"{appels_created} contact(s) ajouté(s) aux appels")
    if updated_presence:
        parts.append(f"{updated_presence} présence(s) mise(s) à jour")
    if not_found:
        parts.append(f"{not_found} apprenant(s) non trouvé(s) en base")
    if skipped:
        parts.append(f"{skipped} ligne(s) ignorée(s)")
    if not parts:
        parts.append("Aucune modification effectuée")

    messages.success(request, "Import présence terminé : " + ", ".join(parts) + ".")
    if errors:
        messages.warning(request, "Détails des erreurs : " + " | ".join(errors[:10]))
    return redirect("appels_index")
