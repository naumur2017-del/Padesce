import re
from functools import lru_cache
from pathlib import Path
from typing import Iterable

from django.conf import settings
from django.core.cache import cache
from django.db import OperationalError, ProgrammingError

from App_PADESCE.apprenants.models import Apprenant

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

PRESENCE_CONTROLS_CACHE_KEY = "presence_controls_v1"
PRESENCE_CONTROLS_DB_SYNC_TOKEN_KEY = "presence_controls_db_sync_token_v1"
VALID_MARKERS = {"PR", "AB"}
CONTROL_KEYS = ("c1", "c2", "c3", "c4", "c5", "c6")
EXCEL_CONTROLS_XLSX = "Fichier pour plateforme de satisfaction.xlsx"
EXCEL_CONTROLS_FALLBACK_XLSX = "fichier_concatene (1) 1 (1).xlsx"
EXCEL_CONTROLS_SHEET = "Rapport Presence"


def _auto_sync_enabled() -> bool:
    value = str(getattr(settings, "PADESCE_AUTO_SYNC_PRESENCE_CONTROLS", "") or "").strip()
    return value.lower() in {"1", "true", "yes", "on"}


def _normalize_identifier(value: str) -> str:
    return str(value or "").strip().upper()


def _normalized_marker(value: str) -> str:
    marker = _normalize_identifier(value)
    if marker in {"PRESENT", "PRESENCE", "PR", "PRÉSENT"}:
        return "PR"
    if marker in {"ABSENT", "AB"}:
        return "AB"
    return marker if marker in VALID_MARKERS else ""


def _normalize_control_type(value: str) -> str:
    normalized = _normalize_identifier(value)
    match = re.search(r"C([1-6])", normalized)
    if not match:
        return ""
    return f"c{match.group(1)}"


def _controls_xlsx_candidates() -> list[Path]:
    data_dir = Path(settings.BASE_DIR) / "data"
    return [
        data_dir / EXCEL_CONTROLS_XLSX,
        data_dir / EXCEL_CONTROLS_FALLBACK_XLSX,
    ]


def _controls_xlsx_path() -> Path:
    for candidate in _controls_xlsx_candidates():
        if candidate.exists():
            return candidate
    return _controls_xlsx_candidates()[0]


def _excel_cache_token() -> tuple[str, float]:
    path = _controls_xlsx_path()
    if not path.exists():
        return (str(path), 0.0)
    return (str(path), path.stat().st_mtime)


@lru_cache(maxsize=2)
def _load_excel_presence_payload(_token: tuple[str, float]) -> dict[str, dict]:
    if load_workbook is None:
        return {}
    path = Path(_token[0])
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        ws = wb[EXCEL_CONTROLS_SHEET] if EXCEL_CONTROLS_SHEET in wb.sheetnames else wb.active
        payload: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            apprenant_id = _normalize_identifier(row[0] if len(row) > 0 else "") or (
                _normalize_identifier(row[1] if len(row) > 1 else "")
            )
            if not apprenant_id:
                continue
            current = payload.setdefault(apprenant_id, {})
            if EXCEL_CONTROLS_SHEET in wb.sheetnames:
                for index, control_type in enumerate(CONTROL_KEYS, start=10):
                    presence_marker = _normalized_marker(row[index] if len(row) > index else "")
                    if presence_marker:
                        current[control_type] = presence_marker
            else:
                presence_marker = _normalized_marker(row[11] if len(row) > 11 else "")
                control_type = _normalize_control_type(row[39] if len(row) > 39 else "")
                if control_type and presence_marker:
                    current[control_type] = presence_marker
        return payload
    except Exception:
        return {}
    finally:
        wb.close()


def _presence_from_controls(controls: dict | None) -> dict:
    controls = controls or {}
    values = {key: _normalized_marker(controls.get(key)) for key in CONTROL_KEYS}
    completed_count = sum(1 for key in CONTROL_KEYS if values[key] in VALID_MARKERS)
    present_count = sum(1 for key in CONTROL_KEYS if values[key] == "PR")
    values["taux_presence"] = (
        round((present_count / completed_count) * 100, 2) if completed_count else 0
    )
    return values


def _has_known_controls(controls: dict | None) -> bool:
    controls = controls or {}
    return any(_normalized_marker(controls.get(key)) for key in CONTROL_KEYS)


def _with_presence_metadata(
    controls: dict,
    *,
    source: str,
    excel_found: bool,
    excel_controls_found: list[str],
) -> dict:
    payload = _presence_from_controls(controls)
    found = set(excel_controls_found)
    missing_controls = [key for key in CONTROL_KEYS if key not in found]
    payload.update(
        {
            "source": source,
            "excel_found": excel_found,
            "excel_complete": len(found) == len(CONTROL_KEYS),
            "excel_controls_found": excel_controls_found,
            "excel_missing_controls": missing_controls,
            **{f"{key}_from_excel": key in found for key in CONTROL_KEYS},
        }
    )
    return payload


def _get_excel_controls(apprenant_id: str) -> dict:
    key = _normalize_identifier(apprenant_id)
    if not key:
        return {}
    return _load_excel_presence_payload(_excel_cache_token()).get(key, {})


def _default_presence() -> dict:
    return _presence_from_controls({key: "" for key in CONTROL_KEYS})


def _sync_controls_from_excel_to_db(force: bool = False) -> None:
    if not force and not _auto_sync_enabled():
        return

    token = _excel_cache_token()
    token_key = f"{token[0]}::{token[1]}"
    if not force and cache.get(PRESENCE_CONTROLS_DB_SYNC_TOKEN_KEY) == token_key:
        return

    payload = _load_excel_presence_payload(token)
    try:
        apprenants = list(Apprenant.objects.only("id", "code", *CONTROL_KEYS))
    except (OperationalError, ProgrammingError):
        return

    to_update = []
    for apprenant in apprenants:
        controls = payload.get(_normalize_identifier(apprenant.code), {})
        normalized = _presence_from_controls(controls)
        has_changed = any(getattr(apprenant, key) != normalized[key] for key in CONTROL_KEYS)
        if not has_changed:
            continue
        for key in CONTROL_KEYS:
            setattr(apprenant, key, normalized[key])
        to_update.append(apprenant)

    try:
        if to_update:
            Apprenant.objects.bulk_update(to_update, list(CONTROL_KEYS), batch_size=1000)
        cache.set(PRESENCE_CONTROLS_DB_SYNC_TOKEN_KEY, token_key, timeout=None)
    except (OperationalError, ProgrammingError):
        return


def get_presence_controls(apprenant_id: str, fallback_seed: str = "") -> dict:
    del fallback_seed
    _sync_controls_from_excel_to_db(force=False)
    key = _normalize_identifier(apprenant_id)

    if key:
        try:
            apprenant = Apprenant.objects.only(*CONTROL_KEYS).filter(code__iexact=key).first()
        except (OperationalError, ProgrammingError):
            apprenant = None
        if apprenant is not None:
            db_controls = {
                control_key: getattr(apprenant, control_key) for control_key in CONTROL_KEYS
            }
            if _has_known_controls(db_controls):
                return _with_presence_metadata(
                    db_controls,
                    source="database",
                    excel_found=False,
                    excel_controls_found=[],
                )

            excel_controls = _get_excel_controls(key)
            if excel_controls:
                return _with_presence_metadata(
                    excel_controls,
                    source="excel_readonly",
                    excel_found=True,
                    excel_controls_found=[
                        control_key
                        for control_key in CONTROL_KEYS
                        if _normalized_marker(excel_controls.get(control_key))
                    ],
                )

    payload = cache.get(PRESENCE_CONTROLS_CACHE_KEY) or {}
    if key and key in payload:
        return _with_presence_metadata(
            payload[key],
            source="cache",
            excel_found=False,
            excel_controls_found=[],
        )

    excel_controls = _get_excel_controls(key)
    if excel_controls:
        return _with_presence_metadata(
            excel_controls,
            source="excel_readonly",
            excel_found=True,
            excel_controls_found=[
                control_key
                for control_key in CONTROL_KEYS
                if _normalized_marker(excel_controls.get(control_key))
            ],
        )

    return _with_presence_metadata(
        _default_presence(),
        source="default_ab",
        excel_found=False,
        excel_controls_found=[],
    )


def upsert_presence_controls(entries: Iterable[dict]) -> int:
    payload = cache.get(PRESENCE_CONTROLS_CACHE_KEY) or {}
    normalized_entries: dict[str, dict] = {}
    updated = 0
    for entry in entries:
        key = _normalize_identifier(entry.get("apprenant_id"))
        if not key:
            continue
        normalized = _presence_from_controls(
            {control_key: entry.get(control_key) for control_key in CONTROL_KEYS}
        )
        payload[key] = normalized
        normalized_entries[key] = normalized
        updated += 1
    cache.set(PRESENCE_CONTROLS_CACHE_KEY, payload, timeout=None)

    if normalized_entries:
        try:
            apprenants = list(Apprenant.objects.only("id", "code", *CONTROL_KEYS))
            to_update = []
            for apprenant in apprenants:
                key = _normalize_identifier(apprenant.code)
                entry = normalized_entries.get(key)
                if not entry:
                    continue
                has_changed = any(
                    getattr(apprenant, field) != entry[field] for field in CONTROL_KEYS
                )
                if not has_changed:
                    continue
                for field in CONTROL_KEYS:
                    setattr(apprenant, field, entry[field])
                to_update.append(apprenant)
            if to_update:
                Apprenant.objects.bulk_update(to_update, list(CONTROL_KEYS), batch_size=1000)
        except (OperationalError, ProgrammingError):
            pass

    return updated
