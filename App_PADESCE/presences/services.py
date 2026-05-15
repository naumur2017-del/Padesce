import csv
import io
from decimal import Decimal
from html import escape
from urllib.parse import quote

import requests
from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from App_PADESCE.core.apprenant_lookup import get_local_apprenant_identifier
from App_PADESCE.reporting.network_excel import (
    build_descente_channel_index,
    normalize_network_lookup,
)

GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"
GRAPH_TOKEN_URL_TEMPLATE = "https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"

PRESENCE_EXPORT_HEADERS = [
    "#",
    "Apprenant ID",
    "Nom complet",
    "Date",
    "Jour",
    "Heure",
    "Prestataire",
    "Bénéficiaire",
    "Inspecteur",
    "Classe",
    "ENQ",
    "Présence",
    "Statut",
    "Moyen",
    "Genre",
    "Âge",
    "Fonction",
    "Qualif.",
    "Exp.",
    "Type Form.",
    "Fenêtre",
    "Ville",
    "Arrondissement",
    "Département",
    "Région",
    "Lieux",
    "Téléphone",
    "Cohorte",
    "Code",
    "Prix Total",
    "",
    "Thème de la séance",
    "H. Début",
    "H. Fin",
    "Conformité",
    "Statut Appr.",
    "Classe Origine",
    "Type Contrôle",
    "Prestation ID",
    "Type de contrôle",
    "Durée Prevu de la formation",
    "Nom du formateur 1",
    "Telephone formateur1",
    "Nom de la formation 2",
    "Telephone du formateur 2",
    "Nom du formateur 3",
    "Telephone du formateur 3",
]


class TeamsSendError(RuntimeError):
    pass


def _time_text(value) -> str:
    if not value:
        return ""
    return value.strftime("%H:%M") if hasattr(value, "strftime") else str(value)


def _date_text(value) -> str:
    if not value:
        return ""
    return value.strftime("%d/%m/%Y") if hasattr(value, "strftime") else str(value)


def _money_total(prestation) -> str:
    if not prestation:
        return ""
    total = (prestation.montant_formation_psoaf_ttc or Decimal("0")) + (
        prestation.montant_mcdc_ttc or Decimal("0")
    )
    return str(total)


def presence_control_export_filename(control, extension: str = "csv") -> str:
    number = "".join(ch for ch in str(control.type_controle or "") if ch.isdigit()) or "1"
    date_part = (
        control.date.strftime("%Y%m%d") if control.date else timezone.now().strftime("%Y%m%d")
    )
    return f"Presence_{control.classe.code}_ENQ{number}_{date_part}.{extension}"


def presence_control_export_rows(control) -> list[list]:
    rows = []
    classe = control.classe
    prestation = getattr(classe, "prestation", None)
    beneficiaire = getattr(prestation, "beneficiaire", None) if prestation else None
    prestataire = getattr(prestation, "prestataire", None) if prestation else None
    lieu = getattr(classe, "lieu", None)
    inspecteur = control.inspecteur
    enqueteur = control.enqueteur

    presences = (
        control.presences.select_related("apprenant")
        .all()
        .order_by("apprenant__nom_complet", "apprenant__code")
    )
    for index, presence in enumerate(presences, start=1):
        apprenant = presence.apprenant
        rows.append(
            [
                index,
                get_local_apprenant_identifier(apprenant),
                apprenant.nom_complet,
                _date_text(control.date),
                control.jour,
                _time_text(presence.heure_presence),
                apprenant.prestataire or getattr(prestataire, "raison_sociale", ""),
                apprenant.beneficiaire or getattr(beneficiaire, "nom_structure", ""),
                getattr(inspecteur, "nom_complet", "") if inspecteur else "",
                classe.code,
                getattr(enqueteur, "username", "") if enqueteur else "",
                presence.presence,
                presence.get_statut_display(),
                presence.moyen_enregistrement,
                apprenant.genre,
                apprenant.age or "",
                apprenant.fonction,
                apprenant.qualification,
                apprenant.nb_annees_experience,
                apprenant.intitule_formation_dispensee or str(getattr(classe, "formation", "")),
                apprenant.fenetre or classe.fenetre,
                apprenant.ville_residence or getattr(lieu, "ville", ""),
                apprenant.arrondissement or getattr(lieu, "arrondissement", ""),
                apprenant.departement or getattr(lieu, "departement", ""),
                apprenant.region or getattr(lieu, "region", ""),
                apprenant.lieu_formation or getattr(lieu, "nom_lieu", ""),
                apprenant.telephone1 or apprenant.telephone2 or "",
                apprenant.cohorte or classe.cohorte,
                apprenant.code,
                _money_total(prestation),
                "",
                control.theme,
                _time_text(control.heure_debut),
                _time_text(control.heure_fin),
                control.get_conformite_display() if control.conformite else "",
                "Actif" if apprenant.actif else "Inactif",
                getattr(apprenant.classe, "code", ""),
                control.type_controle,
                getattr(prestation, "code", ""),
                control.type_controle,
                control.duree_prevue_formation or "",
                control.formateur1_nom,
                control.formateur1_telephone,
                control.formateur2_nom,
                control.formateur2_telephone,
                control.formateur3_nom,
                control.formateur3_telephone,
            ]
        )
    return rows


def build_presence_control_csv_bytes(control) -> bytes:
    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output, delimiter=";")
    writer.writerow(PRESENCE_EXPORT_HEADERS)
    writer.writerows(presence_control_export_rows(control))
    return output.getvalue().encode("utf-8")


def build_presence_control_xlsx_bytes(control) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Présences"
    ws.append(PRESENCE_EXPORT_HEADERS)
    for row in presence_control_export_rows(control):
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="0F8B8D")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 34)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _graph_token() -> str:
    tenant_id = settings.MICROSOFT_GRAPH_TENANT_ID
    client_id = settings.MICROSOFT_GRAPH_CLIENT_ID
    client_secret = settings.MICROSOFT_GRAPH_CLIENT_SECRET
    if not tenant_id or not client_id or not client_secret:
        raise TeamsSendError("Configuration Microsoft Graph incomplete.")

    response = requests.post(
        GRAPH_TOKEN_URL_TEMPLATE.format(tenant_id=tenant_id),
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "scope": "https://graph.microsoft.com/.default",
            "grant_type": "client_credentials",
        },
        timeout=30,
    )
    if response.status_code >= 400:
        detail = response.text[:300]
        try:
            payload = response.json()
        except ValueError:
            payload = {}
        description = str(payload.get("error_description") or "")
        if "AADSTS7000215" in description or "Invalid client secret" in description:
            detail = (
                "Secret Microsoft Graph invalide. Dans Azure, copiez la valeur du "
                "client secret dans MICROSOFT_GRAPH_CLIENT_SECRET, pas le Secret ID."
            )
        raise TeamsSendError(f"Token Graph refuse: {response.status_code} {detail}")
    token = response.json().get("access_token")
    if not token:
        raise TeamsSendError("Token Graph absent dans la reponse.")
    return token


def _resolve_teams_channel(classe) -> dict:
    channels = build_descente_channel_index()
    channel_info = channels.get(normalize_network_lookup(classe.code), {})
    team_id = channel_info.get("teams_team_id") or settings.MICROSOFT_TEAMS_DEFAULT_TEAM_ID
    channel_id = channel_info.get("teams_channel_id")
    if not channel_id and channel_info.get("teams_channels"):
        first = channel_info["teams_channels"][0]
        channel_id = first.get("channel_id")
        team_id = first.get("team_id") or team_id
    if not team_id or not channel_id:
        raise TeamsSendError(f"Canal Teams introuvable pour la classe {classe.code}.")
    return {"team_id": team_id, "channel_id": channel_id}


def _graph_headers(token: str, content_type: str | None = None) -> dict:
    headers = {"Authorization": f"Bearer {token}"}
    if content_type:
        headers["Content-Type"] = content_type
    return headers


def send_presence_control_csv_to_teams(control) -> dict:
    token = _graph_token()
    channel = _resolve_teams_channel(control.classe)
    encoded_channel_id = quote(channel["channel_id"], safe="")
    team_id = channel["team_id"]

    folder_response = requests.get(
        f"{GRAPH_BASE_URL}/teams/{team_id}/channels/{encoded_channel_id}/filesFolder",
        headers=_graph_headers(token),
        timeout=30,
    )
    if folder_response.status_code >= 400:
        raise TeamsSendError(
            f"Dossier Teams inaccessible: {folder_response.status_code} {folder_response.text[:300]}"
        )
    folder = folder_response.json()
    drive_id = (folder.get("parentReference") or {}).get("driveId")
    folder_id = folder.get("id")
    if not drive_id or not folder_id:
        raise TeamsSendError("Graph n'a pas retourne le dossier de fichiers du canal.")

    filename = presence_control_export_filename(control, "csv")
    upload_response = requests.put(
        f"{GRAPH_BASE_URL}/drives/{drive_id}/items/{folder_id}:/{quote(filename)}:/content",
        headers=_graph_headers(token, "text/csv"),
        data=build_presence_control_csv_bytes(control),
        timeout=60,
    )
    if upload_response.status_code >= 400:
        raise TeamsSendError(
            f"Upload Teams refuse: {upload_response.status_code} {upload_response.text[:300]}"
        )
    drive_item = upload_response.json()
    file_url = drive_item.get("webUrl", "")
    message_status = "Fichier depose dans le canal Teams."

    message_html = (
        f"<p>Contrôle de présence {escape(control.type_controle)} - "
        f"{escape(control.classe.code)} du {escape(_date_text(control.date))}</p>"
    )
    if file_url:
        message_html += f'<p><a href="{escape(file_url)}">{escape(filename)}</a></p>'

    message_response = requests.post(
        f"{GRAPH_BASE_URL}/teams/{team_id}/channels/{encoded_channel_id}/messages",
        headers=_graph_headers(token, "application/json"),
        json={"subject": None, "body": {"contentType": "html", "content": message_html}},
        timeout=30,
    )
    if message_response.status_code >= 400:
        message_status = (
            "Fichier depose dans le canal Teams, mais publication du message refusee: "
            f"{message_response.status_code}"
        )

    control.mark_teams_sent(file_url=file_url, message=message_status)
    return {"filename": filename, "file_url": file_url, "message": message_status}
