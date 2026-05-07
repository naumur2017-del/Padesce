from datetime import date
from pathlib import Path
import unicodedata

from django.db import migrations, models


def _normalize(value: str) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).upper()


def _load_cutoff_terminated_prestations(base_dir: Path) -> set[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return set()

    candidates = [
        base_dir / "data" / "network_excel_bundle" / "network-fichier-consolide-cutoff.xlsm",
        base_dir / "data" / "network_excel_cache" / "network-fichier-consolide-cutoff.xlsm",
    ]
    workbook_path = next((path for path in candidates if path.exists()), None)
    if workbook_path is None:
        return set()

    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        if "Prestations" not in wb.sheetnames:
            return set()

        ws = wb["Prestations"]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        headers_norm = [_normalize(h) for h in headers]

        idx_id = -1
        for i, h in enumerate(headers_norm):
            if "ID PRESTATION" in h:
                idx_id = i
                break
        
        if idx_id == -1:
            return set()

        status_headers = {"STATUT DE LA PRESTATION", "STATUT AVEC TAUX", "STATUT"}
        idx_status = next((i for i, h in enumerate(headers_norm) if h in status_headers), None)
        
        terminated_keywords = ("TERMIN", "ARRET")
        presta_codes: set[str] = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_code = str(row[idx_id] or "").strip().upper()
            if not raw_code:
                continue
            
            is_terminated = False
            if idx_status is not None:
                raw_status = _normalize(row[idx_status] or "")
                if any(keyword in raw_status for keyword in terminated_keywords):
                    is_terminated = True
            
            # Si on ne trouve pas le statut, on l'ajoute quand même par sécurité pour Phase 1
            if idx_status is None or is_terminated:
                presta_codes.add(raw_code)
        return presta_codes
    except Exception:
        return set()


def _fix_phase_assignments(apps, schema_editor):
    Phase = apps.get_model("formations", "Phase")
    Prestation = apps.get_model("formations", "Prestation")
    Classe = apps.get_model("formations", "Classe")
    Formateur = apps.get_model("formations", "Formateur")
    Apprenant = apps.get_model("apprenants", "Apprenant")
    AppelFormateur = apps.get_model("appels", "AppelFormateur")

    phase1 = Phase.objects.filter(id_phase=1).first()
    phase2 = Phase.objects.filter(id_phase=2).first()
    
    if not phase1 or not phase2:
        return

    # 1. Reset baseline : Tout le monde en Phase 1 (Vague 1) par défaut
    Prestation.objects.all().update(phase=phase1)
    Classe.objects.all().update(phase=phase1)
    Apprenant.objects.all().update(phase=phase1)
    # Formateur FORMA001 reste sans phase (test), les autres en Phase 1
    Formateur.objects.exclude(code="FORMA001").update(phase=phase1)

    # 2. Identifier la Phase 2 basée sur les dates si possible
    # Toute donnée créée après le 1er Mars 2026 est potentiellement Phase 2
    cutoff_date = date(2026, 3, 1)
    
    # Apprenants via date de création (si disponible) ou via Classe
    # (On va plutôt se baser sur les prestations et classes d'abord)
    
    # 3. Tenter d'utiliser le fichier Excel Cutoff (chemin corrigé)
    base_dir = Path(__file__).resolve().parents[3]
    phase1_presta_codes = _load_cutoff_terminated_prestations(base_dir)

    if phase1_presta_codes:
        # Si on a le fichier, on peut être plus précis.
        # Tout ce qui n'est PAS dans le cutoff de Février est Phase 2.
        phase1_presta_ids = set(
            Prestation.objects.filter(code__in=phase1_presta_codes).values_list("id", flat=True)
        )
        Prestation.objects.exclude(id__in=phase1_presta_ids).update(phase=phase2)
    else:
        # Fallback sans fichier : Utiliser la date de création pour les nouvelles prestations
        # Note: TimeStampedModel ajoute created_at
        from django.utils import timezone
        import datetime
        dt_cutoff = timezone.make_aware(datetime.datetime(2026, 3, 1))
        Prestation.objects.filter(created_at__gte=dt_cutoff).update(phase=phase2)

    # Propager aux classes
    phase2_presta_ids = Prestation.objects.filter(phase=phase2).values_list("id", flat=True)
    Classe.objects.filter(prestation_id__in=phase2_presta_ids).update(phase=phase2)
    
    # Propager aux apprenants
    phase2_class_ids = Classe.objects.filter(phase=phase2).values_list("id", flat=True)
    Apprenant.objects.filter(classe_id__in=phase2_class_ids).update(phase=phase2)

    # Formateurs : se baser sur les appels récents
    # Si un formateur n'a fait des appels qu'après le cutoff, il est Phase 2
    recent_phones = set(
        AppelFormateur.objects.filter(session_date__gte=cutoff_date)
        .values_list("telephone", flat=True)
    )
    old_phones = set(
        AppelFormateur.objects.filter(session_date__lt=cutoff_date)
        .values_list("telephone", flat=True)
    )
    
    # Phase 2 = seulement ceux qui n'ont pas d'appels anciens
    phase2_only_phones = recent_phones - old_phones
    if phase2_only_phones:
        Formateur.objects.filter(telephone__in=phase2_only_phones).exclude(code="FORMA001").update(phase=phase2)


class Migration(migrations.Migration):

    dependencies = [
        ("formations", "0009_phase_and_phase_fk"),
        ("apprenants", "0010_apprenant_phase_and_seed"),
        ("appels", "0001_initial"), # Pour s'assurer que AppelFormateur est là
    ]

    operations = [
        migrations.RunPython(_fix_phase_assignments, migrations.RunPython.noop),
    ]
