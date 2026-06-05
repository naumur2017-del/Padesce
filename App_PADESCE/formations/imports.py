import hashlib
import re
import unicodedata
from dataclasses import dataclass
from decimal import Decimal

from django.db import IntegrityError, transaction

from App_PADESCE.apprenants.models import Apprenant
from App_PADESCE.formations.models import (
    Beneficiaire,
    Classe,
    Formation,
    Lieu,
    Phase,
    Prestataire,
    Prestation,
)


@dataclass
class ReferenceImportResult:
    formations_created: int = 0
    formations_updated: int = 0
    prestataires_created: int = 0
    prestataires_updated: int = 0
    beneficiaires_created: int = 0
    beneficiaires_updated: int = 0
    prestations_created: int = 0
    prestations_updated: int = 0
    lieux_created: int = 0
    lieux_updated: int = 0
    classes_created: int = 0
    classes_updated: int = 0
    apprenants_created: int = 0
    apprenants_updated: int = 0
    apprenants_skipped: int = 0

    @property
    def summary(self) -> str:
        return (
            f"Classes: {self.classes_created} creee(s), {self.classes_updated} mise(s) a jour. "
            f"Apprenants: {self.apprenants_created} cree(s), "
            f"{self.apprenants_updated} mis a jour, {self.apprenants_skipped} ignore(s)."
        )


def _normalize_header(value) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _clean(value) -> str:
    return str(value or "").strip()


def _as_int(value, default: int = 0) -> int:
    try:
        text = _clean(value)
        if not text or text.upper() in {"N/D", "ND"}:
            return default
        return int(Decimal(text))
    except Exception:
        return default


def _as_decimal(value) -> Decimal:
    try:
        text = _clean(value)
        if not text or text.upper() in {"N/D", "ND"}:
            return Decimal("0")
        return Decimal(text)
    except Exception:
        return Decimal("0")


def _status(value) -> str:
    text = _normalize_header(value)
    if "cours" in text:
        return "en_cours"
    if "termine" in text or "arret" in text:
        return "termine"
    return "non_demarre"


def _rows_by_header(workbook, sheet_name: str):
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header_map = {_normalize_header(value): idx for idx, value in enumerate(header)}

    def get(row, *keys):
        for key in keys:
            idx = header_map.get(_normalize_header(key))
            if idx is not None and idx < len(row):
                return row[idx]
        return ""

    parsed_rows = []
    for row in rows:
        if not row or not any(cell not in (None, "") for cell in row):
            continue
        parsed_rows.append((row, get))
    return parsed_rows


def _save_counter(obj, created: bool, created_attr: str, updated_attr: str, result):
    if created:
        setattr(result, created_attr, getattr(result, created_attr) + 1)
    else:
        setattr(result, updated_attr, getattr(result, updated_attr) + 1)
    return obj


def _update_or_create(
    model,
    *,
    lookup: dict,
    defaults: dict,
    created_attr: str,
    updated_attr: str,
    result,
):
    obj, created = model.objects.update_or_create(**lookup, defaults=defaults)
    return _save_counter(obj, created, created_attr, updated_attr, result)


def _load_formations(workbook, result):
    for row, get in _rows_by_header(workbook, "Formations"):
        code = _clean(get(row, "FormationID"))
        name = _clean(get(row, "NomFormation"))
        if not code or not name:
            continue
        _update_or_create(
            Formation,
            lookup={"code": code},
            defaults={
                "nom": name,
                "nom_harmonise": _clean(get(row, "Appellation_harmonisee")),
                "statut": "termine",
                "actif": True,
            },
            created_attr="formations_created",
            updated_attr="formations_updated",
            result=result,
        )


def _load_prestataires(workbook, result):
    for row, get in _rows_by_header(workbook, "Prestataires"):
        code = _clean(get(row, "PrestataireID"))
        name = _clean(get(row, "NomPrestataire"))
        if not code or not name:
            continue
        _update_or_create(
            Prestataire,
            lookup={"code": code},
            defaults={
                "raison_sociale": name,
                "type_structure": _clean(get(row, "Nom Simplifier")),
                "telephone": _clean(get(row, "Telephone1")),
                "email": _clean(get(row, "Email")),
                "actif": True,
            },
            created_attr="prestataires_created",
            updated_attr="prestataires_updated",
            result=result,
        )


def _load_beneficiaires(workbook, result):
    for row, get in _rows_by_header(workbook, "Beneficiaires"):
        code = _clean(get(row, "BeneficiaireID"))
        name = _clean(get(row, "NomBeneficiaire"))
        pk = _beneficiaire_pk_from_code(code)
        if not code or not name or not pk:
            continue
        _update_or_create(
            Beneficiaire,
            lookup={"id": pk},
            defaults={
                "nom_structure": name,
                "type_structure": _clean(get(row, "Nom Simplifier")),
                "contact": _clean(get(row, "Telephone")),
                "email": _clean(get(row, "Email")),
                "ville": _clean(get(row, "Ville")),
                "departement": _clean(get(row, "Departement")),
                "region": _clean(get(row, "Region")),
                "actif": True,
            },
            created_attr="beneficiaires_created",
            updated_attr="beneficiaires_updated",
            result=result,
        )


def _beneficiaire_pk_from_code(code: str) -> int:
    digits = "".join(ch for ch in str(code or "") if ch.isdigit())
    return int(digits or 0) or None


def _beneficiaire_by_code(code: str, fallback_name: str = ""):
    pk = _beneficiaire_pk_from_code(code)
    if pk:
        found = Beneficiaire.objects.filter(pk=pk).first()
        if found:
            return found
    if fallback_name:
        return Beneficiaire.objects.filter(nom_structure__iexact=fallback_name).first()
    return None


def _load_lieux(workbook, result):
    for row, get in _rows_by_header(workbook, "Lieux"):
        code = _clean(get(row, "LieuID"))
        name = _clean(get(row, "NomLieu"))
        if not code or not name:
            continue
        _update_or_create(
            Lieu,
            lookup={"code": code},
            defaults={
                "nom_lieu": name,
                "ville": _clean(get(row, "Ville")),
                "precision": _clean(get(row, "Precisions", "Précisions")),
                "arrondissement": _clean(get(row, "Arrondissement")),
                "departement": _clean(get(row, "Departement")),
                "region": _clean(get(row, "Region")),
                "latitude": _clean(get(row, "GPS_Lat")),
                "longitude": _clean(get(row, "GPS_Lon")),
                "actif": True,
            },
            created_attr="lieux_created",
            updated_attr="lieux_updated",
            result=result,
        )


def _load_prestations(workbook, result, phase: Phase | None = None):
    for row, get in _rows_by_header(workbook, "Prestations"):
        code = _clean(get(row, "ID Prestation"))
        prestataire_code = _clean(get(row, "ID Prestataire"))
        formation_code = _clean(get(row, "ID Formation"))
        beneficiaire_code = _clean(get(row, "ID Beneficaire", "ID Bénéficaire"))
        prestataire = Prestataire.objects.filter(code=prestataire_code).first()
        formation = Formation.objects.filter(code=formation_code).first()
        if not code or not prestataire or not formation:
            continue
        _update_or_create(
            Prestation,
            lookup={"code": code},
            defaults={
                "prestataire": prestataire,
                "formation": formation,
                "beneficiaire": _beneficiaire_by_code(
                    beneficiaire_code,
                    _clean(get(row, "Nom du beneficiaire")),
                ),
                "effectif_a_former": _as_int(get(row, "Nombre D'apprenants Objectifs PADESCE")),
                "femmes": _as_int(get(row, "Objectif d'apprenants femme Inscrit")),
                "cout_unitaire_psoaf": _as_decimal(
                    get(row, "Cout unitaire affecte par la PSOAF")
                ),
                "montant_formation_psoaf_ttc": _as_decimal(
                    get(row, "Montant formation accorde par la PSOAF TTC")
                ),
                "cout_unitaire_mcdc_ttc": _as_decimal(
                    get(row, "Cout unitaire Subvention MCDC TTC")
                ),
                "montant_mcdc_ttc": _as_decimal(get(row, "Montant Total Subvention MCDC TTC")),
                "ville": _clean(get(row, "Region")),
                "actif": "annul" not in _normalize_header(get(row, "Statut de la prestation")),
                "phase": phase,
            },
            created_attr="prestations_created",
            updated_attr="prestations_updated",
            result=result,
        )


def _lieu_for_class(name: str, ville: str = ""):
    if not name:
        return None
    found = Lieu.objects.filter(nom_lieu__iexact=name).first()
    if found:
        return found
    digest = hashlib.sha1(f"{name}|{ville}".encode("utf-8")).hexdigest()[:10].upper()
    code = f"LIE-AUTO-{digest}"
    return Lieu.objects.get_or_create(
        code=code,
        defaults={"nom_lieu": name, "ville": ville, "actif": True},
    )[0]


def _load_classes(workbook, result, phase: Phase | None = None):
    for row, get in _rows_by_header(workbook, "Classes"):
        code = _clean(get(row, "Classe ID"))
        prestation = Prestation.objects.filter(code=_clean(get(row, "Prestation ID"))).first()
        if not code or not prestation:
            continue
        _update_or_create(
            Classe,
            lookup={"code": code},
            defaults={
                "prestation": prestation,
                "lieu": _lieu_for_class(_clean(get(row, "Lieux")), _clean(get(row, "Ville"))),
                "formation": prestation.formation,
                "intitule_formation": _clean(get(row, "FORMATION")) or prestation.formation.nom,
                "fenetre": getattr(prestation, "fenetre", ""),
                "cohorte": _as_int(get(row, "Cohorte"), 1) or 1,
                "statut": _status(get(row, "Statut de la prestation")),
                "actif": "annul" not in _normalize_header(get(row, "Statut de la prestation")),
                "phase": phase,
            },
            created_attr="classes_created",
            updated_attr="classes_updated",
            result=result,
        )


def _load_apprenants(workbook, result, phase: Phase | None = None):
    for row, get in _rows_by_header(workbook, "Apprenants"):
        code = _clean(get(row, "ApprenantID"))
        name = _clean(get(row, "Nom_Individu"))
        classe = Classe.objects.filter(code=_clean(get(row, "Classe ID"))).first()
        if not code or not name or not classe:
            result.apprenants_skipped += 1
            continue
        defaults = {
            "numero": _clean(get(row, "N°", "N")),
            "classe": classe,
            "formation": classe.formation,
            "nom_complet": name,
            "beneficiaire": _clean(get(row, "Nom_Beneficiaire")),
            "genre": _clean(get(row, "Sexe")),
            "cohorte": _clean(get(row, "Cohorte")) or str(classe.cohorte),
            "fenetre": classe.fenetre,
            "prestataire": getattr(classe.prestation.prestataire, "raison_sociale", ""),
            "intitule_formation_dispensee": classe.intitule_formation,
            "ville_formation": getattr(classe.lieu, "ville", "") if classe.lieu else "",
            "lieu_formation": getattr(classe.lieu, "nom_lieu", "") if classe.lieu else "",
            "actif": _normalize_header(get(row, "Statut Apprenant")) != "inactif",
            "phase": phase,
        }
        try:
            obj, created = Apprenant.objects.update_or_create(code=code, defaults=defaults)
        except IntegrityError:
            obj = Apprenant.objects.filter(classe=classe, nom_complet=name).first()
            if not obj:
                result.apprenants_skipped += 1
                continue
            for key, value in defaults.items():
                setattr(obj, key, value)
            obj.save()
            created = False
        _save_counter(
            obj,
            created,
            "apprenants_created",
            "apprenants_updated",
            result,
        )


@transaction.atomic
def import_reference_workbook(file_obj, *, phase: Phase) -> ReferenceImportResult:
    import openpyxl

    if phase is None:
        raise ValueError("Choisissez une vague avant de lancer l'import.")

    workbook = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    result = ReferenceImportResult()
    _load_formations(workbook, result)
    _load_prestataires(workbook, result)
    _load_beneficiaires(workbook, result)
    _load_lieux(workbook, result)
    _load_prestations(workbook, result, phase)
    _load_classes(workbook, result, phase)
    _load_apprenants(workbook, result, phase)
    return result
