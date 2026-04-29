import unicodedata
from pathlib import Path
from typing import Any

from django.conf import settings

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def _workbook_candidates() -> list[Path]:
    base_dir = Path(settings.BASE_DIR)
    return [
        base_dir / "Decompte et facturation.xlsm",
        base_dir / "data" / "network_excel_cache" / "network-fichier-consolide.xlsm",
        base_dir / "data" / "network_excel_cache" / "network-fichier-consolide-cutoff.xlsm",
        base_dir / "data" / "network_excel_bundle" / "network-fichier-consolide-cutoff.xlsm",
    ]


def _normalize_label(value: Any) -> str:
    text = " ".join(str(value or "").strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _is_blank_indicator(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        cleaned = value.strip().upper()
        return not cleaned or cleaned in {"N/D", "ND", "-", "NAN"}
    return False


def _as_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if _is_blank_indicator(cleaned):
            return None
        cleaned = cleaned.replace("%", "").replace(",", ".")
        try:
            number = float(cleaned)
        except ValueError:
            return None
        return number / 100 if "%" in value or number > 1 else number
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number


def _as_count(value: Any) -> int | str:
    if _is_blank_indicator(value):
        return "-"
    number = _as_number(value)
    return int(round(number or 0))


def _find_column(labels: list[str], *needles: str, suffixes: tuple[str, ...] = ()) -> int | None:
    normalized_needles = tuple(_normalize_label(needle) for needle in needles)
    normalized_suffixes = tuple(_normalize_label(suffix) for suffix in suffixes)
    for index, label in enumerate(labels):
        normalized = _normalize_label(label)
        if not all(needle in normalized for needle in normalized_needles):
            continue
        if normalized_suffixes and not normalized.endswith(normalized_suffixes):
            continue
        return index
    return None


def _excel_displays_dash(value: Any, number_format: str = "") -> bool:
    if _is_blank_indicator(value):
        return True
    if value != 0:
        return False
    sections = str(number_format or "").split(";")
    return len(sections) >= 3 and '"-"' in sections[2]


def _read_decompte_global_row(
    prestation_code: str,
) -> tuple[list[str], tuple[Any, ...], tuple[str, ...]] | None:
    if load_workbook is None:
        return None

    target = str(prestation_code or "").strip().upper()
    if not target:
        return None

    for workbook in _workbook_candidates():
        if not workbook.exists():
            continue
        try:
            wb = load_workbook(workbook, read_only=False, data_only=True)
        except Exception:
            continue
        try:
            sheet_name = next(
                (name for name in wb.sheetnames if "decompte" in _normalize_label(name)),
                None,
            )
            if not sheet_name:
                continue
            ws = wb[sheet_name]
            header_top = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            header_bottom = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            labels = []
            current_group = ""
            for top, bottom in zip(header_top, header_bottom):
                if top is not None:
                    current_group = str(top).strip()
                parts = []
                if current_group:
                    parts.append(current_group)
                if bottom is not None:
                    parts.append(str(bottom).strip())
                labels.append(" ".join(parts))
            prestation_index = _find_column(labels, "prestation id")
            if prestation_index is None:
                continue
            for cells in ws.iter_rows(min_row=3):
                row = tuple(cell.value for cell in cells)
                if str(row[prestation_index] or "").strip().upper() == target:
                    excel_row = cells[0].row
                    formats = (
                        tuple(
                            ws.cell(excel_row, column=index + 1).number_format
                            for index in range(len(labels))
                        )
                        if excel_row
                        else tuple("" for _ in labels)
                    )
                    return labels, row, formats
        finally:
            wb.close()
    return None


def get_prestation_indicators_from_db(prestation_code: str) -> dict[str, float]:
    """
    Recupere les indicateurs de la feuille Decompte Global pour une prestation.
    Les taux sont retournes sous forme de ratio: 0.86 = 86%.
    """
    payload = _read_decompte_global_row(prestation_code)
    if payload is None:
        return {
            "total_participants": 0,
            "taux_personnes_formees": 0,
            "taux_participation": 0,
            "taux_presence_globale": 0,
        }

    labels, row, formats = payload
    participants_index = _find_column(labels, "projection sur personnes suivie")

    available_index = _find_column(
        labels,
        "apprenants suivis avec",
        suffixes=(" total", " t"),
    )
    formed_rate_index = _find_column(labels, "taux", "personnes form")
    presence_rate_index = _find_column(labels, "taux", "presence moyen")

    participants_raw = row[participants_index] if participants_index is not None else None
    available_raw = row[available_index] if available_index is not None else None
    participants = (
        "-"
        if participants_index is None
        or _excel_displays_dash(participants_raw, formats[participants_index])
        else _as_count(participants_raw)
    )
    available = _as_number(available_raw) if available_index is not None else None
    formed_raw = row[formed_rate_index] if formed_rate_index is not None else None
    presence_raw = row[presence_rate_index] if presence_rate_index is not None else None
    formed_rate = _as_number(formed_raw)
    presence_rate = _as_number(presence_raw)
    participation_rate = (
        "-"
        if participants == "-" or _is_blank_indicator(available_raw)
        else ((participants / available) if available and available > 0 else 0)
    )

    taux_personnes_formees = (
        "-" if _is_blank_indicator(formed_raw) or participants == "-" else (formed_rate or 0)
    )

    return {
        "total_participants": participants,
        "taux_personnes_formees": taux_personnes_formees,
        "taux_participation": participation_rate,
        "taux_presence_globale": "-" if _is_blank_indicator(presence_raw) else (presence_rate or 0),
    }


def get_participant_count_for_prestation(prestation_code: str) -> int:
    indicators = get_prestation_indicators_from_db(prestation_code)
    value = indicators.get("total_participants")
    return 0 if value == "-" else int(value or 0)
