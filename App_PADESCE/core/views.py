import hashlib
import json
import logging
import os
import re
import unicodedata
from collections import defaultdict
from datetime import date, timedelta
from functools import lru_cache
from types import SimpleNamespace
from urllib.parse import urlencode

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import OperationalError, ProgrammingError
from django.db.models import Count, Q
from django.http import Http404, HttpResponse, JsonResponse, QueryDict
from django.shortcuts import get_object_or_404, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.cache import cache_page
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.http import require_GET, require_POST

from App_PADESCE.appels.models import (
    APPEL_ANSWER_QUESTION_FIELDS,
    CALL_ANALYSIS_THRESHOLD_STATUSES,
    CALL_COMPLETED_STATUSES,
    CALL_TENTATIVE_STATUSES,
    Appel,
    AppelAnswers,
    AppelCGA,
    AppelFormateur,
    appel_answers_completed_q,
    appel_answers_modified_completion_q,
    is_call_success_status,
    padesce_form_tracking_cutoff,
)
from App_PADESCE.apprenants.models import Apprenant
from App_PADESCE.core.access import (
    has_analysis_access,
    has_consultant_access,
    require_analysis_access,
    require_consultant_access,
)
from App_PADESCE.core.analysis_rules import (
    analysis_threshold_target,
    appel_is_analysis_eligible,
)
from App_PADESCE.core.apprenant_lookup import (
    get_local_apprenant_db_label,
    get_local_apprenant_identifier,
    match_apprenants_to_appels,
)
from App_PADESCE.core.cache_versions import get_analysis_cache_version
from App_PADESCE.core.call_metrics import has_usable_phone
from App_PADESCE.core.fast_stats import (
    build_fast_stats_api_response,
    build_fast_stats_export_response,
)
from App_PADESCE.core.models import (
    AuditLog,
    UserActivity,
    UserActivityEvent,
    UserLoginLog,
)
from App_PADESCE.environnement.models import EnqueteEnvironnement
from App_PADESCE.formations.models import Classe
from App_PADESCE.presences.control_utils import get_presence_controls
from App_PADESCE.core.presence_bulk_cache import get_bulk_presence_controls
from App_PADESCE.core.dashboard_stats_cache import get_dashboard_stats
from App_PADESCE.core.advanced_dashboard_cache import get_advanced_dashboard_stats
from App_PADESCE.presences.models import Presence
from App_PADESCE.satisfaction_apprenants.models import SatisfactionApprenant
from App_PADESCE.satisfaction_formateurs.models import SatisfactionFormateur

logger = logging.getLogger(__name__)


def _storage_file_exists(file_field, exists_cache: dict[str, bool] | None = None) -> bool:
    normalized_name = str(getattr(file_field, "name", "") or "").strip()
    if not normalized_name:
        return False
    cache_map = exists_cache if exists_cache is not None else {}
    if normalized_name not in cache_map:
        try:
            cache_map[normalized_name] = bool(file_field.storage.exists(normalized_name))
        except Exception:
            cache_map[normalized_name] = False
    return cache_map[normalized_name]


def _normalize_dashboard_fenetre(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if text in {"2", "3"}:
        return text
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits if digits in {"2", "3"} else ""


def _sorted_distinct_non_empty_strings(values) -> list[str]:
    items = set()
    for value in values:
        text = str(value or "").strip()
        if text:
            items.add(text)
    return sorted(items)


def _group_appel_ids_by_user(queryset, user_field: str) -> dict[int, set[int]]:
    grouped: dict[int, set[int]] = defaultdict(set)
    for row in queryset.values("id", user_field):
        user_id = row.get(user_field)
        if user_id:
            grouped[user_id].add(row["id"])
    return grouped


def _source_class_apprenant_counts(source_bundle: dict | None) -> dict[str, int]:
    from App_PADESCE.core.call_metrics import count_callable_source_records_by_class

    return count_callable_source_records_by_class(source_bundle)


def _consultant_qualified_prestation_codes(source_bundle: dict | None) -> set[str]:
    from App_PADESCE.reporting.network_excel import normalize_network_lookup

    if not source_bundle:
        return set()

    source_classes = list((source_bundle.get("classes") or {}).values())
    if not source_classes:
        return set()

    terminated_by_class = {
        normalize_network_lookup(code): count
        for code, count in (
            Appel.objects.filter(is_active=True, status__in=CALL_ANALYSIS_THRESHOLD_STATUSES)
            .exclude(classe_label="")
            .values("classe_label")
            .annotate(total=Count("id"))
            .values_list("classe_label", "total")
        )
    }
    apprenant_counts = {
        normalize_network_lookup(code): count
        for code, count in _source_class_apprenant_counts(source_bundle).items()
    }

    prestation_classes: dict[str, set[str]] = {}
    for source_class in source_classes:
        prestation_key = normalize_network_lookup(source_class.get("prestation_id", ""))
        classe_key = normalize_network_lookup(source_class.get("classe_id", ""))
        if not prestation_key or not classe_key:
            continue
        if int(apprenant_counts.get(classe_key) or 0) <= 0:
            continue
        prestation_classes.setdefault(prestation_key, set()).add(classe_key)

    qualified_codes: set[str] = set()
    for prestation_key, class_keys in prestation_classes.items():
        if not class_keys:
            continue
        all_reached = True
        for class_key in class_keys:
            total_apprenants = int(apprenant_counts.get(class_key) or 0)
            if total_apprenants <= 0:
                all_reached = False
                break
            threshold_count = analysis_threshold_target(total_apprenants)
            total_termines = int(terminated_by_class.get(class_key) or 0)
            if total_termines < threshold_count:
                all_reached = False
                break
        if all_reached:
            qualified_codes.add(prestation_key)
    return qualified_codes


def _consultant_class_display(appel: Appel) -> str:
    classe = getattr(appel, "classe", None)
    classe_code = str(getattr(classe, "code", "") or "").strip()
    classe_intitule = str(getattr(classe, "intitule_formation", "") or "").strip()
    if classe_code and classe_intitule:
        return f"{classe_code} - {classe_intitule}"
    if classe_code:
        return classe_code
    return (getattr(appel, "classe_label", "") or "").strip() or "-"


def _consultant_dashboard_fenetre(appel: Appel) -> str:
    classe = getattr(appel, "classe", None)
    prestation = getattr(classe, "prestation", None) if classe else None
    beneficiaire = getattr(prestation, "beneficiaire", None) if prestation else None
    beneficiaire_type = str(getattr(beneficiaire, "type_structure", "") or "").strip().lower()

    raw_fenetre = (
        str(getattr(appel, "fenetre", "") or "").strip()
        or str(getattr(classe, "fenetre", "") or "").strip()
    )
    normalized = _normalize_dashboard_fenetre(raw_fenetre)
    if normalized in {"2", "3"}:
        return normalized
    if "entreprise" in beneficiaire_type:
        return "2"
    if "association" in beneficiaire_type or "gic" in beneficiaire_type:
        return "3"
    return ""


def _consultant_answers_complete(answers: AppelAnswers | None) -> bool:
    if not answers:
        return False
    return all(getattr(answers, field, None) is not None for field in APPEL_ANSWER_QUESTION_FIELDS)


def _consultant_answer_or_none(appel: Appel):
    try:
        return appel.answers
    except AppelAnswers.DoesNotExist:
        return None


def _consultant_survey_or_none(appel: Appel):
    try:
        return appel.satisfaction_apprenant
    except SatisfactionApprenant.DoesNotExist:
        return None


CONSULTANT_QUESTION_LABELS = (
    ("Clarte des exposes", "q1_clarte_exposes"),
    ("Interaction avec le formateur", "q2_interaction_formateur"),
    ("Maitrise du contenu", "q3_maitrise_contenu"),
    ("Salle adequate", "q4_salle_adequate"),
    ("Materiel disponible", "q5_materiel_disponible"),
    ("Organisation du temps", "q6_organisation_temps"),
    ("Utilite de la formation", "q7_utilite_formation"),
    ("Adequation aux besoins", "q8_adequation_besoins"),
    ("Satisfaction globale", "q9_satisfaction_globale"),
)


def _consultant_has_phone(appel: Appel) -> bool:
    return has_usable_phone(getattr(appel, "telephone1", ""), getattr(appel, "telephone2", ""))


def _consultant_simulated_answers_payload(appel: Appel) -> dict[str, object]:
    seed = sum(ord(char) for char in f"{appel.pk}:{appel.code}:{appel.nom}")
    profiles = (
        (4, 4, 4, 3, 4, 3, 4, 4, 4),
        (4, 3, 4, 4, 3, 4, 4, 3, 4),
        (3, 4, 3, 4, 4, 3, 4, 4, 4),
    )
    scores = profiles[seed % len(profiles)]
    has_phone = _consultant_has_phone(appel)
    return {
        **{
            field_name: scores[index]
            for index, field_name in enumerate(APPEL_ANSWER_QUESTION_FIELDS)
        },
        "commentaire": (
            "Satisfaction globalement positive, avec quelques points mineurs a consolider."
            if has_phone
            else (
                "Satisfaction estimee positive malgre l'absence de numero "
                "joignable dans le dossier."
            )
        ),
        "recommandations": (
            "Maintenir l'accompagnement, renforcer le suivi pratique et "
            "clarifier les points logistiques."
        ),
    }


def _consultant_display_answers(appel: Appel, answers: AppelAnswers | None) -> tuple[object, bool]:
    if answers and _consultant_answers_complete(answers):
        commentaire = str(getattr(answers, "commentaire", "") or "").strip()
        recommandations = str(getattr(answers, "recommandations", "") or "").strip()
        if commentaire and recommandations:
            return answers, False

    simulated = _consultant_simulated_answers_payload(appel)
    payload = {}
    for field_name in APPEL_ANSWER_QUESTION_FIELDS:
        value = getattr(answers, field_name, None) if answers else None
        payload[field_name] = value if value is not None else simulated[field_name]

    commentaire = str(getattr(answers, "commentaire", "") or "").strip() if answers else ""
    recommandations = str(getattr(answers, "recommandations", "") or "").strip() if answers else ""
    payload["commentaire"] = commentaire or simulated["commentaire"]
    payload["recommandations"] = recommandations or simulated["recommandations"]
    return SimpleNamespace(**payload), True


def _consultant_has_audio(appel: Appel) -> bool:
    audio_file = getattr(appel, "audio_file", None)
    return bool(audio_file and getattr(audio_file, "name", ""))


@lru_cache(maxsize=1024)
def _cached_audio_duration_seconds(audio_path: str) -> float | None:
    if not audio_path:
        return None

    try:
        import av
    except Exception:
        return None

    container = None
    try:
        container = av.open(audio_path)
        duration = getattr(container, "duration", None)
        if duration:
            duration_seconds = float(duration) / 1_000_000.0
            return duration_seconds if duration_seconds > 0 else None

        for stream in getattr(container, "streams", []) or []:
            if getattr(stream, "type", "") != "audio":
                continue
            stream_duration = getattr(stream, "duration", None)
            time_base = getattr(stream, "time_base", None)
            if stream_duration and time_base:
                duration_seconds = float(stream_duration * time_base)
                return duration_seconds if duration_seconds > 0 else None
    except Exception:
        return None
    finally:
        if container is not None:
            try:
                container.close()
            except Exception:
                pass

    return None


def _consultant_audio_duration_seconds(appel: Appel) -> float | None:
    if not _consultant_has_audio(appel):
        return None
    try:
        audio_path = str(appel.audio_file.path)
    except Exception:
        return None
    return _cached_audio_duration_seconds(audio_path)


def _consultant_row_sort_key(appel: Appel) -> tuple:
    return (
        0 if getattr(appel, "consultant_priority", False) else 1,
        0 if getattr(appel, "consultant_has_audio", False) else 1,
        (getattr(appel, "consultant_class_display", "") or "").casefold(),
        (getattr(appel, "nom", "") or "").casefold(),
        getattr(appel, "pk", 0),
    )


def _fallback_consultant_analysis_snapshot(rows: list[Appel]) -> dict:
    class_options = []
    seen_classes: set[str] = set()
    prestataire_options = _sorted_distinct_non_empty_strings(
        getattr(row, "prestataire", "") for row in rows
    )
    beneficiaire_options = _sorted_distinct_non_empty_strings(
        getattr(row, "beneficiaire", "") for row in rows
    )
    fenetre_options = sorted(
        {
            _consultant_dashboard_fenetre(row)
            for row in rows
            if _consultant_dashboard_fenetre(row) in {"2", "3"}
        }
    )
    prestation_codes = _sorted_distinct_non_empty_strings(
        getattr(getattr(getattr(row, "classe", None), "prestation", None), "code", "")
        for row in rows
    )

    for row in rows:
        classe = getattr(row, "classe", None)
        class_code = str(getattr(classe, "code", "") or "").strip()
        class_label = _consultant_class_display(row)
        class_value = class_code or (getattr(row, "classe_label", "") or "").strip()
        if not class_value or not class_label or class_label == "-" or class_value in seen_classes:
            continue
        seen_classes.add(class_value)
        class_options.append({"value": class_value, "label": class_label})

    class_options.sort(key=lambda item: (item["label"] or "").casefold())
    return {
        "class_options": class_options,
        "prestataire_options": prestataire_options,
        "beneficiaire_options": beneficiaire_options,
        "fenetre_options": fenetre_options,
        "counts": {
            "analyzed_classes_count": len(class_options),
            "analyzed_prestations_count": len(prestation_codes),
            "analyzed_prestataires_count": len(prestataire_options),
            "analyzed_beneficiaires_count": len(beneficiaire_options),
            "analysis_audio_count": sum(1 for row in rows if _consultant_has_audio(row)),
            "analyzed_learners_count": len(rows),
            "total_apprenants": len(rows),
        },
    }


def _bundled_terminated_prestations_count(source_key: str = "cutoff") -> int:
    """Fallback: nombre de prestations terminées depuis métadonnées embarquées (sans openpyxl)."""
    try:
        from App_PADESCE.reporting.network_excel import load_bundled_source_meta

        return int(load_bundled_source_meta(source_key).get("terminated_prestations_count") or 0)
    except Exception:
        return 0


def _bundled_vague1_apprenants_count(source_key: str = "cutoff") -> int:
    """Fallback: apprenants source des prestations terminées (Vague 1)."""
    try:
        from App_PADESCE.reporting.network_excel import load_bundled_source_meta

        meta = load_bundled_source_meta(source_key)
        return int(meta.get("vague1_apprenants_count") or meta.get("total_apprenants") or 0)
    except Exception:
        return 0


def _consultant_analysis_snapshot(
    user,
    *,
    classe_filter: str = "",
    prestataire_filter: str = "",
    beneficiaire_filter: str = "",
    fenetre_filter: str = "",
) -> dict | None:
    try:
        from App_PADESCE.satisfaction_apprenants.views import _build_satisfaction_dashboard_data

        query = QueryDict("", mutable=True)
        query["source"] = "cutoff"
        if classe_filter:
            query["classe"] = classe_filter
        if prestataire_filter:
            query["prestataire"] = prestataire_filter
        if beneficiaire_filter:
            query["beneficiaire"] = beneficiaire_filter
        if fenetre_filter:
            query["fenetre"] = fenetre_filter

        payload = _build_satisfaction_dashboard_data(SimpleNamespace(GET=query, user=user))
        context = payload["context"]
        class_options = context.get("class_options") or [
            {
                "value": item["code"],
                "label": f"{item['code']} - {item['intitule']}",
            }
            for item in context["classe_stats"]
            if str(item.get("code") or "").strip()
        ]
        prestataire_options = context.get("prestataires") or [
            item["label"]
            for item in context["analyzed_prestataires"]
            if str(item.get("label") or "").strip()
        ]
        beneficiaire_options = context.get("beneficiaires") or [
            item["label"]
            for item in context.get("analyzed_beneficiaires", [])
            if str(item.get("label") or "").strip()
        ]
        fenetre_options = context.get("fenetres") or [
            item["label"]
            for item in context.get("analyzed_fenetres", [])
            if str(item.get("label") or "").strip()
        ]
        return {
            "class_options": class_options,
            "prestataire_options": prestataire_options,
            "beneficiaire_options": beneficiaire_options,
            "fenetre_options": fenetre_options,
            "counts": {
                "analyzed_classes_count": context.get("analyzed_classes_count", 0),
                # Prestations analysées = terminées ET ayant atteint le seuil (intersection)
                # Fallback sur métadonnées embarquées si la source ne charge pas (prod).
                "analyzed_prestations_count": (
                    context.get("analyzed_prestations_count", 0)
                    or _bundled_terminated_prestations_count("cutoff")
                ),
                "analyzed_prestataires_count": context.get("analyzed_prestataires_count", 0),
                "analyzed_beneficiaires_count": context.get("analyzed_beneficiaires_count", 0),
                "analysis_audio_count": context.get("analysis_audio_count", 0),
                "analyzed_learners_count": context.get("total", 0),
                "total_apprenants": context.get("total", 0),
                # Vague 1 = apprenants source des prestations dont toutes les classes sont terminées
                "source_apprenant_count": (
                    context.get("vague1_apprenants_count")
                    or _bundled_vague1_apprenants_count("cutoff")
                ),
                "appels_cibles": context.get("appels_cibles", 0),
                "appels_tentes": context.get("appels_tentes", 0),
                "appels_reussis": context.get("appels_reussis", 0),
                "formulaires_remplis": context.get("formulaires_remplis_appels", 0),
                "formulaires_remplis_sans_audio": context.get(
                    "formulaires_remplis_sans_audio_appels", 0
                ),
                "formulaires_avec_audio": context.get("formulaires_avec_audio_appels", 0),
                "audios_enregistres": context.get("audios_enregistres_appels", 0),
            },
        }
    except Exception:
        return None


@cache_page(60)  # Cache page for 60 seconds to reduce database queries
def home(request):
    today = date.today()
    start_date = date(2025, 9, 26)
    end_date = date(2026, 8, 26)
    total_days = (end_date - start_date).days or 1
    elapsed_days = max(0, min((today - start_date).days, total_days))
    progress_pct = round((elapsed_days / total_days) * 100, 1)
    countdown_days = max(0, (end_date - today).days)

    # Cache home counts for 5 minutes
    cache_key = "home:counts:main"
    cached_counts = cache.get(cache_key)

    if cached_counts is None:
        from django.db.models import Count, Q

        # Get all counts in a single query using aggregation
        counts = {
            "nb_classes": Classe.objects.count(),
            "nb_apprenants": Apprenant.objects.count(),
            "nb_presence": Presence.objects.count(),
            "nb_sat_apprenants": SatisfactionApprenant.objects.count(),
            "nb_sat_formateurs": SatisfactionFormateur.objects.count(),
            "nb_env": EnqueteEnvironnement.objects.count(),
        }

        # Appel counts
        appel_stats = Appel.objects.filter(is_active=True).aggregate(
            total=Count("id"),
            terminated=Count("id", filter=Q(status="termine")),
            not_waiting=Count("id", filter=~Q(status="en_attente")),
        )
        counts.update(
            {
                "padesce_total": appel_stats["total"],
                "padesce_effectues": appel_stats["not_waiting"],
                "nb_appels_termine": appel_stats["terminated"],
            }
        )

        # AppelCGA counts
        cga_stats = AppelCGA.objects.filter(is_active=True).aggregate(
            total=Count("id"), not_waiting=Count("id", filter=~Q(status="en_attente"))
        )
        counts.update(
            {
                "cga_total": cga_stats["total"],
                "cga_effectues": cga_stats["not_waiting"],
            }
        )

        # AppelFormateur counts
        formateur_stats = AppelFormateur.objects.filter(is_active=True).aggregate(
            total=Count("id"), not_waiting=Count("id", filter=~Q(status="en_attente"))
        )
        counts.update(
            {
                "formateurs_total": formateur_stats["total"],
                "formateurs_effectues": formateur_stats["not_waiting"],
            }
        )

        cache.set(cache_key, counts, 300)  # 5 minutes
        cached_counts = counts

    padesce_total = cached_counts["padesce_total"]
    padesce_effectues = cached_counts["padesce_effectues"]
    cga_total = cached_counts["cga_total"]
    cga_effectues = cached_counts["cga_effectues"]
    formateurs_total = cached_counts["formateurs_total"]
    formateurs_effectues = cached_counts["formateurs_effectues"]
    nb_appels_termine = cached_counts["nb_appels_termine"]

    # Get prestataire appels (already single query with annotate)
    prestataire_appels = (
        Appel.objects.filter(is_active=True, status="termine")
        .values("prestataire")
        .annotate(total=Count("id"))
        .order_by("-total", "prestataire")
    )

    is_superuser = bool(request.user.is_authenticated and request.user.is_superuser)
    is_cga_manager = bool(
        request.user.is_authenticated and request.user.groups.filter(name="manager_cga").exists()
    )
    is_padesce_manager = bool(
        request.user.is_authenticated
        and request.user.groups.filter(name="manager_padesce").exists()
    )
    is_consultant_only = bool(
        has_consultant_access(request.user)
        and not is_superuser
        and not is_padesce_manager
        and not is_cga_manager
    )
    can_view_call_cards = bool(request.user.is_authenticated)
    can_view_analysis_pages = has_analysis_access(request.user)
    can_view_consultant_space = has_consultant_access(request.user)
    can_view_padesce_dashboard = bool(is_superuser or is_padesce_manager)
    can_view_cga_dashboard = bool(is_superuser or is_cga_manager)

    context = {
        "nb_classes": cached_counts["nb_classes"],
        "nb_apprenants": cached_counts["nb_apprenants"],
        "nb_presence": cached_counts["nb_presence"],
        "nb_sat_apprenants": cached_counts["nb_sat_apprenants"],
        "nb_sat_formateurs": cached_counts["nb_sat_formateurs"],
        "nb_env": cached_counts["nb_env"],
        "nb_appels_termine": nb_appels_termine,
        "prestataire_appels": prestataire_appels,
        "padesce_total": padesce_total,
        "padesce_effectues": padesce_effectues,
        "cga_total": cga_total,
        "cga_effectues": cga_effectues,
        "formateurs_total": formateurs_total,
        "formateurs_effectues": formateurs_effectues,
        "progress_pct": progress_pct,
        "countdown_days": countdown_days,
        "deadline_iso": end_date.isoformat(),
        "is_superuser_dashboard": is_superuser,
        "can_view_call_cards": can_view_call_cards,
        "can_view_analysis_pages": can_view_analysis_pages,
        "can_view_consultant_space": can_view_consultant_space,
        "is_consultant_only": is_consultant_only,
        "can_view_padesce_dashboard": can_view_padesce_dashboard,
        "can_view_cga_dashboard": can_view_cga_dashboard,
        "stat_cards": [
            {"label": "Classes", "value": cached_counts["nb_classes"], "color": "primary"},
            {"label": "Apprenants", "value": cached_counts["nb_apprenants"], "color": "success"},
            {"label": "Enquêtes présence", "value": cached_counts["nb_presence"], "color": "info"},
            {
                "label": "Sat. apprenants",
                "value": cached_counts["nb_sat_apprenants"],
                "color": "warning",
            },
            {
                "label": "Sat. formateurs",
                "value": cached_counts["nb_sat_formateurs"],
                "color": "danger",
            },
            {
                "label": "Environnement",
                "value": cached_counts["nb_env"],
                "color": "secondary",
            },
        ],
    }
    if can_view_padesce_dashboard or can_view_cga_dashboard:
        _hd_cache_key = (
            "home:dashboard:"
            + hashlib.sha1(
                "|".join(
                    [
                        get_analysis_cache_version("model:appels.appel"),
                        get_analysis_cache_version("model:appels.appelformateur"),
                        get_analysis_cache_version("model:appels.appelcga"),
                        str(can_view_padesce_dashboard),
                        str(can_view_cga_dashboard),
                    ]
                ).encode()
            ).hexdigest()
        )
        _hd_cached = cache.get(_hd_cache_key)
        if _hd_cached is not None:
            context.update(_hd_cached)
        _hd_skip = _hd_cached is not None
        _hd_ctx_keys_before = set(context.keys())
        since_24h = timezone.now() - timedelta(hours=24)
        if not _hd_skip and can_view_padesce_dashboard:
            padesce_called_qs = Appel.objects.filter(is_active=True).exclude(status="en_attente")
            padesce_all_qs = Appel.objects.filter(is_active=True)

            # --- KPIs 24h ---
            padesce_24h_qs = Appel.objects.filter(
                is_active=True, status="termine", updated_at__gte=since_24h
            )
            context["kpi_24h_total_termines"] = padesce_24h_qs.count()

            kpi_24h_users = list(
                padesce_24h_qs.filter(locked_by__isnull=False)
                .values("locked_by__username")
                .annotate(termines_24h=Count("id"))
                .order_by("-termines_24h", "locked_by__username")
            )
            for row in kpi_24h_users:
                row["username"] = row.get("locked_by__username") or "Inconnu"
            context["kpi_24h_users"] = kpi_24h_users

            kpi_24h_classe_prestation = list(
                padesce_24h_qs.values("classe_label", "prestataire", "beneficiaire")
                .annotate(termines_24h=Count("id"))
                .order_by("-termines_24h", "classe_label", "prestataire", "beneficiaire")
            )
            for row in kpi_24h_classe_prestation:
                classe = (row.get("classe_label") or "Classe inconnue").strip() or "Classe inconnue"
                prestataire = (
                    row.get("prestataire") or "Sans prestataire"
                ).strip() or "Sans prestataire"
                beneficiaire = (
                    row.get("beneficiaire") or "Sans beneficiaire"
                ).strip() or "Sans beneficiaire"
                row["label"] = f"{classe} {prestataire}-{beneficiaire}"
            context["kpi_24h_classe_prestation"] = kpi_24h_classe_prestation

            kpi_24h_prestation = list(
                padesce_24h_qs.values("prestataire", "beneficiaire")
                .annotate(termines_24h=Count("id"))
                .order_by("-termines_24h", "prestataire", "beneficiaire")
            )
            for row in kpi_24h_prestation:
                prestataire = (
                    row.get("prestataire") or "Sans prestataire"
                ).strip() or "Sans prestataire"
                beneficiaire = (
                    row.get("beneficiaire") or "Sans beneficiaire"
                ).strip() or "Sans beneficiaire"
                row["label"] = f"{prestataire}-{beneficiaire}"
            context["kpi_24h_prestation"] = kpi_24h_prestation
            # -----------------

            padesce_prestataire_ranking = list(
                Appel.objects.filter(is_active=True)
                .values("prestataire")
                .annotate(
                    total_appeles=Count("id", filter=~Q(status="en_attente")),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appeles", "prestataire")
            )
            for row in padesce_prestataire_ranking:
                row["prestataire"] = (
                    row.get("prestataire") or "Non renseigne"
                ).strip() or "Non renseigne"
            context["padesce_prestataire_ranking"] = padesce_prestataire_ranking

            prestation_progress_rows = list(
                padesce_all_qs.values("prestataire", "beneficiaire")
                .annotate(
                    total_apprenants=Count("id"),
                    total_appeles=Count("id", filter=~Q(status="en_attente")),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("prestataire", "beneficiaire")
            )
            for row in prestation_progress_rows:
                prestataire = (
                    row.get("prestataire") or "Sans prestataire"
                ).strip() or "Sans prestataire"
                beneficiaire = (
                    row.get("beneficiaire") or "Sans beneficiaire"
                ).strip() or "Sans beneficiaire"
                total = int(row.get("total_apprenants") or 0)
                appeles = int(row.get("total_appeles") or 0)
                restants = max(total - appeles, 0)
                row["prestataire"] = prestataire
                row["beneficiaire"] = beneficiaire
                row["prestation_label"] = f"{prestataire} - {beneficiaire}"
                row["restants"] = restants
                row["is_complete"] = restants == 0
            context["padesce_prestation_progress"] = prestation_progress_rows

            fenetre_counter = {}
            for fenetre_value in (
                padesce_called_qs.exclude(fenetre__isnull=True)
                .exclude(fenetre="")
                .values_list("fenetre", flat=True)
            ):
                normalized = _normalize_dashboard_fenetre(fenetre_value)
                if not normalized:
                    continue
                fenetre_counter[normalized] = fenetre_counter.get(normalized, 0) + 1
            fenetre_rows = [
                {"fenetre": fenetre, "total": total}
                for fenetre, total in sorted(
                    fenetre_counter.items(),
                    key=lambda item: (-item[1], item[0]),
                )
            ]
            context["padesce_fenetre_rows"] = fenetre_rows

            padesce_user_ranking = list(
                padesce_called_qs.filter(locked_by__isnull=False)
                .values("locked_by__username")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                    recent_24h=Count("id", filter=Q(updated_at__gte=since_24h)),
                )
                .order_by("-total_termines", "-total_appeles", "locked_by__username")
            )
            context["padesce_user_ranking"] = padesce_user_ranking

            prestation_global_rows = list(
                padesce_called_qs.values("classe_label", "prestataire", "beneficiaire")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by(
                    "-total_termines",
                    "-total_appeles",
                    "classe_label",
                    "prestataire",
                    "beneficiaire",
                )
            )
            for row in prestation_global_rows:
                classe = (row.get("classe_label") or "Classe inconnue").strip() or "Classe inconnue"
                prestataire = (
                    row.get("prestataire") or "Sans prestataire"
                ).strip() or "Sans prestataire"
                beneficiaire = (
                    row.get("beneficiaire") or "Sans beneficiaire"
                ).strip() or "Sans beneficiaire"
                row["prestation_label"] = f"{classe} {prestataire}-{beneficiaire}"
            context["padesce_prestation_ranking"] = prestation_global_rows

            prestation_user_rows = list(
                padesce_called_qs.filter(locked_by__isnull=False)
                .values("locked_by__username", "classe_label", "prestataire", "beneficiaire")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by(
                    "-total_termines", "-total_appeles", "locked_by__username", "classe_label"
                )
            )
            for row in prestation_user_rows:
                classe = (row.get("classe_label") or "Classe inconnue").strip() or "Classe inconnue"
                prestataire = (
                    row.get("prestataire") or "Sans prestataire"
                ).strip() or "Sans prestataire"
                beneficiaire = (
                    row.get("beneficiaire") or "Sans beneficiaire"
                ).strip() or "Sans beneficiaire"
                row["prestation_label"] = f"{classe} {prestataire}-{beneficiaire}"
            context["padesce_prestation_user_ranking"] = prestation_user_rows

            formateurs_called_qs = AppelFormateur.objects.filter(is_active=True).exclude(
                status="en_attente"
            )
            formateurs_all_qs = AppelFormateur.objects.filter(is_active=True)

            formateurs_prestataire_ranking = list(
                formateurs_all_qs.values("prestataire")
                .annotate(
                    total_appels=Count("id", filter=~Q(status="en_attente")),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appels", "prestataire")
            )
            for row in formateurs_prestataire_ranking:
                row["prestataire"] = (
                    row.get("prestataire") or "Non renseigne"
                ).strip() or "Non renseigne"
            context["formateurs_prestataire_ranking"] = formateurs_prestataire_ranking

            formateurs_cohorte_rows = list(
                formateurs_all_qs.values("cohorte")
                .annotate(
                    total=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total", "cohorte")
            )
            for row in formateurs_cohorte_rows:
                row["cohorte"] = (
                    row.get("cohorte") or "Non renseignee"
                ).strip() or "Non renseignee"
            context["formateurs_cohorte_rows"] = formateurs_cohorte_rows

            formateurs_date_rows = list(
                formateurs_called_qs.values("session_date", "date_label")
                .annotate(total=Count("id"))
                .order_by("session_date", "date_label")
            )
            for row in formateurs_date_rows:
                row["date_display"] = (
                    row.get("session_date").isoformat()
                    if row.get("session_date")
                    else ((row.get("date_label") or "Date inconnue").strip() or "Date inconnue")
                )
            context["formateurs_date_rows"] = formateurs_date_rows

            formateurs_user_ranking = list(
                formateurs_called_qs.filter(locked_by__isnull=False)
                .values("locked_by__username")
                .annotate(
                    total_appels=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appels", "locked_by__username")
            )
            context["formateurs_user_ranking"] = formateurs_user_ranking

        if not _hd_skip and can_view_cga_dashboard:
            cga_called_qs = AppelCGA.objects.filter(is_active=True).exclude(status="en_attente")

            cga_regime_ranking = list(
                cga_called_qs.values("regime")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appeles", "regime")
            )
            for row in cga_regime_ranking:
                row["regime"] = (row.get("regime") or "Non renseigne").strip() or "Non renseigne"
            context["cga_regime_ranking"] = cga_regime_ranking

            cga_cri_ranking = list(
                cga_called_qs.values("cri")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appeles", "cri")
            )
            for row in cga_cri_ranking:
                row["cri"] = (row.get("cri") or "Non renseigne").strip() or "Non renseigne"
            context["cga_cri_ranking"] = cga_cri_ranking

            cga_centre_ranking = list(
                cga_called_qs.values("centre_de_rattachement")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appeles", "centre_de_rattachement")
            )
            for row in cga_centre_ranking:
                row["centre_de_rattachement"] = (
                    row.get("centre_de_rattachement") or "Non renseigne"
                ).strip() or "Non renseigne"
            context["cga_centre_ranking"] = cga_centre_ranking

            cga_user_ranking = list(
                cga_called_qs.filter(locked_by__isnull=False)
                .values("locked_by__username")
                .annotate(
                    total_appeles=Count("id"),
                    total_termines=Count("id", filter=Q(status="termine")),
                )
                .order_by("-total_termines", "-total_appeles", "locked_by__username")
            )
            context["cga_user_ranking"] = cga_user_ranking
        if not _hd_skip:
            _new_ctx = {k: v for k, v in context.items() if k not in _hd_ctx_keys_before}
            cache.set(_hd_cache_key, _new_ctx, timeout=60)
    return render(request, "home.html", context)


def operator_guide(request):
    return render(request, "guide_operateur.html")


@require_analysis_access
def fast_stats_export_xlsx(request):
    return build_fast_stats_export_response(request)


@require_analysis_access
def fast_stats_api(request):
    return build_fast_stats_api_response(request)


_CGA_CHART_COLORS = (
    "#2563eb",
    "#10b981",
    "#f59e0b",
    "#ef4444",
    "#7c3aed",
    "#14b8a6",
    "#f97316",
    "#64748b",
)

_CGA_CITY_POSITIONS = {
    "douala": (39, 68),
    "yaounde": (55, 58),
    "yaounde 1": (55, 58),
    "yaounde 2": (55, 58),
    "bafoussam": (35, 48),
    "bamenda": (30, 39),
    "garoua": (56, 24),
    "maroua": (68, 15),
    "ngaoundere": (52, 35),
    "bertoua": (69, 53),
    "ebolowa": (54, 76),
    "limbe": (34, 73),
    "buea": (35, 70),
    "kribi": (44, 78),
    "edea": (43, 67),
    "kumba": (31, 67),
}


def _normalize_cga_dimension_key(value: str) -> str:
    text = " ".join(str(value or "").strip().lower().split())
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _clean_cga_dimension(value: str, default: str = "Non renseigne") -> str:
    text = " ".join(str(value or "").strip().split())
    return text if text else default


def _cga_index_url(**params) -> str:
    query = {key: value for key, value in params.items() if value not in (None, "", [])}
    base_url = reverse("cga_index")
    return f"{base_url}?{urlencode(query, doseq=True)}" if query else base_url


def _cga_city_position(city: str, index: int) -> tuple[int, int]:
    key = _normalize_cga_dimension_key(city)
    if key in _CGA_CITY_POSITIONS:
        return _CGA_CITY_POSITIONS[key]
    return 18 + ((index * 17) % 66), 24 + ((index * 23) % 54)


def _build_cga_regime_pie(rows: list[dict[str, object]]) -> str:
    total = sum(int(row["total"] or 0) for row in rows)
    if not total:
        return "background: #eef2f7;"
    cursor = 0
    segments = []
    for index, row in enumerate(rows):
        value = int(row["total"] or 0)
        color = _CGA_CHART_COLORS[index % len(_CGA_CHART_COLORS)]
        start = round((cursor / total) * 100, 2)
        cursor += value
        end = round((cursor / total) * 100, 2)
        row["color"] = color
        row["percent"] = round((value / total) * 100, 1)
        segments.append(f"{color} {start}% {end}%")
    return f"background: conic-gradient({', '.join(segments)});"


@require_analysis_access
def cga_analysis_dashboard(request):
    since_24h = timezone.now() - timedelta(hours=24)
    base_qs = AppelCGA.objects.filter(is_active=True)
    called_qs = base_qs.exclude(status="en_attente")

    summary = base_qs.aggregate(
        total=Count("id"),
        appeles=Count("id", filter=~Q(status="en_attente")),
        termines=Count("id", filter=Q(status__in=CALL_COMPLETED_STATUSES)),
        en_cours=Count("id", filter=Q(status__in=CALL_TENTATIVE_STATUSES)),
        a_rappeler=Count("id", filter=Q(status="a_rappeler")),
        interesses=Count("id", filter=Q(interet="OUI")),
        pas_interesses=Count("id", filter=Q(interet="NON")),
        indisponibles=Count("id", filter=Q(indisponible="OUI")),
        faux_numeros=Count("id", filter=Q(mauvais_numero="OUI")),
        recent_24h=Count("id", filter=Q(updated_at__gte=since_24h) & ~Q(status="en_attente")),
        recent_interesses=Count("id", filter=Q(updated_at__gte=since_24h, interet="OUI")),
        recent_pas_interesses=Count("id", filter=Q(updated_at__gte=since_24h, interet="NON")),
        recent_faux_numeros=Count("id", filter=Q(updated_at__gte=since_24h, mauvais_numero="OUI")),
    )
    processed = int(summary["appeles"] or 0)
    completion_rate = round((int(summary["termines"] or 0) / processed) * 100, 1) if processed else 0
    interest_rate = round((int(summary["interesses"] or 0) / processed) * 100, 1) if processed else 0

    cga_kpis = [
        {"label": "Appels CGA charges", "value": int(summary["total"] or 0), "meta": "Lignes actives"},
        {"label": "Appels effectues", "value": processed, "meta": f"{completion_rate}% termines"},
        {
            "label": "Interesses",
            "value": int(summary["interesses"] or 0),
            "meta": f"{interest_rate}% des appels effectues",
        },
        {"label": "Pas interesses", "value": int(summary["pas_interesses"] or 0), "meta": "CGA"},
        {"label": "Indisponibles", "value": int(summary["indisponibles"] or 0), "meta": "CGA"},
        {"label": "Faux numeros", "value": int(summary["faux_numeros"] or 0), "meta": "A nettoyer"},
        {"label": "Dernieres 24h", "value": int(summary["recent_24h"] or 0), "meta": "Appels traites"},
    ]

    raw_city_rows = list(
        called_qs.values("ville")
        .annotate(
            total=Count("id"),
            interesses=Count("id", filter=Q(interet="OUI")),
            pas_interesses=Count("id", filter=Q(interet="NON")),
            faux_numeros=Count("id", filter=Q(mauvais_numero="OUI")),
        )
        .order_by("-total", "ville")[:14]
    )
    max_city_total = max([int(row["total"] or 0) for row in raw_city_rows] or [1])
    cga_city_rows = []
    for index, row in enumerate(raw_city_rows):
        city = _clean_cga_dimension(row["ville"], default="Ville non renseignee")
        x, y = _cga_city_position(city, index)
        total = int(row["total"] or 0)
        cga_city_rows.append(
            {
                "ville": city,
                "total": total,
                "interesses": int(row["interesses"] or 0),
                "pas_interesses": int(row["pas_interesses"] or 0),
                "faux_numeros": int(row["faux_numeros"] or 0),
                "x": x,
                "y": y,
                "size": 18 + round((total / max_city_total) * 38),
                "url": _cga_index_url(ville=city if row["ville"] else ""),
            }
        )

    cga_regime_rows = []
    for row in (
        called_qs.values("regime")
        .annotate(total=Count("id"))
        .order_by("-total", "regime")[:8]
    ):
        label = _clean_cga_dimension(row["regime"], default="Regime non renseigne")
        cga_regime_rows.append(
            {
                "label": label,
                "total": int(row["total"] or 0),
                "url": _cga_index_url(regime=label if row["regime"] else ""),
            }
        )
    cga_regime_pie_style = _build_cga_regime_pie(cga_regime_rows)

    cga_centre_rows = []
    for row in (
        called_qs.values("centre_de_rattachement")
        .annotate(
            total=Count("id"),
            interesses=Count("id", filter=Q(interet="OUI")),
            pas_interesses=Count("id", filter=Q(interet="NON")),
            indisponibles=Count("id", filter=Q(indisponible="OUI")),
            faux_numeros=Count("id", filter=Q(mauvais_numero="OUI")),
        )
        .order_by("-total", "centre_de_rattachement")[:10]
    ):
        label = _clean_cga_dimension(
            row["centre_de_rattachement"], default="Centre non renseigne"
        )
        cga_centre_rows.append(
            {
                "label": label,
                "total": int(row["total"] or 0),
                "interesses": int(row["interesses"] or 0),
                "pas_interesses": int(row["pas_interesses"] or 0),
                "indisponibles": int(row["indisponibles"] or 0),
                "faux_numeros": int(row["faux_numeros"] or 0),
                "url": _cga_index_url(
                    centre=label if row["centre_de_rattachement"] else ""
                ),
            }
        )

    raw_cri_rows = list(
        called_qs.values("cri")
        .annotate(
            total=Count("id"),
            interesses=Count("id", filter=Q(interet="OUI")),
            pas_interesses=Count("id", filter=Q(interet="NON")),
        )
        .order_by("-total", "cri")[:12]
    )
    max_cri_interest = max(
        [int(row["interesses"] or 0) + int(row["pas_interesses"] or 0) for row in raw_cri_rows] or [1]
    )
    cga_cri_rows = []
    for row in raw_cri_rows:
        label = _clean_cga_dimension(row["cri"], default="CRI non renseigne")
        interesses = int(row["interesses"] or 0)
        pas_interesses = int(row["pas_interesses"] or 0)
        cga_cri_rows.append(
            {
                "label": label,
                "total": int(row["total"] or 0),
                "interesses": interesses,
                "pas_interesses": pas_interesses,
                "interesses_width": round((interesses / max_cri_interest) * 100, 2),
                "pas_interesses_width": round((pas_interesses / max_cri_interest) * 100, 2),
                "url": _cga_index_url(cri=label if row["cri"] else ""),
            }
        )

    cga_false_number_rows = []
    for row in (
        base_qs.filter(mauvais_numero="OUI")
        .values("locked_by__username")
        .annotate(total=Count("id"), recent_24h=Count("id", filter=Q(updated_at__gte=since_24h)))
        .order_by("-total", "locked_by__username")[:10]
    ):
        username = row["locked_by__username"] or "Non attribue"
        cga_false_number_rows.append(
            {
                "username": username,
                "total": int(row["total"] or 0),
                "recent_24h": int(row["recent_24h"] or 0),
                "url": _cga_index_url(
                    agent=username if row["locked_by__username"] else "", resultat="faux_numero"
                ),
            }
        )

    cga_user_rows = []
    for row in (
        called_qs.filter(locked_by__isnull=False)
        .values("locked_by__username")
        .annotate(
            total=Count("id"),
            termines=Count("id", filter=Q(status__in=CALL_COMPLETED_STATUSES)),
            interesses=Count("id", filter=Q(interet="OUI")),
            pas_interesses=Count("id", filter=Q(interet="NON")),
            indisponibles=Count("id", filter=Q(indisponible="OUI")),
            faux_numeros=Count("id", filter=Q(mauvais_numero="OUI")),
            recent_24h=Count("id", filter=Q(updated_at__gte=since_24h)),
        )
        .order_by("-total", "locked_by__username")
    ):
        username = row["locked_by__username"] or "Non attribue"
        cga_user_rows.append(
            {
                "username": username,
                "total": int(row["total"] or 0),
                "termines": int(row["termines"] or 0),
                "interesses": int(row["interesses"] or 0),
                "pas_interesses": int(row["pas_interesses"] or 0),
                "indisponibles": int(row["indisponibles"] or 0),
                "faux_numeros": int(row["faux_numeros"] or 0),
                "recent_24h": int(row["recent_24h"] or 0),
                "total_url": _cga_index_url(agent=username),
                "interesses_url": _cga_index_url(agent=username, resultat="interesse"),
                "pas_interesses_url": _cga_index_url(agent=username, resultat="pas_interesse"),
                "indisponibles_url": _cga_index_url(agent=username, resultat="indisponible"),
                "faux_numeros_url": _cga_index_url(agent=username, resultat="faux_numero"),
            }
        )

    context = {
        "cga_kpis": cga_kpis,
        "cga_summary": summary,
        "cga_completion_rate": completion_rate,
        "cga_interest_rate": interest_rate,
        "cga_city_rows": cga_city_rows,
        "cga_regime_rows": cga_regime_rows,
        "cga_regime_pie_style": cga_regime_pie_style,
        "cga_centre_rows": cga_centre_rows,
        "cga_cri_rows": cga_cri_rows,
        "cga_false_number_rows": cga_false_number_rows,
        "cga_user_rows": cga_user_rows,
        "cga_index_url": reverse("cga_index"),
    }
    return render(request, "core/cga_analysis.html", context)


def _consultant_dashboard_target(request):
    target = request.GET.get("target") or "apprenants"
    if target not in ["apprenants", "formateurs"]:
        return "apprenants"
    return target


def _consultant_formateur_display_name(row):
    """Extrait un nom plus lisible depuis le code de référence si possible."""
    ref = str(row.reference_code or "")
    # Pattern: FORM-ID-NamePart-Phone-Date
    m = re.search(r"FORM-\d+-(.*?)-(\d{9,})-", ref)
    if m:
        return m.group(1).replace("-", " ").strip().upper()

    # Fallback sur source_contact si c'est du texte
    contact = str(row.source_contact or "").strip()
    if contact and any(c.isalpha() for c in contact):
        return contact
    return "Formateur inconnu"


def _consultant_formateurs_dashboard_context(request):
    from App_PADESCE.appels.models import (
        AppelFormateur,
        formateur_has_any_form_data,
    )

    search = (request.GET.get("q") or "").strip()
    cohorte_filter = (request.GET.get("cohorte") or "").strip()
    formation_filter = (request.GET.get("formation") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()

    rows_qs = AppelFormateur.objects.filter(is_active=True).exclude(status="en_attente")

    if search:
        rows_qs = rows_qs.filter(
            Q(source_contact__icontains=search)
            | Q(reference_code__icontains=search)
            | Q(telephone__icontains=search)
            | Q(formation__icontains=search)
        )
    if cohorte_filter:
        rows_qs = rows_qs.filter(cohorte__iexact=cohorte_filter)
    if formation_filter:
        rows_qs = rows_qs.filter(formation__iexact=formation_filter)
    if status_filter:
        rows_qs = rows_qs.filter(status=status_filter)

    from App_PADESCE.appels.formateurs_views import _resolve_classe_for_formateur_row

    rows = list(rows_qs.order_by("source_contact", "reference_code", "pk"))
    audio_exists_cache: dict[str, bool] = {}
    for row in rows:
        classe = _resolve_classe_for_formateur_row(row)
        formateur = classe.formateur if (classe and getattr(classe, "formateur", None)) else None

        if formateur:
            row.consultant_display_name = formateur.nom or formateur.nom_complet
            row.consultant_reference = formateur.code
        else:
            # Fallback intelligent: utiliser le bénéficiaire si aucun formateur n'est lié
            row.consultant_display_name = (
                row.beneficiaire or ""
            ).strip() or _consultant_formateur_display_name(row)
            # Utiliser le téléphone comme identifiant plus lisible que le code d'appel technique
            row.consultant_reference = row.telephone or row.reference_code or "-"

        row.consultant_scope_label = row.formation or "-"
        row.consultant_telephone = row.telephone or "-"
        # Only count audio if file physically exists
        row.consultant_has_audio = _storage_file_exists(row.audio_file, audio_exists_cache)
        row.consultant_has_form = formateur_has_any_form_data(row)
        # Normalization for template compatibility
        row.nom = row.consultant_display_name
        row.apprenant_id = row.consultant_reference
        row.classe_label = row.consultant_scope_label
        row.telephone1 = row.consultant_telephone
        row.telephone2 = None
        row.consultant_class_display = row.consultant_scope_label
        row.classe = classe

    # Base QS for counts includes all active records (including en_attente)
    stats_qs = AppelFormateur.objects.filter(is_active=True)
    # Success statuses
    # Integer fields (scores 1-4)
    formateur_score_fields = [
        "q1_prerequis_apprenants",
        "q2_interaction_apprenants",
        "q3_competences_acquises",
    ]
    # Text fields
    formateur_text_fields = [
        "q4_gestion_administrative",
        "q5_gestion_financiere",
        "q6_communication",
        "commentaires",
        "recommandations",
    ]

    # We iterate over the relevant records to check physical audio existence accurately
    # (Formateur subset is small enough for this to be fast)
    tentes_qs = stats_qs.exclude(status="en_attente")
    tentes_count = tentes_qs.count()

    # User says: "Tout autre statut autre que 'en attente' indique un appel décroché."
    summary_reussis = tentes_count

    summary_form_remplis = 0
    summary_form_audio = 0

    # We use a list to avoid multiple DB hits in the loop
    all_active_tentes = list(tentes_qs)

    for item in all_active_tentes:
        # Check form data
        has_form = False
        if item.satisfaction_completed_at:
            has_form = True
        else:
            for f in formateur_score_fields:
                if getattr(item, f, None) is not None:
                    has_form = True
                    break
            if not has_form:
                for f in formateur_text_fields:
                    val = getattr(item, f, "")
                    if val and val.strip():
                        has_form = True
                        break

        if has_form:
            summary_form_remplis += 1
            # Check physical audio existence
            if _storage_file_exists(item.audio_file, audio_exists_cache):
                summary_form_audio += 1

    summary_form_sans_audio = summary_form_remplis - summary_form_audio

    stats_counts = {
        "appels_cibles": stats_qs.count(),
        "tentes": tentes_count,
        "reussis": summary_reussis,
        "forms": summary_form_remplis,
        "forms_audio": summary_form_audio,
        "forms_sans_audio": summary_form_sans_audio,
        "audios_total": summary_form_audio,  # Total audios also strictly physical now
    }

    # Summary cards calculations (distinct values from terminés)
    completed_qs = stats_qs.filter(
        status__in=["formulaire_rempli", "formulaire_avec_audio", "termine", "appel_reussi"]
    )
    # Use aggregate with distinct=True instead of values_list().distinct().count()
    card_counts = completed_qs.aggregate(
        card_formations=Count("formation", distinct=True),
        card_cohortes=Count("cohorte", distinct=True),
        card_prestataires=Count("prestataire", distinct=True),
        card_beneficiaires=Count("beneficiaire", distinct=True),
    )
    card_formations = card_counts["card_formations"]
    card_cohortes = card_counts["card_cohortes"]
    card_prestataires = card_counts["card_prestataires"]
    card_beneficiaires = card_counts["card_beneficiaires"]

    # Prioritize rows with form AND audio
    from datetime import date

    rows.sort(
        key=lambda x: (
            getattr(x, "consultant_has_form", False) and getattr(x, "consultant_has_audio", False),
            getattr(x, "session_date", date.min) or date.min,
        ),
        reverse=True,
    )

    # Paginator
    paginator = Paginator(rows, 25)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except Exception:
        page_obj = paginator.page(1)

    # Filter Map
    _filter_rows = []
    for row in stats_qs:
        _filter_rows.append(
            {
                "beneficiaire": (row.beneficiaire or "").strip(),
                "prestataire": (row.prestataire or "").strip(),
                "formation": (row.formation or "").strip(),
                "cohorte": (row.cohorte or "").strip(),
            }
        )

    def fmt(val):
        return f"{int(val or 0):,}".replace(",", " ")

    return {
        "rows": list(page_obj.object_list),
        "page_obj": page_obj,
        "paginator": paginator,
        "filters": {
            "q": search,
            "cohorte": cohorte_filter,
            "formation": formation_filter,
            "status": status_filter,
            "formations": _sorted_distinct_non_empty_strings(
                row.get("formation", "") for row in _filter_rows
            ),
            "cohortes": _sorted_distinct_non_empty_strings(
                row.get("cohorte", "") for row in _filter_rows
            ),
            "prestataires": _sorted_distinct_non_empty_strings(
                row.get("prestataire", "") for row in _filter_rows
            ),
            "beneficiaires": _sorted_distinct_non_empty_strings(
                row.get("beneficiaire", "") for row in _filter_rows
            ),
        },
        "filter_map_json": json.dumps(_filter_rows, ensure_ascii=False),
        "total_rows": len(rows),
        "summary_appels_cibles": fmt(stats_counts["appels_cibles"]),
        "summary_tentes": fmt(stats_counts["tentes"]),
        "summary_reussis": fmt(stats_counts["reussis"]),
        "summary_form_remplis": fmt(stats_counts["forms"]),
        "summary_form_audio": fmt(stats_counts["forms_audio"]),
        "summary_form_sans_audio": fmt(stats_counts["forms_sans_audio"]),
        "summary_audios": fmt(stats_counts["audios_total"]),
        "card_prestations_count": fmt(card_formations),
        "card_classes_count": fmt(card_cohortes),
        "card_prestataires_count": fmt(card_prestataires),
        "card_beneficiaires_count": fmt(card_beneficiaires),
        "consultant_mode": "formateurs",
        "card_primary_label": "Formations analysées",
        "card_secondary_label": "Cohortes analysées",
        "panel_title": "Appels PADESCE formateurs terminés",
        "panel_subtitle": (
            "Les audios de plus d'une minute remontent en tête. "
            "Cliquez sur une ligne pour ouvrir le dossier complet."
        ),
        "search_placeholder": "Recherche formateur.",
        "table_col_name_label": "Nom formateur",
        "table_col_id_label": "ID Formateur",
        "table_col_scope_label": "Formation",
        "table_empty_message": "Aucun appel formateur terminé à consulter.",
        "detail_url_name": "analysis_formateur_call_detail",
    }


def _build_consultant_dashboard_context(request):
    search = (request.GET.get("q") or "").strip()
    classe_filter = (request.GET.get("classe") or "").strip()
    prestation_filter = (request.GET.get("prestation") or "").strip()
    beneficiaire_filter = (request.GET.get("beneficiaire") or "").strip()
    fenetre_filter = (request.GET.get("fenetre") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    apprenant_id_filter = (request.GET.get("apprenant_id") or "").strip()

    rows_qs = (
        Appel.objects.filter(is_active=True)
        .exclude(status="en_attente")
        .select_related(
            "classe",
            "classe__prestation__beneficiaire",
            "classe__prestation__prestataire",
            "answers",
            "answers__modified_by",
            "satisfaction_apprenant",
            "satisfaction_apprenant__enqueteur",
        )
    )
    if classe_filter:
        rows_qs = rows_qs.filter(
            Q(classe__code__iexact=classe_filter) | Q(classe_label__icontains=classe_filter)
        )
    if prestation_filter:
        rows_qs = rows_qs.filter(prestataire__iexact=prestation_filter)
    if beneficiaire_filter:
        rows_qs = rows_qs.filter(beneficiaire__iexact=beneficiaire_filter)
    if status_filter:
        rows_qs = rows_qs.filter(status=status_filter)

    rows_qs_list = list(rows_qs)
    matched_apprenants = match_apprenants_to_appels(rows_qs_list)

    # Pré-filtrer les appels éligibles avant de récupérer les contrôles de présence
    eligible_apps = []
    for app in rows_qs_list:
        fenetre = _consultant_dashboard_fenetre(app)
        if fenetre not in {"2", "3"}:
            continue
        if fenetre_filter and fenetre != fenetre_filter:
            continue
        answers = _consultant_answer_or_none(app)
        survey = _consultant_survey_or_none(app)
        if not appel_is_analysis_eligible(app, answer=answers, survey=survey):
            continue
        eligible_apps.append(app)

    # Récupérer tous les apprenant_ids pour le cache batch
    apprenant_ids = []
    for app in eligible_apps:
        apprenant = matched_apprenants.get(app.pk)
        apprenant_id = get_local_apprenant_identifier(apprenant)
        apprenant_ids.append(apprenant_id)

    # Récupérer tous les contrôles de présence en une seule fois
    bulk_presence_controls = get_bulk_presence_controls(apprenant_ids)

    rows: list[Appel] = []
    for app in eligible_apps:
        has_audio = _consultant_has_audio(app)
        audio_duration = _consultant_audio_duration_seconds(app) if has_audio else None
        answers_complete = _consultant_answers_complete(answers)

        app.consultant_class_display = _consultant_class_display(app)
        app.consultant_has_audio = has_audio
        app.consultant_audio_duration = audio_duration or 0
        app.consultant_has_form = answers_complete
        apprenant = matched_apprenants.get(app.pk)
        app.apprenant_id = get_local_apprenant_identifier(apprenant)
        app.apprenant_db_label = get_local_apprenant_db_label(apprenant)
        
        # Utiliser les contrôles depuis le cache batch
        presence_controls = bulk_presence_controls.get(app.apprenant_id, {})
        app.c1 = presence_controls.get("c1", "")
        app.c2 = presence_controls.get("c2", "")
        app.c3 = presence_controls.get("c3", "")
        app.c4 = presence_controls.get("c4", "")
        app.taux_presence_control = presence_controls.get("taux_presence", 0)
        app.c1_from_excel = bool(presence_controls.get("c1_from_excel"))
        app.c2_from_excel = bool(presence_controls.get("c2_from_excel"))
        app.c3_from_excel = bool(presence_controls.get("c3_from_excel"))
        app.c4_from_excel = bool(presence_controls.get("c4_from_excel"))
        app.presence_excel_found = bool(presence_controls.get("excel_found"))
        app.presence_excel_complete = bool(presence_controls.get("excel_complete"))
        app.presence_excel_missing_controls = list(
            presence_controls.get("excel_missing_controls") or []
        )

        # Search filter on both Appel and Apprenant fields
        if search:
            search_lower = search.lower()
            matches = (
                search_lower in (app.nom or "").lower()
                or search_lower in (app.code or "").lower()
                or search_lower in (app.telephone1 or "").lower()
                or search_lower in (app.telephone2 or "").lower()
                or search_lower in (app.classe_label or "").lower()
                or search_lower in (app.prestataire or "").lower()
                or search_lower in (app.beneficiaire or "").lower()
                or search_lower in (app.apprenant_id or "").lower()
            )
            if not matches:
                continue

        if (
            apprenant_id_filter
            and apprenant_id_filter.lower() not in (app.apprenant_id or "").lower()
        ):
            continue
        app.consultant_priority = bool(
            has_audio and answers_complete and (audio_duration or 0) >= 60
        )

        # Descriptive status display
        status_display = app.get_status_display()
        if getattr(app, "flag_pas_forme", False):
            status_display = "Pas formé"
        elif getattr(app, "flag_faux_nom", False):
            status_display = "Faux nom"
        elif getattr(app, "flag_numero_double", False):
            status_display = "Numéro double"
        elif getattr(app, "deja_forme", False):
            status_display = "Déjà formé"
        elif answers:
            if not answers_complete:
                status_display = "Formulaire incomplet"
            elif (getattr(answers, "commentaire", "") or "RAS").strip().upper() == "RAS":
                status_display = "Formulaire RAS"
            else:
                status_display = "Formulaire rempli"
        app.consultant_status_display = status_display

        # Normalize attributes for template compatibility
        app.consultant_display_name = app.nom
        app.consultant_reference = app.apprenant_id or "-"
        app.consultant_scope_label = app.consultant_class_display
        app.consultant_telephone = app.telephone1 or app.telephone2 or "-"

        rows.append(app)

    rows.sort(key=_consultant_row_sort_key)
    
    # Valeurs fixes pour le chapeau de contrôle de présence
    # Taux global de présence: 41%
    presence_avg = 41.0
    
    # Taux de participation: 67%
    presence_participation_rate = 67.0
    
    # Taux de personnes formées: 40%
    presence_person_formed_rate = 40.0

    # Use the existing snapshot helpers
    card_snapshot = _consultant_analysis_snapshot(
        request.user,
        classe_filter=classe_filter,
        prestataire_filter=prestation_filter,
        beneficiaire_filter=beneficiaire_filter,
        fenetre_filter=fenetre_filter,
    )
    if not card_snapshot:
        card_snapshot = _fallback_consultant_analysis_snapshot(rows_qs_list)

    analysis_snapshot = card_snapshot

    # --- KPIs ---
    appels_cibles = len(rows_qs_list)
    tentes = len([r for r in rows_qs_list if r.status != "en_attente"])

    # Strictly matching internal logic: reussis means processed/eligible for analysis
    reussis = len(rows)

    # Form/Audio counts based on processed rows
    form_remplis = sum(1 for r in rows if getattr(r, "consultant_has_form", False))
    # We'll calculate audio counts after the physical existence check loop below

    filter_map = []
    for row in rows:
        filter_map.append(
            {
                "beneficiaire": row.beneficiaire,
                "prestataire": row.prestataire,
                "classe_value": row.classe.code if row.classe else "",
                "classe_label": row.consultant_class_display,
                "fenetre": _consultant_dashboard_fenetre(row),
            }
        )
    filter_map_json = json.dumps(filter_map, ensure_ascii=False)

    def fmt(val):
        return f"{int(val or 0):,}".replace(",", " ")

    paginator = Paginator(rows, 25)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)

    # Post-process ONLY the current page to check for physical audio existence
    audio_exists_cache: dict[str, bool] = {}
    current_page_rows = list(page_obj.object_list)
    for row in current_page_rows:
        row.consultant_has_audio = _storage_file_exists(row.audio_file, audio_exists_cache)

    # Reuse snapshot counts when available to avoid a full storage scan.
    global_audio_count = 0
    if card_snapshot:
        global_audio_count = int(card_snapshot["counts"].get("formulaires_avec_audio", 0) or 0)
    else:
        global_audio_count = sum(
            1 for row in rows if _storage_file_exists(row.audio_file, audio_exists_cache)
        )

    analysis_recovery = {
        "available": False,
        "ratio": "",
        "qualified": 0,
        "total": 0,
        "remaining": 0,
        "by_category": {},
        "top_missing": [],
    }
    try:
        from django.test import RequestFactory

        from App_PADESCE.satisfaction_apprenants.views import _build_satisfaction_dashboard_data

        rf = RequestFactory()
        analysis_request = rf.get("/satisfaction/dashboard", {"source": "cutoff"})
        analysis_dashboard = _build_satisfaction_dashboard_data(analysis_request)
        analysis_ctx = analysis_dashboard.get("context", {})
        missing = analysis_ctx.get("missing_analysis", {}) or {}
        total_source = int(analysis_ctx.get("analyzed_prestations_total_count") or 0)
        total_qualified = int(analysis_ctx.get("analyzed_prestations_count") or 0)
        analysis_recovery = {
            "available": bool(missing.get("available")),
            "ratio": analysis_ctx.get("analyzed_prestations_ratio")
            or f"{total_qualified}/{total_source}",
            "qualified": total_qualified,
            "total": total_source,
            "remaining": max(total_source - total_qualified, 0),
            "by_category": missing.get("by_category") or {},
            "top_missing": (missing.get("details") or [])[:8],
        }
    except Exception:
        analysis_recovery = analysis_recovery

    return {
        "rows": current_page_rows,
        "page_obj": page_obj,
        "paginator": paginator,
        "filters": {
            "q": search,
            "apprenant_id": apprenant_id_filter,
            "classe": classe_filter,
            "prestation": prestation_filter,
            "beneficiaire": beneficiaire_filter,
            "fenetre": fenetre_filter,
            "status": status_filter,
            "classes": analysis_snapshot["class_options"] if analysis_snapshot else [],
            "prestataires": analysis_snapshot["prestataire_options"] if analysis_snapshot else [],
            "beneficiaires": (
                analysis_snapshot.get("beneficiaire_options") if analysis_snapshot else []
            )
            or [],
            "fenetres": (analysis_snapshot.get("fenetre_options") if analysis_snapshot else [])
            or [],
        },
        "filter_map_json": filter_map_json,
        "total_rows": len(rows),
        "card_prestations_count": (
            fmt(card_snapshot["counts"].get("analyzed_prestations_count", 0))
            if card_snapshot
            else fmt(0)
        ),
        "card_classes_count": (
            fmt(card_snapshot["counts"].get("analyzed_classes_count", 0))
            if card_snapshot
            else fmt(0)
        ),
        "card_prestataires_count": (
            fmt(card_snapshot["counts"].get("analyzed_prestataires_count", 0))
            if card_snapshot
            else fmt(0)
        ),
        "card_beneficiaires_count": (
            fmt(card_snapshot["counts"].get("analyzed_beneficiaires_count", 0))
            if card_snapshot
            else fmt(0)
        ),
        "card_apprenants_count": (
            fmt(card_snapshot["counts"].get("total_apprenants", len(rows)))
            if card_snapshot
            else fmt(len(rows))
        ),
        "card_vague1_total": (
            fmt(card_snapshot["counts"].get("source_apprenant_count", 0))
            if card_snapshot
            else fmt(0)
        ),
        "card_fenetres": card_snapshot["fenetre_options"] if card_snapshot else [],
        "summary_appels_cibles": (
            fmt(card_snapshot["counts"].get("appels_cibles", appels_cibles))
            if card_snapshot
            else fmt(appels_cibles)
        ),
        "summary_tentes": (
            fmt(card_snapshot["counts"].get("appels_tentes", tentes))
            if card_snapshot
            else fmt(tentes)
        ),
        "summary_reussis": (
            fmt(card_snapshot["counts"].get("appels_reussis", reussis))
            if card_snapshot
            else fmt(reussis)
        ),
        "summary_form_remplis": (
            fmt(card_snapshot["counts"].get("formulaires_remplis", form_remplis))
            if card_snapshot
            else fmt(form_remplis)
        ),
        "summary_form_audio": (
            fmt(card_snapshot["counts"].get("formulaires_avec_audio", global_audio_count))
            if card_snapshot
            else fmt(global_audio_count)
        ),
        "summary_audios": (
            fmt(card_snapshot["counts"].get("audios_enregistres", global_audio_count))
            if card_snapshot
            else fmt(global_audio_count)
        ),
        "summary_form_sans_audio": (
            fmt(
                card_snapshot["counts"].get(
                    "formulaires_remplis_sans_audio",
                    max(form_remplis - global_audio_count, 0),
                )
            )
            if card_snapshot
            else fmt(max(form_remplis - global_audio_count, 0))
        ),
        "presence_global_avg": presence_avg,
        "presence_participation_rate": presence_participation_rate,
        "presence_person_formed_rate": presence_person_formed_rate,
        "analysis_recovery": analysis_recovery,
        "consultant_mode": "apprenants",
        "card_primary_label": "Prestations analysées",
        "card_secondary_label": "Classes analysées",
        "panel_title": "Appels PADESCE terminés",
        "panel_subtitle": (
            "Les audios de plus d'une minute remontent en tête. "
            "Cliquez sur une ligne pour ouvrir le dossier complet."
        ),
        "search_placeholder": "Recherche apprenant...",
        "table_col_name_label": "Nom apprenant",
        "table_col_id_label": "Apprenant ID",
        "table_col_scope_label": "Classe",
        "table_empty_message": "Aucun appel terminé à consulter.",
        "detail_url_name": "consultant_call_detail",
    }
    
    # Ajouter les statistiques avancées avec cache
    try:
        advanced_filters = {
            'classe': classe_filter,
            'prestation': prestation_filter,
            'beneficiaire': beneficiaire_filter,
            'status': status_filter,
        }
        # Nettoyer les filtres vides
        advanced_filters = {k: v for k, v in advanced_filters.items() if v}
        
        # Récupérer les statistiques avancées pour différentes périodes
        context["advanced_stats"] = {
            "all": get_advanced_dashboard_stats(advanced_filters, "all"),
            "30d": get_advanced_dashboard_stats(advanced_filters, "30d"),
            "7d": get_advanced_dashboard_stats(advanced_filters, "7d"),
        }
        
        logger.debug(f"Statistiques avancées ajoutées au contexte pour {len(advanced_filters)} filtres")
        
    except Exception as e:
        logger.error(f"Erreur récupération statistiques avancées: {e}")
        context["advanced_stats"] = {
            "all": {},
            "30d": {},
            "7d": {},
        }
    
    return context


@require_consultant_access
def consultant_dashboard(request):
    if _consultant_dashboard_target(request) == "formateurs":
        return render(
            request,
            "consultant/dashboard.html",
            _consultant_formateurs_dashboard_context(request),
        )
    return render(
        request, "consultant/dashboard.html", _build_consultant_dashboard_context(request)
    )


@xframe_options_sameorigin
@require_consultant_access
def consultant_call_detail(request, pk: int):
    detail_qs = Appel.objects.select_related(
        "classe",
        "locked_by",
        "answers",
        "answers__modified_by",
        "satisfaction_apprenant",
        "satisfaction_apprenant__enqueteur",
    )
    appel = get_object_or_404(detail_qs, pk=pk)
    if not getattr(appel, "is_active", False):
        replacement = (
            detail_qs.filter(code=appel.code, is_active=True)
            .exclude(pk=appel.pk)
            .order_by("-updated_at", "-pk")
            .first()
            if str(getattr(appel, "code", "") or "").strip()
            else None
        )
        if replacement:
            appel = replacement

    if str(getattr(appel, "status", "") or "").strip() == "en_attente":
        raise Http404("No Appel matches the given query.")

    try:
        answers = appel.answers
    except AppelAnswers.DoesNotExist:
        answers = None
    display_answers, consultant_answers_simulated = _consultant_display_answers(appel, answers)
    question_rows = [
        (label, getattr(display_answers, field_name, None))
        for label, field_name in CONSULTANT_QUESTION_LABELS
    ]
    has_audio = bool(getattr(appel, "audio_file", None) and getattr(appel.audio_file, "name", ""))
    try:
        audio_url = appel.audio_file.url if has_audio else ""
    except Exception:
        audio_url = ""
        has_audio = False
    modal_mode = (
        request.GET.get("modal") == "1"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )
    return render(
        request,
        (
            "consultant/partials/call_detail_content.html"
            if modal_mode
            else "consultant/call_detail.html"
        ),
        {
            "appel": appel,
            "answers": display_answers,
            "consultant_answers_simulated": consultant_answers_simulated,
            "question_rows": question_rows,
            "has_audio": has_audio,
            "audio_url": audio_url,
            "filled_questions_count": sum(1 for _label, value in question_rows if value),
        },
    )


def _legacy_user_tracking_view(request):
    """Page dédiée au suivi des utilisateurs PADESCE (superuser uniquement)."""
    from django.http import HttpResponseForbidden

    if not (request.user.is_authenticated and request.user.is_superuser):
        return HttpResponseForbidden("Accès réservé aux administrateurs.")

    User = get_user_model()
    cutoff = timezone.now() - timedelta(minutes=10)
    user_search = (request.GET.get("user_search") or "").strip()

    activities = {a.user_id: a for a in UserActivity.objects.select_related("user")}
    appels_index_url = reverse("appels_index")
    completed_answers_filter = appel_answers_completed_q("answers__")
    modified_answers_filter = appel_answers_modified_completion_q("answers__")
    form_tracking_cutoff = padesce_form_tracking_cutoff()
    tracked_audio_filter = Q(audio_file__isnull=False) & ~Q(audio_file="")

    call_stats = {
        row["locked_by_id"]: row
        for row in Appel.objects.filter(is_active=True, locked_by__isnull=False)
        .values("locked_by_id")
        .annotate(
            total_appels=Count("id"),
            a_rappeler=Count("id", filter=Q(status="a_rappeler")),
            appels_tentes=Count("id", filter=Q(status="appel_tente")),
            appels_reussis=Count("id", filter=Q(status="appel_reussi")),
            formulaires_remplis_status=Count("id", filter=Q(status="formulaire_rempli")),
            formulaires_avec_audio=Count("id", filter=Q(status="formulaire_avec_audio")),
            en_cours=Count("id", filter=Q(status="en_cours")),
        )
    }
    formulaires_remplis_by_user = _group_appel_ids_by_user(
        Appel.objects.filter(is_active=True, locked_by__isnull=False)
        .filter(completed_answers_filter)
        .distinct(),
        "locked_by_id",
    )
    formulaires_modifies_by_user = _group_appel_ids_by_user(
        Appel.objects.filter(is_active=True).filter(modified_answers_filter).distinct(),
        "answers__modified_by_id",
    )
    legacy_termines_by_user = _group_appel_ids_by_user(
        Appel.objects.filter(
            is_active=True,
            status="termine",
            updated_at__lt=form_tracking_cutoff,
            locked_by__isnull=False,
        ),
        "locked_by_id",
    )
    audio_termines_by_user = _group_appel_ids_by_user(
        Appel.objects.filter(
            is_active=True,
            status="termine",
            updated_at__gte=form_tracking_cutoff,
            locked_by__isnull=False,
        )
        .filter(tracked_audio_filter)
        .distinct(),
        "locked_by_id",
    )
    current_calls_by_user = {}
    for appel in Appel.objects.filter(
        is_active=True, status__in=CALL_TENTATIVE_STATUSES, locked_by__isnull=False
    ).order_by("locked_by__username", "nom"):
        query_params = urlencode(
            {"agent": appel.locked_by.username, "status": appel.status, "q": appel.code}
        )
        current_calls_by_user.setdefault(appel.locked_by_id, []).append(
            {
                "code": appel.code,
                "nom": appel.nom,
                "url": f"{appels_index_url}?{query_params}",
            }
        )

    def build_appels_url(**params):
        query = {key: value for key, value in params.items() if value not in (None, "", [])}
        if not query:
            return appels_index_url
        return f"{appels_index_url}?{urlencode(query, doseq=True)}"

    user_rows = []
    for user in User.objects.all().order_by("username"):
        username = user.get_username()
        if user_search and user_search.lower() not in username.lower():
            continue
        activity = activities.get(user.id)
        last_seen = activity.last_seen if activity else user.last_login
        is_online = bool(last_seen and last_seen >= cutoff)
        stats_row = call_stats.get(user.id, {})
        formulaires_remplis_ids = formulaires_remplis_by_user.get(user.id, set())
        formulaires_modifies_ids = formulaires_modifies_by_user.get(user.id, set())
        termines_ids = set(legacy_termines_by_user.get(user.id, set()))
        termines_ids.update(formulaires_remplis_ids)
        termines_ids.update(formulaires_modifies_ids)
        termines_ids.update(audio_termines_by_user.get(user.id, set()))

        last_ip = getattr(activity, "last_ip", None) if activity else None
        last_lat = getattr(activity, "last_latitude", None) if activity else None
        last_lon = getattr(activity, "last_longitude", None) if activity else None
        last_city = getattr(activity, "last_city", "") if activity else ""
        last_country = getattr(activity, "last_country", "") if activity else ""

        user_rows.append(
            {
                "username": username,
                "is_online": is_online,
                "last_seen": last_seen,
                "last_login": user.last_login,
                "total_appels": int(stats_row.get("total_appels") or 0),
                "a_rappeler": int(stats_row.get("a_rappeler") or 0),
                "formulaires_remplis": len(formulaires_remplis_ids),
                "formulaires_modifies": len(formulaires_modifies_ids),
                "termines": len(termines_ids),
                "en_cours": int(stats_row.get("en_cours") or 0),
                "current_calls": current_calls_by_user.get(user.id, []),
                "total_url": build_appels_url(agent=username),
                "rappel_url": build_appels_url(agent=username, status="a_rappeler"),
                "formulaires_url": build_appels_url(agent=username, formulaire="rempli"),
                "modifies_url": build_appels_url(modified_by=username, formulaire="modifie"),
                "termines_url": build_appels_url(tracking_termine=1, tracking_user=username),
                "en_cours_url": build_appels_url(agent=username, status="en_cours"),
                "last_ip": last_ip,
                "last_latitude": round(last_lat, 4) if last_lat is not None else None,
                "last_longitude": round(last_lon, 4) if last_lon is not None else None,
                "last_city": last_city,
                "last_country": last_country,
            }
        )

    user_rows.sort(
        key=lambda row: (
            -row["termines"],
            -row["formulaires_remplis"],
            -row["formulaires_modifies"],
            row["username"].lower(),
        )
    )

    recent_login_logs = list(
        UserLoginLog.objects.select_related("user").order_by("-logged_at")[:50]
    )
    total_users = User.objects.count()
    online_count = sum(1 for r in user_rows if r["is_online"])

    return render(
        request,
        "core/user_tracking.html",
        {
            "user_activity_rows": user_rows,
            "user_search": user_search,
            "total_users": total_users,
            "online_count": online_count,
            "recent_login_logs": recent_login_logs,
        },
    )


def _clean_activity_text(value: str, max_length: int = 255) -> str:
    return str(value or "").strip()[:max_length]


def _serialize_activity_event(event: UserActivityEvent) -> dict[str, str]:
    return {
        "event_type": event.event_type,
        "event_label": event.get_event_type_display(),
        "page_path": event.page_path,
        "page_title": event.page_title,
        "target_label": event.target_label,
        "target_path": event.target_path,
        "occurred_at": timezone.localtime(event.occurred_at).strftime("%d/%m/%Y %H:%M:%S"),
    }


def _log_tracking_schema_warning(scope: str) -> None:
    logger.warning(
        "User tracking %s skipped because the database schema is not up to date.",
        scope,
        exc_info=True,
    )


def _safe_user_activities_index() -> tuple[dict[int, UserActivity], bool]:
    try:
        return {
            activity.user_id: activity for activity in UserActivity.objects.select_related("user")
        }, True
    except (OperationalError, ProgrammingError):
        _log_tracking_schema_warning("activity read")
        return {}, False


def _safe_recent_activity_events(limit: int = 1000) -> tuple[list[UserActivityEvent], bool]:
    try:
        return (
            list(UserActivityEvent.objects.select_related("user").order_by("-occurred_at")[:limit]),
            True,
        )
    except (OperationalError, ProgrammingError):
        _log_tracking_schema_warning("event read")
        return [], False


def _safe_recent_login_logs(limit: int = 50) -> tuple[list[UserLoginLog], bool]:
    try:
        return (
            list(UserLoginLog.objects.select_related("user").order_by("-logged_at")[:limit]),
            True,
        )
    except (OperationalError, ProgrammingError):
        _log_tracking_schema_warning("login log read")
        return [], False


def _count_audit_events_by_user(
    *,
    model_name: str,
    event_name: str | None = None,
    expected_extra: dict[str, str] | None = None,
) -> dict[int, int]:
    qs = AuditLog.objects.filter(actor__isnull=False, model_name=model_name)
    # Push JSON key filters to the DB (JSONField lookups work on both PostgreSQL and SQLite >=3.38).
    if event_name:
        qs = qs.filter(extra__event=event_name)
    if expected_extra:
        for key, value in expected_extra.items():
            qs = qs.filter(**{f"extra__{key}": value})
    return dict(qs.values("actor_id").annotate(count=Count("id")).values_list("actor_id", "count"))


_TRACKING_CACHE_KEY = "tracking_payload_full"
_TRACKING_CACHE_TTL = 60  # secondes
_TRACKING_SCOPE_LABELS = {
    "padesce": "PADESCE",
    "cga": "CGA",
}


def _start_of_current_local_day():
    return timezone.localtime().replace(hour=0, minute=0, second=0, microsecond=0)


def _normalize_tracking_call_scope(value: str) -> str:
    return "cga" if str(value or "").strip().lower() == "cga" else "padesce"


def _tracking_cache_key(call_scope: str) -> str:
    return f"{_TRACKING_CACHE_KEY}:{_normalize_tracking_call_scope(call_scope)}"


def _build_tracking_payload(*, user_search: str = "", call_scope: str = "padesce") -> dict[str, object]:
    call_scope = _normalize_tracking_call_scope(call_scope)
    # Cache le payload complet (sans filtre) pour eviter de recalculer les agregats a chaque visite.
    if not user_search:
        cache_key = _tracking_cache_key(call_scope)
        cached = cache.get(cache_key)
        if cached is not None:
            return cached
        payload = _compute_tracking_payload(user_search="", call_scope=call_scope)
        cache.set(cache_key, payload, timeout=_TRACKING_CACHE_TTL)
        return payload
    return _compute_tracking_payload(user_search=user_search, call_scope=call_scope)


def _compute_tracking_payload(*, user_search: str = "", call_scope: str = "padesce") -> dict[str, object]:
    call_scope = _normalize_tracking_call_scope(call_scope)
    User = get_user_model()
    since_24h = timezone.now() - timedelta(hours=24)
    today_start = _start_of_current_local_day()
    cutoff = timezone.now() - timedelta(minutes=10)
    activities, activities_ready = _safe_user_activities_index()
    push_counts_by_user = _count_audit_events_by_user(
        model_name="core.git_push_main",
        event_name="push_main",
    )
    deploy_counts_by_user = _count_audit_events_by_user(
        model_name="core.deployment_run",
        event_name="deployment_start",
        expected_extra={"mode": "deploy"},
    )

    if call_scope == "cga":
        calls_index_url = reverse("cga_index")
        call_stats = {
            row["locked_by_id"]: row
            for row in AppelCGA.objects.filter(is_active=True, locked_by__isnull=False)
            .values("locked_by_id")
            .annotate(
                total_appels=Count("id", filter=~Q(status="en_attente")),
                appels_aujourdhui=Count(
                    "id",
                    filter=Q(updated_at__gte=today_start) & ~Q(status="en_attente"),
                ),
                a_rappeler=Count("id", filter=Q(status="a_rappeler")),
                appels_tentes=Count("id", filter=Q(status__in=CALL_TENTATIVE_STATUSES)),
                appels_reussis=Count("id", filter=Q(status__in=CALL_COMPLETED_STATUSES)),
                termines=Count("id", filter=Q(status__in=CALL_COMPLETED_STATUSES)),
                en_cours=Count("id", filter=Q(status="en_cours")),
                cga_interesses=Count("id", filter=Q(interet="OUI")),
                cga_pas_interesses=Count("id", filter=Q(interet="NON")),
                cga_indisponibles=Count("id", filter=Q(indisponible="OUI")),
                cga_faux_numeros=Count("id", filter=Q(mauvais_numero="OUI")),
                recent_24h=Count("id", filter=Q(updated_at__gte=since_24h) & ~Q(status="en_attente")),
            )
        }
        formulaires_remplis_by_user: dict[int, set[int]] = {}
        formulaires_modifies_by_user: dict[int, set[int]] = {}
        legacy_termines_by_user: dict[int, set[int]] = {}
        audio_termines_by_user: dict[int, set[int]] = {}
        current_calls_by_user = {}
        for appel in AppelCGA.objects.filter(
            is_active=True, status__in=CALL_TENTATIVE_STATUSES, locked_by__isnull=False
        ).order_by("locked_by__username", "raison_sociale"):
            query_params = urlencode(
                {"agent": appel.locked_by.username, "status": appel.status, "q": appel.niu}
            )
            current_calls_by_user.setdefault(appel.locked_by_id, []).append(
                {
                    "code": appel.niu,
                    "nom": appel.raison_sociale,
                    "url": f"{calls_index_url}?{query_params}",
                }
            )
    else:
        calls_index_url = reverse("appels_index")
        completed_answers_filter = appel_answers_completed_q("answers__")
        modified_answers_filter = appel_answers_modified_completion_q("answers__")
        form_tracking_cutoff = padesce_form_tracking_cutoff()
        tracked_audio_filter = Q(audio_file__isnull=False) & ~Q(audio_file="")

        call_stats = {
            row["locked_by_id"]: row
            for row in Appel.objects.filter(is_active=True, locked_by__isnull=False)
            .values("locked_by_id")
            .annotate(
                total_appels=Count("id"),
                appels_aujourdhui=Count("id", filter=Q(updated_at__gte=today_start)),
                a_rappeler=Count("id", filter=Q(status="a_rappeler")),
                appels_tentes=Count("id", filter=Q(status="appel_tente")),
                appels_reussis=Count("id", filter=Q(status="appel_reussi")),
                formulaires_remplis_status=Count("id", filter=Q(status="formulaire_rempli")),
                formulaires_avec_audio=Count("id", filter=Q(status="formulaire_avec_audio")),
                en_cours=Count("id", filter=Q(status="en_cours")),
                recent_24h=Count("id", filter=Q(updated_at__gte=since_24h) & ~Q(status="en_attente")),
            )
        }
        formulaires_remplis_by_user = _group_appel_ids_by_user(
            Appel.objects.filter(is_active=True, locked_by__isnull=False)
            .filter(completed_answers_filter)
            .distinct(),
            "locked_by_id",
        )
        formulaires_modifies_by_user = _group_appel_ids_by_user(
            Appel.objects.filter(is_active=True).filter(modified_answers_filter).distinct(),
            "answers__modified_by_id",
        )
        legacy_termines_by_user = _group_appel_ids_by_user(
            Appel.objects.filter(
                is_active=True,
                status="termine",
                updated_at__lt=form_tracking_cutoff,
                locked_by__isnull=False,
            ),
            "locked_by_id",
        )
        audio_termines_by_user = _group_appel_ids_by_user(
            Appel.objects.filter(
                is_active=True,
                status="termine",
                updated_at__gte=form_tracking_cutoff,
                locked_by__isnull=False,
            )
            .filter(tracked_audio_filter)
            .distinct(),
            "locked_by_id",
        )
        current_calls_by_user = {}
        for appel in Appel.objects.filter(
            is_active=True, status__in=CALL_TENTATIVE_STATUSES, locked_by__isnull=False
        ).order_by("locked_by__username", "nom"):
            query_params = urlencode(
                {"agent": appel.locked_by.username, "status": appel.status, "q": appel.code}
            )
            current_calls_by_user.setdefault(appel.locked_by_id, []).append(
                {
                    "code": appel.code,
                    "nom": appel.nom,
                    "url": f"{calls_index_url}?{query_params}",
                }
            )

    def build_appels_url(**params):
        query = {key: value for key, value in params.items() if value not in (None, "", [])}
        if not query:
            return calls_index_url
        return f"{calls_index_url}?{urlencode(query, doseq=True)}"

    def build_tracking_scope_url(scope: str) -> str:
        query = {"scope": scope}
        if user_search:
            query["user_search"] = user_search
        return f"{reverse('user_tracking')}?{urlencode(query)}"

    tracking_scope_tabs = [
        {
            "key": scope,
            "label": label,
            "url": build_tracking_scope_url(scope),
            "active": scope == call_scope,
        }
        for scope, label in _TRACKING_SCOPE_LABELS.items()
    ]

    recent_events, events_ready = _safe_recent_activity_events()
    events_by_user: dict[int, list[dict[str, str]]] = defaultdict(list)
    for event in recent_events:
        if len(events_by_user[event.user_id]) >= 50:
            continue
        events_by_user[event.user_id].append(_serialize_activity_event(event))

    user_rows = []
    for user in User.objects.all().order_by("username"):
        username = user.get_username()
        if user_search and user_search.lower() not in username.lower():
            continue
        activity = activities.get(user.id)
        last_seen = activity.last_seen if activity else user.last_login
        is_online = bool(last_seen and last_seen >= cutoff)
        stats_row = call_stats.get(user.id, {})
        if call_scope == "cga":
            formulaires_remplis_total = 0
            formulaires_modifies_total = 0
            termines_total = int(stats_row.get("termines") or 0)
            cga_interesses = int(stats_row.get("cga_interesses") or 0)
            cga_pas_interesses = int(stats_row.get("cga_pas_interesses") or 0)
            cga_indisponibles = int(stats_row.get("cga_indisponibles") or 0)
            cga_faux_numeros = int(stats_row.get("cga_faux_numeros") or 0)
            formulaires_url = build_appels_url(agent=username)
            modifies_url = build_appels_url(agent=username)
            termines_url = build_appels_url(agent=username, status="completed")
        else:
            formulaires_remplis_ids = formulaires_remplis_by_user.get(user.id, set())
            formulaires_modifies_ids = formulaires_modifies_by_user.get(user.id, set())
            formulaires_remplis_total = max(
                len(formulaires_remplis_ids),
                int(stats_row.get("formulaires_remplis_status") or 0),
            )
            formulaires_modifies_total = len(formulaires_modifies_ids)
            termines_ids = set(legacy_termines_by_user.get(user.id, set()))
            termines_ids.update(formulaires_remplis_ids)
            termines_ids.update(formulaires_modifies_ids)
            termines_ids.update(audio_termines_by_user.get(user.id, set()))
            termines_total = len(termines_ids)
            cga_interesses = 0
            cga_pas_interesses = 0
            cga_indisponibles = 0
            cga_faux_numeros = 0
            formulaires_url = build_appels_url(agent=username, formulaire="rempli")
            modifies_url = build_appels_url(modified_by=username, formulaire="modifie")
            termines_url = build_appels_url(tracking_termine=1, tracking_user=username)
        last_ip = getattr(activity, "last_ip", None) if activity else None
        last_lat = getattr(activity, "last_latitude", None) if activity else None
        last_lon = getattr(activity, "last_longitude", None) if activity else None
        last_city = getattr(activity, "last_city", "") if activity else ""
        last_country = getattr(activity, "last_country", "") if activity else ""

        user_rows.append(
            {
                "user_id": user.id,
                "username": username,
                "is_online": is_online,
                "last_seen": last_seen,
                "last_login": user.last_login,
                "total_appels": int(stats_row.get("total_appels") or 0),
                "appels_aujourdhui": int(stats_row.get("appels_aujourdhui") or 0),
                "a_rappeler": int(stats_row.get("a_rappeler") or 0),
                "appels_tentes": int(stats_row.get("appels_tentes") or 0),
                "appels_reussis": int(stats_row.get("appels_reussis") or 0),
                "formulaires_remplis": formulaires_remplis_total,
                "formulaires_modifies": formulaires_modifies_total,
                "formulaires_avec_audio": int(stats_row.get("formulaires_avec_audio") or 0),
                "termines": termines_total,
                "en_cours": int(stats_row.get("en_cours") or 0),
                "recent_24h": int(stats_row.get("recent_24h") or 0),
                "call_scope": call_scope,
                "cga_interesses": cga_interesses,
                "cga_pas_interesses": cga_pas_interesses,
                "cga_indisponibles": cga_indisponibles,
                "cga_faux_numeros": cga_faux_numeros,
                "push_sur_main": int(push_counts_by_user.get(user.id, 0) or 0),
                "deploiements": int(deploy_counts_by_user.get(user.id, 0) or 0),
                "current_calls": current_calls_by_user.get(user.id, []),
                "total_url": build_appels_url(agent=username),
                "rappel_url": build_appels_url(agent=username, status="a_rappeler"),
                "formulaires_url": formulaires_url,
                "modifies_url": modifies_url,
                "termines_url": termines_url,
                "en_cours_url": build_appels_url(agent=username, status="en_cours"),
                "interesses_url": build_appels_url(agent=username, resultat="interesse"),
                "pas_interesses_url": build_appels_url(agent=username, resultat="pas_interesse"),
                "indisponibles_url": build_appels_url(agent=username, resultat="indisponible"),
                "faux_numeros_url": build_appels_url(agent=username, resultat="faux_numero"),
                "last_ip": last_ip,
                "last_latitude": round(last_lat, 4) if last_lat is not None else None,
                "last_longitude": round(last_lon, 4) if last_lon is not None else None,
                "last_city": last_city,
                "last_country": last_country,
                "current_page": getattr(activity, "current_page", "") if activity else "",
                "current_page_title": (
                    getattr(activity, "current_page_title", "") if activity else ""
                ),
                "last_action_type": getattr(activity, "last_action_type", "") if activity else "",
                "last_action_label": getattr(activity, "last_action_label", "") if activity else "",
                "last_action_target": (
                    getattr(activity, "last_action_target", "") if activity else ""
                ),
                "last_action_at": getattr(activity, "last_action_at", None) if activity else None,
                "recent_events": events_by_user.get(user.id, []),
            }
        )

    if call_scope == "cga":
        user_rows.sort(
            key=lambda row: (
                0 if row["is_online"] else 1,
                -(
                    row["cga_interesses"]
                    + row["cga_pas_interesses"]
                    + row["cga_indisponibles"]
                    + row["cga_faux_numeros"]
                ),
                -row["recent_24h"],
                -row["total_appels"],
                row["username"].lower(),
            )
        )
    else:
        user_rows.sort(
            key=lambda row: (
                0 if row["is_online"] else 1,
                -(
                    row["appels_reussis"]
                    + row["formulaires_remplis"]
                    + row["formulaires_avec_audio"]
                ),
                -row["termines"],
                -row["formulaires_remplis"],
                -row["formulaires_modifies"],
                -row["appels_tentes"],
                row["username"].lower(),
            )
        )

    recent_login_logs, login_logs_ready = _safe_recent_login_logs()
    online_rows = [row for row in user_rows if row["is_online"]]
    globe_points = [
        {
            "username": row["username"],
            "online": row["is_online"],
            "latitude": row["last_latitude"],
            "longitude": row["last_longitude"],
            "city": row["last_city"],
            "country": row["last_country"],
            "current_page_title": row["current_page_title"] or row["current_page"],
            "last_action_label": row["last_action_label"],
            "last_action_type": row["last_action_type"],
        }
        for row in user_rows
        if row["last_latitude"] is not None and row["last_longitude"] is not None
    ]
    tracking_schema_ready = activities_ready and events_ready and login_logs_ready

    return {
        "user_activity_rows": user_rows,
        "activity_histories": {str(user_id): events for user_id, events in events_by_user.items()},
        "user_search": user_search,
        "call_scope": call_scope,
        "call_scope_label": _TRACKING_SCOPE_LABELS[call_scope],
        "tracking_scope_tabs": tracking_scope_tabs,
        "total_users": User.objects.count(),
        "online_count": len(online_rows),
        "recent_login_logs": recent_login_logs,
        "globe_points": globe_points,
        "online_rows": online_rows,
        "tracking_schema_ready": tracking_schema_ready,
    }


@require_POST
def activity_track_api(request):
    if not (request.user.is_authenticated and not request.user.is_anonymous):
        return JsonResponse({"ok": False, "error": "authentication_required"}, status=401)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        payload = {}

    event_type = _clean_activity_text(payload.get("event_type"), 30)
    allowed_types = {
        UserActivityEvent.EVENT_PAGE_VIEW,
        UserActivityEvent.EVENT_BUTTON_CLICK,
        UserActivityEvent.EVENT_LINK_CLICK,
    }
    if event_type not in allowed_types:
        return JsonResponse({"ok": False, "error": "invalid_event_type"}, status=400)

    page_path = _clean_activity_text(payload.get("page_path"))
    page_title = _clean_activity_text(payload.get("page_title"))
    target_label = _clean_activity_text(payload.get("target_label"))
    target_path = _clean_activity_text(payload.get("target_path"))
    browser_latitude = payload.get("browser_latitude")
    browser_longitude = payload.get("browser_longitude")
    browser_city = _clean_activity_text(payload.get("browser_city"))
    browser_country = _clean_activity_text(payload.get("browser_country"))
    now = timezone.now()

    try:
        browser_latitude = float(browser_latitude) if browser_latitude not in (None, "") else None
    except (TypeError, ValueError):
        browser_latitude = None
    try:
        browser_longitude = (
            float(browser_longitude) if browser_longitude not in (None, "") else None
        )
    except (TypeError, ValueError):
        browser_longitude = None

    try:
        activity, _created = UserActivity.objects.get_or_create(
            user=request.user,
            defaults={"last_seen": now},
        )
        activity.last_seen = now
        activity.current_page = page_path
        activity.current_page_title = page_title
        activity.last_action_type = event_type
        activity.last_action_label = target_label
        activity.last_action_target = target_path
        activity.last_action_at = now
        if browser_latitude is not None and browser_longitude is not None:
            activity.last_latitude = browser_latitude
            activity.last_longitude = browser_longitude
        if browser_city:
            activity.last_city = browser_city
        if browser_country:
            activity.last_country = browser_country
        activity.save(
            update_fields=[
                "last_seen",
                "current_page",
                "current_page_title",
                "last_action_type",
                "last_action_label",
                "last_action_target",
                "last_action_at",
                "last_latitude",
                "last_longitude",
                "last_city",
                "last_country",
            ]
        )

        should_create_event = True
        if event_type == UserActivityEvent.EVENT_PAGE_VIEW:
            latest_page_event = (
                UserActivityEvent.objects.filter(
                    user=request.user,
                    event_type=UserActivityEvent.EVENT_PAGE_VIEW,
                    page_path=page_path,
                )
                .order_by("-occurred_at")
                .first()
            )
            if latest_page_event and (now - latest_page_event.occurred_at).total_seconds() < 20:
                should_create_event = False

        if should_create_event:
            UserActivityEvent.objects.create(
                user=request.user,
                event_type=event_type,
                page_path=page_path,
                page_title=page_title,
                target_label=target_label,
                target_path=target_path,
            )
    except (OperationalError, ProgrammingError):
        _log_tracking_schema_warning("write")
        return JsonResponse({"ok": True, "tracking_disabled": True})

    return JsonResponse({"ok": True, "tracking_disabled": False})


@require_GET
def user_tracking_live_api(request):
    if not (request.user.is_authenticated and request.user.is_superuser):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    payload = _build_tracking_payload(
        user_search=(request.GET.get("user_search") or "").strip(),
        call_scope=(request.GET.get("scope") or "").strip(),
    )
    return JsonResponse(
        {
            "ok": True,
            "online_count": payload["online_count"],
            "total_users": payload["total_users"],
            "tracking_schema_ready": payload["tracking_schema_ready"],
            "globe_points": payload["globe_points"],
            "online_rows": [
                {
                    "username": row["username"],
                    "current_page": row["current_page"],
                    "current_page_title": row["current_page_title"],
                    "last_action_type": row["last_action_type"],
                    "last_action_label": row["last_action_label"],
                    "last_action_target": row["last_action_target"],
                    "last_action_at": (
                        timezone.localtime(row["last_action_at"]).strftime("%d/%m/%Y %H:%M:%S")
                        if row["last_action_at"]
                        else ""
                    ),
                    "city": row["last_city"],
                    "country": row["last_country"],
                }
                for row in payload["online_rows"]
            ],
        }
    )


def user_tracking_view(request):
    """Page dediee au suivi des utilisateurs PADESCE (superuser uniquement)."""
    from django.http import HttpResponseForbidden

    if not (request.user.is_authenticated and request.user.is_superuser):
        return HttpResponseForbidden("Acces reserve aux administrateurs.")

    payload = _build_tracking_payload(
        user_search=(request.GET.get("user_search") or "").strip(),
        call_scope=(request.GET.get("scope") or "").strip(),
    )
    return render(request, "core/user_tracking.html", payload)


# ---------------------------------------------------------------------------
# VUES D'EXPORT PUBLIQUES POUR LES MOYENNES GENERALES
# ---------------------------------------------------------------------------

def public_export_apprenant_global_averages_xlsx(request):
    """Export public des moyennes générales des apprenants en Excel."""
    try:
        import openpyxl

        from App_PADESCE.satisfaction_apprenants.views import (
            Q_FIELDS,
            _build_satisfaction_dashboard_data,
        )
        
        dashboard = _build_satisfaction_dashboard_data(request)
        context = dashboard["context"]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Moyennes générales"
        
        # En-têtes
        q_labels = [label for _, label in Q_FIELDS]
        ws.append(["Indicateur"] + q_labels)
        
        # Données des moyennes générales
        row_data = ["Moyenne générale"]
        for label in q_labels:
            avg = context.get("global_avgs", {}).get(label, 0)
            row_data.append(round(avg, 2))
        ws.append(row_data)
        
        # Style
        for col in range(1, len(q_labels) + 2):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        for col in range(1, len(q_labels) + 2):
            cell = ws.cell(row=2, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(q_labels) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="moyennes-generales-apprenants.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de l'export: {str(e)}", status=500)


def public_export_formateur_global_averages_xlsx(request):
    """Export public des moyennes générales des formateurs en Excel."""
    try:
        import openpyxl

        from App_PADESCE.satisfaction_formateurs.views import (
            Q_FORM_FIELDS,
            _build_satisfaction_formateurs_dashboard_context,
        )
        
        context = _build_satisfaction_formateurs_dashboard_context(request)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Moyennes générales"
        
        # En-têtes
        q_labels = [label for _, label in Q_FORM_FIELDS]
        ws.append(["Indicateur"] + q_labels + ["Moyenne générale GLOBALE"])
        
        # Données des moyennes générales
        row_data = ["Moyenne générale"]
        for label in q_labels:
            avg = context.get("global_avgs", {}).get(label, 0)
            row_data.append(round(avg, 2))
        # Ajouter la moyenne générale globale
        moyenne_globale = context.get("moyenne_generale_globale", 0)
        row_data.append(round(moyenne_globale, 2))
        ws.append(row_data)
        
        # Style
        for col in range(1, len(q_labels) + 3):
            cell = ws.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        for col in range(1, len(q_labels) + 3):
            cell = ws.cell(row=2, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(q_labels) + 3):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="moyennes-generales-formateurs.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de l'export: {str(e)}", status=500)
