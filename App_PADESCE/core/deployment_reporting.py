from __future__ import annotations

import json
import os
import time
import uuid
from html import escape
from pathlib import Path
from typing import Any
from urllib.parse import urlsplit, urlunsplit

from django.conf import settings
from django.core.mail import EmailMultiAlternatives, get_connection

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - handled at runtime if dependency is missing
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = get_column_letter = None


REPORTS_DIRNAME = "reports"
HISTORY_FILENAME = "history.json"
HISTORY_WORKBOOK_FILENAME = "deployments-history.xlsx"

PURPLE_DARK = "5B21B6"
PURPLE_MAIN = "7C3AED"
PURPLE_SOFT = "F5F3FF"
PURPLE_LIGHT = "EDE9FE"
GREEN_SOFT = "DCFCE7"
GREEN_TEXT = "166534"
AMBER_SOFT = "FEF3C7"
AMBER_TEXT = "92400E"
ROSE_SOFT = "FFE4E6"
ROSE_TEXT = "BE123C"
BORDER_COLOR = "DDD6FE"


def deployments_root() -> Path:
    path = Path(settings.BASE_DIR) / "logs" / "deployments"
    path.mkdir(parents=True, exist_ok=True)
    return path


def reports_dir() -> Path:
    path = deployments_root() / REPORTS_DIRNAME
    path.mkdir(parents=True, exist_ok=True)
    return path


def history_path() -> Path:
    return reports_dir() / HISTORY_FILENAME


def history_workbook_path() -> Path:
    return reports_dir() / HISTORY_WORKBOOK_FILENAME


def _save_json(path: Path, payload: dict[str, Any] | list[Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    content = json.dumps(payload, ensure_ascii=True, indent=2)
    last_error: Exception | None = None
    for _attempt in range(5):
        tmp_path = path.with_suffix(path.suffix + f".{uuid.uuid4().hex}.tmp")
        try:
            tmp_path.write_text(content, encoding="utf-8")
            os.replace(tmp_path, path)
            return
        except PermissionError as exc:
            last_error = exc
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass
            time.sleep(0.15)
    if last_error is not None:
        raise last_error
    path.write_text(content, encoding="utf-8")


def _load_history() -> list[dict[str, Any]]:
    if not history_path().exists():
        return []
    try:
        payload = json.loads(history_path().read_text(encoding="utf-8"))
    except Exception:
        return []
    return payload if isinstance(payload, list) else []


def load_history() -> list[dict[str, Any]]:
    return _load_history()


def load_history_entry(report_id: str) -> dict[str, Any] | None:
    normalized_id = str(report_id or "").strip()
    if not normalized_id:
        return None
    for item in _load_history():
        if str(item.get("id", "") or "").strip() == normalized_id:
            return item
    return None


def load_report(report_id: str) -> dict[str, Any] | None:
    path = reports_dir() / f"{str(report_id or '').strip()}.json"
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return payload if isinstance(payload, dict) else None


def deployment_recipients() -> list[str]:
    raw = str(getattr(settings, "DEPLOYMENT_REPORT_EMAIL_TO", "") or "").strip()
    if not raw:
        raw = str(getattr(settings, "REPORT_EMAIL_TO", "") or "").strip()
    return [item.strip() for item in raw.split(",") if item.strip()]


def _status_label(status: str) -> str:
    mapping = {
        "completed": "Succes",
        "failed": "Erreur",
        "running": "En cours",
        "pending": "En attente",
        "ok": "OK",
        "error": "Erreur",
        "warning": "Attention",
        "info": "Information",
        "not_required": "Non requis",
    }
    return mapping.get(str(status or "").strip().lower(), str(status or "-"))


def _mode_label(mode: str) -> str:
    return {
        "deploy": "Deploiement reel",
        "preview": "Previsualisation",
    }.get(str(mode or "").strip().lower(), str(mode or "-"))


def _base_site_url(url: str) -> str:
    parsed = urlsplit(str(url or "").strip())
    if not parsed.scheme or not parsed.netloc:
        return ""
    return urlunsplit((parsed.scheme, parsed.netloc, "", "", ""))


def build_links(run: dict[str, Any]) -> dict[str, str]:
    config = run.get("config", {}) or {}
    verify_url = str(config.get("verify_url", "") or "").strip()
    base_url = _base_site_url(verify_url)
    site_url = base_url or verify_url
    return {
        "site_url": site_url,
        "deployment_page_url": f"{base_url}/deploiement/" if base_url else "",
        "live_status_url": f"{base_url}/deploiement/live/" if base_url else "",
    }


def build_checks(run: dict[str, Any]) -> list[dict[str, Any]]:
    summary = run.get("summary", {}) or {}
    verification = run.get("verification", {}) or {}
    live_refresh = verification.get("live_refresh", {}) or {}
    http_check = verification.get("http_check", {}) or {}
    checks: list[dict[str, Any]] = [
        {
            "label": "Configuration du deploiement",
            "status": "ok" if (run.get("config", {}) or {}).get("ready") else "error",
            "detail": (
                "Les acces serveur necessaires etaient disponibles."
                if (run.get("config", {}) or {}).get("ready")
                else "Une information serveur manquait au lancement."
            ),
        },
        {
            "label": "Resultat final",
            "status": "ok" if run.get("status") == "completed" else "error",
            "detail": (
                "Le pipeline s'est termine sans erreur bloquante."
                if run.get("status") == "completed"
                else (run.get("error") or "Le pipeline s'est arrete avant la fin.")
            ),
        },
        {
            "label": "Changements prepares",
            "status": "info",
            "detail": (
                f"{summary.get('additions', 0)} ajout(s), "
                f"{summary.get('modifications', 0)} modification(s), "
                f"{summary.get('deletions', 0)} suppression(s)."
            ),
        },
        {
            "label": "Fichiers deja presents mais non encore suivis",
            "status": "warning" if int(summary.get("remote_untracked", 0) or 0) else "info",
            "detail": f"{summary.get('remote_untracked', 0)} element(s) restent hors suivi automatique pour le moment.",
        },
        {
            "label": "Manifeste distant relu",
            "status": "ok" if verification.get("remote_manifest") else "warning",
            "detail": (
                "La reference distante a bien ete relue apres transfert."
                if verification.get("remote_manifest")
                else "La reference distante n'a pas pu etre relue apres transfert."
            ),
        },
    ]

    if live_refresh:
        if live_refresh.get("required") is False:
            live_status = "not_required"
        else:
            live_status = "ok" if live_refresh.get("reloaded") else "error"
        checks.append(
            {
                "label": "Serveur Python recharge",
                "status": live_status,
                "detail": str(live_refresh.get("message", "") or "Aucune information disponible."),
            }
        )

    if http_check:
        checks.append(
            {
                "label": "Site public joignable",
                "status": "ok" if http_check.get("ok") else "error",
                "detail": (
                    f"HTTP {http_check.get('status_code')} via {http_check.get('final_url', '-')}"
                    if http_check.get("status_code")
                    else str(http_check.get("message", "Verification web indisponible."))
                ),
            }
        )

    errors = verification.get("errors", []) or []
    checks.append(
        {
            "label": "Erreurs de controle",
            "status": "ok" if not errors else "error",
            "detail": (
                "Aucune erreur de controle."
                if not errors
                else " ; ".join(str(item) for item in errors)
            ),
        }
    )
    return checks


def build_plain_language(run: dict[str, Any], links: dict[str, str]) -> dict[str, Any]:
    summary = run.get("summary", {}) or {}
    verification = run.get("verification", {}) or {}
    live_refresh = verification.get("live_refresh", {}) or {}
    http_check = verification.get("http_check", {}) or {}

    headline = (
        "Le deploiement est termine."
        if run.get("status") == "completed"
        else "Le deploiement a rencontre un probleme."
    )
    change_sentence = (
        f"Le site devait recevoir {summary.get('additions', 0)} nouveaute(s), "
        f"{summary.get('modifications', 0)} mise(s) a jour et "
        f"{summary.get('deletions', 0)} suppression(s) suivie(s)."
    )

    if live_refresh.get("required") is False:
        refresh_sentence = (
            "Le serveur Python n'avait pas besoin d'etre recharge pour cette operation."
        )
    elif live_refresh.get("reloaded"):
        refresh_sentence = (
            "Le serveur Python a confirme qu'il avait bien recharge la nouvelle version."
        )
    else:
        refresh_sentence = (
            "Le serveur Python n'a pas encore confirme son rechargement automatique. "
            "Le site peut donc ne pas afficher toutes les mises a jour."
        )

    if http_check.get("status_code"):
        website_sentence = (
            f"Le site public a repondu avec le code HTTP {http_check.get('status_code')}."
        )
    else:
        website_sentence = str(
            http_check.get("message", "La verification du site public n'a pas abouti.")
        )

    untracked_count = int(summary.get("remote_untracked", 0) or 0)
    untracked_sentence = (
        f"{untracked_count} fichier(s) deja presents sur le serveur restent pour l'instant hors suivi automatique."
        if untracked_count
        else "Tous les fichiers connus par ce pipeline etaient bien sous controle."
    )

    sentences = [change_sentence, refresh_sentence, website_sentence, untracked_sentence]
    if links.get("deployment_page_url"):
        sentences.append(f"Le suivi detaille reste consultable sur {links['deployment_page_url']}")

    if run.get("status") != "completed":
        next_action = "Verifier le journal et relancer un deploiement une fois la cause corrigee."
    elif live_refresh.get("required") and not live_refresh.get("reloaded"):
        next_action = "Verifier le rechargement du serveur Python avant de communiquer que le site est a jour."
    else:
        next_action = "Aucune action urgente n'est necessaire."

    return {
        "headline": headline,
        "sentences": sentences,
        "next_action": next_action,
    }


def build_report(run: dict[str, Any]) -> dict[str, Any]:
    links = build_links(run)
    report = {
        "id": run.get("id", ""),
        "mode": run.get("mode", ""),
        "mode_label": _mode_label(run.get("mode", "")),
        "status": run.get("status", ""),
        "status_label": _status_label(run.get("status", "")),
        "started_at": run.get("started_at", ""),
        "completed_at": run.get("completed_at", ""),
        "progress_pct": run.get("progress_pct", 0),
        "error": run.get("error", ""),
        "summary": run.get("summary", {}) or {},
        "verification": run.get("verification", {}) or {},
        "checks": build_checks(run),
        "steps": run.get("steps", []) or [],
        "diff": run.get("diff", {}) or {},
        "logs": run.get("logs", []) or [],
        "links": links,
    }
    report["plain_language"] = build_plain_language(run, links)
    return report


def render_report_markdown(report: dict[str, Any]) -> str:
    lines: list[str] = []
    lines.append(f"# Rapport de deploiement {report.get('id', '')}")
    lines.append("")
    lines.append(f"- Operation: {report.get('mode_label', '-')}")
    lines.append(f"- Resultat: {report.get('status_label', '-')}")
    lines.append(f"- Debut: {report.get('started_at', '-')}")
    lines.append(f"- Fin: {report.get('completed_at', '-')}")
    lines.append(f"- Progression: {report.get('progress_pct', 0)}%")
    if report.get("links", {}).get("deployment_page_url"):
        lines.append(f"- Page de suivi: {report['links']['deployment_page_url']}")
    if report.get("error"):
        lines.append(f"- Erreur: {report.get('error')}")
    lines.append("")
    lines.append("## En clair")
    lines.append("")
    lines.append(report.get("plain_language", {}).get("headline", ""))
    lines.append("")
    for sentence in report.get("plain_language", {}).get("sentences", []) or []:
        lines.append(f"- {sentence}")
    lines.append(f"- Suite recommandee: {report.get('plain_language', {}).get('next_action', '-')}")
    lines.append("")
    lines.append("## Controles")
    lines.append("")
    for item in report.get("checks", []):
        lines.append(
            f"- {item.get('label')}: {_status_label(item.get('status', ''))} | {item.get('detail', '')}"
        )
    lines.append("")
    lines.append("## Ce qui change sur le site")
    lines.append("")
    summary = report.get("summary", {}) or {}
    lines.append(f"- Ajouts: {summary.get('additions', 0)}")
    lines.append(f"- Modifications: {summary.get('modifications', 0)}")
    lines.append(f"- Suppressions: {summary.get('deletions', 0)}")
    lines.append(f"- Fichiers non suivis: {summary.get('remote_untracked', 0)}")
    lines.append(f"- Racine distante: {summary.get('remote_path', '-')}")
    lines.append("")
    lines.append("## Etapes")
    lines.append("")
    for step in report.get("steps", []):
        lines.append(
            f"- {step.get('label')}: {_status_label(step.get('status', ''))} | {step.get('message', '')} | "
            f"debut={step.get('started_at', '-')} | fin={step.get('completed_at', '-')}"
        )
    lines.append("")
    lines.append("## Fichiers")
    lines.append("")
    for label, key in (
        ("Ajouts", "additions"),
        ("Modifications", "modifications"),
        ("Suppressions", "deletions"),
        ("Distants non suivis", "remote_untracked"),
    ):
        lines.append(f"### {label}")
        values = (report.get("diff", {}) or {}).get(key, []) or []
        if not values:
            lines.append("- Aucun element")
        else:
            for item in values:
                lines.append(f"- {item}")
        lines.append("")
    lines.append("## Journal")
    lines.append("")
    for entry in report.get("logs", []):
        lines.append(
            f"- [{entry.get('at', '-')}] {entry.get('level', 'info')}: {entry.get('message', '')}"
        )
    lines.append("")
    return "\n".join(lines)


def _workbook_available() -> bool:
    return (
        Workbook is not None
        and Alignment is not None
        and Border is not None
        and Font is not None
        and PatternFill is not None
        and Side is not None
    )


def _cell_fill(color: str):
    return PatternFill(fill_type="solid", fgColor=color)


def _thin_border():
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def _style_title(sheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = _cell_fill(PURPLE_DARK)
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 2


def _style_section(sheet, row: int, title: str) -> int:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = sheet.cell(row=row, column=1, value=title)
    cell.fill = _cell_fill(PURPLE_LIGHT)
    cell.font = Font(color=PURPLE_DARK, bold=True)
    cell.border = _thin_border()
    return row + 1


def _write_label_value(sheet, row: int, label: str, value: Any, *, url: str = "") -> int:
    label_cell = sheet.cell(row=row, column=1, value=label)
    label_cell.font = Font(bold=True, color=PURPLE_DARK)
    label_cell.fill = _cell_fill(PURPLE_SOFT)
    label_cell.border = _thin_border()
    value_cell = sheet.cell(row=row, column=2, value=str(value or "-"))
    value_cell.border = _thin_border()
    value_cell.alignment = Alignment(wrap_text=True, vertical="top")
    if url:
        value_cell.value = url
        value_cell.hyperlink = url
        value_cell.style = "Hyperlink"
    return row + 1


def _write_table_header(sheet, row: int, headers: list[str]) -> int:
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=index, value=header)
        cell.fill = _cell_fill(PURPLE_MAIN)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return row + 1


def _status_style(status: str) -> tuple[str, str]:
    normalized = str(status or "").strip().lower()
    if normalized in {"ok", "completed", "success", "not_required"}:
        if normalized == "not_required":
            return PURPLE_SOFT, PURPLE_DARK
        return GREEN_SOFT, GREEN_TEXT
    if normalized in {"warning", "info", "running", "pending"}:
        return AMBER_SOFT, AMBER_TEXT
    return ROSE_SOFT, ROSE_TEXT


def _fit_columns(sheet, *, max_width: int = 52) -> None:
    for index, column in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
        letter = get_column_letter(index)
        length = 0
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        sheet.column_dimensions[letter].width = min(max(length + 2, 12), max_width)


def _build_run_workbook(report: dict[str, Any]) -> Workbook | None:
    if not _workbook_available():
        return None

    workbook = Workbook()
    resume_sheet = workbook.active
    resume_sheet.title = "Resume"
    row = 1
    row = _style_title(resume_sheet, row, f"Rapport de deploiement {report.get('id', '')}")
    row = _style_section(resume_sheet, row, "Vue simple")
    row = _write_label_value(resume_sheet, row, "Operation", report.get("mode_label", "-"))
    row = _write_label_value(resume_sheet, row, "Resultat", report.get("status_label", "-"))
    row = _write_label_value(resume_sheet, row, "Debut", report.get("started_at", "-"))
    row = _write_label_value(resume_sheet, row, "Fin", report.get("completed_at", "-"))
    row = _write_label_value(
        resume_sheet,
        row,
        "Page de suivi",
        report.get("links", {}).get("deployment_page_url", "-"),
        url=report.get("links", {}).get("deployment_page_url", ""),
    )
    row = _write_label_value(
        resume_sheet,
        row,
        "Site public",
        report.get("links", {}).get("site_url", "-"),
        url=report.get("links", {}).get("site_url", ""),
    )
    row += 1
    row = _style_section(resume_sheet, row, "Explication simple")
    headline_cell = resume_sheet.cell(
        row=row, column=1, value=report.get("plain_language", {}).get("headline", "")
    )
    headline_cell.font = Font(bold=True, color=PURPLE_DARK, size=12)
    resume_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    for sentence in report.get("plain_language", {}).get("sentences", []):
        resume_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = resume_sheet.cell(row=row, column=1, value=f"- {sentence}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    resume_sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    next_action_cell = resume_sheet.cell(
        row=row,
        column=1,
        value=f"Suite recommandee: {report.get('plain_language', {}).get('next_action', '-')}",
    )
    next_action_cell.fill = _cell_fill(PURPLE_SOFT)
    next_action_cell.border = _thin_border()
    next_action_cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 2
    row = _style_section(resume_sheet, row, "Ce qui change sur le site")
    summary = report.get("summary", {}) or {}
    row = _write_label_value(resume_sheet, row, "Nouveautes", summary.get("additions", 0))
    row = _write_label_value(resume_sheet, row, "Mises a jour", summary.get("modifications", 0))
    row = _write_label_value(resume_sheet, row, "Suppressions suivies", summary.get("deletions", 0))
    row = _write_label_value(
        resume_sheet, row, "Elements non encore suivis", summary.get("remote_untracked", 0)
    )
    total_row = row
    row = _write_label_value(resume_sheet, row, "Total des changements suivis", "")
    resume_sheet.cell(row=total_row, column=2, value=f"=SUM(B{total_row - 4}:B{total_row - 2})")
    row += 1
    row = _style_section(resume_sheet, row, "Ce qui a ete controle")
    row = _write_table_header(resume_sheet, row, ["Controle", "Resultat", "Explication"])
    for item in report.get("checks", []):
        fill_color, text_color = _status_style(item.get("status", ""))
        resume_sheet.cell(row=row, column=1, value=item.get("label", "")).border = _thin_border()
        status_cell = resume_sheet.cell(
            row=row, column=2, value=_status_label(item.get("status", ""))
        )
        status_cell.fill = _cell_fill(fill_color)
        status_cell.font = Font(color=text_color, bold=True)
        status_cell.border = _thin_border()
        detail_cell = resume_sheet.cell(row=row, column=3, value=item.get("detail", ""))
        detail_cell.alignment = Alignment(wrap_text=True, vertical="top")
        detail_cell.border = _thin_border()
        row += 1

    controles_sheet = workbook.create_sheet("Controles")
    row = 1
    row = _style_title(controles_sheet, row, "Controles automatiques")
    row = _write_table_header(controles_sheet, row, ["Controle", "Resultat", "Explication"])
    for item in report.get("checks", []):
        fill_color, text_color = _status_style(item.get("status", ""))
        controles_sheet.cell(row=row, column=1, value=item.get("label", "")).border = _thin_border()
        status_cell = controles_sheet.cell(
            row=row, column=2, value=_status_label(item.get("status", ""))
        )
        status_cell.fill = _cell_fill(fill_color)
        status_cell.font = Font(color=text_color, bold=True)
        status_cell.border = _thin_border()
        detail_cell = controles_sheet.cell(row=row, column=3, value=item.get("detail", ""))
        detail_cell.border = _thin_border()
        detail_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    files_sheet = workbook.create_sheet("Fichiers")
    row = 1
    row = _style_title(files_sheet, row, "Liste des fichiers concernes")
    row = _write_table_header(files_sheet, row, ["Type", "Element"])
    diff = report.get("diff", {}) or {}
    for label, key in (
        ("Ajout", "additions"),
        ("Modification", "modifications"),
        ("Suppression", "deletions"),
        ("Non suivi", "remote_untracked"),
    ):
        values = diff.get(key, []) or []
        if not values:
            files_sheet.cell(row=row, column=1, value=label).border = _thin_border()
            files_sheet.cell(row=row, column=2, value="Aucun element").border = _thin_border()
            row += 1
            continue
        for value in values:
            files_sheet.cell(row=row, column=1, value=label).border = _thin_border()
            file_cell = files_sheet.cell(row=row, column=2, value=value)
            file_cell.border = _thin_border()
            row += 1

    steps_sheet = workbook.create_sheet("Etapes")
    row = 1
    row = _style_title(steps_sheet, row, "Pipeline detaille")
    row = _write_table_header(steps_sheet, row, ["Etape", "Statut", "Message", "Debut", "Fin"])
    for step in report.get("steps", []):
        fill_color, text_color = _status_style(step.get("status", ""))
        steps_sheet.cell(row=row, column=1, value=step.get("label", "")).border = _thin_border()
        status_cell = steps_sheet.cell(
            row=row, column=2, value=_status_label(step.get("status", ""))
        )
        status_cell.fill = _cell_fill(fill_color)
        status_cell.font = Font(color=text_color, bold=True)
        status_cell.border = _thin_border()
        for column, value in enumerate(
            [step.get("message", ""), step.get("started_at", ""), step.get("completed_at", "")],
            start=3,
        ):
            cell = steps_sheet.cell(row=row, column=column, value=value)
            cell.border = _thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    logs_sheet = workbook.create_sheet("Journal")
    row = 1
    row = _style_title(logs_sheet, row, "Journal du deploiement")
    row = _write_table_header(logs_sheet, row, ["Date", "Niveau", "Message"])
    for entry in report.get("logs", []):
        logs_sheet.cell(row=row, column=1, value=entry.get("at", "")).border = _thin_border()
        level_cell = logs_sheet.cell(row=row, column=2, value=entry.get("level", "info"))
        fill_color, text_color = _status_style(
            "error"
            if entry.get("level") == "error"
            else "warning" if entry.get("level") == "warning" else "info"
        )
        level_cell.fill = _cell_fill(fill_color)
        level_cell.font = Font(color=text_color, bold=True)
        level_cell.border = _thin_border()
        message_cell = logs_sheet.cell(row=row, column=3, value=entry.get("message", ""))
        message_cell.alignment = Alignment(wrap_text=True, vertical="top")
        message_cell.border = _thin_border()
        row += 1

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        _fit_columns(sheet)
    return workbook


def _build_history_workbook(history: list[dict[str, Any]]) -> Workbook | None:
    if not _workbook_available():
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Historique"
    row = 1
    row = _style_title(sheet, row, "Historique des deploiements")
    row = _write_table_header(
        sheet,
        row,
        [
            "Run",
            "Operation",
            "Resultat",
            "Debut",
            "Fin",
            "Ajouts",
            "Modifications",
            "Suppressions",
            "Serveur recharge",
            "Page de suivi",
        ],
    )
    for entry in history:
        summary = entry.get("summary", {}) or {}
        live_refresh = entry.get("live_refresh", {}) or {}
        fill_color, text_color = _status_style(
            "ok" if entry.get("status") == "completed" else "error"
        )
        values = [
            entry.get("id", ""),
            _mode_label(entry.get("mode", "")),
            _status_label(entry.get("status", "")),
            entry.get("started_at", ""),
            entry.get("completed_at", ""),
            summary.get("additions", 0),
            summary.get("modifications", 0),
            summary.get("deletions", 0),
            (
                "Oui"
                if live_refresh.get("reloaded")
                else "Non" if live_refresh.get("required", True) else "Non requis"
            ),
            entry.get("deployment_page_url", ""),
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = _thin_border()
            if column == 3:
                cell.fill = _cell_fill(fill_color)
                cell.font = Font(color=text_color, bold=True)
            if column == 10 and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
        row += 1
    sheet.freeze_panes = "A2"
    _fit_columns(sheet)
    return workbook


def write_report_files(report: dict[str, Any]) -> dict[str, str]:
    report_id = str(report.get("id", "") or "unknown")
    json_path = reports_dir() / f"{report_id}.json"
    md_path = reports_dir() / f"{report_id}.md"
    xlsx_path = reports_dir() / f"{report_id}.xlsx"
    _save_json(json_path, report)
    md_path.write_text(render_report_markdown(report), encoding="utf-8")
    workbook = _build_run_workbook(report)
    if workbook is not None:
        workbook.save(xlsx_path)
    return {
        "json_path": str(json_path),
        "markdown_path": str(md_path),
        "xlsx_path": str(xlsx_path) if xlsx_path.exists() else "",
    }


def update_history(report: dict[str, Any], report_paths: dict[str, str]) -> dict[str, str]:
    history = _load_history()
    verification = report.get("verification", {}) or {}
    live_refresh = verification.get("live_refresh", {}) or {}
    links = report.get("links", {}) or {}
    entry = {
        "id": report.get("id", ""),
        "mode": report.get("mode", ""),
        "status": report.get("status", ""),
        "started_at": report.get("started_at", ""),
        "completed_at": report.get("completed_at", ""),
        "summary": report.get("summary", {}) or {},
        "error": report.get("error", ""),
        "json_path": report_paths.get("json_path", ""),
        "markdown_path": report_paths.get("markdown_path", ""),
        "xlsx_path": report_paths.get("xlsx_path", ""),
        "deployment_page_url": links.get("deployment_page_url", ""),
        "site_url": links.get("site_url", ""),
        "live_refresh": live_refresh,
    }
    history = [item for item in history if item.get("id") != entry["id"]]
    history.insert(0, entry)
    _save_json(history_path(), history)
    workbook = _build_history_workbook(history)
    workbook_path = history_workbook_path()
    if workbook is not None:
        workbook.save(workbook_path)
    return {
        "json_path": str(history_path()),
        "xlsx_path": str(workbook_path) if workbook_path.exists() else "",
    }


def render_email_body(report: dict[str, Any]) -> str:
    plain = report.get("plain_language", {}) or {}
    links = report.get("links", {}) or {}
    body: list[str] = []
    body.append(f"Rapport de deploiement {report.get('id', '')}")
    body.append("")
    body.append(plain.get("headline", ""))
    body.append("")
    body.append(f"Operation: {report.get('mode_label', '-')}")
    body.append(f"Resultat: {report.get('status_label', '-')}")
    body.append(f"Debut: {report.get('started_at', '-')}")
    body.append(f"Fin: {report.get('completed_at', '-')}")
    body.append("")
    body.append("En clair:")
    for sentence in plain.get("sentences", []):
        body.append(f"- {sentence}")
    body.append(f"- Suite recommandee: {plain.get('next_action', '-')}")
    body.append("")
    body.append("Controles:")
    for item in report.get("checks", []):
        body.append(
            f"- {item.get('label')}: {_status_label(item.get('status', ''))} - {item.get('detail', '')}"
        )
    body.append("")
    if links.get("deployment_page_url"):
        body.append(f"Page de suivi: {links['deployment_page_url']}")
    if links.get("site_url"):
        body.append(f"Site public: {links['site_url']}")
    body.append("")
    body.append("Les rapports JSON, Markdown et Excel sont joints a cet email.")
    return "\n".join(body)


def render_email_html(report: dict[str, Any]) -> str:
    plain = report.get("plain_language", {}) or {}
    links = report.get("links", {}) or {}
    checks_html = "".join(
        (
            f"<tr>"
            f'<td style="padding:10px 12px;border-bottom:1px solid #ede9fe;color:#1f2937;font-weight:600;">{escape(str(item.get("label", "")))}</td>'
            f'<td style="padding:10px 12px;border-bottom:1px solid #ede9fe;color:#5b21b6;font-weight:700;">{escape(_status_label(item.get("status", "")))}</td>'
            f'<td style="padding:10px 12px;border-bottom:1px solid #ede9fe;color:#4b5563;">{escape(str(item.get("detail", "")))}</td>'
            f"</tr>"
        )
        for item in report.get("checks", [])
    )
    sentences_html = "".join(
        f'<li style="margin:0 0 8px;color:#374151;">{escape(str(sentence))}</li>'
        for sentence in plain.get("sentences", [])
    )
    logs_html = "".join(
        f'<li style="margin:0 0 6px;color:#4b5563;">[{escape(str(entry.get("at", "-")))}] {escape(str(entry.get("message", "")))}</li>'
        for entry in (report.get("logs", []) or [])[-8:]
    )
    buttons: list[str] = []
    if links.get("deployment_page_url"):
        buttons.append(
            f'<a href="{escape(links["deployment_page_url"])}" '
            'style="display:inline-block;padding:10px 16px;border-radius:999px;background:#6d28d9;color:#ffffff;text-decoration:none;font-weight:700;">Ouvrir la page de suivi</a>'
        )
    if links.get("site_url"):
        buttons.append(
            f'<a href="{escape(links["site_url"])}" '
            'style="display:inline-block;padding:10px 16px;border-radius:999px;background:#ede9fe;color:#5b21b6;text-decoration:none;font-weight:700;">Voir le site public</a>'
        )
    buttons_html = "&nbsp;".join(buttons)
    return f"""
<!DOCTYPE html>
<html lang="fr">
  <body style="margin:0;padding:24px;background:#f5f3ff;font-family:Arial,sans-serif;color:#1f2937;">
    <div style="max-width:760px;margin:0 auto;background:#ffffff;border:1px solid #ddd6fe;border-radius:24px;overflow:hidden;">
      <div style="padding:28px;background:linear-gradient(135deg,#4c1d95 0%,#7c3aed 100%);color:#ffffff;">
        <div style="font-size:13px;letter-spacing:0.08em;text-transform:uppercase;opacity:0.88;">PADESCE</div>
        <h1 style="margin:8px 0 8px;font-size:28px;line-height:1.2;">Rapport de deploiement</h1>
        <p style="margin:0;font-size:16px;line-height:1.5;">{escape(str(plain.get("headline", "")))}</p>
      </div>
      <div style="padding:24px;">
        <div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:20px;">
          <div style="flex:1 1 180px;padding:16px;border-radius:18px;background:#faf5ff;border:1px solid #ede9fe;">
            <div style="font-size:12px;text-transform:uppercase;letter-spacing:0.08em;color:#7c3aed;">Operation</div>
            <div style="margin-top:6px;font-size:20px;font-weight:700;color:#4c1d95;">{escape(str(report.get("mode_label", "-")))}</div>
          </div>
          <div style="flex:1 1 180px;padding:16px;border-radius:18px;background:#faf5ff;border:1px solid #ede9fe;">
            <div style="font-size:12px;text-transform:uppercase;letter-spacing:0.08em;color:#7c3aed;">Resultat</div>
            <div style="margin-top:6px;font-size:20px;font-weight:700;color:#4c1d95;">{escape(str(report.get("status_label", "-")))}</div>
          </div>
          <div style="flex:1 1 180px;padding:16px;border-radius:18px;background:#faf5ff;border:1px solid #ede9fe;">
            <div style="font-size:12px;text-transform:uppercase;letter-spacing:0.08em;color:#7c3aed;">Reference</div>
            <div style="margin-top:6px;font-size:20px;font-weight:700;color:#4c1d95;">{escape(str(report.get("id", "-")))}</div>
          </div>
        </div>
        <div style="padding:20px;border-radius:20px;background:#fdfbff;border:1px solid #ede9fe;margin-bottom:20px;">
          <h2 style="margin:0 0 12px;color:#5b21b6;font-size:18px;">Explication simple</h2>
          <ul style="margin:0 0 12px 18px;padding:0;">{sentences_html}</ul>
          <p style="margin:0;color:#4b5563;"><strong>Suite recommandee:</strong> {escape(str(plain.get("next_action", "-")))}</p>
        </div>
        <div style="margin-bottom:20px;">{buttons_html}</div>
        <div style="padding:20px;border-radius:20px;background:#ffffff;border:1px solid #ede9fe;margin-bottom:20px;">
          <h2 style="margin:0 0 12px;color:#5b21b6;font-size:18px;">Controles automatiques</h2>
          <table style="width:100%;border-collapse:collapse;border:1px solid #ede9fe;">
            <thead>
              <tr style="background:#7c3aed;color:#ffffff;">
                <th style="padding:10px 12px;text-align:left;">Controle</th>
                <th style="padding:10px 12px;text-align:left;">Resultat</th>
                <th style="padding:10px 12px;text-align:left;">Explication</th>
              </tr>
            </thead>
            <tbody>{checks_html}</tbody>
          </table>
        </div>
        <div style="padding:20px;border-radius:20px;background:#ffffff;border:1px solid #ede9fe;">
          <h2 style="margin:0 0 12px;color:#5b21b6;font-size:18px;">Journal recent</h2>
          <ul style="margin:0 0 12px 18px;padding:0;">{logs_html or '<li style="color:#6b7280;">Aucun journal disponible.</li>'}</ul>
          <p style="margin:0;color:#6b7280;">Les fichiers JSON, Markdown et Excel sont joints a cet email pour archivage.</p>
        </div>
      </div>
    </div>
  </body>
</html>
""".strip()


def send_report_email(
    report: dict[str, Any], report_paths: dict[str, str], history_files: dict[str, str]
) -> dict[str, Any]:
    recipients = deployment_recipients()
    if not recipients:
        return {"sent": False, "error": "Aucun destinataire configure."}

    subject = (
        f"[PADESCE] {_status_label(report.get('status', ''))} - "
        f"{report.get('mode_label', report.get('mode', ''))} - {report.get('id', '')}"
    )
    connection = get_connection(
        backend=getattr(
            settings, "DEPLOYMENT_EMAIL_BACKEND", "django.core.mail.backends.smtp.EmailBackend"
        ),
        host=getattr(settings, "EMAIL_HOST", ""),
        port=getattr(settings, "EMAIL_PORT", 0),
        username=getattr(settings, "EMAIL_HOST_USER", ""),
        password=getattr(settings, "EMAIL_HOST_PASSWORD", ""),
        use_tls=getattr(settings, "EMAIL_USE_TLS", False),
        fail_silently=False,
    )
    message = EmailMultiAlternatives(
        subject=subject,
        body=render_email_body(report),
        from_email=getattr(settings, "DEFAULT_FROM_EMAIL", "")
        or getattr(settings, "EMAIL_HOST_USER", ""),
        to=recipients,
        connection=connection,
    )
    message.attach_alternative(render_email_html(report), "text/html")
    for key in ("json_path", "markdown_path", "xlsx_path"):
        path = report_paths.get(key)
        if path:
            message.attach_file(path)
    history_xlsx = history_files.get("xlsx_path", "")
    if history_xlsx:
        message.attach_file(history_xlsx)
    message.send(fail_silently=False)
    return {"sent": True, "recipients": recipients, "subject": subject}


def record_and_notify_deployment(run: dict[str, Any]) -> dict[str, Any]:
    report = build_report(run)
    report_paths = write_report_files(report)
    history_files = update_history(report, report_paths)
    try:
        email_result = send_report_email(report, report_paths, history_files)
    except Exception as exc:
        email_result = {"sent": False, "error": str(exc)}
    return {
        "report_id": report.get("id", ""),
        "json_path": report_paths.get("json_path", ""),
        "markdown_path": report_paths.get("markdown_path", ""),
        "xlsx_path": report_paths.get("xlsx_path", ""),
        "history_path": history_files.get("json_path", ""),
        "history_xlsx_path": history_files.get("xlsx_path", ""),
        "links": report.get("links", {}),
        "email": email_result,
    }
