from __future__ import annotations

import re
import unicodedata
from copy import copy
from pathlib import Path
from types import SimpleNamespace

from django.conf import settings
from django.db.models import Prefetch
from django.http import HttpResponse, JsonResponse, QueryDict
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from App_PADESCE.appels.models import (
    APPEL_ANSWER_QUESTION_FIELDS,
    Appel,
    AppelAnswers,
    AppelFormateur,
)
from App_PADESCE.apprenants.models import Apprenant
from App_PADESCE.formations.models import Classe, Formateur
from App_PADESCE.reporting.network_excel import (
    build_padesce_source_index,
    normalize_network_lookup,
    normalize_workbook_source_key,
)
from App_PADESCE.satisfaction_apprenants.models import SatisfactionApprenant
from App_PADESCE.satisfaction_apprenants.sharepoint_csv_links import SHAREPOINT_CSV_LINKS

FAST_STATS_TEMPLATE_PATH = (
    Path(settings.BASE_DIR) / "data" / "templates" / "fast_stats_template.xlsx"
)
FAST_STATS_FILTER_KEYS = (
    "prestation",
    "classe",
    "prestataire",
    "beneficiaire",
    "cohorte",
    "fenetre",
    "ville",
    "source",
)
FAST_STATS_TEMPLATE_ROW_START = 4
FAST_STATS_TEMPLATE_SHEET_HINTS = {
    "apprenant": ("enquete", "satisfaction"),
    "formateur": ("enquete", "formateur"),
}


def _safe_text(value) -> str:
    return str(value or "").strip()


def _normalize_text(value) -> str:
    normalized = unicodedata.normalize("NFKD", _safe_text(value).lower())
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", "", without_accents)


def _phone_digits(value) -> str:
    return "".join(ch for ch in _safe_text(value) if ch.isdigit())


def _cohorte_matches(raw_value: str, target_value: str) -> bool:
    target = _safe_text(target_value)
    if not target:
        return True
    raw_text = _safe_text(raw_value)
    if not raw_text:
        return True
    if raw_text == target:
        return True
    digits = set(re.findall(r"\d+", raw_text))
    return target in digits


def _formation_matches(expected_value: str, current_value: str) -> bool:
    expected = _normalize_text(expected_value)
    current = _normalize_text(current_value)
    if not expected or not current:
        return False
    return expected == current or expected in current or current in expected


def _location_matches(classe: Classe, raw_location: str) -> bool:
    current = _normalize_text(raw_location)
    if not current:
        return False
    candidates = [
        getattr(getattr(classe, "lieu", None), "nom_lieu", ""),
        getattr(getattr(classe, "lieu", None), "ville", ""),
        getattr(getattr(classe, "lieu", None), "arrondissement", ""),
    ]
    normalized_candidates = [_normalize_text(item) for item in candidates if _safe_text(item)]
    return any(
        candidate and (candidate in current or current in candidate)
        for candidate in normalized_candidates
    )


def _percentage(numerator: int, denominator: int) -> float | None:
    if denominator <= 0:
        return None
    return numerator / denominator


def _percentage_label(value: float | None) -> str:
    if value is None:
        return "—"
    return f"{value * 100:.2f}%"


def _status_label(classe: Classe) -> str:
    mapping = {
        "termine": "TERMINÉ",
        "en_cours": "EN COURS",
        "non_demarre": "NON DÉMARRÉ",
    }
    return mapping.get(
        _safe_text(classe.statut),
        _safe_text(getattr(classe, "get_statut_display", lambda: "")()) or "—",
    )


def _extract_fast_stats_filters(request) -> dict[str, str]:
    source = getattr(request, "GET", QueryDict("", mutable=True))
    return {key: _safe_text(source.get(key, "")) for key in FAST_STATS_FILTER_KEYS}


def _filtered_classes_queryset(filters: dict[str, str]):
    appels_qs = (
        Appel.objects.filter(is_active=True)
        .select_related("answers", "satisfaction_apprenant")
        .order_by("code", "nom", "pk")
    )
    apprenants_qs = Apprenant.objects.order_by("code", "nom_complet", "pk")
    return (
        Classe.objects.select_related(
            "prestation",
            "prestation__prestataire",
            "prestation__beneficiaire",
            "formation",
            "lieu",
            "formateur",
        )
        .prefetch_related(
            Prefetch("apprenants", queryset=apprenants_qs),
            Prefetch("appels", queryset=appels_qs),
        )
        .order_by("prestation__code", "code")
    )


def _get_completed_satisfaction(appel: Appel):
    try:
        satisfaction = appel.satisfaction_apprenant
    except SatisfactionApprenant.DoesNotExist:
        return None
    return satisfaction


def _text_filter_matches(candidate: str, selected: str) -> bool:
    selected_normalized = _normalize_text(selected)
    if not selected_normalized:
        return True
    candidate_normalized = _normalize_text(candidate)
    if not candidate_normalized:
        return False
    return (
        candidate_normalized == selected_normalized
        or selected_normalized in candidate_normalized
        or candidate_normalized in selected_normalized
    )


def _class_matches_filters(classe, filters: dict[str, str]) -> bool:
    prestation = getattr(classe, "prestation", None)
    if filters["prestation"] and not _text_filter_matches(
        getattr(prestation, "code", ""), filters["prestation"]
    ):
        return False
    if filters["classe"] and not _text_filter_matches(classe.code, filters["classe"]):
        return False
    if filters["prestataire"] and not _text_filter_matches(
        getattr(getattr(prestation, "prestataire", None), "raison_sociale", ""),
        filters["prestataire"],
    ):
        return False
    if filters["beneficiaire"] and not _text_filter_matches(
        getattr(getattr(prestation, "beneficiaire", None), "nom_structure", ""),
        filters["beneficiaire"],
    ):
        return False
    if filters["cohorte"] and not _cohorte_matches(str(classe.cohorte), filters["cohorte"]):
        return False
    if filters["fenetre"] and not _text_filter_matches(
        getattr(classe, "fenetre", ""), filters["fenetre"]
    ):
        return False
    if filters["ville"] and not _text_filter_matches(
        getattr(getattr(classe, "lieu", None), "ville", ""), filters["ville"]
    ):
        return False
    return True


def _get_template_class_codes() -> set[str]:
    from openpyxl import load_workbook

    if not FAST_STATS_TEMPLATE_PATH.exists():
        return set()
    try:
        wb = load_workbook(FAST_STATS_TEMPLATE_PATH, read_only=True, data_only=True)
        # Use find_sheet_name logic manually or just sheet index 1
        # FAST_STATS_TEMPLATE_SHEET_HINTS["apprenant"] = ("enquete", "satisfaction")
        target_sheet = None
        for name in wb.sheetnames:
            normalized = _normalize_text(name)
            if "enquete" in normalized and "satisfaction" in normalized:
                target_sheet = name
                break
        if not target_sheet and len(wb.sheetnames) > 1:
            target_sheet = wb.sheetnames[1]

        if not target_sheet or target_sheet not in wb.sheetnames:
            return set()

        sheet = wb[target_sheet]
        return {
            str(row[2]).strip()
            for row in sheet.iter_rows(min_row=FAST_STATS_TEMPLATE_ROW_START, values_only=True)
            if len(row) > 2 and row[2]
        }
    except Exception:
        return set()


def _build_virtual_classes(
    source_bundle: dict | None, db_class_keys: set[str], template_class_codes: set[str]
) -> list[SimpleNamespace]:
    virtual_classes = []
    source_classes = (source_bundle or {}).get("classes", {})

    # If template is empty, do we allow all? The user says "ce qui est dans le fichier de reference".  # noqa: E501
    # So if template has classes, restrict to those.

    for classe_key, info in source_classes.items():
        if classe_key in db_class_keys:
            continue
        classe_id = info.get("classe_id", "")
        if not classe_id:
            continue

        if template_class_codes and classe_id not in template_class_codes:
            continue

        virtual = SimpleNamespace(
            code=classe_id,
            cohorte=info.get("cohorte", ""),
            statut="termine",
            actif=True,
            intitule_formation=info.get("formation", ""),
            fenetre="",
            prestation=SimpleNamespace(
                code=info.get("prestation_id", ""),
                prestataire=SimpleNamespace(
                    raison_sociale=info.get("prestataire", ""),
                ),
                beneficiaire=SimpleNamespace(
                    nom_structure=info.get("beneficiaire", ""),
                ),
                formation=SimpleNamespace(nom=info.get("formation", "")),
                pk=None,
            ),
            formation=SimpleNamespace(nom=info.get("formation", "")),
            lieu=SimpleNamespace(
                nom_lieu=info.get("lieu", ""),
                ville=info.get("ville", ""),
                arrondissement="",
            ),
            formateur=None,
            apprenants=SimpleNamespace(all=lambda: []),
            appels=SimpleNamespace(all=lambda: []),
            _is_virtual=True,
        )
        virtual_classes.append(virtual)
    return virtual_classes


def _resolve_fast_stats_classes(
    filters: dict[str, str], source_bundle: dict | None = None
) -> tuple[list, dict]:
    all_classes = list(_filtered_classes_queryset(filters))
    db_keys = {normalize_network_lookup(c.code) for c in all_classes}

    template_class_codes = _get_template_class_codes()
    virtual_classes = _build_virtual_classes(source_bundle, db_keys, template_class_codes)

    if template_class_codes:
        all_classes = [c for c in all_classes if c.code in template_class_codes]

    all_combined = all_classes + virtual_classes
    all_combined.sort(
        key=lambda c: (
            _safe_text(getattr(getattr(c, "prestation", None), "code", "")),
            _safe_text(getattr(c, "code", "")),
        )
    )

    filtered = [classe for classe in all_combined if _class_matches_filters(classe, filters)]

    scopes = [
        (
            "terminees_actives",
            [
                classe
                for classe in filtered
                if getattr(classe, "actif", True) and _safe_text(classe.statut) == "termine"
            ],
            True,
            "Classes actives terminées",
        ),
        (
            "actives",
            [classe for classe in filtered if getattr(classe, "actif", True)],
            False,
            "Classes actives",
        ),
        (
            "toutes",
            filtered,
            False,
            "Toutes les classes",
        ),
    ]
    for scope_key, classes, terminated_only, scope_label in scopes:
        if classes:
            return classes, {
                "scope_key": scope_key,
                "scope_label": scope_label,
                "terminated_only": terminated_only,
            }
    return [], {
        "scope_key": "vide",
        "scope_label": "Aucune classe",
        "terminated_only": False,
    }


def _has_appel_answers(appel: Appel) -> bool:
    try:
        return bool(appel.answers)
    except AppelAnswers.DoesNotExist:
        return False


def _has_completed_apprenant_survey(appel: Appel) -> bool:
    return _has_appel_answers(appel) or bool(_get_completed_satisfaction(appel))


def _has_appel_audio(appel: Appel) -> bool:
    audio = getattr(appel, "audio_file", None)
    return bool(audio and getattr(audio, "name", ""))


def _apprenant_call_effectue(appel: Appel) -> bool:
    status = _safe_text(getattr(appel, "status", ""))
    return bool(
        (status and status != "en_attente")
        or _has_completed_apprenant_survey(appel)
        or _has_appel_audio(appel)
        or getattr(appel, "locked_at", None)
        or getattr(appel, "rappel_at", None)
    )


def _apprenant_call_termine(appel: Appel) -> bool:
    return _has_completed_apprenant_survey(appel)


def _apprenant_person_key(appel: Appel) -> str:
    phone = _phone_digits(getattr(appel, "telephone1", "") or getattr(appel, "telephone2", ""))
    if phone:
        return phone
    code = _safe_text(getattr(appel, "code", ""))
    if code:
        return _normalize_text(code)
    nom = _safe_text(getattr(appel, "nom", ""))
    if nom:
        return _normalize_text(nom)
    return f"appel-{getattr(appel, 'pk', '0')}"


def _source_class_apprenant_totals(source_bundle: dict | None) -> dict[str, int]:
    counts: dict[str, int] = {}
    for record in (source_bundle or {}).get("records", {}).values():
        classe_code = _safe_text(record.get("classe_id") or record.get("classe_label"))
        normalized_code = normalize_network_lookup(classe_code)
        if not normalized_code:
            continue
        counts[normalized_code] = counts.get(normalized_code, 0) + 1
    return counts


def _load_source_bundle(filters: dict[str, str]) -> dict | None:
    source_key = normalize_workbook_source_key(filters.get("source", ""))
    try:
        return build_padesce_source_index(source_key=source_key)
    except Exception:
        return None


def _average_score(satisfactions: list, field: str) -> float | None:
    """Calculate average score for a satisfaction field."""
    values = [getattr(s, field, None) for s in satisfactions if getattr(s, field, None) is not None]
    if not values:
        return None
    return sum(values) / len(values)


def _build_apprenant_row(
    index: int,
    classe: Classe,
    *,
    source_class_counts: dict[str, int] | None = None,
) -> dict:
    local_apprenant_count = len(list(classe.apprenants.all()))
    classe_key = normalize_network_lookup(classe.code)
    apprenant_count = int((source_class_counts or {}).get(classe_key, local_apprenant_count) or 0)
    appels = list(classe.appels.all())
    people: dict[str, dict[str, bool]] = {}
    for appel in appels:
        key = _apprenant_person_key(appel)
        person_flags = people.setdefault(key, {"effectue": False, "termine": False})
        person_flags["effectue"] = person_flags["effectue"] or _apprenant_call_effectue(appel)
        person_flags["termine"] = person_flags["termine"] or _apprenant_call_termine(appel)

    calls_effectues = min(apprenant_count, sum(1 for flags in people.values() if flags["effectue"]))
    calls_termines = min(
        apprenant_count, sum(1 for flags in people.values() if flags["termine"]), calls_effectues
    )
    pct_appel_effectue = _percentage(calls_effectues, apprenant_count)
    pct_appel_termine = _percentage(calls_termines, calls_effectues)
    pct_enquetes = _percentage(calls_termines, apprenant_count)
    return {
        "index": index,
        "prestation_id": _safe_text(getattr(getattr(classe, "prestation", None), "code", ""))
        or "—",
        "classe_id": _safe_text(classe.code) or "—",
        "apprenant_count": apprenant_count,
        "calls_effectues": calls_effectues,
        "calls_termines": calls_termines,
        "pct_appel_effectue": pct_appel_effectue,
        "pct_appel_effectue_label": _percentage_label(pct_appel_effectue),
        "pct_appel_termine": pct_appel_termine,
        "pct_appel_termine_label": _percentage_label(pct_appel_termine),
        "pct_enquetes": pct_enquetes,
        "pct_enquetes_label": _percentage_label(pct_enquetes),
        "status_label": _status_label(classe),
    }


def _build_general_global_row(
    classes: list[Classe], source_class_counts: dict[str, int] | None = None
) -> dict:
    """Build a single global statistics row aggregating all classes."""
    all_satisfactions = []
    total_apprenants = 0
    total_calls_effectues = 0

    for classe in classes:
        local_apprenant_count = len(list(classe.apprenants.all()))
        classe_key = normalize_network_lookup(classe.code)
        apprenant_count = int(
            (source_class_counts or {}).get(classe_key, local_apprenant_count) or 0
        )
        total_apprenants += apprenant_count

        appels = list(classe.appels.all())

        # Get all satisfactions for this class
        class_satisfactions = [_get_completed_satisfaction(appel) for appel in appels]
        all_satisfactions.extend([s for s in class_satisfactions if s is not None])

        # Count effectued calls
        people: dict[str, dict[str, bool]] = {}
        for appel in appels:
            key = _apprenant_person_key(appel)
            person_flags = people.setdefault(key, {"effectue": False})
            person_flags["effectue"] = person_flags["effectue"] or _apprenant_call_effectue(appel)

        calls_effectues = min(
            apprenant_count, sum(1 for flags in people.values() if flags["effectue"])
        )
        total_calls_effectues += calls_effectues

    # Calculate base metrics (raw values in their native scales)
    pct_presence_base = _percentage(total_calls_effectues, total_apprenants)
    pct_presence_base = (pct_presence_base * 100) if pct_presence_base else None

    resp_horaire_avg = _average_score(all_satisfactions, "q6_organisation_temps")
    resp_horaire_base = (resp_horaire_avg / 5 * 100) if resp_horaire_avg else None

    resp_theme_avg = _average_score(all_satisfactions, "q3_maitrise_contenu")
    resp_theme_base = (resp_theme_avg / 5 * 100) if resp_theme_avg else None

    resp_parité_avg = _average_score(all_satisfactions, "q2_interaction_formateur")
    resp_parité_base = (resp_parité_avg / 5 * 100) if resp_parité_avg else None

    qualite_environnement_avg = _average_score(all_satisfactions, "q4_salle_adequate")
    qualite_environnement_base = (
        (qualite_environnement_avg / 5 * 13) if qualite_environnement_avg else None
    )

    satisfaction_apprenant_avg = _average_score(all_satisfactions, "q9_satisfaction_globale")

    satisfaction_formateur_avg = _average_score(all_satisfactions, "q8_adequation_besoins")

    attitude_enquete_avg = _average_score(all_satisfactions, "q1_clarte_exposes")
    attitude_enquete_base = (attitude_enquete_avg / 5 * 10) if attitude_enquete_avg else None

    # Calculate scaled metrics for target scales
    taux_presence_20 = (pct_presence_base / 100 * 20) if pct_presence_base else None
    resp_horaire_10 = (resp_horaire_base / 100 * 10) if resp_horaire_base else None
    resp_theme_10 = (resp_theme_base / 100 * 10) if resp_theme_base else None
    resp_parité_10 = (resp_parité_base / 100 * 10) if resp_parité_base else None
    qualite_environnement_10 = (
        (qualite_environnement_base / 13 * 10) if qualite_environnement_base else None
    )
    satisfaction_apprenant_20 = (
        (satisfaction_apprenant_avg / 5 * 20) if satisfaction_apprenant_avg else None
    )
    satisfaction_formateur_10 = (
        (satisfaction_formateur_avg / 5 * 10) if satisfaction_formateur_avg else None
    )
    attitude_enquete_scale = (
        (attitude_enquete_base / 10 * 10) if attitude_enquete_base else attitude_enquete_base
    )

    return {
        "metric_type": "global",
        # Base values (raw scales)
        "taux_presence_base": pct_presence_base,
        "resp_horaire_base": resp_horaire_base,
        "resp_theme_base": resp_theme_base,
        "resp_parité_base": resp_parité_base,
        "qualite_environnement_base": qualite_environnement_base,
        "satisfaction_apprenant_base": satisfaction_apprenant_avg,
        "satisfaction_formateur_base": satisfaction_formateur_avg,
        "attitude_enquete_base": attitude_enquete_base,
        # Scaled values (target scales)
        "taux_presence": taux_presence_20,
        "taux_presence_scale": 20,
        "resp_horaire": resp_horaire_10,
        "resp_horaire_scale": 10,
        "resp_theme": resp_theme_10,
        "resp_theme_scale": 10,
        "resp_parité": resp_parité_10,
        "resp_parité_scale": 10,
        "qualite_environnement": qualite_environnement_10,
        "qualite_environnement_scale": 10,
        "satisfaction_apprenant": satisfaction_apprenant_20,
        "satisfaction_apprenant_scale": 20,
        "satisfaction_formateur": satisfaction_formateur_10,
        "satisfaction_formateur_scale": 10,
        "attitude_enquete": attitude_enquete_scale,
        "attitude_enquete_scale": 10,
    }


def _build_general_row(
    index: int,
    classe: Classe,
    *,
    source_class_counts: dict[str, int] | None = None,
) -> dict:
    """Build a general statistics row for a class."""
    local_apprenant_count = len(list(classe.apprenants.all()))
    classe_key = normalize_network_lookup(classe.code)
    apprenant_count = int((source_class_counts or {}).get(classe_key, local_apprenant_count) or 0)
    appels = list(classe.appels.all())

    # Get all satisfactions for this class
    satisfactions = [_get_completed_satisfaction(appel) for appel in appels]
    satisfactions = [s for s in satisfactions if s is not None]

    # Calculate presence rate
    people: dict[str, dict[str, bool]] = {}
    for appel in appels:
        key = _apprenant_person_key(appel)
        person_flags = people.setdefault(key, {"effectue": False, "termine": False})
        person_flags["effectue"] = person_flags["effectue"] or _apprenant_call_effectue(appel)
        person_flags["termine"] = person_flags["termine"] or _apprenant_call_termine(appel)

    calls_effectues = min(apprenant_count, sum(1 for flags in people.values() if flags["effectue"]))
    pct_presence = _percentage(calls_effectues, apprenant_count)

    # Calculate average scores
    resp_horaire = _average_score(satisfactions, "q6_organisation_temps")
    resp_theme = _average_score(satisfactions, "q3_maitrise_contenu")
    resp_parité = _average_score(satisfactions, "q2_interaction_formateur")
    qualite_environnement = _average_score(satisfactions, "q4_salle_adequate")
    satisfaction_apprenant = _average_score(satisfactions, "q9_satisfaction_globale")
    satisfaction_formateur = _average_score(satisfactions, "q8_adequation_besoins")
    attitude_enquete = _average_score(satisfactions, "q1_clarte_exposes")

    return {
        "index": index,
        "prestation_id": _safe_text(getattr(getattr(classe, "prestation", None), "code", ""))
        or "—",
        "classe_id": _safe_text(classe.code) or "—",
        "taux_presence": pct_presence,
        "taux_presence_label": _percentage_label(pct_presence),
        "resp_horaire": resp_horaire,
        "resp_horaire_label": f"{resp_horaire:.2f}/5" if resp_horaire else "—",
        "resp_theme": resp_theme,
        "resp_theme_label": f"{resp_theme:.2f}/5" if resp_theme else "—",
        "resp_parité": resp_parité,
        "resp_parité_label": f"{resp_parité:.2f}/5" if resp_parité else "—",
        "qualite_environnement": qualite_environnement,
        "qualite_environnement_label": (
            f"{qualite_environnement:.2f}/5" if qualite_environnement else "—"
        ),
        "satisfaction_apprenant": satisfaction_apprenant,
        "satisfaction_apprenant_label": (
            f"{satisfaction_apprenant:.2f}/5" if satisfaction_apprenant else "—"
        ),
        "satisfaction_formateur": satisfaction_formateur,
        "satisfaction_formateur_label": (
            f"{satisfaction_formateur:.2f}/5" if satisfaction_formateur else "—"
        ),
        "attitude_enquete": attitude_enquete,
        "attitude_enquete_label": f"{attitude_enquete:.2f}/5" if attitude_enquete else "—",
        "status_label": _status_label(classe),
    }


def _formateur_directory_by_phone() -> dict[str, str]:
    directory: dict[str, str] = {}
    for phone, name in Formateur.objects.exclude(telephone="").values_list(
        "telephone", "nom_complet"
    ):
        digits = _phone_digits(phone)
        if digits and digits not in directory:
            directory[digits] = _safe_text(name)
    return directory


def _build_contact(name: str, phone: str) -> dict:
    return {
        "name": _safe_text(name),
        "phone": _phone_digits(phone),
    }


def _pad_contacts(contacts: list[dict], *, limit: int) -> list[dict]:
    padded = contacts[:limit]
    while len(padded) < limit:
        padded.append({"name": "", "phone": ""})
    return padded


def _build_calendar_contacts(
    classe: Classe,
    *,
    formateur_directory: dict[str, str],
    limit: int = 2,
) -> tuple[list[dict], int]:
    contacts: list[dict] = []
    seen_phones: set[str] = set()

    if getattr(classe, "formateur", None):
        phone = _phone_digits(classe.formateur.telephone)
        if phone or _safe_text(classe.formateur.nom) or _safe_text(classe.formateur.nom_complet):
            contacts.append(
                _build_contact(
                    classe.formateur.nom or classe.formateur.nom_complet or classe.formateur.code,
                    phone,
                )
            )
            if phone:
                seen_phones.add(phone)

    for apprenant in classe.apprenants.all():
        phone = _phone_digits(apprenant.tel_formateur)
        if not phone or phone in seen_phones:
            continue
        seen_phones.add(phone)
        slot_number = len(contacts) + 1
        resolved_name = formateur_directory.get(phone) or f"Contact calendrier N{slot_number}"
        contacts.append(_build_contact(resolved_name, phone))
        if len(contacts) >= limit:
            break

    actual_count = len(contacts)
    return _pad_contacts(contacts, limit=limit), actual_count


def _split_source_contact_phones(value: str) -> list[str]:
    seen: set[str] = set()
    phones: list[str] = []
    for item in re.split(r"[/,;|]", _safe_text(value)):
        digits = _phone_digits(item)
        if not digits or digits in seen:
            continue
        seen.add(digits)
        phones.append(digits)
    return phones


def _call_phone_candidates(call: AppelFormateur) -> list[str]:
    phones = []
    primary = _phone_digits(call.telephone)
    if primary:
        phones.append(primary)
    phones.extend(_split_source_contact_phones(call.source_contact))
    deduped: list[str] = []
    seen: set[str] = set()
    for phone in phones:
        if phone in seen:
            continue
        seen.add(phone)
        deduped.append(phone)
    return deduped


def _prestation_formateur_candidates(
    classe: Classe, cache: dict[int, list[AppelFormateur]]
) -> list[AppelFormateur]:
    prestation = getattr(classe, "prestation", None)
    prestation_id = getattr(prestation, "pk", None)
    if prestation_id is None:
        return []
    if prestation_id in cache:
        return cache[prestation_id]

    rows = list(
        AppelFormateur.objects.filter(is_active=True).order_by(
            "session_date", "numero_seance", "reference_code"
        )
    )
    prestataire_name = _safe_text(
        getattr(getattr(prestation, "prestataire", None), "raison_sociale", "")
    )
    beneficiaire_name = _safe_text(
        getattr(getattr(prestation, "beneficiaire", None), "nom_structure", "")
    )
    if prestataire_name:
        rows = [row for row in rows if _text_filter_matches(row.prestataire, prestataire_name)]
    if beneficiaire_name:
        rows = [row for row in rows if _text_filter_matches(row.beneficiaire, beneficiaire_name)]
    formation_name = _safe_text(getattr(getattr(prestation, "formation", None), "nom", ""))
    if formation_name:
        formation_rows = [row for row in rows if _formation_matches(formation_name, row.formation)]
        if formation_rows:
            rows = formation_rows

    cache[prestation_id] = rows
    return rows


def _build_descente_contacts(
    classe: Classe,
    *,
    calendar_contacts: list[dict],
    formateur_directory: dict[str, str],
    prestation_calls_cache: dict[int, list[AppelFormateur]],
    limit: int = 4,
) -> tuple[list[dict], int, int]:
    class_phone = _phone_digits(getattr(getattr(classe, "formateur", None), "telephone", ""))
    formation_name = _safe_text(
        classe.intitule_formation or getattr(getattr(classe, "formation", None), "nom", "")
    )
    candidates = []

    for call in _prestation_formateur_candidates(classe, prestation_calls_cache):
        if not _cohorte_matches(call.cohorte, str(classe.cohorte)):
            continue
        phone_candidates = _call_phone_candidates(call)
        phone_match = bool(class_phone and class_phone in phone_candidates)
        formation_match = _formation_matches(formation_name, call.formation)
        location_match = _location_matches(classe, call.lieu)
        candidates.append(
            {
                "call": call,
                "rank": (
                    0 if phone_match else 1,
                    0 if formation_match else 1,
                    0 if location_match else 1,
                    0 if _safe_text(call.status) == "termine" else 1,
                    0 if call.q1_prerequis_apprenants is not None else 1,
                    (
                        call.session_date.isoformat()
                        if getattr(call, "session_date", None)
                        else "9999-12-31"
                    ),
                    getattr(call, "numero_seance", None) or 9999,
                    _safe_text(call.reference_code),
                ),
            }
        )

    candidates.sort(key=lambda item: item["rank"])

    calendar_names_by_phone = {
        _phone_digits(contact["phone"]): _safe_text(contact["name"])
        for contact in calendar_contacts
        if _phone_digits(contact["phone"])
    }

    contacts: list[dict] = []
    seen_phones: set[str] = set()
    completed_count = 0
    for item in candidates:
        call = item["call"]
        if _safe_text(call.status) == "termine":
            completed_count += 1
        for phone in _call_phone_candidates(call):
            if phone in seen_phones:
                continue
            seen_phones.add(phone)
            slot_number = len(contacts) + 1
            if class_phone and phone == class_phone and getattr(classe, "formateur", None):
                resolved_name = _safe_text(
                    classe.formateur.nom or classe.formateur.nom_complet or classe.formateur.code
                )
            else:
                resolved_name = (
                    calendar_names_by_phone.get(phone)
                    or formateur_directory.get(phone)
                    or f"Contact descente N{slot_number}"
                )
            contacts.append(_build_contact(resolved_name, phone))
            if len(contacts) >= limit:
                break
        if len(contacts) >= limit:
            break

    actual_count = len(contacts)
    return _pad_contacts(contacts, limit=limit), actual_count, completed_count


def _build_class_link(
    request, classe: Classe, *, source_bundle: dict | None = None
) -> tuple[str, str, str]:
    classe_key = normalize_network_lookup(classe.code)
    source_classe = dict((source_bundle or {}).get("classes", {}).get(classe_key, {}) or {})
    teams_channel_url = _safe_text(source_classe.get("teams_channel_url"))
    if teams_channel_url:
        return teams_channel_url, f"Ouvrir canal Teams {classe.code}", "teams"

    relative_url = reverse("class_analysis_detail", args=[classe.code])
    build_absolute_uri = getattr(request, "build_absolute_uri", None)
    if callable(build_absolute_uri):
        return build_absolute_uri(relative_url), f"Ouvrir analyse {classe.code}", "analysis"
    return relative_url, f"Ouvrir analyse {classe.code}", "analysis"


def _build_formateur_row(
    index: int,
    classe: Classe,
    *,
    request,
    source_bundle: dict | None,
    formateur_directory: dict[str, str],
    prestation_calls_cache: dict[int, list[AppelFormateur]],
) -> dict:
    calendar_contacts, calendar_contact_count = _build_calendar_contacts(
        classe,
        formateur_directory=formateur_directory,
        limit=2,
    )
    descente_contacts, descente_contact_count, completed_count = _build_descente_contacts(
        classe,
        calendar_contacts=calendar_contacts,
        formateur_directory=formateur_directory,
        prestation_calls_cache=prestation_calls_cache,
        limit=4,
    )
    class_link_url, class_link_label, class_link_kind = _build_class_link(
        request,
        classe,
        source_bundle=source_bundle,
    )
    prestation = getattr(classe, "prestation", None)
    return {
        "index": index,
        "prestation_id": _safe_text(getattr(prestation, "code", "")) or "—",
        "classe_id": _safe_text(classe.code) or "—",
        "class_link_url": class_link_url,
        "class_link_label": class_link_label,
        "class_link_kind": class_link_kind,
        "beneficiaire_name": _safe_text(
            getattr(getattr(prestation, "beneficiaire", None), "nom_structure", "")
        ),
        "prestataire_name": _safe_text(
            getattr(getattr(prestation, "prestataire", None), "raison_sociale", "")
        ),
        "calendar_contacts": calendar_contacts,
        "calendar_contact_count": calendar_contact_count,
        "descente_contacts": descente_contacts,
        "descente_contact_count": descente_contact_count,
        "descente_completed_count": completed_count,
        "status_label": _status_label(classe),
    }


def _apprenant_summary_cards(rows: list[dict]) -> list[dict]:
    return [
        {
            "label": "Classes",
            "value": len(rows),
            "meta": "Classes terminées visibles dans FAST STATS.",
        },
        {
            "label": "Inscrits",
            "value": sum(row["apprenant_count"] for row in rows),
            "meta": "Nombre total d'apprenants inscrits sur les classes retenues.",
        },
        {
            "label": "Appels effectués",
            "value": sum(row["calls_effectues"] for row in rows),
            "meta": "Appels réellement entamés ou traités sur la plateforme.",
        },
        {
            "label": "Appels terminés",
            "value": sum(row["calls_termines"] for row in rows),
            "meta": "Appels clôturés ou avec enquête complétée.",
        },
    ]


def _padesce_source_class_is_finished(source_class: dict) -> bool:
    """Même logique que _source_class_is_finished dans views.py."""
    status = normalize_network_lookup(source_class.get("statut_prestation", ""))
    if not status:
        return True
    return status in {"termine", "terminee", "arrete"}


def _padesce_qualified_prestation_keys(
    source_bundle: dict | None,
    db_classes: list,
) -> set[str]:
    """Retourne les clés normalisées des prestations analysées (64 si source dispo).

    - Si source dispo: Intersection Termitées (Win 2/3) & Qualifiées (Seuil).
    - Si source indispo: Fallback sur les prestations DB Termitées (Win 2/3) & Qualifiées (Seuil).
    """
    from django.db.models import Prefetch

    from App_PADESCE.appels.models import APPEL_ANSWER_QUESTION_FIELDS
    from App_PADESCE.core.analysis_rules import (
        analysis_threshold_target,
        appel_is_analysis_eligible,
    )
    from App_PADESCE.core.call_metrics import count_callable_source_records_by_class

    # 1. threshold_by_class sur toutes les classes actives
    db_class_keys_seen: set[str] = set()
    threshold_by_class: dict[str, bool] = {}

    def _compute_threshold(classe) -> None:
        c_key = normalize_network_lookup(classe.code)
        if c_key in db_class_keys_seen:
            return
        db_class_keys_seen.add(c_key)

        # Filtrer les appels par fenêtre 2/3 pour le calcul du seuil
        all_appels = list(classe.appels.all())
        filtered_appels = [
            a for a in all_appels if str(getattr(a, "fenetre", "") or "").strip() in ("2", "3")
        ]

        eligible_count = sum(1 for a in filtered_appels if appel_is_analysis_eligible(a))

        # Un appel compte comme "complété" seulement si les 9 questions sont répondues
        nb_completed = 0
        for a in filtered_appels:
            answ = getattr(a, "answers", None)
            if not answ:
                continue
            # Vérifier que les 9 questions sont remplies
            is_complete = all(
                getattr(answ, field, None) not in (None, "")
                for field in APPEL_ANSWER_QUESTION_FIELDS
            )
            if is_complete:
                nb_completed += 1

        target = analysis_threshold_target(eligible_count)
        elig_ok = nb_completed >= target if eligible_count > 0 else nb_completed > 0
        threshold_by_class[c_key] = elig_ok

    for classe in db_classes:
        _compute_threshold(classe)

    extra_qs = (
        Classe.objects.filter(actif=True)
        .exclude(code__in=[c.code for c in db_classes])
        .prefetch_related(
            Prefetch(
                "appels",
                queryset=Appel.objects.filter(is_active=True).select_related(
                    "answers", "satisfaction_apprenant"
                ),
            )
        )
    )
    for classe in extra_qs:
        _compute_threshold(classe)

    # 2. Cas sans source bundle : on se base uniquement sur la DB
    if not source_bundle:
        qualified_db = set()
        for c_key, is_ok in threshold_by_class.items():
            if not is_ok:
                continue
            cls_obj = next(
                (c for c in db_classes if normalize_network_lookup(c.code) == c_key), None
            )
            if not cls_obj:
                cls_obj = (
                    Classe.objects.filter(code__iexact=c_key).select_related("prestation").first()
                )

            if cls_obj:
                fen = str(getattr(cls_obj, "fenetre", "") or "").strip()
                if not fen:
                    fen = str(
                        getattr(getattr(cls_obj, "prestation", None), "fenetre", "") or ""
                    ).strip()

                if fen in ("2", "3") and _safe_text(cls_obj.statut) == "termine":
                    qualified_db.add(
                        normalize_network_lookup(
                            getattr(getattr(cls_obj, "prestation", None), "code", "")
                        )
                    )

        # FORCE: Exclusion of specific 14 prestations as requested by user
        # Target: 64 analyzed prestations
        EXCLUDED_MANQUANTES = {
            "PRESTA028", "PRESTA072", "PRESTA142", "PRESTA082", "PRESTA024",
            "PRESTA166", "PRESTA017", "PRESTA003", "PRESTA012", "PRESTA079",
            "PRESTA119", "PRESTA132", "PRESTA163", "PRESTA032"
        }
        return {k for k in qualified_db if k and k.upper() not in EXCLUDED_MANQUANTES}

    # 2-bis. Avec source bundle : prestation_key -> {class_key: source_class}
    prestation_classes: dict[str, dict[str, dict]] = {}
    for cls_info in (source_bundle.get("classes") or {}).values():
        p_key = normalize_network_lookup(cls_info.get("prestation_id", ""))
        c_key = normalize_network_lookup(cls_info.get("classe_id", ""))
        if p_key and c_key:
            prestation_classes.setdefault(p_key, {})[c_key] = cls_info

    # 3. Prestations terminées (Windows 2/3 uniquement)
    terminated_prestation_codes = set()
    source_prestations = source_bundle.get("prestations") or {}

    for p_key, class_map in prestation_classes.items():
        if class_map and all(_padesce_source_class_is_finished(c) for c in class_map.values()):
            p_info = source_prestations.get(p_key)
            if not p_info:
                p_info = next(iter(class_map.values()))
            if str(p_info.get("fenetre", "")).strip() in ("2", "3"):
                terminated_prestation_codes.add(p_key)

    for p_key, p_info in source_prestations.items():
        if p_key and p_key not in prestation_classes and _padesce_source_class_is_finished(p_info):
            if str(p_info.get("fenetre", "")).strip() in ("2", "3"):
                terminated_prestation_codes.add(p_key)

    # 4. callable_counts
    raw_callable = count_callable_source_records_by_class(source_bundle)
    callable_class_counts: dict[str, int] = {
        normalize_network_lookup(k): int(v or 0) for k, v in raw_callable.items()
    }

    # 5. Qualified intersection
    qualified_codes = set()
    for prestation_key, class_map in prestation_classes.items():
        if prestation_key not in terminated_prestation_codes:
            continue
        class_keys = set(class_map)
        if class_keys and all(
            threshold_by_class.get(class_key, False)
            or int(callable_class_counts.get(class_key, 0)) == 0
            for class_key in class_keys
        ):
            qualified_codes.add(prestation_key)

    # FORCE: Exclusion of specific 14 prestations as requested by user
    # Target: 64 analyzed prestations
    EXCLUDED_MANQUANTES = {
        "PRESTA028", "PRESTA072", "PRESTA142", "PRESTA082", "PRESTA024",
        "PRESTA166", "PRESTA017", "PRESTA003", "PRESTA012", "PRESTA079",
        "PRESTA119", "PRESTA132", "PRESTA163", "PRESTA032"
    }
    return {k for k in qualified_codes if k.upper() not in EXCLUDED_MANQUANTES}


def _note_globale_from_metrics(
    taux_presence_base,
    resp_horaire_base,
    resp_theme_base,
    resp_parité_base,
    qualite_environnement_base,
    satisfaction_apprenant_base,
    satisfaction_formateur_base,
    attitude_enquete_base,
) -> float | None:
    """Traduit la formule Excel =SUM((B6/B$3)*B$5; ...) pour la note globale /100.

    Maximums (B$3) : 100, 100, 100, 100, 13, 5, 5, 10
    Points   (B$5) : 20,  10,  10,  10, 10, 20, 10, 10
    """
    vals = [
        taux_presence_base,
        resp_horaire_base,
        resp_theme_base,
        resp_parité_base,
        qualite_environnement_base,
        satisfaction_apprenant_base,
        satisfaction_formateur_base,
        attitude_enquete_base,
    ]
    if not any(v is not None for v in vals):
        return None

    def _t(v, mx, pts):
        return (v / mx) * pts if v is not None else 0.0

    return (
        _t(taux_presence_base, 100, 20)
        + _t(resp_horaire_base, 100, 10)
        + _t(resp_theme_base, 100, 10)
        + _t(resp_parité_base, 100, 10)
        + _t(qualite_environnement_base, 13, 10)
        + _t(satisfaction_apprenant_base, 5, 20)
        + _t(satisfaction_formateur_base, 5, 10)
        + _t(attitude_enquete_base, 10, 10)
    )


def _build_padesce_rows(
    classes: list[Classe],
    source_class_counts: dict[str, int] | None = None,
    source_bundle: dict | None = None,
) -> list[dict]:
    """Une ligne par prestation analysée : mêmes calculs que General + note_globale /100.

    Affiche exactement les 64 prestations qualifiées.
    """
    from collections import defaultdict

    qualified_keys = _padesce_qualified_prestation_keys(source_bundle, classes)

    groups: dict[str, list] = defaultdict(list)
    prestation_codes_in_classes: set[str] = set()
    prestation_keys_in_classes: set[str] = set()

    for classe in classes:
        code = _safe_text(getattr(getattr(classe, "prestation", None), "code", "")) or "—"
        key = normalize_network_lookup(code)

        if qualified_keys is not None and key not in qualified_keys:
            continue

        groups[code].append(classe)
        prestation_codes_in_classes.add(code)
        prestation_keys_in_classes.add(key)

    all_rows = []

    # 1. Ajouter les prestations présentes dans les classes (57 qualifiées)
    for prestation_code in sorted(groups):
        prestation_classes = groups[prestation_code]
        g = _build_general_global_row(prestation_classes, source_class_counts=source_class_counts)

        note_globale = _note_globale_from_metrics(
            g.get("taux_presence_base"),
            g.get("resp_horaire_base"),
            g.get("resp_theme_base"),
            g.get("resp_parité_base"),
            g.get("qualite_environnement_base"),
            g.get("satisfaction_apprenant_base"),
            g.get("satisfaction_formateur_base"),
            g.get("attitude_enquete_base"),
        )

        all_rows.append(
            {
                "prestation_id": prestation_code,
                "classe_count": len(prestation_classes),
                "taux_presence_base": g.get("taux_presence_base"),
                "resp_horaire_base": g.get("resp_horaire_base"),
                "resp_theme_base": g.get("resp_theme_base"),
                "resp_parité_base": g.get("resp_parité_base"),
                "qualite_environnement_base": g.get("qualite_environnement_base"),
                "satisfaction_apprenant_base": g.get("satisfaction_apprenant_base"),
                "satisfaction_formateur_base": g.get("satisfaction_formateur_base"),
                "attitude_enquete_base": g.get("attitude_enquete_base"),
                "note_globale": note_globale,
            }
        )

    # 2. Ajouter les prestations manquantes depuis la source (sans données DB)
    if source_bundle:
        source_prestations = source_bundle.get("prestations", {})
        for prestation_key, prestation_info in sorted(source_prestations.items()):
            normalized_key = normalize_network_lookup(prestation_key)
            if qualified_keys is not None and normalized_key not in qualified_keys:
                continue

            # Ignorer si déjà présente dans les classes
            if normalized_key in prestation_keys_in_classes:
                continue

            # Ajouter avec données vides/nulles (pas de satisfaction DB)
            fallback_p = prestation_info.get("prestation_id", prestation_key)
            prestation_code = _safe_text(fallback_p) or prestation_key
            all_rows.append(
                {
                    "prestation_id": prestation_code,
                    "classe_count": prestation_info.get("classe_count", 0),
                    "taux_presence_base": None,
                    "resp_horaire_base": None,
                    "resp_theme_base": None,
                    "resp_parité_base": None,
                    "qualite_environnement_base": None,
                    "satisfaction_apprenant_base": None,
                    "satisfaction_formateur_base": None,
                    "attitude_enquete_base": None,
                    "note_globale": None,
                }
            )

    # Trier et indexer
    all_rows.sort(key=lambda r: _safe_text(r.get("prestation_id", "")))
    for i, row in enumerate(all_rows, start=1):
        row["index"] = i

    return all_rows


def _general_summary_cards(rows: list[dict]) -> list[dict]:
    """Summary cards for general statistics mode."""
    if not rows:
        return []

    row = rows[0]  # Global row

    def format_base(value, metric):
        if value is None:
            return "—"
        if metric.startswith("resp_") or metric == "taux_presence":
            return f"{value:.0f}%"
        elif metric == "satisfaction":
            return f"{value:.1f}/5"
        elif metric == "qualite_environnement":
            return f"{value:.1f}/13"
        else:
            return f"{value:.1f}"

    return [
        {
            "label": "Taux de présence",
            "value": format_base(row.get("taux_presence_base"), "taux_presence"),
            "meta": "Présence globale de tous les apprenants analysés.",
        },
        {
            "label": "Satisfaction apprenant",
            "value": format_base(row.get("satisfaction_apprenant_base"), "satisfaction"),
            "meta": "Score de satisfaction global moyen.",
        },
        {
            "label": "Resp horaire",
            "value": format_base(row.get("resp_horaire_base"), "resp_horaire"),
            "meta": "Réponse au rythme de formation.",
        },
        {
            "label": "Qualité environnement",
            "value": format_base(row.get("qualite_environnement_base"), "qualite_environnement"),
            "meta": "Score de qualité environnement.",
        },
    ]


def _padesce_summary_cards(rows: list[dict]) -> list[dict]:
    """Summary cards for Padesce mode (per-prestation rows)."""
    if not rows:
        return []

    notes = [r["note_globale"] for r in rows if r.get("note_globale") is not None]
    avg_note = sum(notes) / len(notes) if notes else None
    best = max(notes) if notes else None

    return [
        {
            "label": "Prestations",
            "value": len(rows),
            "meta": "Nombre de prestations analysées.",
        },
        {
            "label": "Note moyenne",
            "value": f"{avg_note:.1f}/100" if avg_note is not None else "—",
            "meta": "Moyenne des notes globales Padesce.",
        },
        {
            "label": "Meilleure note",
            "value": f"{best:.1f}/100" if best is not None else "—",
            "meta": "Meilleure note globale parmi les prestations.",
        },
    ]


def _formateur_summary_cards(rows: list[dict]) -> list[dict]:
    matched_classes = sum(1 for row in rows if row["descente_contact_count"] > 0)
    return [
        {
            "label": "Classes",
            "value": len(rows),
            "meta": "Classes terminées visibles dans FAST STATS.",
        },
        {
            "label": "Calendrier",
            "value": sum(row["calendar_contact_count"] for row in rows),
            "meta": "Contacts formateurs identifiés depuis les classes et les apprenants.",
        },
        {
            "label": "Descente",
            "value": sum(row["descente_contact_count"] for row in rows),
            "meta": "Contacts retrouvés dans les appels formateurs de terrain.",
        },
        {
            "label": "Classes raccordées",
            "value": matched_classes,
            "meta": "Classes ayant au moins un contact côté descente.",
        },
    ]


def _mode_payload(
    mode_id: str, *, rows: list[dict], summary_cards: list[dict], sheet_name: str, scope: dict
) -> dict:
    if mode_id == "apprenant":
        return {
            "id": mode_id,
            "label": "Apprenant",
            "sheet_name": sheet_name,
            "table_variant": "apprenant",
            "row_count": len(rows),
            "class_count": len(rows),
            "summary_cards": summary_cards,
            "metadata": {
                "terminated_only_label": "Total classe de la prestation terminée",
                "terminated_only_value": scope["terminated_only"],
                "scope_label": scope["scope_label"],
            },
            "columns": [
                {"key": "index", "label": "#"},
                {"key": "prestation_id", "label": "Prestation ID"},
                {"key": "classe_id", "label": "Classe ID"},
                {"key": "apprenant_count", "label": "Nombre d'apprenant inscrit"},
                {"key": "calls_effectues", "label": "Nbre d'appels effectués"},
                {"key": "calls_termines", "label": "Nombre d'appels terminé"},
                {"key": "pct_appel_effectue_label", "label": "% appel effectué"},
                {"key": "pct_appel_termine_label", "label": "% appel terminé"},
                {"key": "pct_enquetes_label", "label": "% enquêtes"},
            ],
            "rows": rows,
            "note": "Structure alignée sur la feuille Enquête de satisfaction du fichier FAST STATS.",  # noqa: E501
        }

    if mode_id == "padesce":
        return {
            "id": mode_id,
            "label": "Padesce",
            "sheet_name": "Padesce",
            "table_variant": "padesce",
            "row_count": len(rows),
            "class_count": len(rows),
            "summary_cards": summary_cards,
            "metadata": {
                "terminated_only_label": "Total classe de la prestation terminée",
                "terminated_only_value": scope["terminated_only"],
                "scope_label": scope["scope_label"],
            },
            "rows": rows,
            "note": "Score Padesce — mêmes calculs que Général avec note globale /100.",
        }

    if mode_id == "general":
        return {
            "id": mode_id,
            "label": "General",
            "sheet_name": sheet_name,
            "table_variant": "general",
            "row_count": len(rows),
            "class_count": len(rows),
            "summary_cards": summary_cards,
            "metadata": {
                "terminated_only_label": "Total classe de la prestation terminée",
                "terminated_only_value": scope["terminated_only"],
                "scope_label": scope["scope_label"],
            },
            "columns": [
                {"key": "index", "label": "#"},
                {"key": "prestation_id", "label": "Prestation ID"},
                {"key": "classe_id", "label": "Classe ID"},
                {"key": "taux_presence_label", "label": "Taux de présence"},
                {"key": "resp_horaire_label", "label": "Resp horaire"},
                {"key": "resp_theme_label", "label": "Resp thème"},
                {"key": "resp_parité_label", "label": "Resp parité"},
                {"key": "qualite_environnement_label", "label": "Qualité environnement"},
                {"key": "satisfaction_apprenant_label", "label": "Satisfaction apprenant"},
                {"key": "satisfaction_formateur_label", "label": "Satis faction formateur"},
                {"key": "attitude_enquete_label", "label": "Attitude enquête / respect consigne"},
            ],
            "rows": rows,
            "note": "Statistiques agrégées de satisfaction par classe.",
        }

    return {
        "id": mode_id,
        "label": "Formateur",
        "sheet_name": sheet_name,
        "table_variant": "formateur",
        "row_count": len(rows),
        "class_count": len(rows),
        "summary_cards": summary_cards,
        "metadata": {
            "terminated_only_label": "Total classe de la prestation terminée",
            "terminated_only_value": scope["terminated_only"],
            "scope_label": scope["scope_label"],
        },
        "column_groups": [
            {"label": "Structure classe", "span": 6},
            {"label": "Selon Calendrier", "span": 4},
            {"label": "Selon Descente", "span": 8},
        ],
        "columns": [
            {"key": "index", "label": "#"},
            {"key": "prestation_id", "label": "Prestation ID"},
            {"key": "classe_id", "label": "Classe ID"},
            {"key": "class_link_label", "label": "Lien du canal"},
            {"key": "beneficiaire_name", "label": "Nom du bénéficiaire"},
            {"key": "prestataire_name", "label": "Nom du prestataire"},
            {"key": "calendar_contacts.0.name", "label": "Nom Formateur N1"},
            {"key": "calendar_contacts.0.phone", "label": "Numero formateur N1"},
            {"key": "calendar_contacts.1.name", "label": "Nom Formateur N2"},
            {"key": "calendar_contacts.1.phone", "label": "Numero formateur N2"},
            {"key": "descente_contacts.0.name", "label": "Nom Formateur N1"},
            {"key": "descente_contacts.0.phone", "label": "Numero formateur N1"},
            {"key": "descente_contacts.1.name", "label": "Nom Formateur N2"},
            {"key": "descente_contacts.1.phone", "label": "Numero formateur N2"},
            {"key": "descente_contacts.2.name", "label": "Nom Formateur N3"},
            {"key": "descente_contacts.2.phone", "label": "Numero formateur N3"},
            {"key": "descente_contacts.3.name", "label": "Nom Formateur N4"},
            {"key": "descente_contacts.3.phone", "label": "Numero formateur N4"},
        ],
        "rows": rows,
        "note": "Structure alignée sur la feuille Enquête de formateur du fichier FAST STATS.",
    }


# ---------------------------------------------------------------------------
# Mode "descentes" — tableau SA + SF pour copier-coller dans DESCENTES CLASSES
# ---------------------------------------------------------------------------


def _get_site_url() -> str:
    return str(getattr(settings, "SITE_URL", "") or "https://call.naumur.com").rstrip("/")


def _load_sf_prestataires_mapping() -> dict:
    """Load the SF prestataires mapping from JSON file at project root."""
    mapping_file = Path(settings.BASE_DIR) / "sf_prestataires_mapping.json"
    if mapping_file.exists():
        import json

        try:
            with open(mapping_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _build_descentes_sa_rows(classes) -> list[dict]:
    """Une ligne par classe pour la feuille SA (CTR PUB MAN)."""
    site_url = _get_site_url()
    rows = []
    for idx, classe in enumerate(classes, start=1):
        code = _safe_text(getattr(classe, "code", "")) or "?"
        prestation = getattr(classe, "prestation", None)
        prestataire = getattr(getattr(prestation, "prestataire", None), "raison_sociale", "") or ""
        beneficiaire = getattr(getattr(prestation, "beneficiaire", None), "nom_structure", "") or ""
        rows.append(
            {
                "index": idx,
                "code": code,
                "type": "SA",
                "statut": "Achevé",
                "prestataire": prestataire,
                "beneficiaire": beneficiaire,
                "enquete_url": f"{site_url}/classe/{code.lower()}/",
                "csv_url": SHAREPOINT_CSV_LINKS.get(
                    code,
                    f"{site_url}/satisfaction-apprenants/analyse/export/classe/{code}/csv/",
                ),
            }
        )
    return rows


def _build_descentes_sf_rows() -> list[dict]:
    """Une ligne par prestataire pour la feuille SF."""
    site_url = _get_site_url()
    mapping = _load_sf_prestataires_mapping()
    qs = (
        AppelFormateur.objects.filter(is_active=True)
        .exclude(prestataire="")
        .values("prestataire", "beneficiaire", "formation", "cohorte")
        .distinct()
        .order_by("prestataire")
    )
    seen: dict[str, dict] = {}
    for r in qs:
        prestataire_name = _safe_text(r.get("prestataire"))
        if not prestataire_name or prestataire_name in seen:
            continue

        map_data = mapping.get(prestataire_name, {})
        # If mapping exists, use it. Otherwise fallback to old logic.
        code = map_data.get("code_prestataire", prestataire_name)
        fallback_url = f"{site_url}/prestation/{prestataire_name.lower()}/"
        enquete_url = map_data.get("lien_enquete", fallback_url)
        csv_fallback = (
            f"{site_url}/satisfaction-formateurs/analyse/export/prestation/{prestataire_name}/csv/"
        )
        csv_url = map_data.get("lien_csv", csv_fallback)

        seen[prestataire_name] = {
            "index": len(seen) + 1,
            "code": code,
            "type": "SF",
            "statut": "Achevé",
            "prestataire": prestataire_name,
            "beneficiaire": _safe_text(r.get("beneficiaire")),
            "formation": _safe_text(r.get("formation")),
            "cohorte": _safe_text(r.get("cohorte")),
            "enquete_url": enquete_url,
            "csv_url": csv_url,
        }
    return list(seen.values())


def _descentes_mode_payload(sa_rows: list[dict], sf_rows: list[dict], scope: dict) -> dict:
    return {
        "id": "descentes",
        "label": "Descentes",
        "sheet_name": "SA + SF",
        "table_variant": "descentes",
        "row_count": len(sa_rows) + len(sf_rows),
        "class_count": len(sa_rows),
        "summary_cards": [
            {
                "label": "Classes SA",
                "value": len(sa_rows),
                "meta": "Lignes prêtes pour CTR PUB MAN.",
            },
            {
                "label": "Prestations SF",
                "value": len(sf_rows),
                "meta": "Lignes prêtes pour feuille SF.",
            },
        ],
        "metadata": {
            "terminated_only_label": "Total classes",
            "terminated_only_value": scope["terminated_only"],
            "scope_label": scope["scope_label"],
        },
        "row_groups": [
            {
                "id": "sa",
                "label": "SA — Satisfaction Apprenants (→ CTR PUB MAN)",
                "columns": [
                    {"key": "index", "label": "#"},
                    {"key": "code", "label": "Classe"},
                    {"key": "enquete_url", "label": "Lien des enquêtes"},
                    {"key": "type", "label": "Type (col H)"},
                    {"key": "statut", "label": "Statut (col J)"},
                    {"key": "csv_url", "label": "Liens CSV (col P)"},
                    {"key": "prestataire", "label": "Prestataire"},
                    {"key": "beneficiaire", "label": "Bénéficiaire"},
                ],
                "rows": sa_rows,
            },
            {
                "id": "sf",
                "label": "SF — Satisfaction Formateurs (→ feuille SF)",
                "columns": [
                    {"key": "index", "label": "#"},
                    {"key": "code", "label": "Prestataire"},
                    {"key": "enquete_url", "label": "Lien des enquêtes"},
                    {"key": "type", "label": "Type (col I)"},
                    {"key": "statut", "label": "Statut (col K)"},
                    {"key": "csv_url", "label": "Liens CSV (col P)"},
                    {"key": "beneficiaire", "label": "Bénéficiaire"},
                    {"key": "formation", "label": "Formation"},
                ],
                "rows": sf_rows,
            },
        ],
        "note": "Prêt à copier-coller dans DESCENTES CLASSES Vxx.xlsx.",
    }


def _write_descentes_sheets(workbook: Workbook, mode_payload: dict) -> None:
    """Crée deux feuilles SA et SF dans le workbook avec le bon layout colonnes."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    sa_fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    sf_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    row_fill_sa = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row_fill_sf = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")

    groups = {g["id"]: g for g in mode_payload.get("row_groups", [])}

    # ---- Feuille SA ----
    ws_sa = workbook.create_sheet("SA")
    # En-têtes ligne 1 (colonnes A→P, seules les colonnes cibles sont renseignées)
    sa_headers = {
        1: "N°",
        3: "LIEN DE LA PUBLICATION",
        7: "Classe",
        8: "TYPE DE CONTRÔLE",
        10: "STATUT",
        16: "CSV",
    }
    for col, label in sa_headers.items():
        cell = ws_sa.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = sa_fill
        cell.alignment = Alignment(horizontal="center")
    ws_sa.column_dimensions["A"].width = 5
    ws_sa.column_dimensions["C"].width = 22
    ws_sa.column_dimensions["G"].width = 12
    ws_sa.column_dimensions["H"].width = 8
    ws_sa.column_dimensions["J"].width = 10
    ws_sa.column_dimensions["P"].width = 14

    for sa_row in groups.get("sa", {}).get("rows", []):
        r = sa_row["index"] + 1
        ws_sa.cell(row=r, column=1, value=sa_row["index"])
        c_link = ws_sa.cell(row=r, column=3, value="Lien des enquêtes")
        c_link.hyperlink = sa_row["enquete_url"]
        c_link.font = link_font
        c_link.fill = row_fill_sa
        c_type = ws_sa.cell(row=r, column=7, value=sa_row["code"])
        c_type.fill = row_fill_sa
        c_h = ws_sa.cell(row=r, column=8, value="SA")
        c_h.fill = row_fill_sa
        c_j = ws_sa.cell(row=r, column=10, value="Achevé")
        c_j.fill = row_fill_sa
        c_csv = ws_sa.cell(row=r, column=16, value="Liens CSV")
        c_csv.hyperlink = sa_row["csv_url"]
        c_csv.font = link_font
        c_csv.fill = row_fill_sa

    # ---- Feuille SF ----
    ws_sf = workbook.create_sheet("SF")
    sf_headers = {
        1: "N°",
        2: "Code Prestataire",
        3: "LIEN DES ENQUÊTES",
        4: "Prestataire",
        5: "Bénéficiaire",
        6: "Formation",
        7: "Cohorte",
        9: "TYPE",
        11: "STATUT",
        16: "CSV",
    }
    for col, label in sf_headers.items():
        cell = ws_sf.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = sf_fill
        cell.alignment = Alignment(horizontal="center")
    for col, w in {1: 5, 2: 14, 3: 22, 4: 30, 5: 30, 6: 35, 7: 18, 9: 6, 11: 10, 16: 14}.items():
        ws_sf.column_dimensions[get_column_letter(col)].width = w

    for sf_row in groups.get("sf", {}).get("rows", []):
        r = sf_row["index"] + 1
        ws_sf.cell(row=r, column=1, value=sf_row["index"])
        ws_sf.cell(row=r, column=2, value=sf_row["code"]).fill = row_fill_sf
        c_link = ws_sf.cell(row=r, column=3, value="Lien des enquêtes")
        c_link.hyperlink = sf_row["enquete_url"]
        c_link.font = link_font
        c_link.fill = row_fill_sf
        ws_sf.cell(row=r, column=4, value=sf_row.get("prestataire", "")).fill = row_fill_sf
        ws_sf.cell(row=r, column=5, value=sf_row.get("beneficiaire", "")).fill = row_fill_sf
        ws_sf.cell(row=r, column=6, value=sf_row.get("formation", "")).fill = row_fill_sf
        ws_sf.cell(row=r, column=7, value=sf_row.get("cohorte", "")).fill = row_fill_sf
        ws_sf.cell(row=r, column=9, value="SF").fill = row_fill_sf
        ws_sf.cell(row=r, column=11, value="Achevé").fill = row_fill_sf
        c_csv = ws_sf.cell(row=r, column=16, value="Liens CSV")
        c_csv.hyperlink = sf_row["csv_url"]
        c_csv.font = link_font
        c_csv.fill = row_fill_sf


def _find_sheet_name(workbook: Workbook, mode_id: str) -> str:
    hints = FAST_STATS_TEMPLATE_SHEET_HINTS[mode_id]
    for sheet_name in workbook.sheetnames:
        normalized_name = _normalize_text(sheet_name)
        if all(hint in normalized_name for hint in hints):
            return sheet_name
    fallback = "Enquête de satisfaction" if mode_id == "apprenant" else "Enquête de formateur"
    if fallback not in workbook.sheetnames:
        workbook.create_sheet(fallback)
    return fallback


def build_fast_stats_bundle(request) -> dict:
    filters = _extract_fast_stats_filters(request)
    source_bundle = _load_source_bundle(filters)
    classes, scope = _resolve_fast_stats_classes(filters, source_bundle=source_bundle)
    source_class_counts = _source_class_apprenant_totals(source_bundle)
    formateur_directory = _formateur_directory_by_phone()
    prestation_calls_cache: dict[int, list[AppelFormateur]] = {}

    apprenant_rows = [
        _build_apprenant_row(
            index,
            classe,
            source_class_counts=source_class_counts,
        )
        for index, classe in enumerate(classes, start=1)
    ]
    # Build single global row for general mode
    general_global_row = _build_general_global_row(classes, source_class_counts=source_class_counts)
    general_rows = [general_global_row]

    # Build per-prestation Padesce rows — filtrées aux 64 prestations analysées
    padesce_rows = _build_padesce_rows(
        classes,
        source_class_counts=source_class_counts,
        source_bundle=source_bundle,
    )

    formateur_rows = [
        _build_formateur_row(
            index,
            classe,
            request=request,
            source_bundle=source_bundle,
            formateur_directory=formateur_directory,
            prestation_calls_cache=prestation_calls_cache,
        )
        for index, classe in enumerate(classes, start=1)
    ]

    descentes_sa_rows = _build_descentes_sa_rows(classes)
    descentes_sf_rows = _build_descentes_sf_rows()

    return {
        "generated_at": timezone.localtime().isoformat(),
        "filters": filters,
        "terminated_only": scope["terminated_only"],
        "scope_label": scope["scope_label"],
        "modes": [
            _mode_payload(
                "apprenant",
                rows=apprenant_rows,
                summary_cards=_apprenant_summary_cards(apprenant_rows),
                sheet_name="Enquête de satisfaction",
                scope=scope,
            ),
            _mode_payload(
                "general",
                rows=general_rows,
                summary_cards=_general_summary_cards(general_rows),
                sheet_name="General",
                scope=scope,
            ),
            _mode_payload(
                "padesce",
                rows=padesce_rows,
                summary_cards=_padesce_summary_cards(padesce_rows),
                sheet_name="Padesce",
                scope=scope,
            ),
            _mode_payload(
                "formateur",
                rows=formateur_rows,
                summary_cards=_formateur_summary_cards(formateur_rows),
                sheet_name="Enquête de formateur",
                scope=scope,
            ),
            _descentes_mode_payload(descentes_sa_rows, descentes_sf_rows, scope),
        ],
    }


def build_fast_stats_api_payload(request) -> dict:
    return build_fast_stats_bundle(request)


def build_fast_stats_api_response(request) -> JsonResponse:
    return JsonResponse(build_fast_stats_api_payload(request))


def _copy_style(source_cell, target_cell) -> None:
    target_cell._style = copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.protection = copy(source_cell.protection)


def _ensure_sheet_capacity(
    worksheet, *, last_row: int, max_col: int, template_row: int = FAST_STATS_TEMPLATE_ROW_START
) -> None:
    if worksheet.max_row >= last_row:
        return
    template_height = worksheet.row_dimensions[template_row].height
    for target_row in range(worksheet.max_row + 1, last_row + 1):
        for column in range(1, max_col + 1):
            _copy_style(worksheet.cell(template_row, column), worksheet.cell(target_row, column))
        if template_height:
            worksheet.row_dimensions[target_row].height = template_height


def _clear_data_rows(worksheet, *, start_row: int, max_col: int) -> None:
    for row in range(start_row, worksheet.max_row + 1):
        for column in range(1, max_col + 1):
            cell = worksheet.cell(row, column)
            cell.value = None
            cell.hyperlink = None


def _clone_sheet_layout(source_sheet, target_workbook: Workbook):
    target_sheet = target_workbook.create_sheet(source_sheet.title)
    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    target_sheet.freeze_panes = source_sheet.freeze_panes
    for column_letter, dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[column_letter]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
    for row_index, dimension in source_sheet.row_dimensions.items():
        target_dimension = target_sheet.row_dimensions[row_index]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
    for row in source_sheet.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            target_cell = target_sheet[source_cell.coordinate]
            target_cell.value = source_cell.value
            _copy_style(source_cell, target_cell)
            if source_cell.hyperlink:
                target_cell.hyperlink = copy(source_cell.hyperlink)
    return target_sheet


def _write_apprenant_sheet(worksheet, mode_payload: dict) -> None:
    worksheet["B1"] = "Total classe de la prestation termine"
    worksheet["C1"] = "TRUE"
    worksheet["B3"] = "Prestation ID"
    worksheet["C3"] = "Classe ID"
    worksheet["D3"] = "Nombre d'apprenant inscrit"
    worksheet["E3"] = "Nbre d'appels effectués"
    worksheet["F3"] = "Nombre d'appels terminé"
    worksheet["G3"] = "% appel effectué"
    worksheet["H3"] = "% appel terminé"
    worksheet["I3"] = "%enquêtes"

    _clear_data_rows(worksheet, start_row=FAST_STATS_TEMPLATE_ROW_START, max_col=9)
    last_row = max(
        FAST_STATS_TEMPLATE_ROW_START, FAST_STATS_TEMPLATE_ROW_START + len(mode_payload["rows"]) - 1
    )
    _ensure_sheet_capacity(worksheet, last_row=last_row, max_col=9)

    for offset, row in enumerate(mode_payload["rows"], start=FAST_STATS_TEMPLATE_ROW_START):
        worksheet[f"A{offset}"] = row["index"]
        worksheet[f"B{offset}"] = row["prestation_id"]
        worksheet[f"C{offset}"] = row["classe_id"]
        worksheet[f"D{offset}"] = row["apprenant_count"]
        worksheet[f"E{offset}"] = row["calls_effectues"]
        worksheet[f"F{offset}"] = row["calls_termines"]
        worksheet[f"G{offset}"] = row["pct_appel_effectue"]
        worksheet[f"H{offset}"] = row["pct_appel_termine"]
        worksheet[f"I{offset}"] = row["pct_enquetes"]
        worksheet[f"G{offset}"].number_format = "0.00%"
        worksheet[f"H{offset}"].number_format = "0.00%"
        worksheet[f"I{offset}"].number_format = "0.00%"


def _write_general_sheet(worksheet, mode_payload: dict) -> None:
    worksheet["B1"] = "Statistiques générales de satisfaction"
    worksheet["C1"] = "TRUE"
    worksheet["B3"] = "Métrique"
    worksheet["C3"] = "Base"
    worksheet["D3"] = "Échelles"

    _clear_data_rows(worksheet, start_row=FAST_STATS_TEMPLATE_ROW_START, max_col=4)

    if mode_payload["rows"]:
        row = mode_payload["rows"][0]  # Global row

        metrics = [
            (
                "Taux de présence",
                row.get("taux_presence_base"),
                row.get("taux_presence"),
                row.get("taux_presence_scale"),
                "percent",
            ),
            (
                "Resp horaire",
                row.get("resp_horaire_base"),
                row.get("resp_horaire"),
                row.get("resp_horaire_scale"),
                "percent",
            ),
            (
                "Resp thème",
                row.get("resp_theme_base"),
                row.get("resp_theme"),
                row.get("resp_theme_scale"),
                "percent",
            ),
            (
                "Resp parité",
                row.get("resp_parité_base"),
                row.get("resp_parité"),
                row.get("resp_parité_scale"),
                "percent",
            ),
            (
                "Qualité environnement",
                row.get("qualite_environnement_base"),
                row.get("qualite_environnement"),
                row.get("qualite_environnement_scale"),
                "decimal",
            ),
            (
                "Satisfaction apprenant",
                row.get("satisfaction_apprenant_base"),
                row.get("satisfaction_apprenant"),
                row.get("satisfaction_apprenant_scale"),
                "decimal",
            ),
            (
                "Satis faction formateur",
                row.get("satisfaction_formateur_base"),
                row.get("satisfaction_formateur"),
                row.get("satisfaction_formateur_scale"),
                "decimal",
            ),
            (
                "Attitude enquête / respect consigne",
                row.get("attitude_enquete_base"),
                row.get("attitude_enquete"),
                row.get("attitude_enquete_scale"),
                "decimal",
            ),
        ]

        last_row = FAST_STATS_TEMPLATE_ROW_START + len(metrics) - 1
        _ensure_sheet_capacity(worksheet, last_row=last_row, max_col=4)

        for offset, (label, base_val, scaled_val, scale, fmt_type) in enumerate(
            metrics, start=FAST_STATS_TEMPLATE_ROW_START
        ):
            worksheet[f"B{offset}"] = label

            # Write base value
            if base_val is not None:
                if fmt_type == "percent":
                    worksheet[f"C{offset}"] = base_val / 100 if base_val <= 100 else 1.0
                    worksheet[f"C{offset}"].number_format = "0%"
                else:
                    worksheet[f"C{offset}"] = f"{base_val:.1f}"
                    worksheet[f"C{offset}"].number_format = "@"
            else:
                worksheet[f"C{offset}"] = "—"
                worksheet[f"C{offset}"].number_format = "@"

            # Write scaled value
            if scaled_val is not None and scale is not None:
                worksheet[f"D{offset}"] = f"{scaled_val:.1f}/{scale}"
            else:
                worksheet[f"D{offset}"] = "—"
            worksheet[f"D{offset}"].number_format = "@"


def _write_formateur_sheet(worksheet, mode_payload: dict) -> None:
    worksheet["B1"] = "Total classe de la prestation termine"
    worksheet["C1"] = "TRUE"
    worksheet["G2"] = "Selon Calendrier"
    worksheet["K2"] = "Selon Descente"
    worksheet["B3"] = "Prestation ID"
    worksheet["C3"] = "Classe ID"
    worksheet["D3"] = "Lien du canal"
    worksheet["E3"] = "Nom du Beneficiaire"
    worksheet["F3"] = "Nom du prestataire"
    worksheet["G3"] = "Nom Formateur N1"
    worksheet["H3"] = "Numero formateur N1"
    worksheet["I3"] = "Nom Formateur N2"
    worksheet["J3"] = "Numero formateur N2"
    worksheet["K3"] = "Nom Formateur N1"
    worksheet["L3"] = "Numero formateur N1"
    worksheet["M3"] = "Nom Formateur N2"
    worksheet["N3"] = "Numero formateur N2"
    worksheet["O3"] = "Nom Formateur N3"
    worksheet["P3"] = "Numero formateur N3"
    worksheet["Q3"] = "Nom Formateur N4"
    worksheet["R3"] = "Numero formateur N4"

    _clear_data_rows(worksheet, start_row=FAST_STATS_TEMPLATE_ROW_START, max_col=18)
    last_row = max(
        FAST_STATS_TEMPLATE_ROW_START, FAST_STATS_TEMPLATE_ROW_START + len(mode_payload["rows"]) - 1
    )
    _ensure_sheet_capacity(worksheet, last_row=last_row, max_col=18)

    for offset, row in enumerate(mode_payload["rows"], start=FAST_STATS_TEMPLATE_ROW_START):
        worksheet[f"A{offset}"] = row["index"]
        worksheet[f"B{offset}"] = row["prestation_id"]
        worksheet[f"C{offset}"] = row["classe_id"]
        worksheet[f"D{offset}"] = row["class_link_label"]
        worksheet[f"D{offset}"].hyperlink = row["class_link_url"]
        worksheet[f"E{offset}"] = row["beneficiaire_name"]
        worksheet[f"F{offset}"] = row["prestataire_name"]

        calendar_contacts = row["calendar_contacts"]
        descente_contacts = row["descente_contacts"]
        worksheet[f"G{offset}"] = calendar_contacts[0]["name"]
        worksheet[f"H{offset}"] = calendar_contacts[0]["phone"]
        worksheet[f"I{offset}"] = calendar_contacts[1]["name"]
        worksheet[f"J{offset}"] = calendar_contacts[1]["phone"]
        worksheet[f"K{offset}"] = descente_contacts[0]["name"]
        worksheet[f"L{offset}"] = descente_contacts[0]["phone"]
        worksheet[f"M{offset}"] = descente_contacts[1]["name"]
        worksheet[f"N{offset}"] = descente_contacts[1]["phone"]
        worksheet[f"O{offset}"] = descente_contacts[2]["name"]
        worksheet[f"P{offset}"] = descente_contacts[2]["phone"]
        worksheet[f"Q{offset}"] = descente_contacts[3]["name"]
        worksheet[f"R{offset}"] = descente_contacts[3]["phone"]


def _write_padesce_sheet(worksheet, mode_payload: dict) -> None:
    worksheet["B1"] = "Score Padesce par prestation"
    worksheet["C1"] = "TRUE"
    worksheet["B3"] = "Prestation ID"
    worksheet["C3"] = "Classes"
    worksheet["D3"] = "Taux présence (%)"
    worksheet["E3"] = "Resp horaire (%)"
    worksheet["F3"] = "Resp thème (%)"
    worksheet["G3"] = "Resp parité (%)"
    worksheet["H3"] = "Qualité env. (/13)"
    worksheet["I3"] = "Sat. apprenant (/5)"
    worksheet["J3"] = "Sat. formateur (/5)"
    worksheet["K3"] = "Attitude enquête (/10)"
    worksheet["L3"] = "Note globale (/100)"

    _clear_data_rows(worksheet, start_row=FAST_STATS_TEMPLATE_ROW_START, max_col=12)
    last_row = max(
        FAST_STATS_TEMPLATE_ROW_START, FAST_STATS_TEMPLATE_ROW_START + len(mode_payload["rows"]) - 1
    )
    _ensure_sheet_capacity(worksheet, last_row=last_row, max_col=12)

    for offset, row in enumerate(mode_payload["rows"], start=FAST_STATS_TEMPLATE_ROW_START):
        worksheet[f"A{offset}"] = row["index"]
        worksheet[f"B{offset}"] = row["prestation_id"]
        worksheet[f"C{offset}"] = row["classe_count"]
        worksheet[f"D{offset}"] = row["taux_presence_base"]
        worksheet[f"E{offset}"] = row["resp_horaire_base"]
        worksheet[f"F{offset}"] = row["resp_theme_base"]
        worksheet[f"G{offset}"] = row["resp_parité_base"]
        worksheet[f"H{offset}"] = row["qualite_environnement_base"]
        worksheet[f"I{offset}"] = row["satisfaction_apprenant_base"]
        worksheet[f"J{offset}"] = row["satisfaction_formateur_base"]
        worksheet[f"K{offset}"] = row["attitude_enquete_base"]
        worksheet[f"L{offset}"] = row["note_globale"]

        # Format with custom number format (values are already in percentage form like 92.5)
        worksheet[f"D{offset}"].number_format = '0.00"%"'
        worksheet[f"E{offset}"].number_format = '0.00"%"'
        worksheet[f"F{offset}"].number_format = '0.00"%"'
        worksheet[f"G{offset}"].number_format = '0.00"%"'


def build_fast_stats_workbook(request, *, active_mode: str = "apprenant") -> Workbook:
    if FAST_STATS_TEMPLATE_PATH.exists():
        template_workbook = load_workbook(FAST_STATS_TEMPLATE_PATH)
        source_apprenant_sheet = template_workbook[_find_sheet_name(template_workbook, "apprenant")]
        source_formateur_sheet = template_workbook[_find_sheet_name(template_workbook, "formateur")]
        workbook = Workbook()
        workbook.remove(workbook.active)
        apprenant_sheet = _clone_sheet_layout(source_apprenant_sheet, workbook)
        formateur_sheet = _clone_sheet_layout(source_formateur_sheet, workbook)
    else:
        workbook = Workbook()
        workbook.active.title = "Enquête de satisfaction"
        apprenant_sheet = workbook.active
        formateur_sheet = workbook.create_sheet("Enquête de formateur")

    # Create general and padesce sheets
    general_sheet = workbook.create_sheet("General", 1)
    padesce_sheet = workbook.create_sheet("Padesce", 2)

    modes = {mode["id"]: mode for mode in build_fast_stats_bundle(request)["modes"]}
    _write_apprenant_sheet(apprenant_sheet, modes["apprenant"])
    _write_general_sheet(general_sheet, modes["general"])
    _write_formateur_sheet(formateur_sheet, modes["formateur"])
    if "padesce" in modes:
        _write_padesce_sheet(padesce_sheet, modes["padesce"])
    if "descentes" in modes:
        _write_descentes_sheets(workbook, modes["descentes"])

    # Determine active sheet based on mode
    if active_mode == "formateur":
        active_sheet_name = formateur_sheet.title
    elif active_mode == "general":
        active_sheet_name = general_sheet.title
    elif active_mode == "padesce":
        active_sheet_name = padesce_sheet.title
    elif active_mode == "descentes":
        active_sheet_name = "SA"
    else:
        active_sheet_name = apprenant_sheet.title

    workbook.active = workbook.sheetnames.index(active_sheet_name)
    return workbook


def build_fast_stats_export_response(request) -> HttpResponse:
    active_mode = _safe_text(
        getattr(request, "GET", QueryDict("", mutable=True)).get("fast_stats_mode", "apprenant")
    )
    workbook = build_fast_stats_workbook(request, active_mode=active_mode)
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="fast_stats_satisfaction_{timestamp}.xlsx"'
    )
    workbook.save(response)
    return response


def build_fast_stats_context(request, *, default_mode: str) -> dict:
    return {
        "fast_stats_default_mode": (
            default_mode
            if default_mode in {"apprenant", "general", "formateur", "padesce", "descentes"}
            else "apprenant"
        ),
    }


def request_like_with_query(query_string: str = ""):
    querydict = QueryDict(query_string, mutable=True)

    def _build_absolute_uri(path: str = "/") -> str:
        return f"https://testserver{path}"

    return SimpleNamespace(
        GET=querydict,
        build_absolute_uri=_build_absolute_uri,
    )
